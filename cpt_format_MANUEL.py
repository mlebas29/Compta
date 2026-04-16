#!/usr/bin/env python3
"""
cpt_format_MANUEL.py - Formatter pour opérations manuelles

Format d'entrée : CSV 9 colonnes (format standard)
Format de sortie : Identique (pass-through)

Usage:
  ./cpt_format_MANUEL.py fichier.csv

Les fichiers CSV doivent être au format 9 colonnes :
  Date;Libellé;Montant;Devise;Equiv;Réf;Catégorie;Compte;Commentaire

Ce formatter est un pass-through : il valide et réaffiche le CSV tel quel.
Utilisé pour les opérations non automatisables : Bijoux Or, Créances, Deblock, etc.
"""

import json
import sys
from pathlib import Path
from datetime import datetime, timedelta
import openpyxl
import inc_categorize
from inc_format import process_files, lines_to_tuples, log_csv_debug as _log_csv_debug, site_name_from_file

SITE = site_name_from_file(__file__)

# Lookup case-insensitive des devises connues (config_cotations.json + EUR)
def _build_devise_lookup():
    cfg_path = Path(__file__).parent / 'config_cotations.json'
    with open(cfg_path) as f:
        codes = list(json.load(f).keys())
    codes.append('EUR')
    return {c.lower(): c for c in codes}

_DEVISE_LOOKUP = _build_devise_lookup()


def normalize_devise(devise_str):
    """Normalise la casse d'une devise contre les devises connues.
    Retourne la forme canonique si trouvée, sinon la valeur originale."""
    canonical = _DEVISE_LOOKUP.get(devise_str.lower())
    if canonical:
        return canonical
    print(f"⚠ [MANUEL] devise inconnue '{devise_str}'", file=sys.stderr)
    return devise_str


# Colonnes attendues
EXPECTED_COLS = ['Date', 'Libellé', 'Montant', 'Devise', 'Equiv', 'Réf', 'Catégorie', 'Compte', 'Commentaire']


def format_manuel(input_file):
    """Formate (pass-through) un fichier CSV manuel."""
    input_path = Path(input_file)

    if not input_path.exists():
        print(f"❌ Fichier introuvable: {input_file}", file=sys.stderr)
        return False

    print(f"📄 Format manuel (pass-through): {input_path.name}", file=sys.stderr)

    lines = []
    with open(input_path, 'r', encoding='utf-8') as f:
        for i, line in enumerate(f):
            line = line.strip()
            if not line:
                continue

            # Première ligne = header
            if i == 0:
                # Vérifier le header (optionnel - peut être absent)
                parts = line.split(';')
                if parts[0] == 'Date':
                    # C'est un header, on le garde
                    lines.append(line)
                    continue
                # Pas de header, c'est une ligne de données

            # Lignes de données
            parts = line.split(';')
            if len(parts) < 8:
                print(f"  ⚠ Ligne {i+1} ignorée (moins de 8 colonnes): {line[:50]}", file=sys.stderr)
                continue

            # Compléter à 9 colonnes si nécessaire
            while len(parts) < 9:
                parts.append('')

            lines.append(';'.join(parts[:9]))

    if not lines:
        print("  ⚠ Aucune ligne valide", file=sys.stderr)
        return False

    # Ajouter header si absent
    if not lines[0].startswith('Date;'):
        lines.insert(0, ';'.join(EXPECTED_COLS))

    # Output
    for line in lines:
        print(line)

    print(f"✓ {len(lines)-1} opération(s)", file=sys.stderr)
    return True


# ============================================================================
# API POUR UPDATE - NOUVELLE INTERFACE
# ============================================================================

def _parse_csv(csv_file):
    """Parse un fichier CSV manuel, retourne liste de tuples 9 champs."""
    operations = []
    with open(csv_file, 'r', encoding='utf-8') as f:
        for i, line in enumerate(f):
            line = line.strip()
            if not line:
                continue

            # Première ligne = header, on l'ignore
            if i == 0:
                parts = line.split(';')
                if parts[0] == 'Date':
                    continue

            # Lignes de données
            parts = line.split(';')
            if len(parts) < 8:
                continue

            # Compléter à 9 colonnes si nécessaire
            while len(parts) < 9:
                parts.append('')

            # Normalisation casse devise
            if parts[3]:
                parts[3] = normalize_devise(parts[3])
            operations.append(tuple(parts[:9]))

    return operations


def _validate_cell(row_num, raw_row, name):
    """Valide les cellules d'une ligne XLSX. Retourne True si OK."""
    errors = []
    date_val = raw_row[0]
    montant_val = raw_row[2] if len(raw_row) > 2 else None
    devise_val = raw_row[3] if len(raw_row) > 3 else None
    compte_val = raw_row[7] if len(raw_row) > 7 else None

    # Date : datetime Excel ou chaîne DD/MM/YYYY
    if hasattr(date_val, 'strftime'):
        pass  # datetime OK
    elif isinstance(date_val, str) and date_val:
        try:
            datetime.strptime(date_val, '%d/%m/%Y')
        except ValueError:
            errors.append(f"date invalide '{date_val}'")
    else:
        errors.append("date manquante")

    # Montant : numérique
    if montant_val is None or montant_val == '':
        errors.append("montant manquant")
    elif not isinstance(montant_val, (int, float)):
        try:
            float(str(montant_val).replace(' ', '').replace(',', '.'))
        except ValueError:
            errors.append(f"montant invalide '{montant_val}'")

    # Devise : non vide
    if not devise_val or (isinstance(devise_val, str) and not devise_val.strip()):
        errors.append("devise manquante")

    # Compte : non vide
    if not compte_val or (isinstance(compte_val, str) and not compte_val.strip()):
        errors.append("compte manquant")

    if errors:
        print(f"⚠ [MANUEL] {name} ligne {row_num} ignorée : {', '.join(errors)}", file=sys.stderr)
        return False
    return True


def _parse_xlsx(xlsx_file):
    """Parse un fichier XLSX manuel, retourne liste de tuples 9 champs."""
    operations = []
    name = Path(xlsx_file).name
    wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
    ws = wb['Import'] if 'Import' in wb.sheetnames else wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 and row[0] == 'Date':
            continue
        vals = [str(v) if v is not None else '' for v in row]
        if len(vals) < 8:
            continue
        while len(vals) < 9:
            vals.append('')
        if not _validate_cell(i + 1, row, name):
            continue
        # Dates Excel → format dd/mm/yyyy
        if hasattr(row[0], 'strftime'):
            vals[0] = row[0].strftime('%d/%m/%Y')
        # Montants numériques → virgule décimale
        for col in (2, 4):  # Montant, Equiv
            if vals[col] and vals[col] != '':
                vals[col] = vals[col].replace('.', ',')
        # Normalisation casse devise
        if vals[3]:
            vals[3] = normalize_devise(vals[3])
        operations.append(tuple(vals[:9]))
    wb.close()
    return operations


def _parse_xlsx_positions(xlsx_file):
    """Parse la feuille Positions d'un XLSX manuel → tuples 4 champs.

    Format feuille Positions : Date, Ligne (titre), Montant, Compte
    Retourne des tuples (date, ligne, montant, compte) pour alimenter PVL.
    """
    positions = []
    name = Path(xlsx_file).name
    wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
    if 'Positions' not in wb.sheetnames:
        wb.close()
        return positions
    ws = wb['Positions']
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 and row[0] == 'Date':
            continue
        if len(row) < 4 or not row[0]:
            continue
        date_val, ligne, montant, compte = row[0], row[1], row[2], row[3]
        # Date → dd/mm/yyyy
        date_str = date_val.strftime('%d/%m/%Y') if hasattr(date_val, 'strftime') else str(date_val)
        # Montant → virgule décimale
        montant_str = str(montant).replace('.', ',') if montant is not None else '0'
        positions.append((date_str, str(ligne), montant_str, str(compte)))
    wb.close()
    if positions:
        print(f"  {len(positions)} position(s) PVL depuis {name}", file=sys.stderr)
    return positions


def format_site(site_dir, verbose=False, logger=None):
    """API pour Update."""
    if logger is None:
        from inc_logging import Logger
        logger = Logger(SITE, verbose=verbose)

    handlers = [
        ('*.xlsx', _parse_xlsx, 'ops'),
        ('*.xlsx', _parse_xlsx_positions, 'pos'),
    ]
    return process_files(site_dir, handlers, verbose, SITE, logger=logger)


def log_csv_debug(operations, positions, site_dir, logger=None):
    """Wrapper vers inc_format.log_csv_debug()"""
    _log_csv_debug(SITE, operations, positions, logger)


if __name__ == '__main__':
    from inc_format import cli_main
    cli_main(format_site)
