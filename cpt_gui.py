#!/usr/bin/env python3
"""
GUI Tkinter pour la comptabilité : exécution, configuration, comptes et catégories.

Usage:
    python3 cpt_gui.py
    python3 cpt_gui.py --config /chemin/vers/config.ini
    python3 cpt_gui.py --xlsx /chemin/vers/comptes.xlsm

Cinq onglets :
  - Exécution : lancement collecte/import, outils système
  - Sites : liste des sites activés/désactivés + détails par site
  - Comptes : visualisation/édition de la feuille Avoirs de comptes.xlsm
  - Catégories : édition des patterns config_category_mappings.json
  - Paramètres : sections general, pairing, comparison de config.ini
"""

import json
import os
import queue
import re
import shutil
from datetime import datetime, timedelta
import signal
import subprocess
import sys
import threading
import tkinter as tk
try:
    from inotify_simple import INotify, flags as iflags
    HAS_INOTIFY = True
except ImportError:
    HAS_INOTIFY = False
from copy import copy
from tkinter import ttk, messagebox
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

import inc_mode
import inc_update
import json
from inc_uno import check_env

with open(Path(__file__).parent / 'config_gui_help.json', encoding='utf-8') as _f:
    FRAME_HELP = json.load(_f)
from inc_excel_schema import (
    SHEET_AVOIRS, SHEET_CONTROLES, SHEET_BUDGET, SHEET_OPERATIONS, SHEET_COTATIONS,
    SHEET_PLUS_VALUE,
    DEVISE_SOURCES,
)
import inc_compta_schema as _schema

# ============================================================================
# CHEMINS PAR DÉFAUT
# ============================================================================

SCRIPT_DIR = Path(__file__).parent
DEFAULT_CONFIG = SCRIPT_DIR / 'config.ini'
DEFAULT_JSON = SCRIPT_DIR / 'config_category_mappings.json'
DEFAULT_ACCOUNTS = SCRIPT_DIR / 'config_accounts.json'


# ============================================================================
# LECTURE / ÉCRITURE CONFIG.INI (préserve commentaires et structure)
# ============================================================================

def read_config_raw(path):
    """Lit config.ini et retourne le texte brut + un dict structuré des valeurs."""
    import configparser
    with open(path, 'r', encoding='utf-8') as f:
        raw = f.read()
    cfg = configparser.ConfigParser()
    cfg.optionxform = str  # préserve la casse
    cfg.read(path, encoding='utf-8')
    return raw, cfg


def write_config_value(raw_text, key, new_value):
    """Remplace la valeur d'une clé dans le texte brut config.ini."""
    pattern = rf'^({re.escape(key)}\s*=\s*).*$'
    return re.sub(pattern, rf'\g<1>{new_value}', raw_text, flags=re.MULTILINE)


def _insert_key_in_section(raw_text, section, key, value):
    """Insère une nouvelle clé=valeur à la fin d'une section."""
    lines = raw_text.split('\n')
    section_header = f'[{section}]'
    section_line = None
    for i, line in enumerate(lines):
        if line.strip() == section_header:
            section_line = i
            break
    if section_line is None:
        return raw_text
    # Trouver la dernière ligne de contenu de cette section
    last_content_line = section_line
    for i in range(section_line + 1, len(lines)):
        stripped = lines[i].strip()
        if re.match(r'^\[.+\]', stripped):
            break
        if stripped and not stripped.startswith('#'):
            last_content_line = i
    lines.insert(last_content_line + 1, f'{key} = {value}')
    return '\n'.join(lines)



def write_config_section_key(raw_text, section, key, new_value):
    """Remplace la valeur d'une clé dans une section spécifique (évite les doublons inter-sections).
    Retourne le texte modifié, ou None si la clé n'existe pas dans la section."""
    lines = raw_text.split('\n')
    in_section = False
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped == f'[{section}]':
            in_section = True
            continue
        if in_section and re.match(r'^\[.+\]', stripped):
            break
        if in_section and re.match(rf'^{re.escape(key)}\s*=', stripped):
            lines[i] = f'{key} = {new_value}'
            return '\n'.join(lines)
    return None


# ============================================================================
# CONFIGS JSON — fonctions ré-exportées depuis inc_config_io (module neutre,
# sans dépendance tkinter pour permettre l'usage depuis LO Python 3.8 etc.)
# ============================================================================

from inc_config_io import (
    read_mappings_json, write_mappings_json,
    read_accounts_json, write_accounts_json,
    accounts_to_site_map, site_map_to_accounts,
    read_cotations_json, write_cotations_json,
)


# ============================================================================
# APPLICATION PRINCIPALE
# ============================================================================

from gui_accounts import AccountsMixin
from gui_budget import BudgetMixin
from gui_categories import CategoriesMixin
from gui_daemon import DaemonClientMixin
from gui_devises import DevisesMixin
from gui_exec import ExecMixin
from gui_params import ParamsMixin
from gui_sites import SitesMixin


class ConfigGUI(AccountsMixin, BudgetMixin, CategoriesMixin, DaemonClientMixin,
                DevisesMixin, ExecMixin, ParamsMixin, SitesMixin):
    def __init__(self, config_path, json_path, xlsx_path=None):
        self.config_path = Path(config_path)
        self.json_path = Path(json_path)
        self.xlsx_path = Path(xlsx_path) if xlsx_path else None

        # Charger les données
        self.config_raw, self.config = read_config_raw(self.config_path)
        self.mappings = read_mappings_json(self.json_path)

        # Détecter le mode depuis config.ini
        self.mode = inc_mode.get_mode(config_path=self.config_path)

        # Couleurs selon le mode
        _MODE_THEMES = {
            'PROD': ('PROD', '#b91c1c', '#ffffff', '#fef2f2', 'cpt_gui_prod'),
            'EX':   ('EX',   '#ca8a04', '#ffffff', '#fefce8', 'cpt_gui_export'),
            'DEV':  ('DEV',  '#1d4ed8', '#ffffff', '#eff6ff', 'cpt_gui'),
        }
        label, bg, fg, accent, wm_class = _MODE_THEMES.get(self.mode, _MODE_THEMES['DEV'])
        self._mode_label = label
        self._mode_bg = bg
        self._mode_fg = fg
        self._mode_accent = accent
        self.root = tk.Tk(className=wm_class)
        from inc_excel_schema import APP_VERSION
        self.root.title(f'Comptabilité v{APP_VERSION} [{self._mode_label}]')
        self.root.geometry('1200x880')
        self.root.minsize(1000, 600)
        self.root.report_callback_exception = self._handle_tk_exception

        # Icône fenêtre et barre des tâches (bleu=test, rouge=prod, jaune=export)
        _ICON_NAMES = {'PROD': 'cpt_gui_prod.png', 'EX': 'cpt_gui_export.png'}
        icon_name = _ICON_NAMES.get(self.mode, 'cpt_gui.png')
        icon_path = Path(__file__).parent / icon_name
        if icon_path.exists():
            try:
                self._icon_img = tk.PhotoImage(file=str(icon_path))
                self.root.iconphoto(True, self._icon_img)
            except tk.TclError as e:
                # Tk < 8.6 (Python système macOS) ne décode pas le PNG → icône ignorée (cosmétique)
                print(f"⚠ icône fenêtre ignorée ({icon_path.name}) : {e}", file=sys.stderr)

        # Police plus grande pour meilleur contraste
        default_font = ('', 11)
        self.root.option_add('*Font', default_font)
        self.root.option_add('*TCombobox*Listbox.font', default_font)

        # Style ttk — contraste amélioré
        style = ttk.Style()
        style.configure('TLabel', font=default_font)
        style.configure('TButton', font=default_font, padding=4)
        style.configure('TCheckbutton', font=default_font)
        style.configure('TNotebook.Tab', font=('', 11, 'bold'), padding=(12, 6))
        style.configure('TLabelframe.Label', font=('', 11, 'bold'))
        style.configure('Treeview', font=default_font, rowheight=26)
        style.configure('Treeview.Heading', font=('', 11, 'bold'))
        style.configure('Hint.TLabel', font=('', 9), foreground='#666')
        style.configure('StatusOK.TLabel', font=('', 9, 'bold'), foreground='#16a34a')
        style.configure('StatusWarn.TLabel', font=('', 9, 'bold'), foreground='#b45309')
        style.configure('StatusError.TLabel', font=('', 9, 'bold'), foreground='#dc2626')
        style.configure('Mode.TLabel', font=('', 12, 'bold'),
                        background=self._mode_bg, foreground=self._mode_fg)
        style.configure('SiteOn.TCheckbutton', font=('', 11, 'bold'),
                        foreground='#16a34a')
        style.configure('SiteOff.TCheckbutton', font=('', 11, 'bold'),
                        foreground='#dc2626')

        # Bandeau mode PROD/TEST
        mode_banner = tk.Frame(self.root, bg=self._mode_bg, height=32)
        mode_banner.pack(fill='x')
        mode_banner.pack_propagate(False)
        display_path = str(self.config_path.parent).replace(str(Path.home()), '~', 1)
        tk.Label(mode_banner, text=f'  Mode {self._mode_label}  \u2014  {display_path}',
                 bg=self._mode_bg, fg=self._mode_fg,
                 font=('', 11, 'bold'), anchor='w').pack(side='left', padx=8)

        # Variables Tk
        self.tk_vars = {}
        self.site_vars = {}

        # Descriptions par site : source unique = DESCRIPTION dans cpt_fetch_<site>.py
        from inc_format import get_all_site_descriptions
        self.site_descriptions_default = get_all_site_descriptions()
        self.site_descriptions = self.site_descriptions_default

        # Mapping compte → site (persisté dans config_accounts.json)
        self.accounts_json_path = self.config_path.parent / 'config_accounts.json'
        self.accounts_json_data = self._load_accounts_json()
        self.account_site_map = accounts_to_site_map(self.accounts_json_data)

        # Métadonnées cotations (persistées dans config_cotations.json)
        self.cotations_json_path = self.config_path.parent / 'config_cotations.json'
        self.cotations_meta = read_cotations_json(self.cotations_json_path)

        # Pipeline config (linked_operations)
        self.pipeline_json_path = self.config_path.parent / 'config_pipeline.json'

        # Charger les données comptes.xlsm si disponible
        self.accounts_data = []
        self.budget_categories = []  # catégories depuis Budget col L
        self.budget_cat_rows = {}    # nom → ligne Excel (1-indexed)
        self.budget_total_row = None # ligne "Total" (1-indexed)
        self.budget_insert_row = None # ligne d'insertion (devant ⚓ bot)
        self.budget_posts = []       # postes budgétaires depuis Budget col A
        self.budget_post_rows = {}   # nom → ligne Excel (1-indexed)
        self.budget_post_types = {}  # nom → "Fixe"|"Variable"
        self.budget_posts_total_row = None  # ligne "Total = épargne"
        self._deleted_accounts = []  # noms des comptes supprimés (purge complète Opérations+Avoirs)
        self._soft_deleted_accounts = []  # noms des comptes à purge partielle (ops non appariées)
        self._deleted_ctrl_rows = []  # ctrl_row des comptes supprimés
        self._last_sweep_count = 0  # lignes balayées dans "Compte clos" (paires orphelines)
        self._excel_loaded = False

        # Tous les sites (config sections hors general/pairing/comparison/paths/sites)
        config_sections = ['general', 'pairing', 'comparison', 'paths', 'sites']
        self.all_sites = [s for s in self.config.sections()
                          if s not in config_sections]

        # État exécution subprocess
        self._exec_process = None
        self._exec_queue = queue.Queue()
        self._active_tooltip = None
        self._pending_help_buttons = []

        # État client daemon (gui_daemon.DaemonClientMixin) — spawn lazy.
        # Sur Linux, HAS_UNO=True : ce process reste à None, jamais utilisé.
        # Sur Mac, HAS_UNO=False : le 1er CRUD UNO le démarre.
        import threading as _threading
        self._daemon_proc = None
        self._daemon_lock = _threading.Lock()

        self.root.protocol('WM_DELETE_WINDOW', self._on_close)
        # Cmd+Q sur Mac ne déclenche PAS WM_DELETE_WINDOW (chemin Apple
        # events distinct). Sans ce binding, Cmd+Q tue le process Tk
        # brutalement → daemon enfant tué par SIGHUP avant que _close_batch
        # puisse flusher les modifs UNO → perte silencieuse du save.
        # Idem 'tk::mac::OnHide' pour la cohérence (l'app peut être hidden
        # par Cmd+H sans qu'on veuille fermer).
        if sys.platform == 'darwin':
            try:
                self.root.createcommand('tk::mac::Quit', self._on_close)
            except tk.TclError:
                pass  # vieille version Tk sans cet handler

        # Notebook (onglets)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=8, pady=(4, 8))

        # site_vars AVANT _build_tab_execution : l'onglet collecte (#107) lit
        # l'état activé depuis site_vars (live) pour construire ses cases sites.
        enabled_str = self.config.get('sites', 'enabled', fallback='')
        enabled_list = [s.strip() for s in enabled_str.split(',') if s.strip()]
        for site in self.all_sites:
            self.site_vars[site] = tk.BooleanVar(value=site in enabled_list)

        # Construction lazy des onglets : seul l'onglet visible est construit
        # immédiatement, les autres sont construits au premier affichage
        self._build_tab_execution()
        self._install_help_buttons()

        # Onglets différés : placeholder frames, construits au premier clic
        self._deferred_tabs = {}  # tab_frame -> builder name
        for tab_name, builder, condition in [
            ('Sites', '_build_tab_sites', True),
            ('Avoirs', '_build_tab_accounts', bool(self.xlsx_path)),
            ('Catégories', '_build_tab_categories', True),
            ('Paramètres', '_build_tab_params', True),
        ]:
            if condition:
                tab = ttk.Frame(self.notebook)
                self.notebook.add(tab, text=tab_name)
                self._deferred_tabs[str(tab)] = builder

        self.notebook.bind('<<NotebookTabChanged>>', self._on_tab_changed)

        # Barre de statut — packée side='bottom' before notebook pour réserver
        # l'espace AVANT que le notebook (expand=True) ne consomme tout le cavity.
        # Sans ça, sur démarrage avec contenu d'onglet haut, status bar clippée
        # jusqu'au prochain resize.
        status_frame = ttk.Frame(self.root)
        status_frame.pack(side='bottom', fill='x', padx=8, pady=(0, 8),
                          before=self.notebook)

        self._coherence_auto_fixes = []
        self._coherence_warnings = []
        self._status_details = []  # détails affichés au clic

        # Zone 1 : Statut fusionné (colorée, cliquable)
        self.status_var = tk.StringVar(value='')
        self.status_label = ttk.Label(status_frame, textvariable=self.status_var,
                                      style='Hint.TLabel', cursor='hand2')
        self.status_label.pack(side='left', padx=(5, 0))
        self.status_label.bind('<Button-1>', self._on_status_click)

        # Séparateur
        self._status_sep = ttk.Label(status_frame, text='  —  ', style='Hint.TLabel')

        # Zone 2 : Total Avoirs (neutre, non cliquable)
        self._status_total_var = tk.StringVar(value='')
        self._status_total_label = ttk.Label(status_frame,
                                             textvariable=self._status_total_var,
                                             style='Hint.TLabel')

        # Bouton ? aide statut
        help_text = FRAME_HELP.get('Statut', '')
        if help_text:
            btn = tk.Label(status_frame, text=' ? ', font=('', 9, 'bold'),
                           fg='#555', bg='#e8e8e8', cursor='hand2',
                           relief='flat', padx=2)
            btn.pack(side='right', padx=(0, 5))
            btn.bind('<Button-1>',
                     lambda e, w=btn, t=help_text: self._show_help_tooltip(w, t))

        # Lecture Contrôles A1 (synthèse) au démarrage
        if self.xlsx_path:
            self._refresh_status_bar()
            self._start_file_watcher()

        # Construction des onglets en arrière-plan (après affichage initial)
        self.root.after(100, self._build_deferred_tabs)

        # Checks démarrage différés (300ms) : sur Mac, une messagebox créée
        # avant que la fenêtre principale soit mappée passe derrière → focus
        # issue. Le delay laisse le mainloop afficher root d'abord.
        self.root.after(300, self._check_startup_health)

        # Workarounds bug Tk Linux X11 :
        # (1) clic externe (autre app) → polling focus_displayof() ferme menus
        # (2) clic interne dans la GUI hors menu → bind <Button-1> global
        # Limitation connue (cf. CLAUDE_todo #41) : ne couvre PAS les combobox
        # ttk dont le popdown garde le focus interne.
        self._popup_menus = getattr(self, '_popup_menus', [])
        self.root.after(500, self._global_focus_watch)

        def _on_internal_click(event):
            # Si le clic est sur un Menu ou un Combobox lui-même, Tk gère
            # (sélection d'item ou ouverture du dropdown).
            if isinstance(event.widget, (tk.Menu, ttk.Combobox)):
                return
            # Sinon : fermer les menus posted et les dropdowns combobox.
            for m in self._popup_menus:
                try:
                    if m.winfo_ismapped():
                        m.unpost()
                except tk.TclError:
                    pass
            self._unpost_comboboxes(self.root)
        self.root.bind('<Button-1>', _on_internal_click, add='+')

    def _unpost_comboboxes(self, widget):
        """Ferme tous les dropdowns combobox ttk encore mappés."""
        for child in widget.winfo_children():
            if isinstance(child, ttk.Combobox):
                try:
                    popdown = child.tk.call(
                        'ttk::combobox::PopdownWindow', str(child))
                    if popdown and child.tk.call(
                            'winfo', 'ismapped', popdown):
                        child.tk.call('ttk::combobox::Unpost', str(child))
                except tk.TclError:
                    pass
            self._unpost_comboboxes(child)

    def _global_focus_watch(self):
        # Eval Tcl direct au lieu de focus_displayof() : ce dernier lève
        # KeyError 'popdown' si le focus est sur le popdown listbox d'un
        # combobox (wrapper Python incapable de résoudre ce widget interne).
        try:
            focus_path = self.root.tk.eval('focus -displayof .')
            if not focus_path:
                for m in self._popup_menus:
                    try:
                        m.unpost()
                    except tk.TclError:
                        pass
        except Exception:
            pass
        try:
            self.root.after(150, self._global_focus_watch)
        except Exception:
            pass

    # ----------------------------------------------------------------
    # FILE WATCHER — rafraîchit la barre d'état sur modification externe
    # ----------------------------------------------------------------

    def _start_file_watcher(self):
        """Démarre un thread inotify surveillant comptes.xlsm."""
        self._inotify_stop = threading.Event()
        if not HAS_INOTIFY or not self.xlsx_path:
            return
        t = threading.Thread(target=self._file_watcher_loop, daemon=True)
        t.start()

    def _file_watcher_loop(self):
        """Thread inotify : surveille CLOSE_WRITE sur le xlsm et ses .bak."""
        ino = INotify()
        watch_dir = str(Path(self.xlsx_path).parent)
        ino.add_watch(watch_dir, iflags.CLOSE_WRITE | iflags.MOVED_TO)
        target_name = Path(self.xlsx_path).name
        while not self._inotify_stop.is_set():
            for event in ino.read(timeout=1000):
                if event.name == target_name:
                    self.root.after(200, self._refresh_status_bar)
        ino.close()

    def _stop_file_watcher(self):
        """Arrête le thread inotify."""
        if hasattr(self, '_inotify_stop'):
            self._inotify_stop.set()

    @staticmethod
    def _parse_nr_ref(ref):
        """`Contrôles!$K$17:$K$34` (ou cellule unique) → `(col_letter, start_row, end_row)`
        ou None. Le préfixe feuille (quoté/accentué) est ignoré."""
        m = re.search(r'\$([A-Z]+)\$(\d+)(?::\$[A-Z]+\$(\d+))?', ref or '')
        return (m.group(1), int(m.group(2)), int(m.group(3) or m.group(2))) if m else None

    def _ctrl_tokens(self, read_cell, lab_ref, ver_ref):
        """7 tokens ✓/✗/⚠ via les named ranges CTRL2type (libellés) / CTRL2affichage
        (verdicts) — robuste au décalage du bloc de synthèse selon le nb de comptes
        (≠ cellules en dur, qui se lisaient VIDES = faux ✓ silencieux). Contrôle
        introuvable → '⚠' (jamais un faux OK) ; None si les NR ne se résolvent pas.
        `read_cell(col_letter, row)` → str de la cellule."""
        lab = self._parse_nr_ref(lab_ref)
        ver = self._parse_nr_ref(ver_ref)
        if not lab or not ver:
            return None
        pairs = []
        for row in range(ver[1], ver[2] + 1):
            label = (read_cell(lab[0], row) or '').strip()
            if label:
                pairs.append((label.upper(), (read_cell(ver[0], row) or '').strip()))
        return [next((v or '✓' for lu, v in pairs if lu.startswith(ctrl)), '⚠')
                for ctrl in self._CTRL_LABELS]

    def _read_status_cells_zip(self):
        """Lecture rapide Contrôles A1 + 7 contrôles + Avoirs L2 via ZIP (~9ms vs ~70ms openpyxl).

        Returns:
            tuple: (ctrl_text, total_value, tokens) — ctrl_text str (synthèse mono-char A1),
                tokens list[str] de longueur 7 (✓/✗/⚠ par contrôle, ✓ par défaut),
                total_value float|None
        """
        import zipfile
        import xml.etree.ElementTree as ET
        ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        rns = 'http://schemas.openxmlformats.org/package/2006/relationships'
        ons = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

        def _read_cell(tree, ref):
            cell = tree.find(f'.//s:sheetData/s:row/s:c[@r="{ref}"]', ns)
            if cell is None:
                return None
            v = cell.find('s:v', ns)
            val = v.text if v is not None else None
            if cell.get('t') == 's' and val:
                val = strings[int(val)]
            return val

        with zipfile.ZipFile(self.xlsx_path) as z:
            with z.open('xl/sharedStrings.xml') as f:
                strings = [(''.join(t.text or '' for t in si.findall('.//s:t', ns)))
                           for si in ET.parse(f).findall('.//s:si', ns)]
            with z.open('xl/workbook.xml') as f:
                _wb = ET.parse(f)
            sheets = [(s.get('name'), s.get(f'{{{ons}}}id'))
                      for s in _wb.findall('.//s:sheet', ns)]
            defined = {dn.get('name'): (dn.text or '')
                       for dn in _wb.findall('.//s:definedName', ns)}
            with z.open('xl/_rels/workbook.xml.rels') as f:
                rel_map = {r.get('Id'): r.get('Target')
                           for r in ET.parse(f).findall(f'.//{{{rns}}}Relationship')}
            sheet_targets = {}
            for name, rid in sheets:
                if name in (SHEET_CONTROLES, SHEET_AVOIRS):
                    sheet_targets[name] = 'xl/' + rel_map[rid]

            # Contrôles A1 = synthèse mono-char (=$K$35), valeur cached MAJ par LO à chaque save.
            # 7 verdicts (✓/✗/⚠) lus via les named ranges CTRL2type/CTRL2affichage
            # (cf. _ctrl_tokens) pour le détail au clic. Le tree XML est parsé une seule fois.
            ctrl = ''
            tokens = ['✓'] * 7
            if SHEET_CONTROLES in sheet_targets:
                with z.open(sheet_targets[SHEET_CONTROLES]) as f:
                    tree = ET.parse(f)
                ctrl = (_read_cell(tree, 'A1') or '').strip()
                toks = self._ctrl_tokens(
                    lambda col, row: _read_cell(tree, f'{col}{row}'),
                    defined.get('CTRL2type'), defined.get('CTRL2affichage'))
                if toks is not None:
                    tokens = toks

            # Avoirs L2 = formule Total (=L81 ou similaire), cached value
            # mise à jour par LO à chaque save. Pas besoin de miroir L1.
            total = None
            if SHEET_AVOIRS in sheet_targets:
                with z.open(sheet_targets[SHEET_AVOIRS]) as f:
                    tree = ET.parse(f)
                val = _read_cell(tree, 'L2')
                if val:
                    try:
                        total = float(val)
                    except (ValueError, TypeError):
                        pass

        return ctrl, total, tokens

    def _refresh_status_bar(self):
        """Lecture Contrôles A1 + Avoirs L2 + cohérence JSON au démarrage."""
        try:
            # Lecture rapide ZIP, fallback openpyxl si erreur
            try:
                ctrl, total, tokens = self._read_status_cells_zip()
            except Exception:
                from openpyxl import load_workbook
                wb = load_workbook(self.xlsx_path, read_only=True, data_only=True)
                ctrl, total = '', None
                tokens = ['✓'] * 7
                if SHEET_CONTROLES in wb.sheetnames:
                    ws = wb[SHEET_CONTROLES]
                    ctrl = str(ws['A1'].value or '').strip()

                    def _dn(name):
                        try:
                            return wb.defined_names[name].value
                        except (KeyError, AttributeError, TypeError):
                            return None
                    toks = self._ctrl_tokens(
                        lambda col, row: str(ws[f'{col}{row}'].value or ''),
                        _dn('CTRL2type'), _dn('CTRL2affichage'))
                    if toks is not None:
                        tokens = toks
                if SHEET_AVOIRS in wb.sheetnames:
                    ws = wb[SHEET_AVOIRS]
                    val = ws['L2'].value
                    if isinstance(val, (int, float)):
                        total = val
                wb.close()

            details = []
            level = 'ok'

            if ctrl and ctrl != '.' and ctrl != '✓':
                details.append(ctrl)
                if '✗' in ctrl or any(k in ctrl for k in ('COMPTES', 'CATÉGORIES', 'INCONNUS')):
                    level = 'error'
                else:
                    level = 'warn'

            # Cohérence JSON ↔ Excel
            self._coherence_auto_fixes, self._coherence_warnings = self._check_coherence()
            if self._coherence_warnings or self._coherence_auto_fixes:
                details.extend(self._coherence_warnings)
                if level == 'ok':
                    level = 'warn'

            # Stocker pour le clic
            self._status_ctrl = ctrl if ctrl and ctrl != '.' else ''
            self._status_tokens = tokens
            self._status_details = details

            # Synthèse zone 1 : A1 brut + indicateur alertes config
            status_text = ctrl or '.'
            if self._coherence_warnings:
                status_text += ' ⚙'
                if level == 'ok':
                    level = 'warn'
            self._set_status(status_text, level)

            # Zone 2 : Total Avoirs
            total_text = ''
            if total is not None:
                total_text = f'Total: {total:,.0f} €'.replace(',', ' ')
            self._status_sep.pack_forget()
            self._status_total_label.pack_forget()
            self._status_total_var.set(total_text)
            if total_text:
                self._status_sep.pack(side='left')
                self._status_total_label.pack(side='left')

        except Exception:
            pass

    def _check_schema_version(self):
        """Vérifie la version du schéma classeur (Contrôles!K2) vs l'application.

        Returns:
            str or None: message d'erreur si incompatible, None si OK.

        Délègue à inc_update.check_schema_compat (probe partagée avec
        upgrade.py). Recalculé à chaque démarrage — le GUI ne défère jamais
        à une sortie figée d'upgrade (cf. inc_update).
        """
        return inc_update.check_schema_compat(self.xlsx_path)

    def _startup_check(self):
        """Check de cohérence au démarrage : charge Excel puis vérifie."""
        out = self._exec_output

        # Vérification version schéma (indépendante du chargement complet)
        version_error = self._check_schema_version()
        if version_error:
            self._coherence_auto_fixes = []
            self._coherence_warnings = [version_error]
            out.configure(state='normal')
            out.delete('1.0', 'end')
            out.insert('end', f'⚠ {version_error}\n')
            out.see('end')
            out.configure(state='disabled')
            self._set_status('VERSION', 'error')
            return

        try:
            self._ensure_excel_loaded()
            # Relire les JSON depuis le disque (ils ont pu être modifiés en externe)
            self.accounts_json_data = self._load_accounts_json()
            self.account_site_map = accounts_to_site_map(self.accounts_json_data)
            self.cotations_meta = read_cotations_json(self.cotations_json_path)
            auto_fixes, warnings = self._check_coherence()
        except Exception as e:
            import traceback
            self._coherence_auto_fixes = []
            self._coherence_warnings = [str(e)]
            out.configure(state='normal')
            out.delete('1.0', 'end')
            out.insert('end', f'❌ Erreur au check de cohérence :\n\n{traceback.format_exc()}')
            out.see('end')
            out.configure(state='disabled')
            self._set_status('ERR', 'error')
            return

        # Mettre à jour les attributs pour la status bar
        self._coherence_auto_fixes = auto_fixes
        self._coherence_warnings = warnings

        # Afficher dans la zone Résultat
        out.configure(state='normal')
        out.delete('1.0', 'end')

        if not auto_fixes and not warnings:
            out.insert('end', '✓ Cohérence vérifiée\n')
        else:
            if auto_fixes:
                out.insert('end', '=== Corrections automatiques ===\n\n')
                for line in auto_fixes:
                    out.insert('end', f'  [AUTO-FIX] {line}\n')
                out.insert('end', '\n')
            if warnings:
                out.insert('end', '=== Problèmes détectés ===\n\n')
                for line in warnings:
                    out.insert('end', f'  [ATTENTION] {line}\n')

        out.see('end')
        out.configure(state='disabled')

        # Rafraîchir la status bar avec les résultats
        if self.xlsx_path:
            self._refresh_status_bar()

    def _check_coherence(self):
        """Vérifie la cohérence entre Excel et les fichiers JSON de config.

        Returns:
            tuple(list[str], list[str]): (auto_fixes effectués, warnings à traiter)
        """
        # Avis config au démarrage (#98/#105) : helper partagé avec cpt.py — 1
        # seule source pour l'ordre, gating mutuellement exclusif (marqueur config
        # en retard → avis upgrade SEUL, sinon filet générique check_config_obsolete).
        # Le démarrage ALERTE, ne mute jamais (upgrade SEUL résout — pas de stamp ici).
        config_warnings = inc_update.startup_config_advice(
            self.config_path, self.config_path.parent)
        if not self._excel_loaded:
            if not self.xlsx_path:
                return [], config_warnings + ['Classeur comptes.xlsm introuvable — copier comptes_template.xlsm ou un classeur existant']
            return [], config_warnings
        auto_fixes = []
        warnings = list(config_warnings)
        avoirs_names = {a['intitule'] for a in self.accounts_data}
        avoirs_by_row = {a['row']: a for a in self.accounts_data}

        # --- JSON absent : détection et auto-correction ---
        if not self.accounts_json_path.exists() and self.accounts_data:
            # config_accounts.json absent mais le classeur a des comptes → reconstruire
            self._save_site_map()
            self.accounts_json_data = read_accounts_json(self.accounts_json_path)
            self.account_site_map = accounts_to_site_map(self.accounts_json_data)
            auto_fixes.append('config_accounts.json reconstruit depuis le classeur')

        mappings_path = self.config_path.parent / 'config_category_mappings.json'
        if not mappings_path.exists():
            # Auto-création d'un JSON vide (l'utilisateur enrichit via l'onglet Catégories)
            write_mappings_json(mappings_path, {})
            self.mappings = {}
            auto_fixes.append('config_category_mappings.json absent — créé vide')
            mappings = {}
        else:
            mappings = read_mappings_json(mappings_path)
            # Warning seulement si l'utilisateur a des comptes (sinon rien à
            # catégoriser — JSON vide est un état neutre, pas une incohérence).
            sites_with_accounts = {s for s in self.account_site_map.values()
                                   if s and s != 'N/A'}
            if not mappings and sites_with_accounts:
                warnings.append(
                    'Aucun pattern de catégorisation — toutes les opérations '
                    'importées seront affectées à « - ». Configure via l\'onglet Catégories.')

        if not self.cotations_meta:
            devises_non_eur = [d for d in getattr(self, 'ACCOUNT_DEVISES', []) if d != 'EUR']
            if devises_non_eur:
                warnings.append(
                    f'config_cotations.json est vide — devises sans cotation : '
                    f'{", ".join(devises_non_eur)}')

        # --- Contrôles → Avoirs : vérification formules ---
        try:
            wb_check = openpyxl.load_workbook(self.xlsx_path, data_only=False, read_only=True)
            ws_ctrl = wb_check[SHEET_CONTROLES]
            ctrl_data_start = self._start_ctrl1 + 1
            for row_idx in range(ctrl_data_start, self._end_ctrl1 + 1):
                cell_val = ws_ctrl.cell(row_idx, self.cr.col('CTRL1compte')).value
                if not cell_val:
                    continue
                cell_str = str(cell_val).strip()
                m = re.match(r"=Avoirs!\$?A(\d+)", cell_str)
                if m:
                    ref_row = int(m.group(1))
                    acct = avoirs_by_row.get(ref_row)
                    if not acct:
                        warnings.append(
                            f'Contrôles ligne {row_idx}: référence Avoirs.A{ref_row} invalide')
                elif cell_str and cell_str not in ('✓', '⚓') and cell_str not in avoirs_names:
                    warnings.append(
                        f'Contrôles ligne {row_idx}: compte « {cell_str} » absent d\'Avoirs')
            wb_check.close()
        except Exception:
            pass

        # --- Comptes Avoirs ↔ mapping sites : nettoyage auto ---
        site_map_names = set(self.account_site_map.keys())
        orphan_sites = site_map_names - avoirs_names
        if orphan_sites:
            for name in orphan_sites:
                del self.account_site_map[name]
            # Supprimer les comptes orphelins du JSON (sans reconstruire tout)
            for site_data in self.accounts_json_data.values():
                if not isinstance(site_data, dict):
                    continue  # clés top-level non-site (ex. transfer_pairs)
                accts = site_data.get('accounts', [])
                site_data['accounts'] = [a for a in accts
                                         if a.get('name') not in orphan_sites]
            write_accounts_json(self.accounts_json_path, self.accounts_json_data)
            auto_fixes.append(
                f'config_accounts.json : {len(orphan_sites)} compte(s) orphelin(s) supprimé(s)')

        def _site_label(s):
            """Nom utilisateur du site (config.ini section name) avec fallback clé."""
            return self.config.get(s, 'name', fallback=s)

        # --- Sites activés vs comptes (warnings informatifs, sans auto-fix) ---
        # Ne PAS désactiver automatiquement : l'utilisateur peut activer un
        # site avant de créer ses comptes (préparation de la config). On
        # signale simplement les deux cas d'incohérence.
        enabled_str = self.config.get('sites', 'enabled', fallback='')
        enabled_sites = {s.strip() for s in enabled_str.split(',') if s.strip()}
        enabled_sites.discard('MANUEL')
        sites_with_accounts = set(self.account_site_map.values()) - {'N/A'}
        for s in sorted(enabled_sites - sites_with_accounts):
            warnings.append(
                f'Site « {_site_label(s)} » activé mais aucun compte associé '
                f'— pas d\'import possible')
        for s in sorted(sites_with_accounts - enabled_sites):
            if s in set(self.all_sites):
                warnings.append(
                    f'Site « {_site_label(s)} » a des comptes mais est désactivé')

        # --- Devises Avoirs sans cotation (hors EUR) — réf = feuille Cotations (classeur) ---
        devises_avoirs = {a['devise'] for a in self.accounts_data if a['devise']}
        devises_avoirs.discard('EUR')
        devises_avoirs.discard('-')
        if devises_avoirs:
            from inc_excel_schema import read_cotations_meta
            wb_cot = openpyxl.load_workbook(self.xlsx_path, data_only=True)
            try:
                cotations_codes = set(read_cotations_meta(wb_cot).keys())
            finally:
                wb_cot.close()
            for d in sorted(devises_avoirs - cotations_codes):
                warnings.append(f'Devise « {d} » utilisée dans Avoirs mais absente des cotations')

        # --- Catégories Budget ↔ règles de catégorisation : warning si orphelins ---
        # Pas de purge silencieuse (la perte de patterns serait invisible) :
        # on prévient, l'utilisateur arbitre via l'onglet Catégories ou le Budget.
        if hasattr(self, 'budget_categories') and self.budget_categories:
            budget_cats = set(self.budget_categories)
            if mappings_path.exists():
                mappings = read_mappings_json(mappings_path)
                json_cats_count = {}
                for rules in mappings.values():
                    for rule in rules:
                        cat = rule.get('category', '')
                        if cat and cat != '-' and not cat.startswith('#'):
                            json_cats_count[cat] = json_cats_count.get(cat, 0) + 1
                orphan_cats = {c: n for c, n in json_cats_count.items()
                               if c not in budget_cats}
                for cat in sorted(orphan_cats):
                    n = orphan_cats[cat]
                    warnings.append(
                        f'Catégorie « {cat} » utilisée par {n} pattern(s) '
                        f'mais absente du Budget — ajouter au Budget ou '
                        f'modifier les patterns via l\'onglet Catégories')

        # --- Sites config.ini ↔ JSONs : cohérence ---
        ini_sites = set(self.all_sites)
        # Sites référencés dans config_accounts.json (hors N/A)
        map_sites = set(v for v in self.account_site_map.values() if v != 'N/A')
        for s in map_sites - ini_sites:
            # Site absent de config.ini : pas de nom convivial, on garde la clé
            warnings.append(f'Site « {s} » dans config_accounts.json mais absent de config.ini')
        # Sites sans description par défaut (variable DESCRIPTION dans le module)
        # MANUEL exclu : saisie manuelle sans collecte ni affichage GUI (caché,
        # cf. gui_sites/gui_exec) → pas de cpt_fetch_MANUEL.py ni de DESCRIPTION attendue. #140
        for s in ini_sites - set(self.site_descriptions_default.keys()):
            if s == 'MANUEL':
                continue
            warnings.append(f'Site « {_site_label(s)} » : pas de DESCRIPTION dans cpt_fetch_{s}.py')

        return auto_fixes, warnings

    # 7 verdicts lus via les named ranges CTRL2type/CTRL2affichage (cf. _ctrl_tokens).
    # A1 = =$K$35 (CTRL2_synthese) = synthèse mono-char globale.
    _CTRL_LABELS = [
        'COMPTES',
        'CATÉGORIES',
        'DIVERS',
        'APPARIEMENTS',
        'BALANCES',
        'INCONNUS',
        'FORMULES',
    ]
    _CTRL_EXPLANATIONS = [
        'Écarts entre soldes calculés et soldes relevés',
        'Opération(s) sans catégorie connue',
        'Date hors période / Ventilation Patrimoine / Cotations incomplètes',
        'Appariements incomplets',
        'Déséquilibre balances',
        'Compte(s) absent(s) de la feuille Avoirs',
        'Synthèse PVL ou Avoirs en erreur (#N/A, #REF!, …)',
    ]

    def _on_status_click(self, event):
        """Affiche le détail des alertes dans la zone Résultat."""
        ctrl = getattr(self, '_status_ctrl', '')
        tokens = getattr(self, '_status_tokens', ['✓'] * 7)
        coherence = getattr(self, '_coherence_warnings', [])
        auto_fixes = getattr(self, '_coherence_auto_fixes', [])
        if not ctrl and not coherence and not auto_fixes:
            return

        out = self._exec_output
        out.configure(state='normal')
        out.delete('1.0', 'end')

        # Section classeur — 6 contrôles, colonnes alignées
        if ctrl:
            out.insert('end', '=== Contrôles classeur ===\n\n')
            for i, label in enumerate(self._CTRL_LABELS):
                token = tokens[i] if i < len(tokens) else '✓'
                if token == '✗':
                    icon = '✗'
                    detail = self._CTRL_EXPLANATIONS[i]
                elif token == '⚠':
                    icon = '⚠'
                    detail = self._CTRL_EXPLANATIONS[i]
                else:
                    icon = '✓'
                    detail = 'OK'
                out.insert('end', f'  {icon}  {label:<22} {detail}\n')
            out.insert('end', '\n')

        # Section corrections automatiques
        if auto_fixes:
            out.insert('end', '=== Corrections automatiques ===\n\n')
            for line in auto_fixes:
                out.insert('end', f'  [AUTO-FIX] {line}\n')
            out.insert('end', '\n')

        # Section configuration
        if coherence:
            out.insert('end', '=== Configuration ===\n\n')
            for line in coherence:
                out.insert('end', f'  [ATTENTION] {line}\n')

        out.configure(state='disabled')
        self.notebook.select(self._tab_execution)

    _STATUS_STYLES = {
        'ok': 'StatusOK.TLabel',
        'warn': 'StatusWarn.TLabel',
        'error': 'StatusError.TLabel',
    }

    _XLSX_BTN_COLORS = {
        'warn': ('#FFEB9C', '#FFF0B3', '#000000'),     # jaune pâle, texte noir
        'error': ('#c62828', '#e53935', '#FFFFFF'),     # rouge, texte blanc
    }

    def _set_status(self, msg, level='ok'):
        """Met à jour la barre de statut avec couleur selon le niveau."""
        self.status_var.set(msg)
        self.status_label.configure(
            style=self._STATUS_STYLES.get(level, 'Hint.TLabel'))
        # Colorer le bouton comptes.xlsm seulement en cas de problème
        btn = getattr(self, '_xlsx_btn', None)
        if btn and level in self._XLSX_BTN_COLORS:
            bg, abg, fg = self._XLSX_BTN_COLORS[level]
            btn.configure(bg=bg, activebackground=abg, fg=fg, activeforeground=fg)
        elif btn:
            btn.configure(bg=self._xlsx_btn_default_bg,
                          activebackground=self._xlsx_btn_default_abg,
                          fg=self._xlsx_btn_default_fg,
                          activeforeground=self._xlsx_btn_default_afg)

    def _build_deferred_tabs(self):
        """Construit les onglets restants en arrière-plan, un par tick event-loop."""
        if not self._deferred_tabs:
            return
        try:
            self._ensure_excel_loaded()
            # Prendre le premier onglet en attente et le construire
            tab_id, builder_name = next(iter(self._deferred_tabs.items()))
            del self._deferred_tabs[tab_id]
            builder = getattr(self, builder_name)
            tab_frame = self.notebook.nametowidget(tab_id)
            builder(tab=tab_frame)
            self._install_help_buttons()
        except Exception:
            import traceback
            traceback.print_exc()
            self._deferred_tabs.clear()
        # Planifier le suivant au prochain tick (laisse le GUI respirer)
        if self._deferred_tabs:
            self.root.after_idle(self._build_deferred_tabs)
        else:
            # Tous les onglets construits → lancer le check de cohérence
            # (appelé aussi sans xlsx pour signaler l'absence du classeur).
            self.root.after_idle(self._startup_check)

    def _on_tab_changed(self, event):
        """Construit un onglet différé si l'utilisateur clique avant le background."""
        tab_id = self.notebook.select()
        builder_name = self._deferred_tabs.pop(tab_id, None)
        if builder_name:
            self._ensure_excel_loaded()
            builder = getattr(self, builder_name)
            tab_frame = self.notebook.nametowidget(tab_id)
            builder(tab=tab_frame)
            self._install_help_buttons()
        # Rafraîchir la combobox Catégories à chaque sélection de cet onglet :
        # account_site_map peut avoir évolué via l'onglet Comptes.
        try:
            tab_text = self.notebook.tab(tab_id, 'text')
        except tk.TclError:
            tab_text = ''
        # #107 « édite-et-pars » : persister l'onglet config QUITTÉ (plus de bouton
        # Enregistrer). Avant le refresh collecte ci-dessous → enabled est à jour.
        prev = getattr(self, '_prev_tab_text', None)
        if prev != tab_text:
            self._autosave_config_tab(prev)
        self._prev_tab_text = tab_text
        if tab_text == 'Catégories' and hasattr(self, '_refresh_cat_groups'):
            self._refresh_cat_groups()
        # #107 : l'onglet collecte suit l'état « Actif » des sites (toggles faits
        # dans l'onglet Sites depuis le dernier passage). Idempotent si rien changé.
        if tab_text == 'Exécution' and hasattr(self, '_rebuild_exec_site_list'):
            self._rebuild_exec_site_list()

    def _autosave_config_tab(self, tab_text):
        """#107 « édite-et-pars » : persiste l'onglet config quitté, sans bouton.
        LEAN — réécriture config.ini (+ petit JSON pipeline pour Paramètres), AUCUNE
        relecture du classeur ; les save sont no-op si rien n'a changé (garde
        interne) → une transition sans édition ne touche pas le disque. Statut
        affiché seulement si une écriture a réellement eu lieu."""
        wrote = False
        if tab_text == 'Sites':
            wrote = self._save_config()
        elif tab_text == 'Paramètres':
            wrote = self._save_config()
            wrote = self._save_pipeline_config() or wrote
            wrote = self._save_transfer_pairs() or wrote
        if wrote:
            self._set_status('Configuration enregistrée')

    def _ensure_excel_loaded(self):
        """Charge les données Excel une seule fois, au premier besoin."""
        if not self._excel_loaded and self.xlsx_path:
            self._load_all_excel_data()
            self._excel_loaded = True

    # ----------------------------------------------------------------
    # CHARGEMENT COMPTES.XLSX (feuille Avoirs)
    # ----------------------------------------------------------------

    # Schéma métier — constantes et helpers extraits dans inc_compta_schema
    # (module neutre, sans dépendance tkinter, partagé avec HeadlessGUI/TNR)
    cours_name = staticmethod(_schema.cours_name)
    ACCOUNT_TYPES = _schema.ACCOUNT_TYPES
    BIEN_TYPES = _schema.BIEN_TYPES
    SOUS_TYPES_BASE = _schema.SOUS_TYPES_BASE
    PV_SECTION_LABELS = _schema.PV_SECTION_LABELS
    PV_SECTION_TOTALS = _schema.PV_SECTION_TOTALS

    # Fond gris pour les comptes non-EUR (objet openpyxl, reste dans la GUI)
    NON_EUR_FILL = PatternFill(start_color='FFDCDCDC', end_color='FFDCDCDC',
                               fill_type='solid')

    ACCOUNT_DEVISES = ['EUR']  # état dynamique, rechargé depuis Avoirs

    def _load_accounts_json(self):
        """Charge config_accounts.json."""
        if self.accounts_json_path.exists():
            return read_accounts_json(self.accounts_json_path)
        return {}

    def _save_site_map(self):
        """Reconstruit le mapping compte→site depuis accounts_data et l'écrit."""
        mapping = {}
        for acct in self.accounts_data:
            site = acct.get('site', '')
            if site:
                mapping[acct['intitule']] = site
        self.accounts_json_data = site_map_to_accounts(mapping, self.accounts_json_data)
        write_accounts_json(self.accounts_json_path, self.accounts_json_data)
        self.account_site_map = mapping

    def _load_all_excel_data(self):
        """Charge toutes les données Excel au démarrage (2 ouvertures au lieu de 5)."""
        wb_formula = openpyxl.load_workbook(self.xlsx_path, data_only=False)
        wb_values = openpyxl.load_workbook(self.xlsx_path, data_only=True)
        try:
            self._load_accounts_data(wb_formula, wb_values)
            self._load_budget_categories(wb_values)
            self._load_pv_titles(wb_values)
        finally:
            wb_formula.close()
            wb_values.close()

    def _load_accounts_data(self, wb_formula=None, wb_values=None):
        """Charge Avoirs (éditable) + Contrôles (sous-comptes) de comptes.xlsm."""
        close_formula = close_values = False
        if wb_formula is None:
            wb_formula = openpyxl.load_workbook(self.xlsx_path, data_only=False)
            close_formula = True
        if wb_values is None:
            wb_values = openpyxl.load_workbook(self.xlsx_path, data_only=True)
            close_values = True
        try:
            self._load_accounts_data_inner(wb_formula, wb_values)
        finally:
            if close_formula:
                wb_formula.close()
            if close_values:
                wb_values.close()

    def _load_accounts_data_inner(self, wb_formula, wb_values):
        """Charge Avoirs + Contrôles depuis les workbooks déjà ouverts."""
        # Bornes via named ranges colonnes (cr.rows retourne start/end 1-indexed)
        from inc_excel_schema import ColResolver
        self.cr = ColResolver.from_openpyxl(wb_formula)
        self._start_avr, self._end_avr = self.cr.rows('AVRintitulé')
        self._start_ctrl1, self._end_ctrl1 = self.cr.rows('CTRL1compte')
        self._start_pvl, self._end_pvl = self.cr.rows('PVLcompte')
        self._start_cot, self._end_cot = self.cr.rows('COTcode')
        self._start_op, _ = self.cr.rows('OPdate')
        self._end_op = None  # OP n'a pas de borne END fixe
        # Fallbacks
        if self._end_avr is None: self._end_avr = 200
        if self._end_ctrl1 is None: self._end_ctrl1 = 100
        if self._end_pvl is None: self._end_pvl = 200
        if self._end_cot is None: self._end_cot = 30

        # --- Avoirs : données éditables + formules ---
        ws_formula = wb_formula[SHEET_AVOIRS]

        # --- Avoirs : comptes éditables (pour sauvegarde) ---
        self.accounts_data = []
        self._accounts_total_row = None

        # Données entre start AVR+1 et end AVR, Total = end AVR+1
        avr_data_start = self._start_avr + 1
        self._accounts_total_row = (self._end_avr + 1) if self._end_avr else None
        for row_idx in range(avr_data_start, self._end_avr or avr_data_start + 200):
            cell_a = ws_formula.cell(row_idx, self.cr.col('AVRintitulé')).value
            if not cell_a or str(cell_a).strip() in ('✓', '⚓'):
                continue

            orig_name = str(cell_a).strip()
            account = {
                'row': row_idx,
                'intitule': orig_name,
                'type': str(ws_formula.cell(row_idx, self.cr.col('AVRtype')).value or '').strip(),
                'domiciliation': str(ws_formula.cell(row_idx, self.cr.col('AVRdomiciliation')).value or '').strip(),
                'sous_type': str(ws_formula.cell(row_idx, self.cr.col('AVRsous_type')).value or '').strip(),
                'devise': str(ws_formula.cell(row_idx, self.cr.col('AVRdevise')).value or '').strip(),
                'titulaire': str(ws_formula.cell(row_idx, self.cr.col('AVRtitulaire')).value or '').strip(),
                'propriete': str(ws_formula.cell(row_idx, self.cr.col('AVRpropriete')).value or '').strip(),
                'date_anter': ws_formula.cell(row_idx, self.cr.col('AVRdate_anter')).value,
                'montant_anter': ws_formula.cell(row_idx, self.cr.col('AVRmontant_anter')).value,
                'formula_j': ws_formula.cell(row_idx, self.cr.col('AVRdate_solde')).value,
                'formula_k': ws_formula.cell(row_idx, self.cr.col('AVRmontant_solde')).value,
                'formula_l': ws_formula.cell(row_idx, self.cr.col('AVRmontant_solde_euro')).value,
                'site': self.account_site_map.get(orig_name, ''),
            }

            self.accounts_data.append(account)

        avoirs_by_row = {a['row']: a for a in self.accounts_data}
        avoirs_by_name = {a['intitule']: a for a in self.accounts_data}

        # Devises : Avoirs col E + Budget header (devises avec colonne Budget)
        # '-' n'est pas une vraie devise (placeholder legacy pour biens non
        # monétisables) — exclu pour ne pas polluer les listes ni les warnings.
        devises_set = {a['devise'] for a in self.accounts_data if a['devise']}
        if hasattr(self, 'budget_devises'):
            devises_set |= self.budget_devises
        devises_set.discard('')
        devises_set.discard('-')
        self.ACCOUNT_DEVISES = sorted(devises_set)

        # Sous-types : base + sous-types existants dans les données
        existing_sous_types = {a.get('sous_type', '') for a in self.accounts_data}
        self.ACCOUNT_SOUS_TYPES = sorted(
            set(self.SOUS_TYPES_BASE) | (existing_sous_types - {''})
        )

        # --- Contrôles : une ligne par (compte, sous-compte) ---
        ws_ctrl = wb_formula[SHEET_CONTROLES]
        ws_ctrl_val = wb_values[SHEET_CONTROLES]

        self.display_accounts = []
        seen_avoirs = set()  # intitulés vus dans Contrôles

        ctrl_data_start = self._start_ctrl1 + 1
        for row_idx in range(ctrl_data_start, self._end_ctrl1 + 1):
            cell_a = ws_ctrl.cell(row_idx, self.cr.col('CTRL1compte')).value
            if not cell_a or str(cell_a).strip() in ('✓', '⚓'):
                continue

            name = str(cell_a).strip()
            m = re.match(r"=Avoirs!\$?A(\d+)", name)
            if m:
                avoirs_row = int(m.group(1))
                acct = avoirs_by_row.get(avoirs_row)
                if acct:
                    name = acct['intitule']
                else:
                    continue

            if name not in avoirs_by_name:
                continue

            avoirs_acct = avoirs_by_name[name]
            seen_avoirs.add(name)
            ctrl_val = str(ws_ctrl_val.cell(row_idx, self.cr.col('CTRL1controle')).value or '').strip()

            self.display_accounts.append({
                'ctrl_row': row_idx,
                'intitule': name,
                'devise': avoirs_acct['devise'],
                'controle': ctrl_val.lower() == 'oui',
                'site': avoirs_acct['site'],
                'type': avoirs_acct['type'],
                'avoirs_ref': avoirs_acct,
            })

        # Biens matériels : pas dans Contrôles, ajout direct
        for acct in self.accounts_data:
            if acct.get('type') == 'Biens matériels' and acct['intitule'] not in seen_avoirs:
                self.display_accounts.append({
                    'ctrl_row': None,
                    'intitule': acct['intitule'],
                    'devise': acct['devise'],
                    'controle': False,
                    'site': acct.get('site', 'N/A'),
                    'type': acct['type'],
                    'avoirs_ref': acct,
                })

    def _load_pv_titles(self, wb_values=None):
        """Wrapper sur inc_compta_schema.load_pv_titles (extrait neutre)."""
        _schema.load_pv_titles(self, wb_values)

    def _load_budget_categories(self, wb_values=None):
        """Wrapper sur inc_compta_schema.load_budget_categories (extrait neutre)."""
        _schema.load_budget_categories(self, wb_values)

    # ----------------------------------------------------------------
    # LANCEMENT
    # ----------------------------------------------------------------
    def _handle_tk_exception(self, exc_type, exc_value, exc_tb):
        """Attrape les exceptions non gérées dans les callbacks Tkinter."""
        import traceback
        msg = ''.join(traceback.format_exception(exc_type, exc_value, exc_tb))
        print(msg, file=sys.stderr)
        try:
            out = self._exec_output
            out.configure(state='normal')
            out.insert('end', f'\n❌ Erreur inattendue :\n\n{msg}')
            out.see('end')
            out.configure(state='disabled')
            # Auto-bascule sur l'onglet Exécution pour rendre l'erreur visible
            try:
                self.notebook.select(self._tab_execution)
            except Exception:
                pass
        except Exception:
            pass  # zone Résultat pas encore construite
        try:
            self._set_status('ERR', 'error')
        except Exception:
            pass

    def _check_startup_health(self):
        """Checks groupés au démarrage GUI, schedulés après mapping root pour
        éviter les messageboxes cachées derrière la fenêtre principale (Mac).

        Limité à `check_env()` (présence du wrapper python3-uno). Les
        exceptions daemon sont désormais loggées dans `logs/journal.log` via
        inc_logging — accessible via le menu Outils/journal — plutôt que
        rejouées en popup intrusif au démarrage suivant.
        """
        # Forcer la fenêtre au premier plan (focus issue Mac)
        try:
            self.root.lift()
            self.root.update_idletasks()
        except tk.TclError:
            pass

        env_ok, env_msg = check_env()
        if not env_ok:
            messagebox.showwarning('Environnement Compta', env_msg,
                                   parent=self.root)

    def run(self):
        self.root.mainloop()


# ============================================================================
# MAIN
# ============================================================================

def main():
    import argparse
    parser = argparse.ArgumentParser(
        description='GUI Comptabilité')
    parser.add_argument('--config', '-c', default=str(DEFAULT_CONFIG),
                        help=f'Chemin config.ini (défaut: {DEFAULT_CONFIG})')
    parser.add_argument('--json', '-j', default=str(DEFAULT_JSON),
                        help=f'Chemin config_category_mappings.json (défaut: {DEFAULT_JSON})')
    parser.add_argument('--xlsx', '-x', default=None,
                        help='Chemin comptes.xlsm (défaut: depuis config.ini [paths] comptes_file)')
    args = parser.parse_args()

    config_path = Path(args.config)
    json_path = Path(args.json)

    if not config_path.exists():
        print(f"Erreur: {config_path} introuvable", file=sys.stderr)
        sys.exit(1)
    # json_path (config_category_mappings.json) : pas d'exit si absent —
    # le check de cohérence créera un fichier vide au démarrage.

    # Résoudre le chemin xlsx
    xlsx_path = None
    if args.xlsx:
        xlsx_path = Path(args.xlsx)
    else:
        import configparser
        cfg = configparser.ConfigParser()
        cfg.optionxform = str
        cfg.read(config_path, encoding='utf-8')
        comptes_file = cfg.get('paths', 'comptes_file', fallback='./comptes.xlsm')
        xlsx_path = (config_path.parent / comptes_file).resolve()

    if xlsx_path and not xlsx_path.exists():
        print(f"Warning: {xlsx_path} introuvable, onglet Comptes désactivé",
              file=sys.stderr)
        xlsx_path = None

    app = ConfigGUI(config_path, json_path, xlsx_path)
    app.run()


if __name__ == '__main__':
    try:
        main()
    except Exception:
        import traceback
        traceback.print_exc()
        try:
            from tkinter import messagebox
            messagebox.showerror('Erreur fatale', traceback.format_exc())
        except Exception:
            pass
        sys.exit(1)
