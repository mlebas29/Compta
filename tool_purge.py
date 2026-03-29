#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
cpt_purge.py - Purge des opérations anciennes dans comptes.xlsm

Objectif:
  - Conserver 1+ an d'historique (opérations récentes)
  - Purger les opérations anciennes pour réduire la taille du fichier

Contraintes:
  1. Ne PAS purger les comptes listés dans Plus_value!A5:A119
     (assurances vie, PEE, portefeuilles, métaux → suivis pour valorisations)
  2. Couper à un point équilibré (pas de paires cassées)
  3. Garder 1 seul #Solde par compte (le dernier avant date de purge)
  4. Créer backup avant modification

Usage:
  ./cpt_purge.py --date 2024-01-01       # Purger avant cette date
  ./cpt_purge.py --keep-months 12        # Garder 12 mois
  ./cpt_purge.py --audit                 # Simuler sans modifier
"""

import sys
import os
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font
import re
import argparse
from copy import copy
from inc_logging import Logger
from inc_excel_schema import (
    OpCol, PvCol, SHEET_OPERATIONS, SHEET_PLUS_VALUE,
    PV_PROTECTED_FIRST_ROW,
)


# ============================================================================
# FONCTIONS UTILITAIRES FORMATAGE
# ============================================================================

def copy_cell_formatting(source_cell, target_cell):
    """Copie tous les attributs de formatage d'une cellule source vers une cellule cible

    Args:
        source_cell: Cellule source (template)
        target_cell: Cellule cible (à formater)

    Note:
        Copie Font, Border, Fill, Alignment et NumberFormat
    """
    # Copier la police
    if source_cell.font:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            underline=source_cell.font.underline,
            color=source_cell.font.color
        )

    # Copier les bordures
    if source_cell.border:
        target_cell.border = copy(source_cell.border)

    # Copier le remplissage
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)

    # Copier l'alignement
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)

    # Copier le format de nombre (date, monétaire, etc.)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format


# ============================================================================
# FONCTIONS PRINCIPALES
# ============================================================================

def load_protected_accounts(ws_plusvalue):
    """Charge la liste des comptes protégés depuis Plus_value colonne A.

    Scanne dynamiquement de PV_PROTECTED_FIRST_ROW jusqu'à la fin du tableau.

    Args:
        ws_plusvalue: worksheet Plus_value

    Returns:
        set: Noms des comptes à ne jamais purger
    """
    protected = set()

    for row in range(PV_PROTECTED_FIRST_ROW, ws_plusvalue.max_row + 1):
        cell_value = ws_plusvalue.cell(row=row, column=PvCol.COMPTE).value
        if cell_value:
            account_name = str(cell_value).strip()
            if account_name:
                protected.add(account_name)

    return protected


def parse_date(date_str):
    """Parse une date au format DD/MM/YYYY

    Args:
        date_str: str (format DD/MM/YYYY)

    Returns:
        datetime ou None
    """
    try:
        return datetime.strptime(date_str, '%d/%m/%Y')
    except:
        return None


def analyze_operations(ws_operations, cutoff_date, protected_accounts):
    """Analyse les opérations et identifie celles à purger

    Args:
        ws_operations: worksheet Opérations
        cutoff_date: datetime (date de coupure)
        protected_accounts: set (comptes à ne jamais purger)

    Returns:
        dict: {
            'total_rows': int,
            'purgeable_rows': [row_num, ...],
            'protected_rows': [row_num, ...],
            'accounts': {
                'Account Name': {
                    'total': int,
                    'purgeable': int,
                    'protected': bool,
                    'last_solde_before_cutoff': row_num or None,
                    'operations_before_cutoff': [(row, date, amount, category), ...],
                    'all_soldes': [(row, date, amount), ...]
                }
            }
        }
    """
    # Headers en ligne 2
    # Colonnes: A=Date, B=Libellé, C=Montant, D=Devise, E=Equiv, F=Réf, G=Catégorie, H=Compte, I=Sous-compte, J=Commentaire

    stats = {
        'total_rows': 0,
        'purgeable_rows': [],
        'protected_rows': [],
        'accounts': {}
    }

    # Parcourir toutes les lignes (skip header row 1-2)
    for row in range(3, ws_operations.max_row + 1):
        date_cell = ws_operations.cell(row=row, column=OpCol.DATE).value
        account_cell = ws_operations.cell(row=row, column=OpCol.COMPTE).value
        category_cell = ws_operations.cell(row=row, column=OpCol.CATEGORIE).value
        amount_cell = ws_operations.cell(row=row, column=OpCol.MONTANT).value

        # Skip lignes vides
        if not date_cell:
            continue

        stats['total_rows'] += 1

        # Parser la date
        if isinstance(date_cell, datetime):
            op_date = date_cell
        else:
            op_date = parse_date(str(date_cell))

        if not op_date:
            continue

        account_name = str(account_cell).strip() if account_cell else ""
        category_name = str(category_cell).strip() if category_cell else ""

        # SKIP lignes fantômes avec compte vide/None
        if not account_name or account_name in ['None', 'nan']:
            continue

        # Initialiser stats compte
        if account_name not in stats['accounts']:
            stats['accounts'][account_name] = {
                'total': 0,
                'purgeable': 0,
                'protected': account_name in protected_accounts,
                'last_solde_before_cutoff': None,
                'operations_before_cutoff': [],
                'all_operations': [],  # TOUTES les opérations (pour backward)
                'all_soldes': []
            }

        stats['accounts'][account_name]['total'] += 1

        # Tracker tous les #Solde
        if '#Solde' in category_name or '#solde' in category_name.lower():
            try:
                amount = float(str(amount_cell).replace(',', '.'))
            except:
                amount = 0.0
            stats['accounts'][account_name]['all_soldes'].append((row, op_date, amount))

        # Tracker TOUTES les opérations (pour calcul backward)
        try:
            amount = float(str(amount_cell).replace(',', '.'))
        except:
            amount = 0.0
        stats['accounts'][account_name]['all_operations'].append((row, op_date, amount, category_name))

        # Vérifier si cette ligne est purgeable
        is_purgeable = False

        if account_name in protected_accounts:
            # Compte protégé → jamais purger
            stats['protected_rows'].append(row)
        elif op_date < cutoff_date:
            # Date ancienne → purgeable
            is_purgeable = True
            stats['purgeable_rows'].append(row)
            stats['accounts'][account_name]['purgeable'] += 1

            # Tracker opérations avant cutoff (pour calculer nouveau solde en mode forward)
            stats['accounts'][account_name]['operations_before_cutoff'].append((row, op_date, amount, category_name))

            # Tracker le dernier #Solde avant cutoff
            if '#Solde' in category_name or '#solde' in category_name.lower():
                stats['accounts'][account_name]['last_solde_before_cutoff'] = row
        else:
            # Date récente → garder
            pass

    return stats


def check_broken_pairs(ws_operations, purgeable_rows, protected_rows):
    """Vérifie les références qui seraient cassées par la purge

    Args:
        ws_operations: worksheet Opérations
        purgeable_rows: [row_num, ...] lignes à purger
        protected_rows: [row_num, ...] lignes protégées (comptes Plus_value)

    Returns:
        dict: {
            'refs_mapping': {ref_str: [row_num, ...]},
            'broken_refs': {ref_str: {'purgeable': [...], 'kept': [...]}},
            'rows_to_save': [row_num, ...]  # Lignes à sauver pour éviter paires cassées
        }
    """
    # Construire mapping ref → lignes
    refs_mapping = {}

    # Parcourir toutes les lignes avec une référence
    for row in range(3, ws_operations.max_row + 1):
        ref_cell = ws_operations.cell(row=row, column=OpCol.REF).value

        if not ref_cell:
            continue

        ref_str = str(ref_cell).strip()

        # Ignorer les refs vides ou "-" (non appariées)
        if not ref_str or ref_str == '-':
            continue

        # Ignorer les #Info
        category_cell = ws_operations.cell(row=row, column=OpCol.CATEGORIE).value
        if category_cell and ('#Info' in str(category_cell) or '#info' in str(category_cell).lower()):
            continue

        if ref_str not in refs_mapping:
            refs_mapping[ref_str] = []

        refs_mapping[ref_str].append(row)

    # Identifier les refs qui seraient cassées
    purgeable_set = set(purgeable_rows)
    protected_set = set(protected_rows)

    broken_refs = {}
    rows_to_save = []

    for ref_str, rows in refs_mapping.items():
        # Classer les lignes de cette ref
        ref_purgeable = [r for r in rows if r in purgeable_set]
        ref_kept = [r for r in rows if r not in purgeable_set]

        # Si la ref a des lignes purgeables ET des lignes conservées → paire cassée!
        if ref_purgeable and ref_kept:
            broken_refs[ref_str] = {
                'purgeable': ref_purgeable,
                'kept': ref_kept
            }

            # Sauver les lignes purgeables de cette ref
            rows_to_save.extend(ref_purgeable)

    return {
        'refs_mapping': refs_mapping,
        'broken_refs': broken_refs,
        'rows_to_save': rows_to_save
    }


def calculate_new_soldes(stats, cutoff_date):
    """Calcule les nouveaux #Solde à la date de coupure

    Args:
        stats: dict from analyze_operations()
        cutoff_date: datetime

    Returns:
        dict: {
            'Account Name': {
                'new_solde_date': datetime,
                'new_solde_amount': float,
                'calculated_from': row_num (dernier #Solde avant cutoff)
            }
        }
    """
    new_soldes = {}

    for account, account_stats in stats['accounts'].items():
        if account_stats['protected']:
            # Compte protégé → skip
            continue

        if account_stats['purgeable'] == 0:
            # Rien à purger → skip
            continue

        # STRATÉGIE: Remonter le temps depuis le premier #Solde conservé après cutoff

        # Trouver le premier #Solde APRÈS cutoff (celui qui sera conservé)
        soldes_after_cutoff = [(row, date, amt) for row, date, amt in account_stats['all_soldes'] if date >= cutoff_date]

        if not soldes_after_cutoff:
            # Pas de #Solde après cutoff
            # Chercher dernier #Solde AVANT cutoff
            soldes_before_cutoff = [(row, date, amt) for row, date, amt in account_stats['all_soldes'] if date < cutoff_date]

            # Vérifier s'il y a des opérations après cutoff (hors #Solde/#Balance)
            operations_after_cutoff = [
                (row, date, amt, cat)
                for row, date, amt, cat in account_stats['all_operations']
                if date >= cutoff_date and not cat.startswith('#')
            ]

            if soldes_before_cutoff:
                # On a un #Solde de référence AVANT cutoff
                soldes_before_cutoff.sort(key=lambda x: x[1])
                last_solde_row, last_solde_date, last_solde_amount = soldes_before_cutoff[-1]

                # Partir du dernier #Solde et ajouter les opérations jusqu'au cutoff
                new_solde_amount = last_solde_amount

                # Ajouter les opérations entre le dernier #Solde et le cutoff (EXCLUS cutoff)
                for row, op_date, amount, category in account_stats['all_operations']:
                    if last_solde_date < op_date < cutoff_date and not category.startswith('#'):
                        new_solde_amount += amount

                calculated_from = f'forward from L{last_solde_row} ({last_solde_date.strftime("%d/%m/%Y")} = {last_solde_amount})'

            elif operations_after_cutoff:
                # Pas de #Solde de référence, mais des opérations après cutoff
                # → Calculer backward depuis TOUTES les opérations
                # En partant du principe que le solde final = somme des opérations (compte sans #Solde historique)

                # Calculer solde fictif depuis LE DÉBUT (toutes les opérations, toutes périodes)
                total_balance = 0.0
                for row, op_date, amount, category in account_stats['all_operations']:
                    if not category.startswith('#'):
                        total_balance += amount

                # Soustraire les opérations après cutoff pour obtenir solde au cutoff-1
                for row, op_date, amount, category in operations_after_cutoff:
                    total_balance -= amount

                new_solde_amount = total_balance
                calculated_from = f'backward from all operations (no #Solde reference, {len(operations_after_cutoff)} ops after cutoff)'

                # AVERTISSEMENT si solde calculé = 0 mais qu'il y a des opérations importantes après cutoff
                if abs(new_solde_amount) < 1000 and any(abs(amt) > 10000 for _, _, amt, _ in operations_after_cutoff):
                    print(f"⚠️  {account}: Solde calculé = {new_solde_amount:.2f}€ mais opérations importantes après cutoff", file=sys.stderr)
                    print(f"   → Historique incomplet (versement initial probablement purgé)", file=sys.stderr)
                    print(f"   → Vérifier manuellement et ajouter #Solde initial si nécessaire", file=sys.stderr)

            else:
                # Pas de #Solde et pas d'opérations après cutoff → solde = 0
                new_solde_amount = 0.0
                calculated_from = 'no historical #Solde (starting from 0)'

        else:
            # MÉTHODE PRINCIPALE: Remonter le temps depuis le premier #Solde conservé
            soldes_after_cutoff.sort(key=lambda x: x[1])
            first_solde_row, first_solde_date, first_solde_amount = soldes_after_cutoff[0]

            # Partir du solde futur
            new_solde_amount = first_solde_amount

            # Soustraire les opérations entre cutoff et ce #Solde futur
            # (car on remonte le temps)
            # FILTRER comme Excel: ignorer catégories commençant par #
            for row, op_date, amount, category in account_stats['all_operations']:
                # Prendre les opérations entre cutoff et first_solde_date (INCLUS)
                if cutoff_date <= op_date <= first_solde_date:
                    # Ignorer toutes les catégories # (comme Excel)
                    if not category.startswith('#'):
                        new_solde_amount -= amount  # Soustraire car on remonte

            calculated_from = f'backward from L{first_solde_row} ({first_solde_date.strftime("%d/%m/%Y")} = {first_solde_amount})'
            last_solde_row = first_solde_row

        # Nouveau #Solde à la veille de la date de coupure
        new_soldes[account] = {
            'new_solde_date': cutoff_date - timedelta(days=1),
            'new_solde_amount': new_solde_amount,
            'calculated_from': calculated_from
        }

    return new_soldes


def generate_purge_plan(stats, ws_operations, cutoff_date):
    """Génère un plan de purge détaillé

    Args:
        stats: dict from analyze_operations()
        ws_operations: worksheet Opérations
        cutoff_date: datetime

    Returns:
        dict: {
            'rows_to_delete': [row_num, ...],
            'new_soldes': {account: {'date': datetime, 'amount': float}},
            'accounts_summary': {account: {...}},
            'broken_pairs': {...}  # Info sur paires cassées
        }
    """
    plan = {
        'rows_to_delete': [],
        'new_soldes': {},
        'accounts_summary': {},
        'broken_pairs': {}
    }

    # Calculer les nouveaux #Solde à la date de coupure
    new_soldes = calculate_new_soldes(stats, cutoff_date)
    plan['new_soldes'] = new_soldes

    # Pour chaque compte purgeable
    for account, account_stats in stats['accounts'].items():
        if account_stats['protected']:
            # Compte protégé → skip
            plan['accounts_summary'][account] = {
                'action': 'PROTECTED',
                'rows_deleted': 0,
                'reason': 'Compte dans Plus_value'
            }
            continue

        if account_stats['purgeable'] == 0:
            # Rien à purger
            plan['accounts_summary'][account] = {
                'action': 'SKIP',
                'rows_deleted': 0,
                'reason': 'Aucune opération ancienne'
            }
            continue

        # Nouveau solde calculé
        new_solde_info = new_soldes.get(account)

        plan['accounts_summary'][account] = {
            'action': 'PURGE',
            'rows_deleted': account_stats['purgeable'],
            'new_solde_amount': new_solde_info['new_solde_amount'] if new_solde_info else None,
            'new_solde_date': new_solde_info['new_solde_date'] if new_solde_info else None
        }

    # Construire la liste initiale des lignes à supprimer
    # TOUTES les opérations purgeable (y compris les anciens #Solde)
    initial_delete = list(stats['purgeable_rows'])

    # Vérifier les paires cassées
    all_kept_rows = set(range(3, ws_operations.max_row + 1)) - set(initial_delete)
    pair_check = check_broken_pairs(ws_operations, initial_delete, stats['protected_rows'])

    plan['broken_pairs'] = pair_check

    # Retirer les lignes à sauver (paires cassées)
    rows_to_save_set = set(pair_check['rows_to_save'])
    plan['rows_to_delete'] = [r for r in initial_delete if r not in rows_to_save_set]

    return plan


def execute_purge(ws_operations, plan, cutoff_date, dry_run=False):
    """Exécute le plan de purge

    Args:
        ws_operations: worksheet Opérations
        plan: dict from generate_purge_plan()
        cutoff_date: datetime
        dry_run: bool (si True, ne modifie pas)

    Returns:
        int: nombre de lignes supprimées
    """
    if dry_run:
        print("🔍 MODE AUDIT - Aucune modification")
        return 0

    # 1. Supprimer d'abord les anciennes lignes
    # Trier en ordre décroissant (pour ne pas casser les indices)
    rows_to_delete = sorted(plan['rows_to_delete'], reverse=True)

    print(f"🗑️  Suppression de {len(rows_to_delete)} lignes...")

    for row_num in rows_to_delete:
        ws_operations.delete_rows(row_num, 1)

    # 2. Créer les nouveaux #Solde à leur position chronologique
    print(f"📝 Création de {len(plan['new_soldes'])} nouveaux #Solde...")

    for account, solde_info in plan['new_soldes'].items():
        # Trouver un #Solde existant du même compte pour copier son formatage
        template_row = None

        # 1. Chercher un #Solde du même compte
        for row in range(3, ws_operations.max_row + 1):
            account_cell = ws_operations.cell(row=row, column=OpCol.COMPTE).value
            cat_cell = ws_operations.cell(row=row, column=OpCol.CATEGORIE).value
            date_cell = ws_operations.cell(row=row, column=OpCol.DATE).value

            if account_cell and str(account_cell).strip() == account:
                if cat_cell and '#Solde' in str(cat_cell):
                    # Préférer un #Solde après cutoff (même devise)
                    if isinstance(date_cell, datetime) and date_cell >= cutoff_date:
                        template_row = row
                        break
                    elif template_row is None:
                        # Sinon prendre n'importe quel #Solde du compte
                        template_row = row

        # 2. Fallback : chercher n'importe quelle ligne EUR si template pas trouvé
        if template_row is None:
            for row in range(ws_operations.max_row, 3, -1):  # Chercher depuis la fin
                row_devise = ws_operations.cell(row, OpCol.DEVISE).value
                if row_devise == 'EUR':
                    template_row = row
                    break

        # 3. Dernier fallback : utiliser la dernière ligne de données
        if template_row is None:
            template_row = ws_operations.max_row

        # Position d'insertion chronologique : juste avant la première opération >= cutoff (tous comptes)
        # Car le #Solde est daté au 31/12/2024
        insert_row = None

        for row in range(3, ws_operations.max_row + 1):
            date_cell = ws_operations.cell(row=row, column=OpCol.DATE).value

            if not date_cell:
                continue

            # Parser date
            if isinstance(date_cell, datetime):
                op_date = date_cell
            else:
                op_date = parse_date(str(date_cell))

            if not op_date:
                continue

            # Première ligne avec date >= cutoff (tous comptes confondus)
            if op_date >= cutoff_date:
                insert_row = row
                break

        # Si aucune opération >= cutoff trouvée, insérer avant #Balance
        if insert_row is None:
            for row in range(3, ws_operations.max_row + 1):
                cat = ws_operations.cell(row=row, column=OpCol.CATEGORIE).value
                if cat and '#Balance' in str(cat):
                    insert_row = row
                    break

        # En dernier recours, insérer à la fin
        if insert_row is None:
            insert_row = ws_operations.max_row + 1

        # Insérer une nouvelle ligne
        ws_operations.insert_rows(insert_row)

        # Remplir la nouvelle ligne (style homogène avec cpt_update.py)
        new_solde_date = solde_info['new_solde_date']

        # Date (déjà datetime)
        cell = ws_operations.cell(insert_row, OpCol.DATE)
        cell.value = new_solde_date
        copy_cell_formatting(ws_operations.cell(template_row, OpCol.DATE), cell)

        # Libellé
        cell = ws_operations.cell(insert_row, OpCol.LABEL)
        cell.value = "Relevé compte"
        copy_cell_formatting(ws_operations.cell(template_row, OpCol.LABEL), cell)

        # Montant (déjà float)
        cell = ws_operations.cell(insert_row, OpCol.MONTANT)
        cell.value = solde_info['new_solde_amount']
        copy_cell_formatting(ws_operations.cell(template_row, OpCol.MONTANT), cell)

        # Devise
        cell = ws_operations.cell(insert_row, OpCol.DEVISE)
        cell.value = "EUR"
        copy_cell_formatting(ws_operations.cell(template_row, OpCol.DEVISE), cell)

        # Equiv
        cell = ws_operations.cell(insert_row, OpCol.EQUIV)
        cell.value = None
        copy_cell_formatting(ws_operations.cell(template_row, OpCol.EQUIV), cell)

        # Réf
        cell = ws_operations.cell(insert_row, OpCol.REF)
        cell.value = None
        copy_cell_formatting(ws_operations.cell(template_row, OpCol.REF), cell)

        # Catégorie
        cell = ws_operations.cell(insert_row, OpCol.CATEGORIE)
        cell.value = "#Solde"
        copy_cell_formatting(ws_operations.cell(template_row, OpCol.CATEGORIE), cell)

        # Compte
        cell = ws_operations.cell(insert_row, OpCol.COMPTE)
        cell.value = account
        copy_cell_formatting(ws_operations.cell(template_row, OpCol.COMPTE), cell)

        # Commentaire
        cell = ws_operations.cell(insert_row, OpCol.COMMENTAIRE)
        cell.value = None
        copy_cell_formatting(ws_operations.cell(template_row, OpCol.COMMENTAIRE), cell)

    return len(rows_to_delete)


def main():
    parser = argparse.ArgumentParser(description='Purge des opérations anciennes')
    parser.add_argument('--date', help='Date de coupure (YYYY-MM-DD)')
    parser.add_argument('--keep-months', type=int, help='Nombre de mois à conserver')
    parser.add_argument('--audit', action='store_true', help='Simuler sans modifier')
    parser.add_argument('-v', '--verbose', action='store_true', help='Mode verbeux')

    args = parser.parse_args()

    logger = Logger(script_name="tool_purge", verbose=args.verbose)

    # Déterminer la date de coupure
    if args.date:
        try:
            cutoff_date = datetime.strptime(args.date, '%Y-%m-%d')
        except ValueError:
            print(f"❌ Date invalide: {args.date} (format attendu: YYYY-MM-DD)", file=sys.stderr)
            return 1
    elif args.keep_months:
        cutoff_date = datetime.now() - timedelta(days=args.keep_months * 30)
    else:
        # Par défaut: conserver toute l'année fiscale précédente (année N-1 complète)
        # Exemple: si on est en 2026 → garder 2025 + 2026 → cutoff = 01/01/2025
        current_year = datetime.now().year
        cutoff_date = datetime(current_year - 1, 1, 1)

    print("=" * 80)
    print("PURGE DES OPÉRATIONS ANCIENNES")
    print("=" * 80)
    print()
    print(f"📅 Date de coupure: {cutoff_date.strftime('%d/%m/%Y')}")
    print(f"   Opérations antérieures à cette date seront purgées")
    print()
    print("⚠️  Contraintes de sécurité:")
    print(f"   • Comptes Plus_value: jamais purgés (valorisations)")
    print(f"   • Références appariées: paires conservées intégralement")
    print(f"   • Dernier #Solde: conservé par compte avant coupure")
    if not args.date and not args.keep_months:
        print(f"   • Année fiscale N-1: conservée intégralement (calculs fiscaux)")
    print()

    # Charger Excel
    excel_file = Path('comptes.xlsm')
    if not excel_file.exists():
        print(f"❌ Fichier introuvable: {excel_file}", file=sys.stderr)
        return 1

    print(f"📖 Lecture de {excel_file}...")
    wb = load_workbook(excel_file, keep_vba=True)

    if SHEET_OPERATIONS not in wb.sheetnames:
        print(f"❌ Feuille '{SHEET_OPERATIONS}' introuvable", file=sys.stderr)
        return 1

    if SHEET_PLUS_VALUE not in wb.sheetnames:
        print(f"❌ Feuille '{SHEET_PLUS_VALUE}' introuvable", file=sys.stderr)
        return 1

    ws_operations = wb[SHEET_OPERATIONS]
    ws_plusvalue = wb[SHEET_PLUS_VALUE]

    # Charger comptes protégés
    print(f"🔒 Lecture des comptes protégés ({SHEET_PLUS_VALUE}!A{PV_PROTECTED_FIRST_ROW}:A{ws_plusvalue.max_row})...")
    protected_accounts = load_protected_accounts(ws_plusvalue)
    print(f"   {len(protected_accounts)} comptes protégés")

    for acc in sorted(protected_accounts):
        logger.verbose(f"  - {acc}")
    print()

    # Analyser opérations
    print("🔍 Analyse des opérations...")
    stats = analyze_operations(ws_operations, cutoff_date, protected_accounts)
    print(f"   {stats['total_rows']} opérations totales")
    print(f"   {len(stats['purgeable_rows'])} opérations purgeables")
    print(f"   {len(stats['protected_rows'])} opérations protégées")
    print()

    # Générer plan de purge
    print("📋 Génération du plan de purge...")
    plan = generate_purge_plan(stats, ws_operations, cutoff_date)
    print(f"   {len(plan['rows_to_delete'])} lignes à supprimer")
    print(f"   {len(plan['new_soldes'])} nouveaux #Solde à créer")

    # Afficher infos sur paires cassées
    broken_pairs = plan['broken_pairs']
    if broken_pairs['broken_refs']:
        num_broken = len(broken_pairs['broken_refs'])
        num_saved = len(broken_pairs['rows_to_save'])
        print(f"   {num_broken} paires sauvées (éviter références cassées)")
        print(f"   {num_saved} lignes conservées pour intégrité")

        logger.verbose("Détail des paires sauvées:")
        for ref, info in sorted(broken_pairs['broken_refs'].items())[:10]:
            purg = ', '.join(f"L{r}" for r in info['purgeable'])
            kept = ', '.join(f"L{r}" for r in info['kept'])
            logger.verbose(f"  {ref}: purgeables={purg} | conservées={kept}")
        if len(broken_pairs['broken_refs']) > 10:
            logger.verbose(f"  ... et {len(broken_pairs['broken_refs']) - 10} autres")
    print()

    # Afficher résumé par compte
    print("📊 Résumé par compte:")
    print()

    for account in sorted(plan['accounts_summary'].keys()):
        summary = plan['accounts_summary'][account]
        action = summary['action']

        if action == 'PROTECTED':
            print(f"  🔒 {account:40} → PROTÉGÉ ({summary['reason']})")
        elif action == 'PURGE':
            rows_del = summary['rows_deleted']
            new_solde = summary.get('new_solde_amount')
            new_date = summary.get('new_solde_date')
            if new_solde is not None and new_date:
                date_str = new_date.strftime('%d/%m/%Y')
                solde_info = f"(#Solde {date_str}: {new_solde:.2f}€)"
            else:
                solde_info = "(pas de #Solde)"
            print(f"  🗑️  {account:40} → {rows_del} lignes {solde_info}")
        elif action == 'SKIP':
            print(f"  ✓  {account:40} → Aucune opération ancienne")

    print()

    # Confirmation ou audit
    if args.audit:
        print("✓ AUDIT TERMINÉ - Aucune modification effectuée")
        return 0

    response = input(f"Supprimer {len(plan['rows_to_delete'])} lignes? [o/N] ")
    if response.lower() not in ['o', 'oui', 'y', 'yes']:
        print("❌ Annulé")
        return 0

    # Créer backup
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_file = Path(f"archives/comptes_BACKUP_PURGE_{timestamp}.xlsm")
    backup_file.parent.mkdir(exist_ok=True)

    print(f"\n💾 Création backup...")
    wb.save(backup_file)
    print(f"   {backup_file}")

    # Exécuter purge
    rows_deleted = execute_purge(ws_operations, plan, cutoff_date, dry_run=False)

    # Sauvegarder
    print("\n💾 Sauvegarde Excel...")
    wb.save(excel_file)
    print(f"   {excel_file}")

    # Recalcul + miroir C1 si lancé depuis la GUI
    if os.environ.get('COMPTA_GUI'):
        from inc_uno import refresh_controles
        refresh_controles(excel_file)

    print()
    print("=" * 80)
    print(f"✓ PURGE TERMINÉE - {rows_deleted} lignes supprimées")
    print("=" * 80)

    return 0


if __name__ == '__main__':
    sys.exit(main())
