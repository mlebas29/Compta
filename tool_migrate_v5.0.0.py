#!/usr/bin/env python3-uno
"""Migration classeur v5.0.0 — fiabilisation alarmes contre #REF! orphelines.

Sections intégrées :

1. Cotations!B{alarme} — IFERROR sur SUMPRODUCT completeness
   Wrappe `SUMPRODUCT((COTcode<>"")*(COTcours=""))` dans `IFERROR(...;1)`.
   Sans ça, un `#REF!` dans une cellule COTcours (typiquement après
   suppression d'une devise mère sans nettoyage des dérivées) propageait
   l'erreur dans `COTcours=""` → SUMPRODUCT plantait → IF retournait l'erreur
   et l'alarme métier n'affichait plus ⚠ visible. IFERROR force le résultat
   à 1 → l'alarme bascule à ⚠.

2. Contrôles!K{Synthèse} — IFERROR par section dans la concat
   Wrappe chaque token K{section} (COMPTES, CATÉGORIES, DIVERS, …) dans
   `IFERROR(K{r};"⚠")`. Sans ça, une section déjà en erreur (#REF! propagé
   depuis une dépendance) cassait la concat → FIND retournait l'erreur →
   `ISNUMBER(#REF!) = FALSE` → la synthèse tombait à ✓ alors qu'une section
   affichait ✗ visiblement. IFERROR convertit chaque section cassée en ⚠.

Idempotent : si déjà migré, ne fait rien. SCHEMA_VERSION inchangé (3) — ces
deux fixes sont des améliorations de formules, pas du structurel.

Inspection / écriture SÉPARÉES (#111 — parité simulé/réel) :
  - `_inspect()` détecte en **openpyxl read-only** (sans LibreOffice) les cellules
    à fiabiliser (critère = absence d'IFERROR, qui EST la marque de migration).
  - l'écriture seule passe par UNO (openpyxl interdit en écriture sur .xlsm).
  → `--dry-run` n'ouvre PAS LibreOffice : un appelant (upgrade --check) peut
  prédire l'effet pour pas cher. Codes retour : 0 = rien à faire / déjà migré ·
  3 = des changements seraient appliqués (dry-run) · 1 = erreur · 2 = LO < 24.8.

Usage:
    python3 tool_migrate_v5.0.0.py ~/Compta/comptes.xlsm
    python3 tool_migrate_v5.0.0.py ~/Compta/comptes.xlsm --dry-run   # sans LibreOffice
"""
import argparse
import shutil
import sys
from pathlib import Path

import openpyxl

from inc_uno import UnoDocument, check_lock_file, require_libreoffice_min
from inc_excel_schema import uno_row


CTRL2_TYPE_COL = 9    # J (0-based, indexation UNO)
CTRL2_DISPL_COL = 10  # K (0-based)
_COL_A = 1            # col A 1-based (openpyxl) — sentinelles ⚓ Cotations
_COL_B = 2            # col B 1-based — alarme Cotations
_COL_J = CTRL2_TYPE_COL + 1   # 10, labels CTRL2 (openpyxl 1-based)
_COL_K = CTRL2_DISPL_COL + 1  # 11, affichage CTRL2

_HEADERS = ['COMPTES', 'CATÉGORIES', 'DIVERS', 'APPARIEMENTS',
            'BALANCES', 'INCONNUS', 'FORMULES']


def _cot_target_row(ws_cot):
    """Ligne de l'alarme completeness Cotations = ligne après la 2e sentinelle ⚓
    en col A (pied de table, layout-agnostic) ; fallback 20."""
    anchors = [r for r in range(2, 100)
               if (ws_cot.cell(row=r, column=_COL_A).value or '').strip() == '⚓']
    return (anchors[1] + 1) if len(anchors) >= 2 else 20


def _section1_formula():
    """Formule cible Cotations!B{alarme} (IFERROR sur la branche completeness)."""
    return (
        '=IF('
        # (a) Devises utilisées non listées
        'SUMPRODUCT((COUNTIF(COTcode;PVLdevise)=0)*(PVLdevise<>""))'
        '+SUMPRODUCT((COUNTIF(COTcode;AVRdevise)=0)*(AVRdevise<>""))'
        # (b) Codes listés mais cours vide / (c) cours en erreur (#REF!)
        '+IFERROR(SUMPRODUCT((COTcode<>"")*(COTcours=""));1)'
        '>0;"⚠";"✓")'
    )


def _inspect(xlsm_path):
    """Détection READ-ONLY (openpyxl, SANS LibreOffice) des cellules à fiabiliser.

    Critère de migration = absence d'IFERROR dans la cellule cible : le wrapper
    IFERROR EST la marque de la migration v5.0.0 (avant : aucune IFERROR ; après :
    présente) → robuste à la représentation des formules (openpyxl `,` vs UNO `;`).
    Retourne une liste de changements (chacun : sheet/col0/row/formula/desc) — la
    formule cible est PRÉ-CONSTRUITE ici, l'écriture UNO n'a plus qu'à la poser.
    """
    wb = openpyxl.load_workbook(xlsm_path, data_only=False)
    changes = []

    # --- Section 1 : Cotations!B{target} ---
    if 'Cotations' in wb.sheetnames:
        ws = wb['Cotations']
        tr = _cot_target_row(ws)
        cur = ws.cell(row=tr, column=_COL_B).value
        if 'IFERROR' not in (cur or ''):
            changes.append({
                'sheet': 'Cotations', 'col0': 1, 'row': tr,
                'formula': _section1_formula(),
                'desc': (f"~ Cotations!B{tr} : IFERROR sur SUMPRODUCT "
                         f"completeness (capte #REF! orphelines)"),
            })

    # --- Section 2 : Contrôles!K{Synthèse} ---
    if 'Contrôles' in wb.sheetnames:
        ws = wb['Contrôles']
        rows = {}
        synth_row = None
        for r in range(1, 200):
            v = (ws.cell(row=r, column=_COL_J).value or '').strip()
            if not v:
                continue
            for lbl in _HEADERS:
                if v == lbl or v.startswith(lbl):
                    rows[lbl] = r
            if synth_row is None and v in ('Synthèse des contrôles', 'Synthèse'):
                synth_row = r

        missing = [h for h in _HEADERS if h not in rows]
        if missing:
            print(f"⚠ Synthèse : headers manquants {missing} — skip "
                  f"(classeur pas migré v4.1.0 ?)")
        elif synth_row is None:
            print("⚠ Synthèse : ligne 'Synthèse [des contrôles]' introuvable — skip")
        else:
            cur = ws.cell(row=synth_row, column=_COL_K).value
            if 'IFERROR' not in (cur or ''):
                k_concat = '&'.join(f'IFERROR(K{rows[h]};"⚠")' for h in _HEADERS)
                synth_k = (
                    f'=IF(ISNUMBER(FIND("✗";{k_concat}));"✗";'
                    f'IF(ISNUMBER(FIND("⚠";{k_concat}));"⚠";"✓"))'
                )
                changes.append({
                    'sheet': 'Contrôles', 'col0': CTRL2_DISPL_COL, 'row': synth_row,
                    'formula': synth_k,
                    'desc': (f"~ Contrôles!K{synth_row} : IFERROR wrapper par "
                             f"section (évite masquage ✓ silencieux)"),
                })
    wb.close()
    return changes


def migrate(xlsm_path, dry_run=False):
    p = Path(xlsm_path).resolve()
    if not p.exists():
        print(f"❌ Fichier introuvable : {p}")
        return 1
    if check_lock_file(p):
        print(f"❌ Fichier verrouillé : {p}")
        print("   Ferme LibreOffice.")
        return 1

    # Détection openpyxl read-only — AUCUN LibreOffice requis ici (dry-run inclus).
    changes = _inspect(p)

    if not changes:
        print(f"✓ {p.name} : déjà migré, rien à faire.")
        return 0

    print(f"Migration {p.name} :")
    for c in changes:
        print(f"  {c['desc']}")

    if dry_run:
        print("\n[dry-run] pas de sauvegarde")
        return 3  # code distinct : des changements seraient appliqués

    # --- écriture seule via UNO (openpyxl corromprait macros/styles d'un .xlsm) ---
    # LO < 24.8 corrompt les XLOOKUP via UNO save (`_xlfn.` illisible) : on refuse.
    require_libreoffice_min(24, 8)
    bak = p.with_suffix('.xlsm.bak')
    shutil.copy2(p, bak)
    print(f"📦 Backup : {bak.name}")

    with UnoDocument(p) as doc:
        for c in changes:
            ws = doc.get_sheet(c['sheet'])
            ws.getCellByPosition(c['col0'], uno_row(c['row'])).setFormula(c['formula'])
        doc.save()
    print(f"\n✓ Sauvé : {p}")
    return 0


def main():
    ap = argparse.ArgumentParser(description=__doc__.split('\n')[0])
    ap.add_argument('xlsm', help='Chemin du classeur à migrer')
    ap.add_argument('--dry-run', action='store_true',
                    help="Détecte et affiche les changements prévus SANS LibreOffice "
                         "ni sauvegarde (code retour 3 si des changements seraient posés)")
    args = ap.parse_args()
    return migrate(args.xlsm, dry_run=args.dry_run)


if __name__ == '__main__':
    sys.exit(main())
