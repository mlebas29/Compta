#!/usr/bin/env python3-uno
"""
Actions GUI en mode CLI (headless) — pour debug et tests automatisés.

Charge le xlsm + configs, exécute une action, sauvegarde.
Même code que la GUI Tkinter mais sans interface graphique.

Usage:
    python3 tool_gui_cli.py <xlsm> add-account <nom> <type> [--devise EUR] [--site N/A]
    python3 tool_gui_cli.py <xlsm> add-bien <nom> <nature> [--devise] [--montant]
    python3 tool_gui_cli.py <xlsm> add-devise <code> <famille>
    python3 tool_gui_cli.py <xlsm> list-accounts
    python3 tool_gui_cli.py <xlsm> check

Exemples:
    python3 tool_gui_cli.py comptes.xlsm add-account "Mon compte" Euros
    python3 tool_gui_cli.py comptes.xlsm add-account "Portfolio" Portefeuilles
    python3 tool_gui_cli.py comptes.xlsm add-bien "Maison" Foncier --montant 200000
    python3 tool_gui_cli.py comptes.xlsm add-bien "Bijoux" Mobilier --devise OrJo
    python3 tool_gui_cli.py comptes.xlsm add-devise USD fiat
    python3 tool_gui_cli.py comptes.xlsm list-accounts
    python3 tool_gui_cli.py comptes.xlsm check
"""

import sys
import argparse
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

from inc_uno import check_env
from inc_excel_schema import (
    SHEET_AVOIRS, SHEET_CONTROLES, SHEET_OPERATIONS, SHEET_PLUS_VALUE,
    SHEET_BUDGET, SHEET_COTATIONS,
    DEVISE_SOURCES, uno_col, uno_row,
)
from inc_config_io import (
    read_cotations_json, write_cotations_json,
    read_accounts_json, write_accounts_json, read_mappings_json,
)
import inc_compta_schema as _schema


from gui_devises import DevisesMixin
from gui_accounts import AccountsMixin
from gui_budget import BudgetMixin
from gui_categories import CategoriesMixin


class HeadlessGUI(DevisesMixin, AccountsMixin, BudgetMixin, CategoriesMixin):
    """ConfigGUI sans Tkinter — charge les données, exécute les actions UNO."""

    def __init__(self, xlsx_path):
        self.xlsx_path = Path(xlsx_path).absolute()
        if not self.xlsx_path.exists():
            raise FileNotFoundError(f"Fichier introuvable : {self.xlsx_path}")

        base = self.xlsx_path.parent

        # Chemins configs
        self.config_path = base / 'config.ini'
        self.json_path = base / 'config_category_mappings.json'
        self.accounts_json_path = base / 'config_accounts.json'
        self.cotations_json_path = base / 'config_cotations.json'
        self.pipeline_json_path = base / 'config_pipeline.json'

        # Charger les métadonnées
        self.cotations_meta = read_cotations_json(self.cotations_json_path)
        self.accounts_json_data = read_accounts_json(self.accounts_json_path) \
            if self.accounts_json_path.exists() else {}

        # Constantes de classe (depuis inc_compta_schema, neutre)
        self.ACCOUNT_TYPES = _schema.ACCOUNT_TYPES
        self.SOUS_TYPES_BASE = _schema.SOUS_TYPES_BASE
        self.PV_SECTION_TOTALS = _schema.PV_SECTION_TOTALS
        self.PV_SECTION_LABELS = _schema.PV_SECTION_LABELS
        self.cours_name = _schema.cours_name

        # Charger les données Excel
        self._load_excel_data()

        # État pour _save_accounts
        self._deleted_accounts = []
        self._soft_deleted_accounts = []
        self._deleted_ctrl_rows = []

    def _load_excel_data(self):
        """Charge les données Avoirs/Contrôles depuis le xlsm."""
        wb_formula = openpyxl.load_workbook(self.xlsx_path, data_only=False)
        wb_values = openpyxl.load_workbook(self.xlsx_path, data_only=True)
        try:
            from inc_excel_schema import ColResolver
            self.cr = ColResolver.from_openpyxl(wb_formula)
            self._start_avr, self._end_avr = self.cr.rows('AVRintitulé')
            self._start_ctrl1, self._end_ctrl1 = self.cr.rows('CTRL1compte')
            self._start_pvl, self._end_pvl = self.cr.rows('PVLcompte')
            self._start_cot, self._end_cot = self.cr.rows('COTcode')
            self._start_op, _ = self.cr.rows('OPdate')
            self._end_op = None
            if self._end_avr is None: self._end_avr = 200
            if self._end_ctrl1 is None: self._end_ctrl1 = 100
            if self._end_pvl is None: self._end_pvl = 200
            if self._end_cot is None: self._end_cot = 30

            # Avoirs
            ws_f = wb_formula[SHEET_AVOIRS]
            ws_v = wb_values[SHEET_AVOIRS]
            self.accounts_data = []
            self._accounts_total_row = (self._end_avr + 1) if self._end_avr else None
            avr_data_start = self._start_avr + 1
            for row_idx in range(avr_data_start, self._end_avr or avr_data_start + 200):
                cell_a = ws_f.cell(row_idx, self.cr.col('AVRintitulé')).value
                if not cell_a or not str(cell_a).strip() or str(cell_a).strip() in ('✓', '⚓'):
                    continue
                self.accounts_data.append({
                    'row': row_idx,
                    'intitule': str(cell_a).strip(),
                    'type': str(ws_f.cell(row_idx, self.cr.col('AVRtype')).value or '').strip(),
                    'domiciliation': str(ws_f.cell(row_idx, self.cr.col('AVRdomiciliation')).value or '').strip(),
                    'sous_type': str(ws_f.cell(row_idx, self.cr.col('AVRsous_type')).value or '').strip(),
                    'devise': str(ws_f.cell(row_idx, self.cr.col('AVRdevise')).value or '').strip(),
                    'titulaire': str(ws_f.cell(row_idx, self.cr.col('AVRtitulaire')).value or '').strip(),
                    'propriete': str(ws_f.cell(row_idx, self.cr.col('AVRpropriete')).value or '').strip(),
                    'formula_j': ws_f.cell(row_idx, self.cr.col('AVRdate_solde')).value,
                })

            # Display accounts (simplifié — pas de sous-comptes)
            self.display_accounts = []
            ws_ctrl_v = wb_values[SHEET_CONTROLES] if SHEET_CONTROLES in wb_values.sheetnames else None
            for acct in self.accounts_data:
                ctrl_row = None
                controle = True
                if ws_ctrl_v:
                    ctrl_data_start = self._start_ctrl1 + 1
                    for crow in range(ctrl_data_start, self._end_ctrl1 + 5):
                        cv = ws_ctrl_v.cell(crow, self.cr.col('CTRL1compte')).value
                        if cv and str(cv).strip() == acct['intitule']:
                            ctrl_row = crow
                            cv_controle = ws_ctrl_v.cell(crow, self.cr.col('CTRL1controle')).value
                            controle = (str(cv_controle or 'Oui').strip().lower() == 'oui')
                            break
                self.display_accounts.append({
                    'intitule': acct['intitule'],
                    'devise': acct.get('devise', ''),
                    'controle': controle,
                    'ctrl_row': ctrl_row,
                    'type': acct.get('type', ''),
                    'avoirs_ref': acct,
                })

            # Budget + CTRL2 : fonctions du module neutre inc_compta_schema
            _schema.load_budget_categories(self, wb_values)
            # pv_titles : nécessaire pour purge_account
            _schema.load_pv_titles(self, wb_values)

        finally:
            wb_formula.close()
            wb_values.close()

    def add_account(self, intitule, acct_type, devise='EUR', sous_type='',
                    domiciliation='', titulaire='', propriete='', site='N/A',
                    date_anter=None, montant_anter=None,
                    date_debut=None, montant_debut=None, equiv_euro_debut=None,
                    date_solde=None, controle=True, doc=None):
        """Ajoute un compte et sauvegarde."""
        existing = {a['intitule'] for a in self.accounts_data}
        if intitule in existing:
            print(f"ERREUR: compte '{intitule}' existe déjà")
            return False

        all_types = self.ACCOUNT_TYPES + ['Biens matériels']
        if acct_type not in all_types:
            print(f"ERREUR: type '{acct_type}' invalide. Choix: {all_types}")
            return False

        new_acct = {
            'row': None,
            '_is_new': True,
            'intitule': intitule,
            'type': acct_type,
            'sous_type': sous_type or ('Euro' if devise == 'EUR' else ''),
            'domiciliation': domiciliation,
            'devise': devise,
            'titulaire': titulaire,
            'propriete': propriete,
            'date_anter': date_anter,
            'montant_anter': montant_anter,
            'date_debut': date_debut,
            'montant_debut': montant_debut,
            'equiv_euro_debut': equiv_euro_debut,
            'date_solde': date_solde,
            'site': site,
        }
        self.accounts_data.append(new_acct)
        self.display_accounts.append({
            'intitule': intitule,
            'devise': devise,
            'controle': controle,
            'ctrl_row': None,
            'type': acct_type,
            'avoirs_ref': new_acct,
        })

        print(f"Compte ajouté: {intitule} ({acct_type}, {devise})")
        self._save_and_reload(doc=doc)
        return True

    def _save_and_reload(self, doc=None):
        """Sauvegarde via UNO puis recharge les données.

        En mode batch (doc fourni) : skip — les comptes sont accumulés en mémoire
        et sauvés en une seule fois à la sortie du batch.
        """
        if doc is not None:
            return  # batch : différer le save
        self._save_accounts()
        self._load_excel_data()

    def _uno_finalize(self, doc):
        """Finalise le document UNO (calcul + sauvegarde)."""
        doc.calculate_all()
        doc.save()

    def batch(self):
        """Context manager pour grouper les opérations UNO en une seule session.

        Usage:
            with gui.batch() as doc:
                gui.add_devise('USD', 'fiat', doc=doc)
                gui.add_account(..., doc=doc)
                gui.add_title(..., doc=doc)
            # calculate + save + reload à la sortie
        """
        return _BatchContext(self)

    def __enter__(self):
        """No-op : permet l'usage `with HeadlessGUI(path) as gui:` symétrique
        avec DaemonGUI dans les TNR (qui sélectionnent un backend via flag)."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        return False

    def list_accounts(self):
        """Affiche les comptes."""
        if not self.accounts_data:
            print("Aucun compte")
            return
        print(f"{'Intitulé':<30} {'Type':<20} {'Devise':<6} {'Row':>4}")
        print("-" * 64)
        for a in self.accounts_data:
            print(f"{a['intitule']:<30} {a['type']:<20} {a['devise']:<6} {a.get('row', '?'):>4}")
        total = self._accounts_total_row
        print(f"\nTotal: {len(self.accounts_data)} comptes, Total row: {total}")

    def add_title(self, account_name, title_name, devise=None, date_init=None, doc=None):
        """Ajoute un titre à un portefeuille."""
        acct = None
        for a in self.accounts_data:
            if a['intitule'] == account_name:
                acct = a
                break
        if not acct:
            print(f"ERREUR: compte '{account_name}' introuvable")
            return False
        if acct['type'] != 'Portefeuilles':
            print(f"ERREUR: '{account_name}' n'est pas un Portefeuille (type={acct['type']})")
            return False

        devise = devise or acct.get('devise') or 'EUR'
        if date_init:
            from datetime import datetime as dt
            date_init = dt.strptime(date_init, '%d/%m/%Y')

        self._insert_pv_title(account_name, title_name, devise, date_init, doc=doc)
        print(f"Titre ajouté: *{title_name}* dans {account_name} ({devise})")
        if doc is None:
            self._load_excel_data()
        return True

    def add_devise(self, code, famille, nom=None,
                    derived_from=None, formula=None, decimals=2, doc=None):
        """Ajoute une devise et sauvegarde."""
        if code in self.cotations_meta:
            print(f"ERREUR: devise '{code}' déjà présente")
            return False
        if famille not in DEVISE_SOURCES:
            print(f"ERREUR: famille '{famille}' invalide. Choix: {list(DEVISE_SOURCES.keys())}")
            return False
        self._save_devise(code, famille, nom=nom,
                          derived_from=derived_from, formula=formula,
                          decimals=decimals, doc=doc)
        print(f"Devise ajoutée: {code} ({famille})")
        if doc is None:
            self._load_excel_data()
        return True

    def add_category(self, name, poste='Divers', alloc_pct=100.0, doc=None):
        """Ajoute une catégorie dans la feuille Budget.

        Args:
            alloc_pct: pourcentage 0-100 (converti en décimal 0-1 pour le worker).
        """
        return self._add_category(name, poste=poste,
                                  alloc_pct=alloc_pct / 100.0, doc=doc)

    def add_poste(self, name, fixe=True, doc=None):
        """Ajoute un poste dans la feuille Budget."""
        return self._add_poste(name, fixe=fixe, doc=doc)

    def rename_category(self, old_name, new_name, doc=None):
        """Renomme une catégorie dans Budget + Opérations."""
        self._rename_budget_category(old_name, new_name, doc=doc)
        # Mémoire mise à jour normalement par _after_budget_cat_modify côté GUI ;
        # ici (HeadlessGUI), on ajuste manuellement pour cohérence batch.
        if old_name in self.budget_categories:
            idx = self.budget_categories.index(old_name)
            self.budget_categories[idx] = new_name
            self.budget_cat_rows[new_name] = self.budget_cat_rows.pop(old_name)
        print(f"Catégorie renommée: {old_name} → {new_name}")
        return True

    def set_category_poste(self, name, poste, doc=None):
        """Change le poste de rattachement d'une catégorie (cellule CATposte)."""
        self._set_category_poste(name, poste, doc=doc)
        print(f"Catégorie '{name}' rattachée au poste '{poste}'")
        return True

    def recategorize(self, doc=None):
        """Re-catégorise les opérations « - » via les patterns courants.

        Retourne le nombre d'opérations recatégorisées."""
        n = self._recategorize_operations(doc=doc)
        print(f"{n} opération(s) recatégorisée(s)")
        return n

    def update_poste(self, old_name, new_name, new_type, doc=None):
        """Modifie un poste budgétaire (renommage et/ou type)."""
        self._update_budget_post(old_name, new_name, new_type, doc=doc)
        if old_name in self.budget_posts:
            idx = self.budget_posts.index(old_name)
            self.budget_posts[idx] = new_name
            if new_name != old_name:
                self.budget_post_rows[new_name] = self.budget_post_rows.pop(old_name)
            self.budget_post_types[new_name] = new_type
            if new_name != old_name and old_name in self.budget_post_types:
                del self.budget_post_types[old_name]
        print(f"Poste modifié: {old_name} → {new_name} ({new_type})")
        return True

    def purge_account(self, intitule, doc=None):
        """Purge un compte : supprime ses opérations et titres, garde la structure.

        Pas d'effet sur les biens matériels (pas d'opérations).
        """
        acct = next((a for a in self.accounts_data if a['intitule'] == intitule), None)
        if not acct:
            print(f"ERREUR: compte '{intitule}' introuvable")
            return False
        if acct.get('type') == 'Biens matériels':
            return True  # rien à purger
        titles = getattr(self, 'pv_titles', {}).get(intitule, [])
        self._purge_account_uno(intitule, titles, doc=doc)
        if doc is None:
            self._load_excel_data()
        suffix = f" ({len(titles)} titres supprimés)" if acct.get('type') == 'Portefeuilles' else ""
        print(f"Compte purgé: {intitule}{suffix}")
        return True

    def delete_account(self, intitule, doc=None):
        """Supprime un compte (Avoirs, Contrôles, PVL, opérations résiduelles)."""
        entry = next((e for e in self.display_accounts if e['intitule'] == intitule), None)
        if not entry:
            print(f"ERREUR: compte '{intitule}' introuvable")
            return False
        self.display_accounts.remove(entry)
        avoirs_ref = entry.get('avoirs_ref')
        if avoirs_ref and avoirs_ref in self.accounts_data:
            self.accounts_data.remove(avoirs_ref)
        self._deleted_accounts.append(intitule)
        if entry.get('ctrl_row') is not None:
            self._deleted_ctrl_rows.append(entry['ctrl_row'])
        self._save_accounts(doc=doc)
        if doc is None:
            self._load_excel_data()
        label = 'Bien matériel' if entry.get('type') == 'Biens matériels' else 'Compte'
        print(f"{label} supprimé: {intitule}")
        return True

    def modify_account(self, intitule, fields, doc=None):
        """Modifie les champs d'un compte existant.

        Args:
            intitule: nom du compte à modifier.
            fields: dict {champ: nouvelle_valeur} pour les champs éditables
                (type, sous_type, domiciliation, devise, titulaire, propriete, site).
        """
        acct = next((a for a in self.accounts_data if a['intitule'] == intitule), None)
        if not acct:
            print(f"ERREUR: compte '{intitule}' introuvable")
            return False
        acct.update(fields)
        entry = next((e for e in self.display_accounts if e['intitule'] == intitule), None)
        if entry:
            for k in ('type', 'devise', 'site'):
                if k in fields:
                    entry[k] = fields[k]
        self._save_accounts(doc=doc)
        if doc is None:
            self._load_excel_data()
        print(f"Compte modifié: {intitule} ({', '.join(fields.keys())})")
        return True

    def rename_account(self, old_intitule, new_intitule, doc=None):
        """Renomme un compte (Avoirs + Opérations + Plus_value)."""
        existing = {a['intitule'] for a in self.accounts_data}
        if old_intitule not in existing:
            print(f"ERREUR: compte '{old_intitule}' introuvable")
            return False
        if new_intitule in existing:
            print(f"ERREUR: compte '{new_intitule}' déjà présent")
            return False
        for acct in self.accounts_data:
            if acct['intitule'] == old_intitule:
                acct['intitule'] = new_intitule
                break
        for entry in self.display_accounts:
            if entry['intitule'] == old_intitule:
                entry['intitule'] = new_intitule
                break
        self._rename_account(old_intitule, new_intitule, doc=doc)
        if doc is None:
            self._load_excel_data()
        print(f"Compte renommé: {old_intitule} → {new_intitule}")
        return True

    def rename_pv_title(self, account_name, old_title, new_title, pv_row, doc=None):
        """Renomme un titre dans Plus_value."""
        self._rename_pv_title(account_name, old_title, new_title, pv_row, doc=doc)
        if doc is None:
            self._load_excel_data()
        print(f"Titre renommé: {account_name}/{old_title} → {new_title}")
        return True

    def delete_pv_title(self, account_name, title_name, pv_row, doc=None):
        """Supprime un titre dans Plus_value."""
        self._delete_pv_title(account_name, title_name, pv_row, doc=doc)
        if doc is None:
            self._load_excel_data()
        print(f"Titre supprimé: {account_name}/{title_name}")
        return True

    def delete_devise(self, code, doc=None):
        """Supprime une devise non-EUR."""
        self._delete_devise(code, doc=doc)
        if doc is None:
            self._load_excel_data()
        print(f"Devise supprimée: {code}")
        return True

    def delete_category(self, name, reassign_to=None, doc=None):
        """Supprime une catégorie du Budget."""
        return self._delete_category(name, reassign_to=reassign_to, doc=doc)

    def delete_poste(self, name, doc=None):
        """Supprime un poste budgétaire du Budget."""
        return self._delete_poste(name, doc=doc)

    def cleanup_patrimoine(self, keep_values=None, doc=None):
        """Supprime les lignes Patrimoine non conservées."""
        self._cleanup_patrimoine(keep_values=keep_values, doc=doc)

    def check(self):
        """Lance tool_check_integrity."""
        from inc_uno import UnoDocument
        from inc_check_integrity import IntegrityChecker
        with UnoDocument(self.xlsx_path, read_only=True) as doc:
            doc.calculate_all()
            checker = IntegrityChecker(doc.document)
            checker.run_all()
            return checker.report()


class _BatchContext:
    """Context manager pour grouper les opérations UNO en une seule session."""

    def __init__(self, gui):
        self.gui = gui
        self.doc = None

    def __enter__(self):
        from inc_uno import UnoDocument
        from inc_check_integrity import validate_structure
        self._uno = UnoDocument(self.gui.xlsx_path)
        self.doc = self._uno.__enter__()
        # Validation structurelle avant toute modification
        ok, errors, warnings = validate_structure(self.doc.document)
        for w in warnings:
            print(f"  ⚠ {w}")
        if not ok:
            for e in errors:
                print(f"  ❌ {e}")
            self._uno.__exit__(None, None, None)
            raise RuntimeError("Structure xlsm invalide — batch annulé")
        return self.doc

    def __exit__(self, *exc_info):
        if self.doc and not exc_info[0]:
            # _save_accounts n'est pas idempotent : ne l'appeler que si l'état
            # comptes a effectivement été modifié — sinon les opérations
            # purement Cotations/Budget polluent CTRL1 (réécriture controle,
            # ré-insertion lignes fantômes).
            g = self.gui
            accounts_dirty = (
                any(a.get('_is_new') for a in g.accounts_data)
                or g._deleted_accounts
                or g._soft_deleted_accounts
                or g._deleted_ctrl_rows
            )
            if accounts_dirty:
                g._save_accounts(doc=self.doc)
            self.gui._uno_finalize(self.doc)
        self._uno.__exit__(*exc_info)
        self.gui._load_excel_data()


# ============================================================================
# SCÉNARIO TEST TEMPLATE
# ============================================================================

# Devises et cotations à ajouter
# (code, famille, nom_long, derived_from, formula)
TEST_COTATIONS_SPOT = [
    ('XAU', 'metal', "Gramme d'or Spot", None, None),
    ('BTC', 'crypto', 'Bitcoin', None, None),
]
TEST_DEVISES = [
    ('USD', 'fiat', 'Dollar US', None, None),
    ('SGD', 'fiat', 'Dollar Singapour', None, None),
    ('OrPr', 'metal', "Gramme d'or Premium (pièces)", 'XAU', '*1.043'),
    ('SAT', 'crypto', 'Satoshi (1 / 100 000 000 Bitcoin)', 'BTC', '/100000000'),
    ('XMR', 'crypto', 'Monero', None, None),
]

# Comptes à créer (intitulé, type, devise, domiciliation, sous_type, titulaire, propriété)
_D = datetime
TEST_ACCOUNTS = [
    # (intitulé, type, devise, dom, sous_type, titulaire, propriété, date_anter, montant_anter)
    # Les #Solde initiaux (2020) sont créés à montant=0 par _append_solde_lines.
    # Les vrais #Solde et opérations viennent de l'import (ref_man).
    ('Assurance vie Alice', 'Portefeuilles', 'EUR', 'Société Générale', 'Titres', 'Alice', 'non', _D(2011, 5, 1), 210000),
    ('PEE Alice', 'Portefeuilles', 'EUR', 'Natixis', 'Titres', 'Alice', 'non', None, None),
    ('Portefeuille BB Titres', 'Portefeuilles', 'EUR', 'BoursoBank', 'Titres', 'Barnabé', 'non', _D(2024, 7, 12), None),
    ('Portefeuille DEGIRO Titres', 'Portefeuilles', 'EUR', 'DEGIRO', 'Titres', 'Barnabé', 'non', _D(2024, 11, 21), None),
    ('Portefeuille eToro USD Titres', 'Portefeuilles', 'USD', 'eToro', 'Titres', 'Barnabé', 'non', _D(2024, 11, 26), None),
    ('Compte Amazon', 'Euros', 'EUR', 'Amazon', 'Euro', 'Barnabé', 'non', None, None),
    ('Compte chèque BB', 'Euros', 'EUR', 'BoursoBank', 'Euro', 'Barnabé', 'non', _D(2024, 1, 1), None),
    ('Compte chèque commun', 'Euros', 'EUR', 'Société Générale', 'Euro', 'Commun', 'non', _D(1977, 1, 1), None),
    ('Espèces', 'Euros', 'EUR', 'Maison', 'Euro', 'Commun', 'oui', None, None),
    ('Livret A Barnabé', 'Euros', 'EUR', 'Société Générale', 'Euro', 'Barnabé', 'non', None, None),
    ('Compte Paypal', 'Euros', 'EUR', 'Paypal', 'Euro', 'Barnabé', 'non', None, None),
    ('Compte Wise EUR', 'Euros', 'EUR', 'Wise', 'Euro', 'Barnabé', 'non', _D(2024, 12, 1), None),
    # Transfert de fonds : pas un compte, calcul de déséquilibre — exclu
    ('Portefeuille DEGIRO Réserve', 'Euros', 'EUR', 'DEGIRO', 'Euro', 'Barnabé', 'non', None, None),
    ('Portefeuille BB Réserve', 'Euros', 'EUR', 'BoursoBank', 'Euro', 'Barnabé', 'non', None, None),
    # Compte Wise CHF : pas dans CTRL1 de la ref (aucune opération) — exclu
    ('Compte Wise USD', 'Devises étrangères', 'USD', 'Wise', 'Dollar US', 'Barnabé', 'non', _D(2024, 12, 27), None),
    ('Compte Wise SGD', 'Devises étrangères', 'SGD', 'Wise', 'Dollar Singapour', 'Barnabé', 'non', _D(2024, 12, 5), None),
    ('Portefeuille eToro USD Réserve', 'Devises étrangères', 'USD', 'eToro', 'Dollar US', 'Barnabé', 'non', None, None),
    ('Créance Francine', 'Créances', 'EUR', 'Francine', 'Euro', 'Commun', 'non', _D(2017, 5, 18), 156270),
    ('BlueWallet BTC', 'Crypto monnaies', 'SAT', 'Blockchain', 'Bitcoin', 'Commun', 'oui', _D(2023, 2, 10), None),
    ('Phoenix Lightning BTC', 'Crypto monnaies', 'SAT', 'Blockchain', 'Bitcoin', 'Commun', 'oui', _D(2023, 2, 10), None),
    ('Compte Kraken BTC', 'Crypto monnaies', 'SAT', 'Kraken', 'Bitcoin', 'Commun', 'non', _D(2025, 7, 17), None),
    ('Cake Wallet XMR', 'Crypto monnaies', 'XMR', 'Blockchain', 'Monero', 'Commun', 'oui', _D(2023, 3, 14), None),
    ('Pièces or', 'Biens matériels', 'OrPr', 'Coffre', 'Mobilier', 'Commun', 'oui', _D(2023, 5, 2), None),
]

# Titres à ajouter aux portefeuilles
TEST_TITLES = [
    ('Portefeuille eToro USD Titres', 'AI.PA', 'USD'),
    ('Portefeuille eToro USD Titres', 'SPX500', 'USD'),
    ('Portefeuille eToro USD Titres', 'BTC/USD', 'USD'),
    ('Portefeuille eToro USD Titres', 'COMT', 'USD'),
    ('Portefeuille eToro USD Titres', 'KWEB', 'USD'),
]


def run_build_example(xlsm_path, source_xlsm=None, reference=None, with_ops=False):
    """Exécute le scénario de construction complet.

    1. Génère un template vierge depuis source_xlsm
    2. Ajoute devises, comptes, titres via CLI headless
    3. Vérifie l'intégrité
    4. Compare vs référence si fournie
    """
    import subprocess
    import time

    xlsm_path = Path(xlsm_path).absolute()
    base = xlsm_path.parent
    errors = 0
    t0 = time.time()

    # --- Étape 1 : template vierge ---
    if source_xlsm:
        source = Path(source_xlsm).absolute()
        if source.resolve() != xlsm_path.resolve():
            shutil.copy2(source, xlsm_path)
        print(f"[1/5] Template copié depuis {source.name}")
    else:
        print("[1/5] Utilisation du xlsm existant comme template")

    # --- Étape 2 : devises ---
    print(f"[2/5] Ajout de {len(TEST_DEVISES)} devises...")
    gui = HeadlessGUI(xlsm_path)
    for code, famille, nom, df, fm in TEST_DEVISES:
        ok = gui.add_devise(code, famille, nom=nom, derived_from=df, formula=fm)
        if not ok:
            print(f"  ✗ Échec add-devise {code}")
            errors += 1

    # --- Étape 3 : comptes ---
    print(f"[3/5] Ajout de {len(TEST_ACCOUNTS)} comptes...")
    gui = HeadlessGUI(xlsm_path)  # recharger après devises
    for intitule, acct_type, devise, dom, st, tit, prop, da, ma in TEST_ACCOUNTS:
        ctrl = not (acct_type == 'Portefeuilles' and st == 'Titres')
        ok = gui.add_account(intitule, acct_type, devise=devise,
                             sous_type=st, domiciliation=dom,
                             titulaire=tit, propriete=prop,
                             date_anter=da, montant_anter=ma,
                             controle=ctrl)
        if not ok:
            print(f"  ✗ Échec add-account {intitule}")
            errors += 1

    # --- Étape 4 : titres ---
    print(f"[4/5] Ajout de {len(TEST_TITLES)} titres...")
    gui = HeadlessGUI(xlsm_path)  # recharger après comptes
    for compte, titre, devise in TEST_TITLES:
        ok = gui.add_title(compte, titre, devise=devise)
        if not ok:
            print(f"  ✗ Échec add-title {titre} dans {compte}")
            errors += 1

    # --- Étape 5 : vérifications ---
    print("[5/5] Vérification intégrité...")
    gui = HeadlessGUI(xlsm_path)
    check_ok = gui.check()
    if not check_ok:
        errors += 1

    # --- Comparaison vs référence ---
    if reference:
        ref = Path(reference).absolute()
        if ref.exists():
            print(f"\nComparaison vs {ref.name}...")
            result = subprocess.run(
                ['python3', str(Path(__file__).parent / 'tool_compare_xlsx.py'),
                 str(xlsm_path), str(ref)],
                capture_output=True, text=True)
            print(result.stdout)
            if result.returncode != 0:
                errors += 1
        else:
            print(f"\n⚠ Référence introuvable : {ref}")

    # --- Rapport ---
    elapsed = time.time() - t0
    print(f"\n{'='*50}")
    if errors == 0:
        print(f"✓ Test template PASSED ({elapsed:.0f}s)")
    else:
        print(f"✗ Test template FAILED — {errors} erreur(s) ({elapsed:.0f}s)")
    print(f"{'='*50}")

    return errors == 0


def _daemon_loop(gui):
    """Mode daemon : lit des requêtes JSON sur stdin, dispatche aux méthodes
    publiques de HeadlessGUI, répond sur stdout en JSON line-based.

    Protocole :
      - Au démarrage : print `{"event": "ready"}` sur stdout (flush)
      - Requête (stdin)  : `{"method": "<name>", "kwargs": {...}}`
      - Réponse (stdout) : `{"ok": true, "result": ..., "stdout": "..."}` ou
                           `{"ok": false, "error": "...", "trace": "...",
                             "stdout": "..."}`
      - Commande spéciale : `{"method": "__quit__"}` → save + exit

    Le stdout interne des méthodes (prints d'info/erreur) est capturé puis
    renvoyé dans la clé `stdout` de la réponse JSON — il ne peut pas se
    mélanger aux réponses JSON sur le canal stdout du processus.

    Un batch (`_BatchContext`) est ouvert lazy au 1er appel UNO et maintenu
    pour amortir le cold start soffice (~6s sur Mac). __quit__ ferme le batch
    proprement (save + cleanup).
    """
    import json
    import os
    import io
    import contextlib
    import inspect
    import signal
    import socket
    import subprocess as _sp
    import traceback

    # Préemption orphelin : si un soffice traîne sur 2002 (parent mort
    # d'une session antérieure), il intercepterait notre UnoDocument et
    # bloquerait aussi tout 'open -a LibreOffice' interactif (single-instance
    # sur Mac). On le tue avant de continuer.
    def _kill_orphan_soffice():
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            if s.connect_ex(('localhost', 2002)) != 0:
                return  # port libre, rien à faire
        # Port occupé : trouver les soffice headless et les tuer
        try:
            out = _sp.run(['pgrep', '-f', 'soffice.*--headless'],
                          capture_output=True, text=True, timeout=5)
        except (FileNotFoundError, _sp.TimeoutExpired):
            return
        for pid_str in out.stdout.strip().split('\n'):
            if not pid_str.strip():
                continue
            try:
                pid = int(pid_str.strip())
                os.kill(pid, signal.SIGTERM)
            except (ValueError, ProcessLookupError, PermissionError):
                pass
        # Laisser au noyau le temps de libérer le port
        import time as _time
        for _ in range(10):
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                if s.connect_ex(('localhost', 2002)) != 0:
                    return
            _time.sleep(0.3)

    _kill_orphan_soffice()

    batch_ctx = None
    batch_doc = None

    def _close_batch():
        nonlocal batch_ctx, batch_doc
        buf = io.StringIO()
        if batch_ctx is not None:
            with contextlib.redirect_stdout(buf):
                batch_ctx.__exit__(None, None, None)
            batch_ctx = None
            batch_doc = None
        return buf.getvalue()

    # Cleanup sur SIGTERM/SIGINT/SIGHUP (kill du daemon par la GUI ou parent-death
    # Mac) : déclenche _close_batch pour éviter d'orphaner soffice et de perdre
    # les modifs UNO en mémoire. Toute exception remonte sur stderr (qui est
    # redirigé vers logs/daemon.err par le client GUI) pour diagnostic.
    def _on_signal(signum, frame):
        try:
            _close_batch()
        except Exception:
            print(f"=== EXCEPTION _on_signal (signum={signum}) → _close_batch ===",
                  file=sys.stderr, flush=True)
            traceback.print_exc(file=sys.stderr)
            sys.stderr.flush()
        sys.exit(0)

    signal.signal(signal.SIGTERM, _on_signal)
    signal.signal(signal.SIGINT, _on_signal)
    # SIGHUP : Mac envoie ce signal au daemon enfant quand le parent meurt
    # (ex: Cmd+Q sur la GUI Tk qui n'a pas invoqué WM_DELETE_WINDOW).
    try:
        signal.signal(signal.SIGHUP, _on_signal)
    except Exception:
        pass

    # Logger : route les exceptions attrapées vers logs/journal.log au lieu
    # de stderr → daemon.err. daemon.err reste réservé aux crashes Python
    # fatals non interceptables (segfault, OOM, etc.).
    from inc_logging import Logger
    journal_file = Path(__file__).parent / 'logs' / 'journal.log'
    journal_file.parent.mkdir(parents=True, exist_ok=True)
    logger = Logger(script_name="daemon", journal_file=journal_file)

    print(json.dumps({'event': 'ready', 'pid': os.getpid()}), flush=True)

    for line in sys.stdin:
        line = line.strip()
        if not line:
            continue
        captured = io.StringIO()
        try:
            req = json.loads(line)
            method_name = req['method']

            if method_name == '__quit__':
                # Trace toute exception de _close_batch sur stderr (DEVNULL
                # ferait perdre le diagnostic — voir logs/daemon.err).
                try:
                    close_out = _close_batch()
                except Exception:
                    print("=== EXCEPTION __quit__ → _close_batch ===",
                          file=sys.stderr, flush=True)
                    traceback.print_exc(file=sys.stderr)
                    sys.stderr.flush()
                    raise
                print(json.dumps({'ok': True, 'result': 'bye',
                                  'stdout': close_out}), flush=True)
                break

            if method_name == '__flush__':
                # Sauvegarde + ferme le batch pour relâcher le lock file.
                # Trace les exceptions sur stderr (cf. __quit__).
                try:
                    close_out = _close_batch()
                except Exception:
                    print("=== EXCEPTION __flush__ → _close_batch ===",
                          file=sys.stderr, flush=True)
                    traceback.print_exc(file=sys.stderr)
                    sys.stderr.flush()
                    raise
                print(json.dumps({'ok': True, 'result': 'flushed',
                                  'stdout': close_out}), flush=True)
                continue

            kwargs = req.get('kwargs', {})

            with contextlib.redirect_stdout(captured):
                # Ouvre le batch au 1er appel UNO (= toute méthode autre que les
                # commandes spéciales)
                if batch_ctx is None:
                    batch_ctx = gui.batch()
                    batch_doc = batch_ctx.__enter__()

                method = getattr(gui, method_name, None)
                if method is None:
                    raise AttributeError(
                        f"HeadlessGUI n'a pas de méthode '{method_name}'")

                # Inject doc si la méthode l'accepte et qu'il n'est pas déjà fourni
                sig = inspect.signature(method)
                if 'doc' in sig.parameters and 'doc' not in kwargs:
                    kwargs['doc'] = batch_doc

                result = method(**kwargs)

            # Sérialiser : types JSON-natifs OK, sinon str() forcé
            if isinstance(result, (bool, int, float, str, list, dict, type(None))):
                serializable = result
            else:
                serializable = str(result)
            print(json.dumps({'ok': True, 'result': serializable,
                              'stdout': captured.getvalue()}), flush=True)
        except Exception as e:
            # Log dans journal.log uniquement (PAS via Logger.warning/.error
            # qui écrivent aussi stdout/stderr — stdout est le canal JSON-RPC
            # du daemon, le polluer casse le parsing client).
            ts = datetime.now().strftime('%H:%M:%S')
            method_name_for_log = req.get('method', '?')
            if isinstance(e, _schema.BusinessError):
                # Refus métier : ⚠️, pas de stack — déjà remonté par JSON-RPC
                # et affiché à l'utilisateur en popup.
                logger.write_to_journal(
                    f"{ts} daemon ⚠️ {method_name_for_log}: {e}")
            else:
                # Vraie panne : ❌ + traceback brut au journal.
                logger.write_to_journal(
                    f"{ts} daemon ❌ {method_name_for_log}: {e}")
                if captured.getvalue():
                    logger.write_to_journal(
                        f"  captured stdout: {captured.getvalue()!r}")
                logger.write_to_journal(traceback.format_exc())
            err = {'ok': False, 'error': str(e),
                   'trace': traceback.format_exc(),
                   'stdout': captured.getvalue()}
            print(json.dumps(err), flush=True)


def main():
    parser = argparse.ArgumentParser(
        description="Actions GUI en mode CLI (headless)")
    parser.add_argument('xlsm', type=Path, help='Fichier comptes.xlsm')
    sub = parser.add_subparsers(dest='action', required=True)

    # add-account
    p_add = sub.add_parser('add-account', help='Ajouter un compte')
    p_add.add_argument('nom', help='Nom du compte')
    p_add.add_argument('type', help='Type (Euros, Portefeuilles, ...)')
    p_add.add_argument('--devise', default='EUR', help='Devise (défaut: EUR)')
    p_add.add_argument('--sous-type', default='', help='Sous-type')
    p_add.add_argument('--domiciliation', default='', help='Domiciliation')
    p_add.add_argument('--titulaire', default='', help='Titulaire')
    p_add.add_argument('--propriete', default='', help='Propriété')
    p_add.add_argument('--site', default='N/A', help='Site (défaut: N/A)')

    # add-bien
    p_bien = sub.add_parser('add-bien', help='Ajouter un bien matériel')
    p_bien.add_argument('nom', help='Nom du bien')
    p_bien.add_argument('nature', choices=['Foncier', 'Mobilier'], help='Foncier ou Mobilier')
    p_bien.add_argument('--devise', default='', help='Devise cotée (vide pour immobilier)')
    p_bien.add_argument('--domiciliation', default='', help='Domiciliation')
    p_bien.add_argument('--titulaire', default='', help='Titulaire')
    p_bien.add_argument('--propriete', default='', help='Propriété')
    p_bien.add_argument('--montant', type=float, default=None, help='Montant initial')

    # add-devise
    p_dev = sub.add_parser('add-devise', help='Ajouter une devise')
    p_dev.add_argument('code', help='Code devise (USD, CHF, ...)')
    p_dev.add_argument('famille', help='Famille (fiat, metal, crypto)')

    # add-title
    p_title = sub.add_parser('add-title', help='Ajouter un titre à un portefeuille')
    p_title.add_argument('compte', help='Nom du portefeuille')
    p_title.add_argument('titre', help='Nom du titre')
    p_title.add_argument('--devise', default=None, help='Devise du titre (défaut: devise du compte)')
    p_title.add_argument('--date', default=None, help='Date initiale JJ/MM/AAAA')

    # list-accounts
    sub.add_parser('list-accounts', help='Lister les comptes')

    # check
    sub.add_parser('check', help='Vérifier l\'intégrité')

    # daemon : mode interactif piloté par JSON sur stdin
    sub.add_parser('daemon',
                   help='Mode daemon : lit des requêtes JSON sur stdin, '
                        'dispatche aux méthodes HeadlessGUI, répond sur stdout')

    # build-example
    p_test = sub.add_parser('build-example', help='Scénario de construction complet sur template vierge')
    p_test.add_argument('--source', default=None,
                        help='Fichier source pour générer le template (défaut: comptes.xlsm DEV)')
    p_test.add_argument('--reference', default=None,
                        help='Fichier référence pour comparaison (défaut: comptes_exemple.xlsm)')
    p_test.add_argument('--with-ops', action='store_true',
                        help='Importer des opérations via cpt_update MANUEL')

    args = parser.parse_args()

    _env_ok, _env_msg = check_env()
    if not _env_ok:
        print(f"⚠️  {_env_msg}")

    if args.action == 'build-example':
        source = args.source or str(Path(__file__).parent / 'comptes.xlsm')
        reference = args.reference
        ok = run_build_example(args.xlsm, source_xlsm=source,
                               reference=reference, with_ops=args.with_ops)
        sys.exit(0 if ok else 1)

    gui = HeadlessGUI(args.xlsm)

    if args.action == 'add-account':
        ok = gui.add_account(
            args.nom, args.type,
            devise=args.devise,
            sous_type=getattr(args, 'sous_type', ''),
            domiciliation=args.domiciliation,
            titulaire=args.titulaire,
            propriete=args.propriete,
            site=args.site,
        )
        sys.exit(0 if ok else 1)

    elif args.action == 'add-bien':
        ok = gui.add_account(
            args.nom, 'Biens matériels',
            devise=args.devise,
            sous_type=args.nature,
            domiciliation=args.domiciliation,
            titulaire=args.titulaire,
            propriete=args.propriete,
            controle=False,
        )
        sys.exit(0 if ok else 1)

    elif args.action == 'add-devise':
        ok = gui.add_devise(args.code, args.famille)
        sys.exit(0 if ok else 1)

    elif args.action == 'add-title':
        ok = gui.add_title(args.compte, args.titre,
                           devise=args.devise, date_init=args.date)
        sys.exit(0 if ok else 1)

    elif args.action == 'list-accounts':
        gui.list_accounts()

    elif args.action == 'check':
        ok = gui.check()
        sys.exit(0 if ok else 1)

    elif args.action == 'daemon':
        _daemon_loop(gui)


if __name__ == '__main__':
    main()
