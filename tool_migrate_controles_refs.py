#!/usr/bin/env python3
"""
Migration one-shot : convertit les noms de comptes en dur dans la feuille
Contrôles en formules =Avoirs!A<row>.

Ainsi, un renommage dans Avoirs se propage automatiquement à Contrôles.

Usage:
    python3 tool_migrate_controles_refs.py [comptes.xlsx]
    python3 tool_migrate_controles_refs.py --dry-run  # affiche sans modifier
"""

import shutil
import sys
from pathlib import Path

import openpyxl


def build_avoirs_map(ws_avoirs):
    """Construit le dict nom_compte → numéro_de_ligne dans Avoirs."""
    avoirs_map = {}
    for row in range(4, 201):
        val = ws_avoirs.cell(row, 1).value
        if val and 'total' in str(val).lower():
            break
        if val:
            avoirs_map[str(val).strip()] = row
    return avoirs_map


def migrate(xlsx_path, dry_run=False):
    wb = openpyxl.load_workbook(xlsx_path, data_only=False, keep_vba=True)
    ws_avoirs = wb['Avoirs']
    ws_ctrl = wb['Contrôles']

    avoirs_map = build_avoirs_map(ws_avoirs)
    print(f"Avoirs : {len(avoirs_map)} comptes (lignes 4-{max(avoirs_map.values())})")

    converted = 0
    already_formula = 0
    skipped = 0

    for row in range(3, 100):
        cell = ws_ctrl.cell(row, 1)
        val = cell.value
        if not val:
            continue

        name = str(val).strip()

        # Déjà une formule
        if name.startswith('='):
            already_formula += 1
            continue

        # En-tête
        if name in ('Compte', 'Compte '):
            continue

        # Chercher dans Avoirs
        if name in avoirs_map:
            avoirs_row = avoirs_map[name]
            formula = f'=Avoirs!A{avoirs_row}'
            sub = ws_ctrl.cell(row, 2).value
            sub_str = f" ({sub})" if sub else ""
            print(f"  L{row:2d}: {name:35s}{sub_str:12s} → {formula}")
            if not dry_run:
                cell.value = formula
            converted += 1
        else:
            print(f"  L{row:2d}: {name:35s} → INTROUVABLE dans Avoirs (ignoré)")
            skipped += 1

    print(f"\nRésultat : {converted} convertis, {already_formula} déjà formules, {skipped} ignorés")

    if not dry_run and converted > 0:
        # Backup
        bak_path = xlsx_path.with_suffix('.xlsm.bak')
        shutil.copy2(xlsx_path, bak_path)
        print(f"Backup : {bak_path}")

        wb.save(xlsx_path)
        print(f"Sauvegardé : {xlsx_path}")
    elif dry_run:
        print("\n(dry-run, aucune modification)")

    wb.close()
    return converted


if __name__ == '__main__':
    dry_run = '--dry-run' in sys.argv
    args = [a for a in sys.argv[1:] if not a.startswith('--')]

    if args:
        xlsx_path = Path(args[0])
    else:
        xlsx_path = Path(__file__).parent / 'comptes.xlsm'

    if not xlsx_path.exists():
        print(f"Erreur: {xlsx_path} introuvable", file=sys.stderr)
        sys.exit(1)

    print(f"{'[DRY-RUN] ' if dry_run else ''}Migration Contrôles → formules Avoirs")
    print(f"Fichier : {xlsx_path}\n")

    migrate(xlsx_path, dry_run=dry_run)
