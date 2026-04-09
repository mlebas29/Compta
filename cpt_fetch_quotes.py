#!/usr/bin/env python3
"""
cpt_fetch_quotes.py - Récupération des cotations financières

Met à jour l'onglet "Cotations" du fichier Excel avec les cours actuels.
Les métadonnées (type, sources API) sont dans config_cotations.json,
les codes et numéros de ligne sont lus depuis la feuille Cotations.

Usage:
    ./cpt_fetch_quotes.py           # Met à jour toutes les cotations
    ./cpt_fetch_quotes.py --dry-run # Affiche les cours sans modifier Excel
"""

import json
import os
import sys
import requests
from datetime import datetime
import openpyxl
import inc_mode
import inc_exchange_rates
from inc_logging import Logger
from inc_excel_schema import SHEET_COTATIONS, ColResolver

# ============================================================================
# CONFIGURATION
# ============================================================================

BASE_DIR = inc_mode.get_base_dir()
COMPTES_FILE = BASE_DIR / "comptes.xlsm"
COTATIONS_JSON = BASE_DIR / "config_cotations.json"

# Conversion once troy → gramme
OZ_TO_GRAM = 31.1035

# Timeout pour les requêtes API
REQUEST_TIMEOUT = 10

# Mappings API (détails d'implémentation — les codes restent dans Excel)
COINGECKO_IDS = {'BTC': 'bitcoin', 'XMR': 'monero'}
YAHOO_TICKERS = {'XAU': 'GC=F', 'XAG': 'SI=F'}
KRAKEN_PAIRS = {'BTC': 'XBTEUR', 'XMR': 'XMREUR'}

# ============================================================================
# LECTURE CONFIG EXCEL
# ============================================================================

logger = None


def _load_cotations_json():
    """Lit config_cotations.json → dict {code: {famille, source1, source2}}."""
    if not COTATIONS_JSON.exists():
        return {}
    with open(COTATIONS_JSON, 'r', encoding='utf-8') as f:
        return json.load(f)


def load_cotations_config():
    """Lit les métadonnées depuis JSON + numéros de ligne depuis Excel.

    Returns:
        list[dict]: [{'code': 'BTC', 'type': 'crypto', 'source_1': 'coingecko',
                      'source_2': 'kraken', 'row': 5}, ...]
        Seules les lignes avec un code ET une source_1 (dans le JSON) sont retournées.
    """
    meta = _load_cotations_json()

    wb = openpyxl.load_workbook(COMPTES_FILE, data_only=True)
    ws = wb[SHEET_COTATIONS]

    cr = ColResolver.from_openpyxl(wb)
    cot_start, _ = cr.rows('COTcode')
    cot_start = cot_start or 3
    config = []
    excel_codes = set()
    for row in range(cot_start + 1, ws.max_row + 1):
        code = ws.cell(row=row, column=cr.col('COTcode')).value
        if not code:
            continue
        code = str(code).strip()
        excel_codes.add(code)
        info = meta.get(code, {})
        source_1 = info.get('source1', '')
        if not source_1:
            continue
        config.append({
            'code': code,
            'type': info.get('famille', ''),
            'source_1': source_1,
            'source_2': info.get('source2', ''),
            'row': row,
        })

    # Vérification cohérence JSON ↔ Excel
    # Exclure formules (dérivés) et immobilier (clé = label, pas un code court)
    json_codes = {k for k, v in meta.items()
                  if v.get('source1') and v['source1'] != 'formule'
                  and v.get('famille') != 'immobilier'}
    in_json_not_excel = json_codes - excel_codes
    in_excel_not_json = excel_codes - set(meta.keys())
    if in_excel_not_json and logger:
        logger.warning(f"Codes dans Excel sans métadonnées JSON (pas de fetch) : {', '.join(sorted(in_excel_not_json))}")
    if in_json_not_excel and logger:
        logger.warning(f"Codes dans JSON sans ligne Excel (orphelins) : {', '.join(sorted(in_json_not_excel))}")

    wb.close()
    return config


# ============================================================================
# FETCHERS API
# ============================================================================

def _fetch_frankfurter(codes):
    """Récupère les taux de change fiat via Frankfurter (BCE).

    Returns:
        dict: {code: cours_en_eur} pour les codes demandés
    """
    try:
        all_rates = inc_exchange_rates.get_rates_for_date()  # 'latest'
        if all_rates is None:
            logger.error("Erreur Frankfurter (via inc_exchange_rates)")
            return {}

        results = {}
        for code in codes:
            if code in all_rates:
                results[code] = round(1 / all_rates[code], 4)
        return results
    except Exception as e:
        logger.error(f"Erreur Frankfurter: {e}")
        return {}


def _fetch_coingecko(codes):
    """Récupère les prix crypto via CoinGecko.

    Returns:
        dict: {code: prix_eur}
    """
    # Construire la liste des IDs CoinGecko pour les codes demandés
    ids_map = {COINGECKO_IDS[c]: c for c in codes if c in COINGECKO_IDS}
    if not ids_map:
        return {}

    url = "https://api.coingecko.com/api/v3/simple/price"
    params = {
        'ids': ','.join(ids_map.keys()),
        'vs_currencies': 'eur'
    }

    try:
        response = requests.get(url, params=params, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        data = response.json()

        results = {}
        for cg_id, code in ids_map.items():
            if cg_id in data and 'eur' in data[cg_id]:
                results[code] = data[cg_id]['eur']
        return results
    except Exception as e:
        logger.error(f"Erreur CoinGecko: {e}")
        return {}


def _fetch_yahoo(codes):
    """Récupère les prix métaux via Yahoo Finance (USD/oz → EUR/g).

    Returns:
        dict: {code: prix_eur_gramme}
    """
    # Taux USD → EUR
    try:
        forex_url = "https://api.frankfurter.app/latest?from=USD&to=EUR"
        response = requests.get(forex_url, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        usd_to_eur = response.json()['rates']['EUR']
    except Exception as e:
        logger.error(f"Erreur taux USD/EUR: {e}")
        return {}

    results = {}
    for code in codes:
        ticker = YAHOO_TICKERS.get(code)
        if not ticker:
            continue
        try:
            url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}"
            headers = {'User-Agent': 'Mozilla/5.0'}
            response = requests.get(url, headers=headers, timeout=REQUEST_TIMEOUT)
            response.raise_for_status()
            data = response.json()

            price_usd_oz = data['chart']['result'][0]['meta']['regularMarketPrice']
            price_eur_gram = (price_usd_oz * usd_to_eur) / OZ_TO_GRAM
            results[code] = round(price_eur_gram, 2)
        except Exception as e:
            logger.error(f"Erreur Yahoo Finance ({code}): {e}")

    return results


def _fetch_kraken(codes):
    """Récupère les prix crypto via Kraken Ticker API.

    Returns:
        dict: {code: prix_eur}
    """
    pairs_map = {KRAKEN_PAIRS[c]: c for c in codes if c in KRAKEN_PAIRS}
    if not pairs_map:
        return {}

    pair_str = ','.join(pairs_map.keys())
    url = f"https://api.kraken.com/0/public/Ticker?pair={pair_str}"

    try:
        response = requests.get(url, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        data = response.json()

        if data.get('error'):
            logger.error(f"Erreur Kraken API: {data['error']}")
            return {}

        results = {}
        for pair, code in pairs_map.items():
            # Kraken peut renommer les paires (XBTEUR → XXBTZEUR)
            for key, info in data.get('result', {}).items():
                if pair in key or key in pair:
                    # 'c' = last trade closed [price, lot-volume]
                    results[code] = float(info['c'][0])
                    break
        return results
    except Exception as e:
        logger.error(f"Erreur Kraken: {e}")
        return {}


# Registre de dispatch source → fetcher
API_FETCHERS = {
    'frankfurter': _fetch_frankfurter,
    'coingecko': _fetch_coingecko,
    'yahoo': _fetch_yahoo,
    'kraken': _fetch_kraken,
}


# ============================================================================
# ORCHESTRATION
# ============================================================================

def fetch_all_quotes(config):
    """Récupère toutes les cotations en suivant la config Excel.

    Pour chaque source_1, appelle le fetcher groupé.
    En cas d'échec, tente source_2 par code.

    Returns:
        dict: {code: prix_eur} ou None si erreur critique
    """
    quotes = {}
    errors = []

    print("\nRécupération des cotations...")

    # Grouper par source_1 (exclure les cours dérivés calculés par formule Excel)
    by_source = {}
    for item in config:
        if item['source_1'] == 'formule':
            continue
        by_source.setdefault(item['source_1'], []).append(item)

    # Labels affichage par source
    source_labels = {
        'yahoo': 'Métaux précieux (Yahoo Finance)',
        'coingecko': 'Cryptomonnaies (CoinGecko)',
        'frankfurter': 'Devises (Frankfurter/BCE)',
        'kraken': 'Cryptomonnaies (Kraken)',
    }

    for source, items in by_source.items():
        fetcher = API_FETCHERS.get(source)
        if not fetcher:
            logger.error(f"Source inconnue: {source}")
            errors.append(source)
            continue

        codes = [it['code'] for it in items]
        label = source_labels.get(source, source)
        print(f"→ {label}...", end=" ", flush=True)

        result = fetcher(codes)
        if result:
            quotes.update(result)
            print("✓")
        else:
            print("✗")
            errors.append(label)

        # Fallback source_2 pour les codes manquants
        for item in items:
            code = item['code']
            if code not in quotes and item['source_2']:
                fallback_fetcher = API_FETCHERS.get(item['source_2'])
                if fallback_fetcher:
                    print(f"  → Fallback {item['source_2']} pour {code}...", end=" ", flush=True)
                    fallback_result = fallback_fetcher([code])
                    if code in fallback_result:
                        quotes[code] = fallback_result[code]
                        print("✓")
                    else:
                        print("✗")

    if errors:
        missing = [it['code'] for it in config if it['code'] not in quotes]
        if missing:
            logger.error(f"Codes sans cotation: {', '.join(missing)}")
        if not quotes:
            return None

    return quotes


# ============================================================================
# MISE À JOUR EXCEL
# ============================================================================

def update_excel(quotes, config, dry_run=False):
    """Met à jour les cotations dans le fichier Excel.

    Utilise config (avec numéros de ligne) au lieu d'un mapping hardcodé.
    """
    if not COMPTES_FILE.exists():
        logger.error(f"Fichier Excel introuvable: {COMPTES_FILE}")
        return False

    if dry_run:
        print("\n=== MODE DRY-RUN (pas de modification) ===")
        print(f"{'Code':<6} {'Cours':>12} {'Ligne'}")
        print("-" * 30)
        for item in config:
            code = item['code']
            if code in quotes:
                print(f"{code:<6} {quotes[code]:>12.4f} €  → ligne {item['row']}")
            else:
                print(f"{code:<6} {'N/A':>12}    → ligne {item['row']}")
        return True

    try:
        wb = openpyxl.load_workbook(COMPTES_FILE, keep_vba=True)
    except Exception as e:
        logger.error(f"Erreur ouverture Excel: {e}")
        return False

    if SHEET_COTATIONS not in wb.sheetnames:
        logger.error(f"Onglet '{SHEET_COTATIONS}' introuvable dans Excel")
        wb.close()
        return False

    ws = wb[SHEET_COTATIONS]
    today = datetime.now().strftime('%d/%m/%Y')
    updated_count = 0

    for item in config:
        code = item['code']
        if code not in quotes:
            continue

        ws.cell(row=item['row'], column=cr.col('COTcours')).value = quotes[code]
        ws.cell(row=item['row'], column=cr.col('COTdate')).value = today
        updated_count += 1

    if updated_count > 0:
        try:
            wb.save(COMPTES_FILE)
            logger.verbose(f"{updated_count} cotation(s) mise(s) à jour")
        except Exception as e:
            logger.error(f"Erreur sauvegarde Excel: {e}")
            wb.close()
            return False
    else:
        logger.info("Aucune cotation à mettre à jour")

    wb.close()

    # Recalcul + miroir C1 si lancé depuis la GUI
    if updated_count > 0 and os.environ.get('COMPTA_GUI'):
        from inc_uno import refresh_controles
        refresh_controles(COMPTES_FILE, logger)

    return True


# ============================================================================
# MAIN
# ============================================================================

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description='Récupération des cotations financières',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemples:
  %(prog)s               # Met à jour les cotations dans Excel
  %(prog)s --dry-run     # Affiche les cours sans modifier Excel

Sources (lues depuis la feuille Cotations) :
  - yahoo: Métaux précieux (or, argent)
  - coingecko: Crypto (BTC, XMR)
  - kraken: Crypto fallback
  - frankfurter: Devises (BCE)
        """)

    parser.add_argument('--dry-run',
                        action='store_true',
                        help='Affiche les cours sans modifier Excel')
    parser.add_argument('-v', '--verbose',
                        action='store_true',
                        help='Mode verbeux')

    args = parser.parse_args()

    global logger
    logger = Logger(script_name="cpt_fetch_quotes", verbose=args.verbose)

    logger.info("Cotations")

    # Charger la config depuis Excel
    config = load_cotations_config()
    if not config:
        logger.error("Aucun asset configuré dans la feuille Cotations")
        sys.exit(1)

    logger.verbose(f"{len(config)} asset(s) configuré(s): {', '.join(it['code'] for it in config)}")

    # Récupérer toutes les cotations
    quotes = fetch_all_quotes(config)

    if not quotes:
        logger.error("Impossible de récupérer les cotations")
        sys.exit(1)

    # Afficher les cotations récupérées
    print("\nCotations récupérées:")
    print("-" * 40)
    for item in config:
        code = item['code']
        if code not in quotes:
            continue
        if item['type'] == 'metal':
            print(f"  {code}: {quotes[code]:.2f} €/g")
        elif item['type'] == 'crypto':
            print(f"  {code}: {quotes[code]:,.2f} €")
        else:
            print(f"  {code}: {quotes[code]:.4f} €")

    # Mettre à jour Excel
    if update_excel(quotes, config, dry_run=args.dry_run):
        sys.exit(0)
    else:
        sys.exit(1)


if __name__ == "__main__":
    main()
