"""Mixin Catégories pour ConfigGUI."""

from tkinter import messagebox
from tkinter import ttk
import openpyxl
import re
import shutil
import sys
import tkinter as tk

from inc_excel_schema import (
    OpCol, col_letter,
    SHEET_BUDGET, SHEET_OPERATIONS,
)


class CategoriesMixin:
    """Onglet Catégories et Postes budgétaires."""

    def _build_tab_categories(self, tab=None):
        if tab is None:
            tab = ttk.Frame(self.notebook)
            self.notebook.add(tab, text='Catégories')

        # Haut : sélection du groupe
        top = ttk.Frame(tab)
        top.pack(fill='x', padx=5, pady=5)
        ttk.Label(top, text='Site :').pack(side='left')

        groups = list(self.mappings.keys())
        # Noms lisibles : GENERIC → Tous, sites → nom config.ini
        display_groups = []
        for g in groups:
            if g == 'GENERIC':
                display_groups.append('Tous')
            else:
                display_groups.append(self.config.get(g, 'name', fallback=g))
        self.cat_group_map = dict(zip(display_groups, groups))

        self.cat_combo = ttk.Combobox(top, values=display_groups,
                                      state='readonly', width=25)
        self.cat_combo.pack(side='left', padx=5)
        self.cat_combo.bind('<<ComboboxSelected>>', self._on_cat_group_selected)

        count_label = ttk.Label(top, text='', foreground='gray')
        count_label.pack(side='left', padx=10)
        self.cat_count_label = count_label

        # Centre : Treeview
        tree_frame = ttk.Frame(tab)
        tree_frame.pack(fill='both', expand=True, padx=5)

        cols = ('pattern', 'category', 'ref')
        self.cat_tree = ttk.Treeview(tree_frame, columns=cols,
                                     show='headings', selectmode='browse')
        self.cat_tree.heading('pattern', text='Pattern (regex)')
        self.cat_tree.heading('category', text='Catégorie')
        self.cat_tree.heading('ref', text='Appariement')
        self.cat_tree.column('pattern', width=400, minwidth=200)
        self.cat_tree.column('category', width=150, minwidth=80)
        self.cat_tree.column('ref', width=90, minwidth=60, anchor='center')

        vsb = ttk.Scrollbar(tree_frame, orient='vertical',
                            command=self.cat_tree.yview)
        self.cat_tree.configure(yscrollcommand=vsb.set)
        self.cat_tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

        # Bas : boutons
        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill='x', padx=5, pady=5)

        ttk.Button(btn_frame, text='\u2795 Ajouter pattern',
                   command=self._cat_add).pack(side='left', padx=2)

        ttk.Separator(btn_frame, orient='vertical').pack(
            side='left', fill='y', padx=8)

        ttk.Button(btn_frame, text='\u25b2 Monter',
                   command=self._cat_move_up).pack(side='left', padx=2)
        ttk.Button(btn_frame, text='\u25bc Descendre',
                   command=self._cat_move_down).pack(side='left', padx=2)

        if self.xlsx_path:
            ttk.Separator(btn_frame, orient='vertical').pack(
                side='left', fill='y', padx=8)
            ttk.Button(btn_frame, text='\u2795 Nouvelle catégorie',
                       command=self._budget_cat_add).pack(side='left', padx=2)
            ttk.Button(btn_frame, text='\U0001f4cb Catégories Budget',
                       command=self._budget_cat_manage).pack(side='left', padx=2)

        # --- Section Postes budgétaires ---
        if self.xlsx_path:
            post_frame = ttk.LabelFrame(tab, text='Postes budgétaires', padding=5)
            post_frame.pack(fill='x', padx=5, pady=(2, 5))
            self._make_help_button(post_frame)

            post_tree_frame = ttk.Frame(post_frame)
            post_tree_frame.pack(fill='x')

            post_cols = ('poste', 'type')
            self.post_tree = ttk.Treeview(
                post_tree_frame, columns=post_cols,
                show='headings', selectmode='browse', height=6)
            self.post_tree.heading('poste', text='Poste')
            self.post_tree.heading('type', text='Type')
            self.post_tree.column('poste', width=300, minwidth=150)
            self.post_tree.column('type', width=80, minwidth=60, anchor='center')

            post_vsb = ttk.Scrollbar(post_tree_frame, orient='vertical',
                                     command=self.post_tree.yview)
            self.post_tree.configure(yscrollcommand=post_vsb.set)
            self.post_tree.pack(side='left', fill='x', expand=True)
            post_vsb.pack(side='right', fill='y')

            # Bouton ajout poste (modifier/supprimer via clic droit)
            post_btn = ttk.Frame(post_frame)
            post_btn.pack(fill='x', pady=(5, 0))
            ttk.Button(post_btn, text='\u2795 Ajouter',
                       command=self._budget_post_add).pack(side='left', padx=2)

            # Menu contextuel postes
            self._post_context_menu = tk.Menu(self.post_tree, tearoff=0)
            self._post_context_menu.add_command(
                label='\u270f Modifier', command=self._budget_post_edit)
            self._post_context_menu.add_command(
                label='\u2716 Supprimer', command=self._budget_post_delete)
            self.post_tree.bind('<Button-3>', self._post_show_context_menu)
            self.post_tree.bind('<Double-1>', lambda e: self._budget_post_edit())

            # Peupler
            self._refresh_post_tree()

        # --- Menu contextuel (clic droit) ---
        self._cat_context_menu = tk.Menu(self.cat_tree, tearoff=0)
        self._cat_context_menu.add_command(
            label='\u270f Modifier pattern', command=self._cat_edit)
        self._cat_context_menu.add_command(
            label='\u2716 Supprimer pattern', command=self._cat_delete)
        if self.xlsx_path:
            self._cat_context_menu.add_separator()
            self._cat_context_menu.add_command(
                label='\u270f Renommer catégorie', command=self._budget_cat_rename_from_menu)
            self._cat_context_menu.add_command(
                label='\u2716 Supprimer catégorie', command=self._budget_cat_delete_from_menu)
        self.cat_tree.bind('<Button-3>', self._cat_show_context_menu)
        self.cat_tree.bind('<Double-1>', lambda e: self._cat_edit())

        if display_groups:
            self.cat_combo.set(display_groups[0])
            self._on_cat_group_selected(None)

    def _current_group_key(self):
        display = self.cat_combo.get()
        return self.cat_group_map.get(display, display)

    def _on_cat_group_selected(self, event):
        group = self._current_group_key()
        entries = self.mappings.get(group, [])
        self.cat_count_label.config(text=f'{len(entries)} pattern(s)')

        self.cat_tree.delete(*self.cat_tree.get_children())
        for entry in entries:
            appariement = 'oui' if entry.get('ref') == '-' else ''
            self.cat_tree.insert('', 'end', values=(
                entry['pattern'], entry['category'], appariement))

    def _cat_show_context_menu(self, event):
        """Affiche le menu contextuel sur clic droit d'un pattern."""
        item = self.cat_tree.identify_row(event.y)
        if item:
            self.cat_tree.selection_set(item)
            # Mettre à jour les labels catégorie si xlsx dispo
            if self.xlsx_path:
                vals = self.cat_tree.item(item)['values']
                cat_name = str(vals[1]) if vals and len(vals) > 1 else ''
                in_budget = cat_name in self.budget_categories
                self._cat_context_menu.entryconfigure(
                    3, label=f'\u270f Renommer catégorie \u00ab{cat_name}\u00bb',
                    state='normal' if in_budget else 'disabled')
                self._cat_context_menu.entryconfigure(
                    4, label=f'\u2716 Supprimer catégorie \u00ab{cat_name}\u00bb',
                    state='normal' if in_budget else 'disabled')
            self._cat_context_menu.tk_popup(event.x_root, event.y_root)

    def _cat_add(self):
        self._cat_dialog('Ajouter un pattern')

    def _cat_edit(self):
        sel = self.cat_tree.selection()
        if not sel:
            return
        item = self.cat_tree.item(sel[0])
        vals = item['values']
        ref_val = '-' if str(vals[2]) == 'oui' else ''
        self._cat_dialog('Modifier le pattern',
                         pattern=str(vals[0]),
                         category=str(vals[1]),
                         ref=ref_val,
                         edit_item=sel[0])

    def _cat_delete(self):
        sel = self.cat_tree.selection()
        if not sel:
            return
        vals = self.cat_tree.item(sel[0])['values']
        if not messagebox.askyesno(
                'Confirmer la suppression',
                f"Supprimer le pattern '{vals[0]}' ?",
                parent=self.root):
            return
        idx = self.cat_tree.index(sel[0])
        group = self._current_group_key()
        del self.mappings[group][idx]
        self._save_mappings()
        self.cat_tree.delete(sel[0])
        self.cat_count_label.config(
            text=f'{len(self.mappings[group])} pattern(s)')

    def _cat_move_up(self):
        sel = self.cat_tree.selection()
        if not sel:
            return
        idx = self.cat_tree.index(sel[0])
        if idx == 0:
            return
        group = self._current_group_key()
        entries = self.mappings[group]
        entries[idx - 1], entries[idx] = entries[idx], entries[idx - 1]
        self._save_mappings()
        self._on_cat_group_selected(None)
        # Re-sélectionner l'élément déplacé
        children = self.cat_tree.get_children()
        self.cat_tree.selection_set(children[idx - 1])
        self.cat_tree.see(children[idx - 1])

    def _cat_move_down(self):
        sel = self.cat_tree.selection()
        if not sel:
            return
        idx = self.cat_tree.index(sel[0])
        group = self._current_group_key()
        entries = self.mappings[group]
        if idx >= len(entries) - 1:
            return
        entries[idx], entries[idx + 1] = entries[idx + 1], entries[idx]
        self._save_mappings()
        self._on_cat_group_selected(None)
        children = self.cat_tree.get_children()
        self.cat_tree.selection_set(children[idx + 1])
        self.cat_tree.see(children[idx + 1])

    def _cat_dialog(self, title, pattern='', category='', ref='',
                    edit_item=None):
        dlg = tk.Toplevel(self.root)
        dlg.title(title)
        dlg.geometry('500x200')
        dlg.transient(self.root)
        dlg.wait_visibility()
        dlg.grab_set()

        ttk.Label(dlg, text='Pattern (regex) :').grid(
            row=0, column=0, sticky='w', padx=10, pady=5)
        pat_var = tk.StringVar(value=pattern)
        pat_entry = ttk.Entry(dlg, textvariable=pat_var, width=50)
        pat_entry.grid(row=0, column=1, padx=10, pady=5)

        ttk.Label(dlg, text='Catégorie :').grid(
            row=1, column=0, sticky='w', padx=10, pady=5)
        cat_var = tk.StringVar(value=category)
        cat_combo = ttk.Combobox(dlg, textvariable=cat_var,
                                 values=self.budget_categories, width=28,
                                 state='readonly')
        cat_combo.grid(row=1, column=1, sticky='w', padx=10, pady=5)

        appariement_var = tk.BooleanVar(value=(ref == '-'))
        ttk.Checkbutton(dlg, text='Appariement',
                        variable=appariement_var).grid(
            row=2, column=0, columnspan=2, sticky='w', padx=10, pady=5)

        # Validation regex en temps réel
        regex_status = ttk.Label(dlg, text='', foreground='green')
        regex_status.grid(row=3, column=0, columnspan=2, padx=10)

        def validate_regex(*args):
            try:
                re.compile(pat_var.get())
                regex_status.config(text='Regex valide', foreground='green')
            except re.error as e:
                regex_status.config(text=f'Regex invalide: {e}',
                                    foreground='red')
        pat_var.trace_add('write', validate_regex)
        validate_regex()

        def on_ok():
            p = pat_var.get().strip()
            c = cat_var.get().strip()
            if not p or not c:
                messagebox.showwarning('Champs requis',
                                       'Pattern et Catégorie sont obligatoires.',
                                       parent=dlg)
                return
            try:
                re.compile(p)
            except re.error:
                messagebox.showerror('Regex invalide',
                                     'Le pattern regex est invalide.',
                                     parent=dlg)
                return

            entry = {'pattern': p, 'category': c}
            if appariement_var.get():
                entry['ref'] = '-'

            group = self._current_group_key()
            if edit_item is not None:
                idx = self.cat_tree.index(edit_item)
                self.mappings[group][idx] = entry
            else:
                self.mappings[group].append(entry)
            self._save_mappings()

            self._on_cat_group_selected(None)
            dlg.destroy()

        btn_frame = ttk.Frame(dlg)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text='OK', command=on_ok).pack(
            side='left', padx=5)
        ttk.Button(btn_frame, text='Annuler',
                   command=dlg.destroy).pack(side='left', padx=5)

        pat_entry.focus()

    # ----------------------------------------------------------------
    # NOUVELLE CATÉGORIE BUDGET
    # ----------------------------------------------------------------
    def _budget_cat_add(self):
        """Ouvre le dialog d'ajout d'une nouvelle catégorie Budget."""
        if not self.budget_insert_row:
            messagebox.showerror('Erreur',
                                 'Impossible de trouver le point d\'insertion dans Budget.',
                                 parent=self.root)
            return
        self._budget_cat_add_dialog()

    def _budget_cat_add_dialog(self):
        dlg = tk.Toplevel(self.root)
        dlg.title('Nouvelle catégorie Budget')
        dlg.geometry('460x200')
        dlg.transient(self.root)
        dlg.wait_visibility()
        dlg.grab_set()

        ttk.Label(dlg, text='Nom :').grid(
            row=0, column=0, sticky='w', padx=10, pady=5)
        name_var = tk.StringVar()
        name_entry = ttk.Entry(dlg, textvariable=name_var, width=30)
        name_entry.grid(row=0, column=1, padx=10, pady=5, sticky='w')

        ttk.Label(dlg, text='Poste budgétaire :').grid(
            row=1, column=0, sticky='w', padx=10, pady=5)
        poste_var = tk.StringVar()
        poste_combo = ttk.Combobox(dlg, textvariable=poste_var,
                                   values=self.budget_posts, width=28,
                                   state='readonly')
        poste_combo.grid(row=1, column=1, padx=10, pady=5, sticky='w')

        ttk.Label(dlg, text='Allocation % :').grid(
            row=2, column=0, sticky='w', padx=10, pady=5)
        alloc_var = tk.StringVar(value='100')
        alloc_entry = ttk.Entry(dlg, textvariable=alloc_var, width=10)
        alloc_entry.grid(row=2, column=1, padx=10, pady=5, sticky='w')

        status_label = ttk.Label(dlg, text='', foreground='red')
        status_label.grid(row=3, column=0, columnspan=2, padx=10)

        def on_ok():
            name = name_var.get().strip()
            poste = poste_var.get().strip()
            alloc_str = alloc_var.get().strip()

            if not name:
                status_label.config(text='Le nom est obligatoire.')
                return
            if name in self.budget_categories:
                status_label.config(text=f'La catégorie "{name}" existe déjà.')
                return
            if not poste:
                status_label.config(text='Le poste budgétaire est obligatoire.')
                return
            try:
                alloc_pct = float(alloc_str)
            except ValueError:
                status_label.config(text='Allocation % doit être un nombre.')
                return

            dlg.destroy()
            self._run_budget_save(name, poste, alloc_pct)

        btn_frame = ttk.Frame(dlg)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text='OK', command=on_ok).pack(
            side='left', padx=5)
        ttk.Button(btn_frame, text='Annuler',
                   command=dlg.destroy).pack(side='left', padx=5)

        name_entry.focus()

    def _run_budget_save(self, name, poste, alloc_pct):
        """Lance _save_budget_category dans un thread avec fenêtre d'attente animée."""
        self._run_uno_operation(
            'Écriture en cours',
            lambda: self._save_budget_category(name, poste, alloc_pct),
            lambda: self._set_status(f'Catégorie "{name}" ajoutée.')
        )

    def _save_budget_category(self, name, poste, alloc_pct=100.0):
        """Insère une nouvelle catégorie dans Budget via BudgetMixin."""
        self._add_category(name, poste=poste, alloc_pct=alloc_pct / 100)

    # ----------------------------------------------------------------
    # GÉRER / RENOMMER / SUPPRIMER CATÉGORIE BUDGET
    # ----------------------------------------------------------------

    def _budget_cat_manage(self):
        """Dialog listant toutes les catégories Budget avec Renommer/Supprimer."""
        dlg = tk.Toplevel(self.root)
        dlg.title('Catégories Budget')
        dlg.geometry('400x420')
        dlg.transient(self.root)

        ttk.Label(dlg, text=f'{len(self.budget_categories)} catégorie(s) dans Budget'
                  ).pack(padx=10, pady=(10, 5), anchor='w')

        # Listbox avec scrollbar
        list_frame = ttk.Frame(dlg)
        list_frame.pack(fill='both', expand=True, padx=10, pady=5)

        lb = tk.Listbox(list_frame, selectmode='browse', exportselection=False)
        vsb = ttk.Scrollbar(list_frame, orient='vertical', command=lb.yview)
        lb.configure(yscrollcommand=vsb.set)
        lb.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

        for cat in self.budget_categories:
            lb.insert('end', cat)

        # Boutons
        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(pady=(5, 10))

        def get_selected():
            sel = lb.curselection()
            if not sel:
                return None
            return lb.get(sel[0])

        def on_rename():
            cat = get_selected()
            if cat:
                dlg.destroy()
                self._budget_cat_rename_dialog(cat)

        def on_delete():
            cat = get_selected()
            if cat:
                dlg.destroy()
                self._budget_cat_delete_confirm(cat)

        ttk.Button(btn_frame, text='\u270f Renommer',
                   command=on_rename).pack(side='left', padx=5)
        ttk.Button(btn_frame, text='\u2716 Supprimer',
                   command=on_delete).pack(side='left', padx=5)
        ttk.Button(btn_frame, text='Fermer',
                   command=dlg.destroy).pack(side='left', padx=5)

        # Context menu
        ctx = tk.Menu(lb, tearoff=0)
        ctx.add_command(label='\u270f Renommer', command=on_rename)
        ctx.add_command(label='\u2716 Supprimer', command=on_delete)
        lb.bind('<Button-3>', lambda e: (
            lb.selection_clear(0, 'end'),
            lb.selection_set(lb.nearest(e.y)),
            ctx.tk_popup(e.x_root, e.y_root)))
        lb.bind('<Double-1>', lambda e: on_rename())

    def _budget_cat_rename_from_menu(self):
        """Renomme une catégorie Budget depuis le menu contextuel patterns."""
        sel = self.cat_tree.selection()
        if not sel:
            return
        vals = self.cat_tree.item(sel[0])['values']
        cat_name = str(vals[1]) if vals and len(vals) > 1 else ''
        if not cat_name or cat_name not in self.budget_categories:
            return
        self._budget_cat_rename_dialog(cat_name)

    def _budget_cat_rename_dialog(self, cat_name):
        """Dialog pour renommer une catégorie Budget."""
        dlg = tk.Toplevel(self.root)
        dlg.title('Renommer catégorie')
        dlg.geometry('400x150')
        dlg.transient(self.root)
        dlg.wait_visibility()
        dlg.grab_set()

        ttk.Label(dlg, text=f'Catégorie actuelle : {cat_name}').grid(
            row=0, column=0, columnspan=2, sticky='w', padx=10, pady=5)
        ttk.Label(dlg, text='Nouveau nom :').grid(
            row=1, column=0, sticky='w', padx=10, pady=5)
        name_var = tk.StringVar(value=cat_name)
        name_entry = ttk.Entry(dlg, textvariable=name_var, width=30)
        name_entry.grid(row=1, column=1, padx=10, pady=5, sticky='w')

        ttk.Label(dlg, text='Les opérations existantes seront mises à jour.',
                  foreground='grey').grid(
            row=2, column=0, columnspan=2, padx=10, sticky='w')

        def on_ok():
            new_name = name_var.get().strip()
            if not new_name or new_name == cat_name:
                dlg.destroy()
                return
            if new_name in self.budget_categories:
                messagebox.showwarning('Doublon',
                    f'La catégorie "{new_name}" existe déjà.',
                    parent=dlg)
                return
            dlg.destroy()
            self._run_uno_operation(
                'Renommage en cours',
                lambda: self._rename_budget_category(cat_name, new_name),
                lambda: self._after_budget_cat_rename(cat_name, new_name)
            )

        btn_frame = ttk.Frame(dlg)
        btn_frame.grid(row=3, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text='OK', command=on_ok).pack(side='left', padx=5)
        ttk.Button(btn_frame, text='Annuler',
                   command=dlg.destroy).pack(side='left', padx=5)
        name_entry.focus()
        name_entry.select_range(0, 'end')

    def _budget_cat_delete_from_menu(self):
        """Supprime une catégorie Budget depuis le menu contextuel."""
        sel = self.cat_tree.selection()
        if not sel:
            return
        vals = self.cat_tree.item(sel[0])['values']
        cat_name = str(vals[1]) if vals and len(vals) > 1 else ''
        if not cat_name or cat_name not in self.budget_categories:
            return

        if len(self.budget_categories) <= 1:
            messagebox.showwarning(
                'Suppression impossible',
                'Impossible de supprimer la dernière catégorie.\n'
                'Elle sert de modèle pour les ajouts futurs.',
                parent=self.root)
            return

        self._budget_cat_delete_confirm(cat_name)

    def _budget_cat_delete_confirm(self, cat_name):
        """Vérifie les ops associées et lance la suppression ou réaffectation."""
        # Compter les opérations qui référencent cette catégorie (col G)
        ops_count = self._count_ops_for_category(cat_name)

        if ops_count > 0:
            # Réaffectation obligatoire avant suppression
            others = [c for c in self.budget_categories if c != cat_name]
            dlg = tk.Toplevel(self.root)
            dlg.title('Réaffecter les opérations')
            dlg.transient(self.root)
            dlg.grab_set()

            ttk.Label(dlg, text=(
                f'{ops_count} opération(s) utilisent la catégorie « {cat_name} ».\n'
                'Choisir une catégorie de remplacement :'),
                wraplength=380).pack(padx=15, pady=(15, 5))

            combo_var = tk.StringVar(value=others[0] if others else '')
            combo = ttk.Combobox(dlg, textvariable=combo_var,
                                 values=others, state='readonly', width=35)
            combo.pack(padx=15, pady=5)

            btn_frame = ttk.Frame(dlg)
            btn_frame.pack(pady=(5, 15))

            def on_ok():
                target = combo_var.get()
                if not target:
                    return
                dlg.destroy()
                self._run_uno_operation(
                    'Suppression en cours',
                    lambda: self._delete_budget_category(cat_name,
                                                         reassign_to=target),
                    lambda: self._after_budget_cat_delete(cat_name))

            ttk.Button(btn_frame, text='Réaffecter et supprimer',
                       command=on_ok).pack(side='left', padx=5)
            ttk.Button(btn_frame, text='Annuler',
                       command=dlg.destroy).pack(side='left', padx=5)

            dlg.geometry('420x160')
            dlg.wait_window()
        else:
            if not messagebox.askyesno(
                    'Confirmer la suppression',
                    f'Supprimer la catégorie « {cat_name} » du Budget ?\n\n'
                    'Aucune opération ne référence cette catégorie.',
                    parent=self.root):
                return

            self._run_uno_operation(
                'Suppression en cours',
                lambda: self._delete_budget_category(cat_name),
                lambda: self._after_budget_cat_delete(cat_name))

    def _count_ops_for_category(self, cat_name):
        """Compte les opérations dans col G qui référencent une catégorie."""
        if not self.xlsx_path:
            return 0
        wb = openpyxl.load_workbook(self.xlsx_path, read_only=True, data_only=True)
        try:
            ws = wb[SHEET_OPERATIONS]
            count = 0
            for row in ws.iter_rows(min_row=3, max_col=OpCol.CATEGORIE):
                val = row[OpCol.CATEGORIE - 1].value
                if val and str(val).strip() == cat_name:
                    count += 1
            return count
        finally:
            wb.close()

    def _rename_budget_category(self, old_name, new_name):
        """Worker UNO : renomme une catégorie dans Budget (col L) + Opérations (col G)."""
        from inc_uno import UnoDocument
        from inc_excel_schema import uno_row, uno_col, OpCol

        bak_path = self.xlsx_path.with_suffix('.xlsm.bak')
        shutil.copy2(self.xlsx_path, bak_path)

        cat_row = self.budget_cat_rows.get(old_name)
        if not cat_row:
            raise ValueError(f'Catégorie "{old_name}" introuvable dans budget_cat_rows')

        with UnoDocument(self.xlsx_path) as doc:
            # 1. Budget — colonne catégories
            ws = doc.get_sheet(SHEET_BUDGET)
            ws.getCellByPosition(uno_col(self.budget_cat_col), uno_row(cat_row)).setString(new_name)

            # 2. Opérations col G — renommer toutes les occurrences
            ws_ops = doc.get_sheet(SHEET_OPERATIONS)
            col_g = uno_col(OpCol.CATEGORIE)
            cursor = ws_ops.createCursor()
            cursor.gotoStartOfUsedArea(False)
            cursor.gotoEndOfUsedArea(True)
            last_row_0 = cursor.getRangeAddress().EndRow
            count = 0
            for r in range(2, last_row_0 + 1):  # row 0=header, 1=header2, data from row 2
                cell = ws_ops.getCellByPosition(col_g, r)
                if cell.getString() == old_name:
                    cell.setString(new_name)
                    count += 1
            if count:
                print(f'Catégorie renommée dans {count} opération(s)', file=sys.stderr)

            self._uno_finalize(doc)

        # Mettre à jour les patterns dans config_category_mappings.json
        for group_entries in self.mappings.values():
            for entry in group_entries:
                if entry.get('category') == old_name:
                    entry['category'] = new_name

    def _after_budget_cat_rename(self, old_name, new_name):
        """Callback après renommage catégorie : met à jour l'état mémoire."""
        idx = self.budget_categories.index(old_name)
        self.budget_categories[idx] = new_name
        self.budget_cat_rows[new_name] = self.budget_cat_rows.pop(old_name)
        # Rafraîchir le Treeview catégories
        self._on_cat_group_selected(None)
        self._save_mappings()
        self._set_status(f'Catégorie "{old_name}" renommée en "{new_name}".')

    def _delete_budget_category(self, name, reassign_to=None):
        """Worker UNO : supprime une catégorie du Budget.

        Délègue à BudgetMixin._delete_category (backup + UNO + mémoire).
        """
        bak_path = self.xlsx_path.with_suffix('.xlsm.bak')
        shutil.copy2(self.xlsx_path, bak_path)
        self._delete_category(name, reassign_to=reassign_to)

    def _after_budget_cat_delete(self, name):
        """Callback après suppression catégorie : GUI + purge mappings."""
        # Purger les patterns orphelins qui référençaient cette catégorie
        purged = 0
        for group_entries in self.mappings.values():
            before = len(group_entries)
            group_entries[:] = [e for e in group_entries if e.get('category') != name]
            purged += before - len(group_entries)
        if purged:
            self._save_mappings()
            print(f'{purged} pattern(s) orphelin(s) purgé(s) pour catégorie "{name}"', file=sys.stderr)

        self._set_status(f'Catégorie "{name}" supprimée du Budget.')
        self._on_cat_group_selected(None)

    # ----------------------------------------------------------------
    # POSTES BUDGÉTAIRES (CRUD)
    # ----------------------------------------------------------------

    def _refresh_post_tree(self):
        """Rafraîchit le Treeview des postes budgétaires."""
        self.post_tree.delete(*self.post_tree.get_children())
        for name in self.budget_posts:
            type_fv = self.budget_post_types.get(name, '')
            self.post_tree.insert('', 'end', values=(name, type_fv))

    def _post_show_context_menu(self, event):
        """Affiche le menu contextuel sur clic droit d'un poste."""
        item = self.post_tree.identify_row(event.y)
        if item:
            self.post_tree.selection_set(item)
            self._post_context_menu.tk_popup(event.x_root, event.y_root)

    def _budget_post_add(self):
        """Ouvre le dialog d'ajout d'un nouveau poste budgétaire."""
        if not self.budget_posts_total_row:
            messagebox.showerror('Erreur',
                                 'Impossible de trouver la ligne "Total" dans Budget.',
                                 parent=self.root)
            return
        self._budget_post_dialog('Nouveau poste budgétaire')

    def _budget_post_edit(self):
        """Ouvre le dialog de modification du poste sélectionné."""
        sel = self.post_tree.selection()
        if not sel:
            return
        vals = self.post_tree.item(sel[0])['values']
        old_name = str(vals[0])
        old_type = str(vals[1])
        self._budget_post_dialog('Modifier le poste', name=old_name,
                                 type_fv=old_type, edit_name=old_name)

    def _budget_post_delete(self):
        """Supprime le poste sélectionné après vérification des dépendances."""
        sel = self.post_tree.selection()
        if not sel:
            return
        vals = self.post_tree.item(sel[0])['values']
        name = str(vals[0])

        # Garder au moins un poste (template de style)
        if len(self.budget_posts) <= 1:
            messagebox.showwarning(
                'Suppression impossible',
                'Impossible de supprimer le dernier poste.\n'
                'Il sert de modèle pour les ajouts futurs.',
                parent=self.root)
            return

        # Vérifier qu'aucune catégorie ne référence ce poste (col AB)
        linked_cats = [cat for cat in self.budget_categories
                       if self._get_cat_post(cat) == name]
        if linked_cats:
            messagebox.showerror(
                'Suppression impossible',
                f'Le poste "{name}" est référencé par {len(linked_cats)} '
                f'catégorie(s) :\n\n' + '\n'.join(f'  - {c}' for c in linked_cats) +
                '\n\nRattachez ces catégories à un autre poste avant de supprimer.',
                parent=self.root)
            return

        if not messagebox.askyesno(
                'Confirmer la suppression',
                f'Supprimer le poste "{name}" ?',
                parent=self.root):
            return

        self._run_uno_operation(
            'Suppression en cours',
            lambda: self._delete_budget_post(name),
            lambda: self._after_budget_post_delete(name)
        )

    def _get_cat_post(self, cat_name):
        """Retourne le poste rattaché à une catégorie (col AB) via openpyxl."""
        cat_row = self.budget_cat_rows.get(cat_name)
        if not cat_row:
            return None
        try:
            wb = openpyxl.load_workbook(self.xlsx_path, data_only=True)
            ws = wb[SHEET_BUDGET]
            last = self.budget_last_devise_col
            poste_col = last + 4  # AB
            val = ws.cell(cat_row, poste_col).value
            wb.close()
            return str(val).strip() if val else ''
        except Exception:
            return ''

    def _budget_post_dialog(self, title, name='', type_fv='', edit_name=None):
        """Dialog pour ajouter ou modifier un poste budgétaire."""
        dlg = tk.Toplevel(self.root)
        dlg.title(title)
        dlg.geometry('400x160')
        dlg.transient(self.root)
        dlg.wait_visibility()
        dlg.grab_set()

        ttk.Label(dlg, text='Nom :').grid(
            row=0, column=0, sticky='w', padx=10, pady=5)
        name_var = tk.StringVar(value=name)
        name_entry = ttk.Entry(dlg, textvariable=name_var, width=30)
        name_entry.grid(row=0, column=1, padx=10, pady=5, sticky='w')

        ttk.Label(dlg, text='Type :').grid(
            row=1, column=0, sticky='w', padx=10, pady=5)
        type_var = tk.StringVar(value=type_fv)
        type_combo = ttk.Combobox(dlg, textvariable=type_var,
                                  values=['Fixe', 'Variable'], width=12,
                                  state='readonly')
        type_combo.grid(row=1, column=1, padx=10, pady=5, sticky='w')

        status_label = ttk.Label(dlg, text='', foreground='red')
        status_label.grid(row=2, column=0, columnspan=2, padx=10)

        def on_ok():
            new_name = name_var.get().strip()
            new_type = type_var.get().strip()

            if not new_name:
                status_label.config(text='Le nom est obligatoire.')
                return
            if not new_type:
                status_label.config(text='Le type est obligatoire.')
                return
            # Vérifier unicité (sauf si c'est le même en édition)
            if new_name != edit_name and new_name in self.budget_posts:
                status_label.config(text=f'Le poste "{new_name}" existe déjà.')
                return

            dlg.destroy()
            if edit_name:
                self._run_uno_operation(
                    'Modification en cours',
                    lambda: self._update_budget_post(edit_name, new_name, new_type),
                    lambda: self._after_budget_post_update(edit_name, new_name, new_type)
                )
            else:
                self._run_uno_operation(
                    'Écriture en cours',
                    lambda: self._save_budget_post(new_name, new_type),
                    lambda: self._after_budget_post_save(new_name, new_type)
                )

        btn_frame = ttk.Frame(dlg)
        btn_frame.grid(row=3, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text='OK', command=on_ok).pack(
            side='left', padx=5)
        ttk.Button(btn_frame, text='Annuler',
                   command=dlg.destroy).pack(side='left', padx=5)
        name_entry.focus()
        if edit_name:
            name_entry.select_range(0, 'end')

    def _save_budget_post(self, name, type_fv):
        """Insère un nouveau poste budgétaire via BudgetMixin."""
        self._add_poste(name, fixe=(type_fv == 'Fixe'))

    def _after_budget_post_save(self, name, type_fv):
        """Callback après ajout poste : refresh GUI (état mémoire mis à jour par le mixin)."""
        self.budget_post_types[name] = type_fv
        self._refresh_post_tree()
        self._set_status(f'Poste "{name}" ajouté.')

    def _update_budget_post(self, old_name, new_name, new_type):
        """Worker UNO : modifie le nom et/ou le type d'un poste."""
        from inc_uno import UnoDocument
        from inc_excel_schema import uno_row, uno_col

        bak_path = self.xlsx_path.with_suffix('.xlsm.bak')
        shutil.copy2(self.xlsx_path, bak_path)

        post_row = self.budget_post_rows.get(old_name)
        if not post_row:
            raise ValueError(f'Poste "{old_name}" introuvable dans budget_post_rows')

        last = self.budget_last_devise_col
        poste_col = last + 4  # AB

        with UnoDocument(self.xlsx_path) as doc:
            ws = doc.get_sheet(SHEET_BUDGET)
            r0 = uno_row(post_row)

            # Mettre à jour nom et type
            if new_name != old_name:
                ws.getCellByPosition(0, r0).setString(new_name)
            if new_type != self.budget_post_types.get(old_name, ''):
                ws.getCellByPosition(1, r0).setString(new_type)

            # Si renommage : propager dans col AB (poste des catégories)
            if new_name != old_name:
                cat_start = min(self.budget_cat_rows.values()) if self.budget_cat_rows else 28
                cat_end = max(self.budget_cat_rows.values()) if self.budget_cat_rows else 50
                for cr in range(cat_start, cat_end + 1):
                    cell = ws.getCellByPosition(uno_col(poste_col), uno_row(cr))
                    val = cell.getString()
                    if val.strip() == old_name:
                        cell.setString(new_name)

            self._uno_finalize(doc)

    def _after_budget_post_update(self, old_name, new_name, new_type):
        """Callback après modification poste : met à jour état mémoire + GUI."""
        idx = self.budget_posts.index(old_name)
        self.budget_posts[idx] = new_name
        if new_name != old_name:
            self.budget_post_rows[new_name] = self.budget_post_rows.pop(old_name)
            self.budget_post_types[new_name] = new_type
            del self.budget_post_types[old_name]
        else:
            self.budget_post_types[new_name] = new_type
        self._refresh_post_tree()
        if new_name != old_name:
            self._set_status(f'Poste "{old_name}" renommé en "{new_name}".')
        else:
            self._set_status(f'Poste "{new_name}" modifié.')

    def _delete_budget_post(self, name):
        """Worker UNO : supprime un poste budgétaire.

        Délègue à BudgetMixin._delete_poste (backup + UNO + mémoire).
        """
        bak_path = self.xlsx_path.with_suffix('.xlsm.bak')
        shutil.copy2(self.xlsx_path, bak_path)
        self._delete_poste(name)

    def _after_budget_post_delete(self, name):
        """Callback après suppression poste : GUI."""
        self._refresh_post_tree()
        self._set_status(f'Poste "{name}" supprimé.')

