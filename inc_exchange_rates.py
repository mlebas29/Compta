#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
cpt_exchange_rates.py - Récupération des taux de change via API ECB

Utilise frankfurter.app (données ECB) pour obtenir les taux de change.
Fournit un cache en mémoire pour éviter les appels répétés.

Usage:
    from inc_exchange_rates import get_rate_to_eur, convert_to_eur

    # Taux CHF → EUR pour une date
    rate = get_rate_to_eur('CHF', '2026-01-15')

    # Conversion directe
    eur_amount = convert_to_eur(2606.56, 'CHF', '2026-01-15')
"""

import sys
import time
import requests
from datetime import datetime
from inc_excel_schema import CotCol, SHEET_COTATIONS, COT_FIRST_ROW

# Cache des taux : {date_str: {currency: rate_to_eur}}
_rates_cache = {}

# Cache des taux Excel (chargé une seule fois)
_excel_fallback_rates = None

# Flag TNR — positionné par cpt_update.py --TNR, jamais par variable d'environnement
TNR_MODE = False

# Flag pour éviter les warnings répétés
_api_warning_shown = False

# Timeout pour les requêtes API
REQUEST_TIMEOUT = 10


def _load_excel_fallback():
    """Charge les taux de change depuis la feuille Cotations de comptes.xlsm.

    Fallback quand l'API Frankfurter est indisponible.
    Les taux sont approximatifs (dernière mise à jour cpt_fetch_quotes).

    Returns:
        dict: {currency: rate} au format Frankfurter (1 EUR = X currency)
              ou None si erreur
    """
    global _excel_fallback_rates

    if _excel_fallback_rates is not None:
        return _excel_fallback_rates

    try:
        import openpyxl
        from inc_mode import get_base_dir

        comptes_path = get_base_dir() / 'comptes.xlsm'
        wb = openpyxl.load_workbook(comptes_path, data_only=True)
        ws = wb[SHEET_COTATIONS]

        rates = {'EUR': 1.0}

        from inc_excel_schema import get_named_ranges, get_table_start
        named = get_named_ranges(wb)
        cot_start = get_table_start(named, 'COT') or COT_FIRST_ROW
        for row in range(cot_start + 1, ws.max_row + 1):
            code = ws.cell(row=row, column=CotCol.CODE).value
            cours_eur = ws.cell(row=row, column=CotCol.COURS_EUR).value
            if code and code.strip() and cours_eur and float(cours_eur) > 0:
                # Inverser : 1 EUR = 1/cours_eur devise
                rates[code.strip()] = 1.0 / float(cours_eur)

        wb.close()

        _excel_fallback_rates = rates
        return rates

    except Exception as e:
        print(f"⚠ Fallback Excel taux de change échoué: {e}", file=sys.stderr)
        return None


def get_rates_for_date(date_str=None):
    """Récupère tous les taux EUR pour une date donnée

    Args:
        date_str: Date au format 'YYYY-MM-DD' ou 'DD/MM/YYYY'
                  Si None, récupère les taux du jour ('latest')

    Returns:
        dict: {currency: rate} où rate est le taux 1 EUR = X currency
              ou None si erreur API
    """
    global _api_warning_shown

    # Normaliser la date en YYYY-MM-DD ou 'latest'
    if date_str is None:
        date_key = 'latest'
    elif '/' in date_str:
        try:
            dt = datetime.strptime(date_str, '%d/%m/%Y')
            date_key = dt.strftime('%Y-%m-%d')
        except ValueError:
            return None
    else:
        date_key = date_str

    # Vérifier le cache
    if date_key in _rates_cache:
        return _rates_cache[date_key]

    # En TNR, utiliser uniquement le fallback Excel (pas d'appel API)
    if TNR_MODE:
        fallback = _load_excel_fallback()
        if fallback:
            _rates_cache[date_key] = fallback
            return fallback
        return None

    # Appeler l'API frankfurter.app (données ECB)
    url = f"https://api.frankfurter.app/{date_key}"

    # Tentative API avec retry (1 retry après 2s)
    last_error = None
    for attempt in range(2):
        try:
            response = requests.get(url, timeout=REQUEST_TIMEOUT)
            response.raise_for_status()
            data = response.json()

            # Format: {"amount": 1, "base": "EUR", "date": "2026-01-15", "rates": {"CHF": 0.94, "USD": 1.03, ...}}
            rates = data.get('rates', {})

            # Ajouter EUR = 1
            rates['EUR'] = 1.0

            # Mettre en cache
            _rates_cache[date_key] = rates

            return rates

        except (requests.exceptions.RequestException, Exception) as e:
            last_error = e
            if attempt == 0:
                time.sleep(2)

    # API échouée après retry — fallback sur cache Excel
    fallback = _load_excel_fallback()
    if fallback:
        if not _api_warning_shown:
            print(f"⚠ API taux de change indisponible, utilisation taux cache Excel", file=sys.stderr)
            _api_warning_shown = True
        _rates_cache[date_key] = fallback
        return fallback

    # Aucun fallback disponible
    if not _api_warning_shown:
        print(f"⚠ API taux de change indisponible: {last_error}", file=sys.stderr)
        print("  Les colonnes Equiv resteront vides pour les changes non-EUR", file=sys.stderr)
        _api_warning_shown = True
    return None


def get_rate_to_eur(currency, date_str):
    """Récupère le taux de conversion currency → EUR

    Args:
        currency: Code devise (CHF, USD, SGD, etc.)
        date_str: Date au format 'YYYY-MM-DD' ou 'DD/MM/YYYY'

    Returns:
        float: Taux pour convertir 1 currency en EUR
               ou None si erreur
    """
    if currency == 'EUR':
        return 1.0

    rates = get_rates_for_date(date_str)
    if rates is None:
        return None

    # rates contient 1 EUR = X currency
    # On veut 1 currency = Y EUR, donc Y = 1/X
    rate_eur_to_currency = rates.get(currency)
    if rate_eur_to_currency is None or rate_eur_to_currency == 0:
        # Devise absente de l'API (crypto/metal) → fallback Excel
        fallback = _load_excel_fallback()
        if fallback:
            rate_eur_to_currency = fallback.get(currency)
        if rate_eur_to_currency is None or rate_eur_to_currency == 0:
            return None

    return 1.0 / rate_eur_to_currency


def convert_to_eur(amount, currency, date_str):
    """Convertit un montant en EUR

    Args:
        amount: Montant dans la devise source
        currency: Code devise source
        date_str: Date de l'opération

    Returns:
        float: Montant équivalent en EUR
               ou None si conversion impossible
    """
    if currency == 'EUR':
        return amount

    rate = get_rate_to_eur(currency, date_str)
    if rate is None:
        return None

    return amount * rate


def get_cross_rate(from_currency, to_currency, date_str):
    """Calcule le taux de change entre deux devises non-EUR

    Args:
        from_currency: Devise source
        to_currency: Devise cible
        date_str: Date de l'opération

    Returns:
        float: Taux from → to
               ou None si erreur
    """
    rates = get_rates_for_date(date_str)
    if rates is None:
        return None

    rate_from = rates.get(from_currency)
    rate_to = rates.get(to_currency)

    if rate_from is None or rate_to is None or rate_from == 0:
        return None

    # from_currency → EUR → to_currency
    # 1 EUR = rate_from from_currency
    # 1 EUR = rate_to to_currency
    # Donc 1 from_currency = (rate_to / rate_from) to_currency
    return rate_to / rate_from


# Test si exécuté directement
if __name__ == '__main__':
    print("Test API taux de change ECB (frankfurter.app)")
    print("=" * 50)

    test_date = '2026-01-15'
    print(f"\nTaux pour {test_date}:")

    rates = get_rates_for_date(test_date)
    if rates:
        for curr in ['USD', 'CHF', 'SGD', 'SEK', 'JPY', 'GBP']:
            if curr in rates:
                rate_to_eur = get_rate_to_eur(curr, test_date)
                print(f"  1 {curr} = {rate_to_eur:.6f} EUR (1 EUR = {rates[curr]:.4f} {curr})")

        print(f"\nConversions:")
        print(f"  2606.56 CHF = {convert_to_eur(2606.56, 'CHF', test_date):.2f} EUR")
        print(f"  3224.47 USD = {convert_to_eur(3224.47, 'USD', test_date):.2f} EUR")
        print(f"  3728.97 SGD = {convert_to_eur(3728.97, 'SGD', test_date):.2f} EUR")
    else:
        print("  API indisponible")
