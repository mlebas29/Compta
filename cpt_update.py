#!/usr/bin/env python3
"""
Script de mise à jour automatique de comptes.xlsm - Phase 2 (corrigé)
Usage:
    cpt_update.py                    # Traite les fichiers dans dropbox/
    cpt_update.py --fallback         # Restaure la dernière sauvegarde
    cpt_update.py --archive-only     # Archive sans importer (après correction manuelle)
"""

import os
import sys
import shutil
import subprocess
import argparse
import configparser
import csv
from pathlib import Path
from datetime import datetime
import importlib
import inc_mode
from inc_logging import Logger
from inc_excel_import import ComptaExcelImport, get_valid_accounts
from inc_excel_schema import SHEET_OPERATIONS, Operation

# Sections config.ini qui ne sont pas des sites
_CONFIG_META_SECTIONS = {'general', 'pairing', 'comparison', 'paths', 'sites'}


def _load_format_modules(config):
    """Découvre les sites depuis config.ini et charge dynamiquement les modules cpt_format_SITE."""
    modules = {}
    for section in config.sections():
        if section in _CONFIG_META_SECTIONS:
            continue
        module_name = f'cpt_format_{section}'
        try:
            modules[section] = importlib.import_module(module_name)
        except ModuleNotFoundError:
            pass  # Site sans formatteur (ne devrait pas arriver)
    return modules

# Supprimer les warnings pdfminer (FontBBox, gray color) non pertinents
import logging
logging.getLogger('pdfminer').setLevel(logging.ERROR)

# ============================================================================
# CONFIGURATION
# ============================================================================

# Utiliser la détection automatique de mode avec vérification de cohérence
BASE_DIR = inc_mode.get_base_dir()
CONFIG_FILE = BASE_DIR / 'config.ini'

# Charger la configuration
config = configparser.ConfigParser()
if not CONFIG_FILE.exists():
    print(f"⚠ config.ini introuvable dans {BASE_DIR}")
    print(f"   Utilisation des valeurs par défaut")
    # Valeurs par défaut
    DEBUG = False
    DROPBOX_DIR = BASE_DIR / "dropbox"
    ARCHIVES_DIR = BASE_DIR / "archives"
    LOGS_DIR = BASE_DIR / "logs"
    DEBUG_DIR = LOGS_DIR / 'debug'  # Debug files go in logs/debug/
    COMPTES_FILE = BASE_DIR / "comptes.xlsm"
    MAX_SESSIONS = 10
else:
    config.read(CONFIG_FILE)

    # Paramètres généraux
    DEBUG = config.getboolean('general', 'DEBUG', fallback=False)

    # Chemins (tous relatifs au BASE_DIR)
    DROPBOX_DIR = BASE_DIR / config.get('paths', 'dropbox', fallback='./dropbox')
    ARCHIVES_DIR = BASE_DIR / config.get('paths', 'archives', fallback='./archives')
    LOGS_DIR = BASE_DIR / config.get('paths', 'logs', fallback='./logs')
    DEBUG_DIR = LOGS_DIR / 'debug'  # Debug files go in logs/debug/
    COMPTES_FILE = BASE_DIR / config.get('paths', 'comptes_file', fallback='./comptes.xlsm')

    # Nombre maximum de sessions à conserver
    MAX_SESSIONS = config.getint('general', 'max_sessions', fallback=10)

# Charger les modules formatteurs dynamiquement depuis config.ini
SITE_FORMAT_MODULES = _load_format_modules(config)


def get_site_dropbox_dir(site):
    """Retourne le répertoire dropbox d'un site (lit 'dossier' dans config.ini)."""
    return DROPBOX_DIR / config.get(site, 'dossier', fallback=site)


def tuples_to_csv(tuples_list, header=None):
    """Convertit une liste de tuples en CSV texte.

    Args:
        tuples_list: Liste de tuples
        header: Header CSV optionnel

    Returns:
        str: Contenu CSV (lignes séparées par \\n)
    """
    lines = []
    if header:
        lines.append(header)
    for t in tuples_list:
        lines.append(';'.join(str(x) for x in t))
    return '\n'.join(lines)


# ============================================================================
# FONCTIONS UTILITAIRES HDS (Horodatage de Session)
# ============================================================================

import re

def add_hds(filename, hds):
    """Ajoute un HDS (Horodatage de Session) au nom d'un fichier

    Args:
        filename: Nom du fichier (ex: "comptes.xlsm", "rapport.tar.gz")
        hds: Horodatage au format YYYYMMDD_HHMMSS

    Returns:
        Nouveau nom avec HDS (ex: "comptes_HDS_20251120_153045.xlsx")

    Exemples:
        add_hds("comptes.xlsm", "20251120_153045")
        → "comptes_HDS_20251120_153045.xlsx"

        add_hds("rapport.tar.gz", "20251120_153045")
        → "rapport_HDS_20251120_153045.tar.gz"

        add_hds("@R0GERAN_00053_20251016_20251111.PDF", "20251120_153045")
        → "@R0GERAN_00053_20251016_20251111_HDS_20251120_153045.PDF"
    """
    # Gérer les extensions composées (.tar.gz, .tar.bz2, .tar.xz)
    if filename.endswith(('.tar.gz', '.tar.bz2', '.tar.xz')):
        parts = filename.rsplit('.', 2)
        name = parts[0]
        ext = '.' + '.'.join(parts[1:])
    else:
        parts = filename.rsplit('.', 1)
        if len(parts) == 2:
            name, ext = parts[0], '.' + parts[1]
        else:
            name, ext = filename, ''

    return f"{name}_HDS_{hds}{ext}"


def extract_hds(filename):
    """Extrait le HDS d'un nom de fichier

    Args:
        filename: Nom du fichier avec HDS

    Returns:
        Le HDS au format YYYYMMDD_HHMMSS ou None si absent

    Exemples:
        extract_hds("comptes_HDS_20251120_153045.xlsx") → "20251120_153045"
        extract_hds("rapport.pdf") → None
    """
    match = re.search(r'_HDS_(\d{8}_\d{6})', filename)
    return match.group(1) if match else None


def has_hds(filename):
    """Vérifie si un fichier a déjà un HDS

    Args:
        filename: Nom du fichier à vérifier

    Returns:
        True si le fichier a un HDS, False sinon
    """
    return extract_hds(filename) is not None


def remove_hds(filename):
    """Retire le HDS d'un nom de fichier

    Args:
        filename: Nom du fichier avec HDS

    Returns:
        Nom du fichier sans HDS, ou le nom original si pas de HDS

    Exemples:
        remove_hds("comptes_HDS_20251120_153045.xlsx") → "comptes.xlsm"
        remove_hds("releve_HDS_20251120_153045.csv") → "releve.csv"
        remove_hds("rapport.pdf") → "rapport.pdf"
    """
    return re.sub(r'_HDS_\d{8}_\d{6}', '', filename)


def parse_filename(filename):
    """Parse un nom de fichier selon la convention type_account_method[N].ext

    Convention de nommage:
        {type}_{account}_{method}[discriminant].{ext}

    Args:
        filename: Nom du fichier (avec ou sans extension)

    Returns:
        dict ou None: {
            'type': 'operations'|'positions'|'soldes',
            'account': 'compte-principal'|'compte-titre'|...,
            'method': 'download'|'parsed',
            'discriminant': None ou numéro (str),
            'ext': 'csv'|'pdf'|'xlsx'
        }
        Retourne None si le nom ne correspond pas à la convention

    Note: Cette fonction est pour la rétrocompatibilité avec l'ancien nommage.
        Les fichiers actuels utilisent les noms originaux (eToroTransactions_*.tsv,
        Account.csv, etc.) et passent par le fallback "mode manuel".

    Exemples (ancienne convention):
        parse_filename("operations_compte-principal_download.csv")
        → {'type': 'operations', 'account': 'compte-principal', ...}

        parse_filename("positions_ass-vie-ebene-cecile_parsed.xlsx")
        → {'type': 'positions', 'account': 'ass-vie-ebene-cecile', ...}

        parse_filename("eToroTransactions_22-12-2024.tsv")
        → None (traité en mode manuel par le site ETORO)
    """
    # Pattern: type_account_method[N].ext
    # - type: operations, positions, soldes
    # - account: peut contenir des tirets (ex: ass-vie-ebene-cecile)
    # - method: download, parsed
    # - discriminant: optionnel, chiffres (ex: 1, 2, 3 pour multi-mois)
    # - ext: csv, pdf, xlsx

    match = re.match(
        r'^(operations|positions|soldes)_([a-z0-9-]+?)_(download|parsed)(\d*)\.([a-zA-Z]+)$',
        filename,
        re.IGNORECASE
    )

    if not match:
        # Cas spécial BG_GESTION: operations_{nom_rapport}_download.pdf
        # Ex: operations_@R0GERAN_00053_20251016_20251111_download.pdf
        match_bg = re.match(
            r'^(operations)_(.+?)_(download)\.([a-zA-Z]+)$',
            filename,
            re.IGNORECASE
        )
        if match_bg:
            return {
                'type': match_bg.group(1).lower(),
                'account': 'bg-gestion',  # Compte virtuel pour BG_GESTION
                'method': match_bg.group(3).lower(),
                'discriminant': None,
                'ext': match_bg.group(4).lower()
            }
        return None

    return {
        'type': match.group(1).lower(),
        'account': match.group(2).lower(),
        'method': match.group(3).lower(),
        'discriminant': match.group(4) if match.group(4) else None,
        'ext': match.group(5).lower()
    }


def validate_account_names_in_formatted_files(logs_dir=None, excel_file=None):
    """
    Validation centralisée des noms de comptes dans tous les fichiers formatés

    Lit tous les fichiers .tmp.csv générés par les format scripts et vérifie
    que tous les noms de comptes utilisés existent dans comptes.xlsm (feuille Avoirs).

    Args:
        logs_dir: Répertoire contenant les fichiers .tmp.csv (défaut: logs/)
        excel_file: Chemin vers comptes.xlsm (défaut: comptes.xlsm)

    Returns:
        tuple: (success: bool, invalid_accounts: dict)
               invalid_accounts = {nom_compte: [fichiers_utilisateurs]}
    """
    if logs_dir is None:
        logs_dir = BASE_DIR / 'logs'
    if excel_file is None:
        excel_file = BASE_DIR / 'comptes.xlsm'

    logs_path = Path(logs_dir)
    if not logs_path.exists():
        return True, {}

    # Lire les comptes valides depuis Excel
    valid_accounts = get_valid_accounts(excel_file, verbose=False)
    if not valid_accounts:
        # Excel non disponible ou erreur → skip validation
        return True, {}

    valid_set = set(valid_accounts)

    # Analyser tous les fichiers .tmp.csv
    tmp_files = list(logs_path.glob("*.tmp.csv"))
    if not tmp_files:
        return True, {}

    # Collecter tous les noms de comptes utilisés
    accounts_usage = {}  # {nom_compte: [fichiers]}

    for tmp_file in tmp_files:
        try:
            with open(tmp_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f, delimiter=';')
                for row in reader:
                    compte = row.get('Compte', '').strip()
                    if compte and compte != '#Solde':  # Ignorer les lignes vides et #Solde
                        if compte not in accounts_usage:
                            accounts_usage[compte] = []
                        if tmp_file.name not in accounts_usage[compte]:
                            accounts_usage[compte].append(tmp_file.name)
        except Exception:
            # Erreur lecture fichier → skip
            continue

    # Identifier les comptes invalides
    invalid_accounts = {
        compte: files
        for compte, files in accounts_usage.items()
        if compte not in valid_set
    }

    return len(invalid_accounts) == 0, invalid_accounts


class ComptaUpdater:
    def __init__(self, verbose=False, all_soldes=False):
        self.verbose = verbose  # Verbose UNIQUEMENT si --verbose explicite
        self.all_soldes = all_soldes

        # Initialiser le logger centralisé
        LOGS_DIR.mkdir(parents=True, exist_ok=True)
        self.journal_file = LOGS_DIR / "journal.log"
        self.logger = Logger(
            script_name="cpt_update",
            journal_file=self.journal_file,
            verbose=self.verbose,
            debug=DEBUG
        )

        self.session_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.session_backup = None  # Sera défini lors du backup
        self.archived_files = []  # Liste des fichiers archivés (pour rollback)

        # Composition : accès Excel délégué à ComptaExcel
        self.excel = ComptaExcelImport(
            comptes_file=COMPTES_FILE,
            verbose=verbose,
            all_soldes=all_soldes,
            logger=self.logger,
        )

        # Marquer le début de session dans le journal avec HDS
        session_header = f"\n=== SESSION HDS_{self.session_timestamp} {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===\n"
        self.logger.write_to_journal(session_header.strip())

    def backup_comptes(self):
        """Sauvegarde comptes.xlsm avec HDS (Horodatage de Session)

        comptes.xlsm reste toujours comptes.xlsm dans le répertoire courant.
        Le backup est archivé dans archives/ avec un HDS unique.
        """
        if not COMPTES_FILE.exists():
            self.logger.error(f"Fichier {COMPTES_FILE} introuvable")
            return False

        # Toujours ajouter le HDS de cette session au backup
        # (comptes.xlsm dans le répertoire courant ne doit jamais avoir de HDS)
        backup_filename = add_hds(COMPTES_FILE.name, self.session_timestamp)
        backup_file = ARCHIVES_DIR / backup_filename

        try:
            # S'assurer que le répertoire existe
            ARCHIVES_DIR.mkdir(parents=True, exist_ok=True)

            shutil.copy2(COMPTES_FILE, backup_file)
            self.session_backup = backup_file.name  # Enregistrer pour le journal
            self.logger.verbose(f"Backup: {backup_file.name}")

            # Nettoyage des anciennes sauvegardes
            self.cleanup_old_backups()
            return True
        except Exception as e:
            self.logger.error(f"Erreur lors de la sauvegarde: {e}")
            return False

    def delete_backup_if_no_changes(self):
        """Supprime le backup si aucun fichier n'a été archivé (= aucune modification)

        Appelé à la fin de process_files() pour éviter de polluer les archives
        avec des backups vides (sessions où tous les fichiers ont échoué).
        """
        if not self.archived_files and self.session_backup:
            backup_file = ARCHIVES_DIR / self.session_backup
            if backup_file.exists():
                try:
                    backup_file.unlink()
                    self.logger.verbose(f"Backup supprimé (aucun fichier traité): {self.session_backup}")
                    self.session_backup = None
                except Exception as e:
                    self.logger.error(f"Erreur suppression backup: {e}")

    def cleanup_old_backups(self):
        """Garde uniquement les MAX_SESSIONS sessions les plus récentes

        Nouvelle logique HDS:
        1. Scanner tous les HDS dans archives/
        2. Trier chronologiquement
        3. Garder les N derniers HDS
        4. Supprimer tous les fichiers avec les vieux HDS
        5. Purger le journal en conséquence
        """
        if not ARCHIVES_DIR.exists():
            return

        # Scanner tous les HDS dans archives/
        all_hds = set()
        files_by_hds = {}  # {hds: [file_paths]}

        for file_path in ARCHIVES_DIR.rglob("*"):
            if file_path.is_file():
                hds = extract_hds(file_path.name)
                if hds:
                    all_hds.add(hds)
                    if hds not in files_by_hds:
                        files_by_hds[hds] = []
                    files_by_hds[hds].append(file_path)

        if not all_hds:
            # Pas de HDS trouvés, rien à purger
            return

        # Trier les HDS chronologiquement (du plus récent au plus ancien)
        sorted_hds = sorted(all_hds, reverse=True)

        # Identifier les HDS à supprimer (au-delà de MAX_SESSIONS)
        hds_to_keep = set(sorted_hds[:MAX_SESSIONS])
        hds_to_delete = set(sorted_hds[MAX_SESSIONS:])

        if not hds_to_delete:
            # Pas de vieilles sessions à purger
            return

        # Supprimer tous les fichiers avec les vieux HDS
        deleted_count = 0
        for hds in hds_to_delete:
            for file_path in files_by_hds[hds]:
                try:
                    file_path.unlink()
                    deleted_count += 1
                    if DEBUG:
                        self.logger.verbose(f"Session HDS_{hds} purgée: {file_path.name}")
                except Exception as e:
                    self.logger.error(f"Impossible de supprimer {file_path.name}: {e}")

        if deleted_count > 0:
            self.logger.verbose(f"{deleted_count} fichier(s) purgé(s) ({len(hds_to_delete)} sessions)")

        # Purger le journal synchronisé avec les HDS gardés
        self.cleanup_old_journal_sessions(hds_to_keep)

        # Nettoyer aussi les vieux logs
        self.cleanup_old_logs()

    def cleanup_old_journal_sessions(self, hds_to_keep):
        """Supprime les sessions du journal qui ne correspondent pas aux HDS gardés

        Args:
            hds_to_keep: Set des HDS à conserver (format: YYYYMMDD_HHMMSS)
        """
        if not self.journal_file.exists():
            return

        try:
            with open(self.journal_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            new_content = []
            current_session_hds = None
            keep_current_session = False

            for line in lines:
                if line.startswith('=== SESSION'):
                    # Extraire HDS depuis le format: "=== SESSION HDS_20251120_153045 2025-11-20 15:30:45 ==="
                    parts = line.split()
                    if len(parts) >= 3:
                        hds_token = parts[2]  # "HDS_20251120_153045"
                        if hds_token.startswith('HDS_'):
                            # Nouveau format avec HDS explicite
                            current_session_hds = hds_token.replace('HDS_', '')
                        else:
                            # Ancien format, reconstituer le HDS depuis la date/heure
                            # "=== SESSION 2025-11-17 15:15:17 ==="
                            if len(parts) >= 4:
                                date_part = parts[2].replace('-', '')  # 20251117
                                time_part = parts[3].replace(':', '').replace('===', '').strip()  # 151517
                                current_session_hds = f"{date_part}_{time_part}"

                        keep_current_session = current_session_hds in hds_to_keep

                if keep_current_session:
                    new_content.append(line)

            # Réécrire le journal filtré
            with open(self.journal_file, 'w', encoding='utf-8') as f:
                f.writelines(new_content)

        except Exception as e:
            self.logger.error(f"Erreur purge journal: {e}")

    def cleanup_old_logs(self):
        """Supprime les anciens logs individuels (format update_YYYYMMDD_HHMMSS.log)

        Ces logs individuels sont obsolètes depuis l'introduction du journal centralisé.
        Garde uniquement les N plus récents (MAX_SESSIONS) par sécurité.
        """
        try:
            old_logs = sorted(LOGS_DIR.glob("update_*.log"), key=lambda p: p.stat().st_mtime, reverse=True)

            # Garder les N plus récents, supprimer les autres
            for old_log in old_logs[MAX_SESSIONS:]:
                try:
                    old_log.unlink()
                    if DEBUG:
                        self.logger.verbose(f"Ancien log supprimé: {old_log.name}")
                except Exception as e:
                    if DEBUG:
                        self.logger.error(f"Impossible de supprimer {old_log.name}: {e}")

        except Exception as e:
            self.logger.error(f"Erreur purge logs: {e}")

    def cleanup_debug_files(self):
        """Supprime tous les fichiers dans logs/debug/

        Appelé au début de chaque session pour nettoyer les fichiers
        intermédiaires de la session précédente (screenshots, HTML, tmp.csv).
        """
        if not DEBUG_DIR.exists():
            return

        try:
            deleted_count = 0
            for debug_file in DEBUG_DIR.glob('*'):
                if debug_file.is_file():
                    try:
                        debug_file.unlink()
                        deleted_count += 1
                    except Exception:
                        pass

            if deleted_count > 0:
                self.logger.verbose(f"Fichiers debug nettoyés: {deleted_count}")

        except Exception as e:
            self.logger.error(f"Erreur nettoyage debug: {e}")

    def restore_fallback(self):
        """Restaure la dernière session en inversant les déplacements HDS

        Logique:
        1. Effacer tous les fichiers dans dropbox/SITE/ (préserve les dossiers)
        2. Scanner archives/ pour trouver tous les HDS
        3. Identifier le dernier HDS (le plus récent)
        4. Restaurer tous les fichiers avec ce HDS vers leur emplacement d'origine
        5. Retirer le HDS de tous les fichiers restaurés (retour à l'état vierge)

        Note: Les fichiers sont supprimés d'archives/ après restauration,
        permettant des fallbacks successifs remontant dans l'historique.
        La traçabilité est assurée par le journal (logs/journal.log).
        """
        # 1. Effacer tous les fichiers dans dropbox/ (préserver les dossiers SITE)
        if DROPBOX_DIR.exists():
            file_count = 0
            for file_path in DROPBOX_DIR.rglob('*'):
                if file_path.is_file():
                    file_path.unlink()
                    file_count += 1
            if file_count > 0:
                self.logger.verbose(f"Purge dropbox/: {file_count} fichier(s) supprimé(s)")
        # Scanner tous les fichiers dans archives/ récursivement
        all_hds = set()
        archived_files = []  # Liste de tuples (hds, file_path)

        if not ARCHIVES_DIR.exists():
            self.logger.error("Répertoire archives/ inexistant")
            return False

        # Collecter tous les fichiers avec HDS
        for file_path in ARCHIVES_DIR.rglob("*"):
            if file_path.is_file():
                hds = extract_hds(file_path.name)
                if hds:
                    all_hds.add(hds)
                    archived_files.append((hds, file_path))

        if not all_hds:
            self.logger.error("Aucune session archivée trouvée (pas de HDS)")
            return False

        # Trouver le dernier HDS (tri chronologique)
        latest_hds = sorted(all_hds, reverse=True)[0]

        # Logger le fallback
        try:
            fallback_header = f"\n=== FALLBACK {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===\n"
            fallback_header += f"Session annulée: HDS_{latest_hds}\n"
            with open(self.journal_file, 'a', encoding='utf-8') as f:
                f.write(fallback_header)
        except Exception:
            pass

        self.logger.verbose(f"Restauration de la session: HDS_{latest_hds}")

        # Restaurer tous les fichiers avec ce HDS
        restored_count = 0
        try:
            for hds, archive_path in archived_files:
                if hds != latest_hds:
                    continue

                # Déterminer l'emplacement de restauration
                # Le HDS est retiré de tous les fichiers (retour à l'état vierge)
                # archives/comptes_HDS_xxx.xlsx → ./comptes.xlsm
                # archives/SG/file_HDS_xxx.csv → dropbox/SG/file.csv
                # archives/BG_GESTION/report_HDS_xxx.pdf → dropbox/BG_GESTION/report.pdf

                filename = archive_path.name
                filename_sans_hds = remove_hds(filename)
                parent_dir = archive_path.parent

                if parent_dir == ARCHIVES_DIR:
                    # Fichier à la racine d'archives/ (comptes*.xlsx)
                    dest_path = BASE_DIR / COMPTES_FILE.name
                elif parent_dir.name in ['PDF', 'CSV']:
                    # Ancienne structure (legacy) - retirer le HDS
                    dest_path = DROPBOX_DIR / filename_sans_hds
                else:
                    # Nouvelle structure: archives/SITE/ → dropbox/SITE/
                    # Retirer le HDS du fichier
                    site_name = parent_dir.name
                    dest_dir = DROPBOX_DIR / site_name
                    dest_dir.mkdir(parents=True, exist_ok=True)
                    dest_path = dest_dir / filename_sans_hds

                try:
                    # Restaurer le fichier (move)
                    shutil.move(str(archive_path), str(dest_path))
                    self.logger.verbose(f"Restauré: {filename}")
                    restored_count += 1
                except Exception as e:
                    self.logger.error(f"Erreur restauration {filename}: {e}")

            if restored_count > 0:
                self.logger.verbose(f"{restored_count} fichier(s) restauré(s) depuis archives/")
                return True
            else:
                self.logger.error("Aucun fichier restauré")
                return False

        except Exception as e:
            self.logger.error(f"Erreur lors de la restauration: {e}")
            return False

    def archive_file(self, file_path):
        """Archive un fichier traité dans le sous-répertoire correspondant

        Préserve la structure: dropbox/SITE/ -> archives/SITE/
        Ajoute un HDS au nom du fichier (une seule fois dans sa vie)
        Enregistre l'opération pour permettre un rollback
        """
        # Déterminer le sous-répertoire source (BG_GESTION, SG, etc.)
        source_subdir = file_path.parent.name

        # Si le fichier est directement dans dropbox/ (ancienne structure)
        if source_subdir == 'dropbox':
            if file_path.suffix.lower() == '.pdf':
                dest_dir = ARCHIVES_DIR / "PDF"
            else:
                dest_dir = ARCHIVES_DIR / "CSV"
        else:
            # Nouvelle structure: archives/SITE/
            dest_dir = ARCHIVES_DIR / source_subdir

        dest_dir.mkdir(parents=True, exist_ok=True)

        # Ajouter HDS si pas déjà présent (un fichier = un HDS à vie)
        current_name = file_path.name
        if has_hds(current_name):
            # Fichier déjà estampillé (après fallback), on garde le HDS original
            dest_filename = current_name
        else:
            # Premier archivage, ajouter le HDS de cette session
            dest_filename = add_hds(current_name, self.session_timestamp)

        dest_file = dest_dir / dest_filename

        try:
            # Enregistrer l'opération pour rollback
            archive_op = {
                'source': str(file_path),
                'destination': str(dest_file)
            }

            shutil.move(str(file_path), str(dest_file))

            # Stocker dans la liste pour rollback
            self.archived_files.append(archive_op)

            self.logger.verbose(f"Archivé: {file_path.name} → {dest_dir.name}/{dest_file.name}")
            return True
        except Exception as e:
            self.logger.error(f"Erreur archivage {file_path.name}: {e}")
            return False

    def process_and_archive(self, files, handler_method):
        """Pattern générique: traiter fichiers puis archiver systématiquement

        Args:
            files: Liste de fichiers à traiter
            handler_method: Méthode à appeler pour traiter les fichiers

        Returns:
            bool: True si succès complet, False sinon

        Note:
            Archive tous les fichiers après traitement (succès ou erreur).
        """
        if not files:
            return True

        success = handler_method(files)

        # Archivage systématique (succès ou erreur)
        for file_path in files:
            self.archive_file(file_path)

        return success

    def collect_files_by_convention(self):
        """Collecte les fichiers dropbox/ selon la convention de nommage

        Scan dropbox/ racine uniquement (les sous-répertoires de sites sont traités par format_site).
        Parse chaque nom de fichier avec parse_filename() et groupe par type.

        Note: Les fichiers dans dropbox/SITE/ sont traités par format_site() dans process_dropbox().
              Cette fonction ne scanne que les fichiers à la racine de dropbox/.

        Returns:
            dict: {
                'positions': [liste de Path],  # Format unifié 5 colonnes
                'soldes': [liste de Path],
                'unknown': [liste de Path]  # Fichiers non reconnus (archivés sans traitement)
            }
        """
        files_by_type = {
            'positions': [],
            'soldes': [],
            'unknown': []
        }

        # Scanner uniquement dropbox/ racine (pas les sous-répertoires de sites)
        for file_path in DROPBOX_DIR.glob("*.*"):
            # Ignorer les répertoires
            if not file_path.is_file():
                continue

            # Parser le nom de fichier selon la convention stricte
            parsed = parse_filename(file_path.name)

            if not parsed:
                # Fichier non reconnu → archiver sans traitement
                files_by_type['unknown'].append(file_path)
                self.logger.verbose(f"Fichier non reconnu: {file_path.name}")
                continue

            file_type = parsed['type']

            # Dispatcher selon le type (positions et soldes seulement, pas operations)
            if file_type == 'positions':
                files_by_type['positions'].append(file_path)
            elif file_type == 'soldes':
                files_by_type['soldes'].append(file_path)
            else:
                # Fichiers operations à la racine → traiter comme unknown (legacy)
                files_by_type['unknown'].append(file_path)
                self.logger.verbose(f"Fichier operations en racine ignoré: {file_path.name}")

        return files_by_type

    def cleanup_temp_files(self):
        """Supprime les fichiers temporaires .tmp.csv dans logs/

        En mode DEBUG, cpt_update.py écrit le stdout des format scripts dans logs/*.tmp.csv
        pour permettre l'inspection manuelle du contenu formaté.

        Ces fichiers sont conservés en mode DEBUG, supprimés sinon.
        """
        if not LOGS_DIR.exists():
            return

        # Trouver tous les .tmp.csv dans logs/
        tmp_files = list(LOGS_DIR.glob('*.tmp.csv'))

        # En mode DEBUG, conserver les fichiers temporaires
        if DEBUG:
            if tmp_files:
                self.logger.verbose(f"Mode DEBUG: {len(tmp_files)} fichier(s) temporaire(s) conservé(s) pour inspection")
            return

        # Sinon, supprimer les .tmp.csv
        if tmp_files:
            for tmp_file in tmp_files:
                try:
                    tmp_file.unlink()
                    self.logger.verbose(f"Supprimé: {tmp_file.name}")
                except Exception as e:
                    self.logger.error(f"Erreur suppression {tmp_file.name}: {e}")

            self.logger.verbose(f"Nettoyage: {len(tmp_files)} fichier(s) temporaire(s) supprimé(s)")

    def deduplicate_files_by_basename(self, files):
        """Déduplique les fichiers par nom de base, garde le plus récent

        Ex: 00050659433.csv, 00050659433(1).csv, 00050659433(2).csv
        → même base "00050659433.csv" → garder le plus récent (mtime)

        Pattern détecté: nom(N).ext ou nom (N).ext ou nom#N.ext ou nom #N.ext

        NOTE: Pour eToro, les PDFs de nature différente doivent être renommés
        lors de la collecte manuelle (ex: eToro_portfolio.pdf, eToro_transactions.pdf)
        """
        import re
        from collections import defaultdict

        if not files:
            return files

        # Grouper par (site, nom_base)
        groups = defaultdict(list)

        for f in files:
            site = f.parent.name  # ETORO, SG, etc.
            name = f.name

            # Extraire le nom de base (sans suffixe de copie)
            # Patterns: nom(1).ext, nom (1).ext, nom#2.ext, nom #2.ext
            base_name = re.sub(r'\s*[\(#]\d+[\)]?\s*(?=\.[^.]+$)', '', name)

            groups[(site, base_name)].append(f)

        # Pour chaque groupe, garder le plus récent
        result = []
        for key, group in groups.items():
            if len(group) == 1:
                result.append(group[0])
            else:
                # Plusieurs fichiers avec même base → garder le plus récent
                group.sort(key=lambda f: f.stat().st_mtime, reverse=True)
                kept = group[0]
                skipped = group[1:]
                result.append(kept)
                # Archiver les fichiers ignorés (doublons)
                for f in skipped:
                    self.logger.verbose(f"  Fichier doublon ignoré: {f.name} (plus récent: {kept.name})")
                    self.archive_file(f)

        return result

    def process_dropbox(self):
        """Traite tous les fichiers dans le répertoire dropbox selon la convention de nommage"""
        # Nettoyage fichiers debug de la session précédente
        self.cleanup_debug_files()

        if not DROPBOX_DIR.exists():
            self.logger.error(f"Répertoire dropbox inexistant: {DROPBOX_DIR}")
            return False

        # Collecter les fichiers par type (utilise parse_filename)
        # Collecter fichiers à la racine de dropbox/ (legacy)
        files_by_type = self.collect_files_by_convention()
        positions_files = files_by_type['positions']
        soldes_files = files_by_type['soldes']
        unknown_files = files_by_type['unknown']

        # Dédupliquer les fichiers positions par nom de base
        positions_files = self.deduplicate_files_by_basename(positions_files)

        # Compter les fichiers (racine + sous-répertoires de sites)
        total_files = len(positions_files) + len(soldes_files) + len(unknown_files)
        for site in SITE_FORMAT_MODULES.keys():
            site_dir = get_site_dropbox_dir(site)
            if site_dir.exists():
                total_files += len(list(site_dir.glob('*.*')))

        if total_files == 0:
            self.logger.verbose("Aucun fichier à traiter dans dropbox/")
            return True

        self.logger.verbose(
            f"Fichiers trouvés: {len(positions_files)} positions (racine), "
            f"{len(soldes_files)} soldes, {len(unknown_files)} non reconnus"
        )

        # Sauvegarde avant toute modification
        if not self.backup_comptes():
            return False

        # Ouvre le fichier Excel
        if not self.excel.open_workbook():
            return False

        success = True
        has_errors = False  # Erreurs non-bloquantes (fichiers individuels)

        # Traite les fichiers d'opérations (conversions + imports)
        # AGRÉGATION: collecter toutes les opérations avant import pour déduplication globale des #Solde
        all_operations = []  # Toutes les Operation (dataclass)

        # Traitement de tous les sites via nouvelle interface Python directe (format_site)
        for site, format_module in SITE_FORMAT_MODULES.items():
            site_dir = get_site_dropbox_dir(site)
            if not site_dir.exists():
                continue

            # Collecter les fichiers à traiter (ZIPs + fichiers de données)
            # Note: glob sur Linux est sensible à la casse, inclure les deux
            zip_files = list(site_dir.glob('*.zip')) + list(site_dir.glob('*.ZIP'))
            data_files = (list(site_dir.glob('*.csv')) + list(site_dir.glob('*.CSV')) +
                          list(site_dir.glob('*.xlsx')) + list(site_dir.glob('*.XLSX')) +
                          list(site_dir.glob('*.pdf')) + list(site_dir.glob('*.PDF')) +
                          list(site_dir.glob('*.txt')) + list(site_dir.glob('*.TXT')) +
                          list(site_dir.glob('*.tsv')) + list(site_dir.glob('*.TSV')))

            if not zip_files and not data_files:
                continue

            self.logger.verbose(f"Traitement {site} (interface Python directe)")
            site_logger = self.logger.with_prefix(site)
            try:
                ops, pos = format_module.format_site(site_dir, verbose=self.verbose, logger=site_logger)
            except Exception as e:
                self.logger.error(f"  ❌ Erreur format_site({site}): {e}")
                has_errors = True
                continue

            # Log CSV debug si DEBUG activé
            if DEBUG and (ops or pos):
                format_module.log_csv_debug(ops, pos, site_dir, self.logger)

            # Convertir tuples → Operation et ajouter aux opérations agrégées
            if ops:
                all_operations.extend(Operation.from_tuple(t) for t in ops)
                self.logger.verbose(f"  {site}: {len(ops)} opérations")

            # Positions → traiter via process_valorisations
            if pos:
                pos_csv = tuples_to_csv(pos, header='Date;Ligne;Montant;Compte;Sous-compte')
                # Écrire dans un fichier temporaire (dans DEBUG_DIR, pas dropbox/)
                DEBUG_DIR.mkdir(parents=True, exist_ok=True)
                temp_pos_file = DEBUG_DIR / f'{site}_positions_temp.csv'
                with open(temp_pos_file, 'w', encoding='utf-8') as f:
                    f.write(pos_csv)
                positions_files.append(temp_pos_file)
                self.logger.verbose(f"  {site}: {len(pos)} positions")

            # Archiver les fichiers traités
            for zip_path in zip_files:
                self.archive_file(zip_path)
                self.logger.verbose(f"  ZIP archivé: {zip_path.name}")

            for data_file in data_files:
                # Ne pas archiver les fichiers temporaires
                if 'temp' not in data_file.name:
                    self.archive_file(data_file)
                    self.logger.verbose(f"  Fichier archivé: {data_file.name}")

        # Import agrégé (une seule fois pour tous les fichiers)
        if all_operations:
            if not self.excel.append_to_comptes(all_operations):
                has_errors = True

        # Traite tous les fichiers positions/supports (format unifié 5 colonnes)
        # Continue même si has_errors (traite les fichiers qui peuvent réussir)
        if positions_files:
            if not self.process_and_archive(positions_files, self.excel.process_valorisations):
                has_errors = True

        # Générer les #Solde manquants pour les comptes avec opérations mais sans solde
        self.excel.generate_missing_soldes()

        # Vérifier que les dates Plus_value ont été mises à jour
        self.excel.verify_plus_value_dates()

        # Archive les fichiers soldes (sans import - contenu déjà dans operations/positions)
        # Continue même si has_errors
        if soldes_files:
            for file_path in soldes_files:
                self.archive_file(file_path)
                self.logger.verbose(f"  Archivé: {file_path.name}")

        # Archive les fichiers non reconnus (sans traitement)
        if unknown_files:
            for file_path in unknown_files:
                self.archive_file(file_path)
                self.logger.verbose(f"  Archivé (non reconnu): {file_path.name}")

        # Validation centralisée des noms de comptes AVANT sauvegarde Excel
        # SEULE erreur BLOQUANTE qui empêche la sauvegarde
        validation_ok, invalid_accounts = validate_account_names_in_formatted_files()
        if not validation_ok:
            self.logger.error("❌ VALIDATION: Noms de comptes invalides détectés")
            self.logger.error("   Ces comptes n'existent pas dans comptes.xlsm (feuille Avoirs):")
            for compte, files in invalid_accounts.items():
                self.logger.error(f"     - '{compte}' utilisé dans:")
                for f in files:
                    self.logger.error(f"         {f}")
            self.logger.error("   → Corriger les scripts Tier 1/2 pour utiliser les bons noms de comptes")
            self.logger.error("   → Consulter la feuille Avoirs de comptes.xlsm pour les noms valides")
            success = False  # Erreur bloquante

        # Sauvegarde et ferme le fichier
        # Sauvegarde MÊME s'il y a des erreurs non-bloquantes (has_errors)
        # Ne sauvegarde PAS s'il y a erreur bloquante (success=False)
        self.excel.close_workbook(save=success)

        # Recalcul + miroir C1 si lancé depuis la GUI
        if success and os.environ.get('COMPTA_GUI'):
            from inc_uno import refresh_controles
            refresh_controles(COMPTES_FILE, self.logger)

        # Supprimer le backup si aucun fichier n'a été archivé (évite sessions vides)
        self.delete_backup_if_no_changes()

        # Vérifie le contrôle (COMPTES = erreur bloquante)
        if success:
            if not self.excel.open_workbook():
                return False
            if not self.excel.check_control_sheet():
                success = False  # ERREUR COMPTES = échec
            self.excel.close_workbook(save=False)

        # Nettoyer les fichiers temporaires .tmp.csv générés par les formatters
        self.cleanup_temp_files()

        # Retourner False seulement si erreur bloquante
        # Les erreurs non-bloquantes sont informatives (fichiers restent dans dropbox)
        return success

    def print_stats(self):
        """Affiche les statistiques du traitement"""
        stats = self.excel.stats
        if stats['lines_added'] > 0:
            print(f"\nOpérations ajoutées:")
            for compte, count in sorted(stats['operations_by_compte'].items()):
                print(f"  {compte}: {count}")
        else:
            print(f"\nAucune nouvelle opération")

        if stats['positions_by_compte']:
            print(f"\nPositions mises à jour:")
            for compte, count in sorted(stats['positions_by_compte'].items()):
                print(f"  {compte}: {count}")

        if stats['errors']:
            print(f"\n⚠️  Erreurs ({len(stats['errors'])}):")
            for error in stats['errors']:
                print(f"  - {error}")

    def print_valorisations_summary(self):
        """Affiche le récapitulatif des valorisations mises à jour"""
        if self.excel.comptes_valorisations:
            print("\nValorisations mises à jour:")
            total_valorisation = 0.0
            for compte, valorisation in self.excel.comptes_valorisations.items():
                # Nom court du compte pour l'affichage
                compte_court = compte.replace('Ass vie ébène', 'Ébène')
                print(f"  {compte_court}: {valorisation:,.2f} €".replace(',', ' '))
                total_valorisation += valorisation
            if len(self.excel.comptes_valorisations) > 1:
                print(f"  Total: {total_valorisation:,.2f} €".replace(',', ' '))

def main():

    parser = argparse.ArgumentParser(
        description='Script de mise à jour de comptes.xlsm',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemples:
  %(prog)s                           # Traite les fichiers dans dropbox/
  %(prog)s --fallback                # Restaure la dernière sauvegarde
  %(prog)s --archive-only            # Archive sans importer (import manuel fait)

Cas d'usage --archive-only:
  Lorsqu'un import échoue et que vous corrigez Excel manuellement :
  1. Corrigez les données dans Excel
  2. Lancez: %(prog)s --archive-only
  3. Les fichiers sont archivés avec entrées ARCHIVE dans le journal
  4. Cohérence maintenue pour le fallback
        """)

    parser.add_argument('--fallback',
                        action='store_true',
                        help='Restaure la dernière sauvegarde de comptes.xlsm')
    parser.add_argument('--archive-only',
                        action='store_true',
                        help='Archive les fichiers de dropbox SANS les importer (import manuel déjà fait)')
    parser.add_argument('--all-soldes',
                        action='store_true',
                        help='Écrire tous les #Solde collectés (pas seulement ceux avec nouvelles opérations)')
    parser.add_argument('--no-pair',
                        action='store_true',
                        help='Ne pas lancer l\'appariement (cpt_pair) après l\'import')
    parser.add_argument('-v', '--verbose',
                        action='store_true',
                        help='Mode verbeux')
    parser.add_argument('--TNR',
                        action='store_true',
                        help='Mode test de non-régression (désactive filtrage date, API exchange)')

    args = parser.parse_args()

    # TNR : positionner les flags module-level (pas de variable d'environnement)
    if args.TNR:
        import inc_format
        import inc_exchange_rates
        inc_format.TNR_MODE = True
        inc_exchange_rates.TNR_MODE = True

    updater = ComptaUpdater(verbose=args.verbose, all_soldes=args.all_soldes)

    timestamp = datetime.now().strftime('%H:%M:%S')
    print(f"{timestamp} Update")

    # Mode fallback
    if args.fallback:
        if updater.restore_fallback():
            updater.logger.info("Restauration effectuée avec succès")
            sys.exit(0)
        else:
            updater.logger.error("Échec de la restauration")
            sys.exit(1)

    # Mode archive-only (archiver sans importer)
    if args.archive_only:
        if not DROPBOX_DIR.exists():
            updater.logger.error("Répertoire dropbox inexistant")
            sys.exit(1)

        # Lister les fichiers (racine + sites configurés uniquement)
        pdf_files = list(DROPBOX_DIR.glob("*.pdf")) + list(DROPBOX_DIR.glob("*.PDF"))
        csv_files = list(DROPBOX_DIR.glob("*.csv")) + list(DROPBOX_DIR.glob("*.CSV"))

        # Fichiers dans les sous-répertoires de sites configurés uniquement
        for site in SITE_FORMAT_MODULES.keys():
            site_dir = get_site_dropbox_dir(site)
            if site_dir.exists() and site_dir.is_dir():
                pdf_files += list(site_dir.glob("*.pdf")) + list(site_dir.glob("*.PDF"))
                csv_files += list(site_dir.glob("*.csv")) + list(site_dir.glob("*.CSV"))

        pdf_files = list(set(pdf_files))
        csv_files = list(set(csv_files))
        all_files = pdf_files + csv_files

        if not all_files:
            print("✓ Aucun fichier à archiver dans dropbox")
            sys.exit(0)

        print(f"Fichiers à archiver: {len(pdf_files)} PDF, {len(csv_files)} CSV")

        # Créer un backup d'abord (au cas où)
        if not updater.backup_comptes():
            updater.logger.error("Impossible de créer le backup")
            sys.exit(1)

        # Archiver tous les fichiers
        archived_count = 0
        for file_path in all_files:
            if updater.archive_file(file_path):
                archived_count += 1

        print(f"\n✓ {archived_count} fichier(s) archivé(s) (import manuel supposé fait)")
        print("Note: Le journal contient les entrées ARCHIVE pour cohérence")

        print(f"\n💡 Lancez: ./cpt_pair.py --pair  (appariement)")

        sys.exit(0)

    # Mode normal: traitement
    success = updater.process_dropbox()
    print(f"\n{datetime.now().strftime('%H:%M:%S')} Import terminé")
    updater.print_valorisations_summary()
    updater.print_stats()

    # Appariement automatique (sauf --no-pair)
    lines_added = updater.excel.stats['lines_added']
    if lines_added > 0 and not args.no_pair:
        print(f"\n{datetime.now().strftime('%H:%M:%S')} Appariement")
        cpt_pair_script = BASE_DIR / 'cpt_pair.py'
        pair_cmd = [sys.executable, str(cpt_pair_script), '--pair']
        if args.verbose:
            pair_cmd.append('-v')
        pair_result = subprocess.run(pair_cmd, cwd=BASE_DIR)
        if pair_result.returncode != 0:
            updater.logger.warning("Erreur lors de l'appariement")
    elif lines_added > 0 and args.no_pair:
        print(f"\n💡 Lancez: ./cpt_pair.py --pair  (appariement des {lines_added} nouvelle(s) ligne(s))")

        # Validation des noms de comptes
        try:
            valid_accounts = set(get_valid_accounts(COMPTES_FILE))
            if valid_accounts:
                import openpyxl
                wb = openpyxl.load_workbook(COMPTES_FILE, data_only=True)
                ws = wb[SHEET_OPERATIONS]
                used_accounts = set()
                for row in ws.iter_rows(min_row=max(2, ws.max_row - lines_added), max_row=ws.max_row, values_only=True):
                    if row[7]:  # Colonne Compte (index 7)
                        used_accounts.add(row[7])
                wb.close()
                invalid = used_accounts - valid_accounts - {''}
                if invalid:
                    updater.logger.warning(f"Comptes non reconnus: {', '.join(sorted(invalid))}")
                else:
                    updater.logger.info("Comptes OK")
        except Exception as e:
            updater.logger.warning(f"Validation comptes: {e}")

    # Comparaison avec l'archive précédente (informatif)
    if lines_added > 0:
        print(f"\n{datetime.now().strftime('%H:%M:%S')} Comparaison avec archive précédente")
        try:
            from inc_compare_xlsx import compare_xlsx_with_prev
            compare_xlsx_with_prev(COMPTES_FILE, ARCHIVES_DIR)
        except Exception as e:
            updater.logger.warning(f"Comparaison archive: {e}")

    if not success:
        updater.logger.error("Le traitement a échoué ou contient des erreurs")
        sys.exit(1)

    sys.exit(0)


if __name__ == "__main__":
    main()
