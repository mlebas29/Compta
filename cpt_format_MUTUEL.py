#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""cpt_format_MUTUEL.py — formateur Crédit Mutuel.

Parse l'export Excel multi-comptes du Crédit Mutuel (menu « Téléchargement » :
un seul `.xlsx` contenant une feuille de synthèse « Vos comptes » + une feuille
par compte mouvementé `Cpt <guichet> <numéro>`) vers le format standard interne.

- Opérations (9 champs) : une par mouvement, depuis chaque feuille compte.
  Montant = Crédit (positif) sinon Débit (déjà signé négatif). Les lignes à
  montant nul (« NOUVEAU TAUX… ») sont ignorées (informatives).
- #Solde (9 champs) : un par compte, depuis la synthèse « Vos comptes » — couvre
  aussi les comptes sans feuille de mouvements (ex. prêts), absents des onglets.

L'onglet « Vos comptes » fait office de **registre** : tous les comptes y sont
déclarés (RIB + nom + solde). Tout est importé.

Mapping comptes : **par RIB → nom de compte classeur** (config privée). La clé de
jointure est le **numéro de compte** (dernier bloc du RIB), présent à la fois
dans la synthèse (`12345 06789 00001234567`) et dans le nom d'onglet
(`Cpt 06789 00001234567`). Source : `config_accounts.json[MUTUEL]["accounts"]`
(liste de `{"rib": "00001234567", "name": "Livret X"}`) ; fallback = nom brut
du relevé. Le RIB désambiguïse les homonymes (ex. plusieurs prêts immobiliers
de même intitulé).
"""

import sys
import json
import re
from pathlib import Path
from datetime import datetime, date

import inc_categorize
from inc_format import process_files, get_file_date, site_name_from_file, base_dir, log_csv_debug as _log_csv_debug

SITE = site_name_from_file(__file__)  # → 'MUTUEL'

try:
    import openpyxl as xl
except ImportError:
    xl = None

SUMMARY_SHEET = 'Vos comptes'
HEADER = 'Date;Libellé;Montant;Devise;Equiv;Réf;Catégorie;Compte;Commentaire'

# Champs de création de compte (onglet Comptes) — hook d'extensibilité GUI
# (inc_format.get_account_fields → gui_accounts._site_account_fields). Un compte
# MUTUEL = 2 champs : le nom (générique) + le RIB (ici), persisté dans
# config_accounts.json et utilisé comme clé de mapping RIB→nom.
ACCOUNT_FIELDS = [('RIB :', 'rib', 'entry', None)]

# Fichiers attendus dans dropbox/MUTUEL/ (l'export natif s'appelle tous_comptes.xlsx)
EXPECTED_FILES = [('tous_comptes*.xlsx', 'glob', '1+')]


def _account_key(rib):
    """Numéro de compte = dernier bloc du RIB (clé de jointure synthèse↔onglet↔config)."""
    return str(rib or '').split()[-1] if str(rib or '').strip() else ''


def _load_rib_map():
    """Mapping `numéro de compte → nom classeur` (config privée, vide par défaut).

    Lit `config_accounts.json[MUTUEL]["accounts"]` (liste de `{rib, name}`).
    """
    p = base_dir() / 'config_accounts.json'
    try:
        with open(p, encoding='utf-8') as f:
            accounts = json.load(f).get(SITE, {}).get('accounts', [])
    except (FileNotFoundError, json.JSONDecodeError):
        return {}
    return {_account_key(a['rib']): a['name']
            for a in accounts if a.get('rib') and a.get('name')}


RIB_MAP = _load_rib_map()


def _account_name(key, raw_name):
    """Nom classeur depuis le numéro de compte ; fallback = nom brut du relevé."""
    return RIB_MAP.get(key, (raw_name or '').strip())


def _fmt_amount(value):
    """float → montant français `'-1234,56'`."""
    return f"{float(value):.2f}".replace('.', ',')


def _fmt_date(value):
    """datetime/date → `'DD/MM/YYYY'` ; sinon best-effort str."""
    if isinstance(value, (datetime, date)):
        return value.strftime('%d/%m/%Y')
    return str(value).strip()


def _parse_summary(ws):
    """Synthèse « Vos comptes » → `(situation_date_str, [ {name, rib, solde, dev} ])`.

    Couvre TOUS les comptes (y compris sans feuille de mouvements, ex. prêts).
    """
    rows = list(ws.iter_rows(values_only=True))
    situation_date = None
    if rows:
        m = re.search(r'(\d{2}/\d{2}/\d{4})', str(rows[0][0] or ''))
        if m:
            situation_date = m.group(1)

    # Localiser l'entête [Compte, R.I.B., Solde, Dev]
    start = None
    for i, row in enumerate(rows):
        if row and str(row[0] or '').strip() == 'Compte':
            start = i + 1
            break

    accounts = []
    if start is not None:
        for row in rows[start:]:
            name = row[0] if row else None
            if name is None or str(name).strip() == '':
                break  # ligne vide = fin de la liste
            accounts.append({
                'name': str(name).strip(),
                'rib': str(row[1]).strip() if len(row) > 1 and row[1] else '',
                'solde': row[2] if len(row) > 2 else None,
                'dev': str(row[3]).strip() if len(row) > 3 and row[3] else 'EUR',
            })
    return situation_date, accounts


def _parse_account_sheet(ws):
    """Feuille `Cpt …` → `(key, raw_name, [ (date_str, libelle, montant_str, dev) ])`.

    `key` = numéro de compte extrait du nom d'onglet (`Cpt 06789 00001234567`).
    """
    key = _account_key(ws.title.replace('Cpt', '', 1))
    rows = list(ws.iter_rows(values_only=True))
    name, devise = None, 'EUR'
    if rows:
        m = re.search(r'Situation de votre compte (.+?) \((\w+)\) au', str(rows[0][0] or ''))
        if m:
            name = m.group(1).strip()
            devise = m.group(2).strip()

    ops = []
    for row in rows:
        if not row:
            continue
        d = row[0]
        if not isinstance(d, (datetime, date)):
            continue  # header / footer / « Solde au… » / « Liste de vos comptes »
        libelle = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
        debit = row[3] if len(row) > 3 else None
        credit = row[4] if len(row) > 4 else None
        if credit not in (None, ''):
            montant = credit
        elif debit not in (None, ''):
            montant = debit
        else:
            continue
        try:
            if float(montant) == 0:
                continue  # ligne informative (NOUVEAU TAUX…)
        except (TypeError, ValueError):
            continue
        dev = str(row[6]).strip() if len(row) > 6 and row[6] else devise
        ops.append((_fmt_date(d), libelle, _fmt_amount(montant), dev))
    return key, name, ops


def process_workbook(file_path, logger=None):
    """Parse l'export multi-comptes → liste de tuples 9 champs (opérations + #Solde)."""
    if xl is None:
        print("❌ openpyxl requis: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = xl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        # 1) Synthèse → date de relevé + soldes de tous les comptes
        situation_date, summary = (None, [])
        if SUMMARY_SHEET in wb.sheetnames:
            situation_date, summary = _parse_summary(wb[SUMMARY_SHEET])
        solde_date = situation_date or get_file_date(file_path)

        out = []

        # 2) Opérations, une feuille par compte (mapping par RIB → nom classeur)
        for ws in wb.worksheets:
            if not ws.title.startswith('Cpt'):
                continue  # ignore « Vos comptes » et les feuilles hidden*
            key, raw_name, ops = _parse_account_sheet(ws)
            compte = _account_name(key, raw_name)
            for date_str, libelle, montant, dev in ops:
                cat, opts = inc_categorize.categorize_operation(libelle, SITE)
                ref = opts.get('ref', '')
                out.append((date_str, libelle, montant, dev, '', ref, cat, compte, ''))

        # 3) #Solde par compte depuis la synthèse (couvre les prêts sans feuille)
        seen = {}
        for acc in summary:
            key = _account_key(acc['rib'])
            compte = _account_name(key, acc['name'])
            if acc['solde'] in (None, ''):
                continue
            if compte in seen and seen[compte] != key:
                msg = (f"deux comptes (n° {seen[compte]} et {key}) retombent sur "
                       f"le même nom « {compte} » → soldes ambigus ; précise le mapping")
                if logger:
                    logger.warning(msg)
                else:
                    print(f"⚠ [{SITE}] {msg}", file=sys.stderr)
            seen[compte] = key
            out.append((solde_date, 'Relevé compte', _fmt_amount(acc['solde']),
                        acc.get('dev', 'EUR'), '', '', '#Solde', compte, ''))

        return out
    finally:
        wb.close()


def format_site(site_dir, verbose=False, logger=None):
    """API pour Update : traite l'export Excel de dropbox/MUTUEL/."""
    if logger is None:
        from inc_logging import Logger
        logger = Logger(SITE, verbose=verbose)

    from inc_format import verify_dropbox_files
    for w in verify_dropbox_files(site_dir, SITE):
        logger.warning(w)

    def _handler(f):
        return process_workbook(f, logger=logger)

    handlers = [('tous_comptes*.xlsx', _handler, 'ops')]
    return process_files(site_dir, handlers, verbose, SITE, logger=logger)


def log_csv_debug(operations, positions, site_dir, logger=None):
    """Wrapper vers inc_format.log_csv_debug()"""
    _log_csv_debug(SITE, operations, positions, logger)


if __name__ == '__main__':
    from inc_format import cli_main
    cli_main(format_site)
