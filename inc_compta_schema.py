"""Schéma comptable Compta : constantes métier et fonctions de chargement xlsm.

Module neutre (pas de dépendance tkinter) — extrait de cpt_gui.py pour permettre
l'usage depuis des environnements headless (LO Python 3.8 sans tkinter, scripts
CLI, TNR).

Constantes :
  - ACCOUNT_TYPES, BIEN_TYPES, SOUS_TYPES_BASE
  - PV_SECTION_LABELS, PV_SECTION_TOTALS
  - AVOIRS_K_FORMATS (alias de FORMATS_DEVISE de inc_formats)

Fonctions :
  - cours_name(code) : nom du named range cours pour une devise
  - load_pv_titles(target, wb_values=None) : charge target.pv_titles depuis Plus_value
  - load_budget_categories(target, wb_values=None) : charge target.budget_* depuis Budget

Les fonctions load_* écrivent leurs résultats sur l'objet `target` (ConfigGUI ou
HeadlessGUI) qui doit exposer : xlsx_path, cr (ColResolver), _start_pvl.
"""

import openpyxl

from inc_formats import FORMATS_DEVISE as AVOIRS_K_FORMATS
from inc_excel_schema import SHEET_BUDGET, SHEET_CONTROLES, SHEET_PLUS_VALUE


# ============================================================================
# EXCEPTIONS MÉTIER
# ============================================================================

class BusinessError(Exception):
    """Refus métier légitime côté worker daemon (garde, invariant, validation).

    À distinguer des vraies pannes : le daemon logge en `journal.log` avec
    niveau ⚠️ (warning) et sans traceback, plutôt que ❌ (error) avec stack.
    Convention : toute nouvelle garde métier côté daemon utilise cette classe.
    """
    pass


# ============================================================================
# CONSTANTES MÉTIER
# ============================================================================

ACCOUNT_TYPES = [
    'Portefeuilles', 'Euros', 'Devises étrangères',
    'Crypto monnaies', 'Créances', 'Dettes',
]

BIEN_TYPES = ['Foncier', 'Mobilier']

SOUS_TYPES_BASE = ['Euro', 'Foncier', 'Titres', 'Mobilier']

# Marqueurs de section Plus_value (col A) : section_id → label
PV_SECTION_LABELS = {
    'portefeuilles': 'Les portefeuilles',
    'métaux': 'Les métaux',
    'crypto': 'Les cryptos',
    'devises': 'Les devises',
}

# Mapping famille cotation → label TOTAL dans Plus_value
PV_SECTION_TOTALS = {
    'metal': 'TOTAL métaux',
    'crypto': 'TOTAL crypto-monnaies',
    'fiat': 'TOTAL devises',
}


# ============================================================================
# HELPERS
# ============================================================================

def cours_name(code):
    """Retourne le nom du named range cours pour une devise (None si EUR)."""
    if not code or code == 'EUR':
        return None
    return f'cours_{code}'


# ============================================================================
# CHARGEMENT FEUILLES XLSM
# ============================================================================

def load_pv_titles(target, wb_values=None):
    """Charge les titres de portefeuille depuis Plus_value.

    Écrit target.pv_titles = {account_name: [(title, devise, row), ...]}.
    """
    target.pv_titles = {}
    if not target.xlsx_path:
        return
    close_wb = False
    if wb_values is None:
        wb_values = openpyxl.load_workbook(target.xlsx_path, data_only=True)
        close_wb = True
    try:
        ws = wb_values[SHEET_PLUS_VALUE]
        pvl_data_start = (target._start_pvl or 5) + 1
        for row_idx in range(pvl_data_start, ws.max_row + 1):
            val_a = str(ws.cell(row_idx, target.cr.col('PVLcompte')).value or '').strip()
            val_b = str(ws.cell(row_idx, target.cr.col('PVLtitre')).value or '').strip()
            if not val_a or not val_b:
                continue
            if val_b.startswith('*') and val_b.endswith('*') and len(val_b) > 2:
                title = val_b[1:-1]  # enlever les *
                devise = str(ws.cell(row_idx, target.cr.col('PVLdevise')).value or '').strip()
                target.pv_titles.setdefault(val_a, []).append(
                    (title, devise, row_idx))
        if close_wb:
            wb_values.close()
    except Exception:
        target.pv_titles = {}


def load_budget_categories(target, wb_values=None):
    """Charge les catégories, postes et métadonnées depuis la feuille Budget.

    Écrit sur target les attributs : budget_categories, budget_cat_rows,
    budget_cat_col, budget_start_row, budget_header_row, budget_total_row,
    budget_insert_row, budget_posts, budget_post_rows, budget_post_types,
    budget_posts_total_row, budget_first_devise_col, budget_last_devise_col,
    budget_devises, ctrl_last_devise_col, ctrl2_header_row.
    """
    close_wb = False
    if wb_values is None:
        wb_values = openpyxl.load_workbook(target.xlsx_path, data_only=True)
        close_wb = True
    try:
        ws = wb_values[SHEET_BUDGET]

        # Résoudre les colonnes/lignes via ColResolver
        cat_col = target.cr.col('CATnom')
        cat_start_row, cat_end_row = target.cr.rows('CATnom')
        postes_start_row, postes_end_row = target.cr.rows('POSTESnom')

        # Catégories : colonne cat_col entre start CAT et end CAT
        cats = []
        cat_rows = {}
        start_row = cat_start_row
        total_row = None
        # Scanner les catégories entre ⚓ top et ⚓ bot (sentinelles dans le NR)
        scan_start = (start_row or 0) + 1
        scan_end = cat_end_row if cat_end_row else scan_start + 200
        for row_idx in range(scan_start, scan_end):
            val = ws.cell(row_idx, cat_col).value
            if not val:
                continue
            name = str(val).strip()
            if name in ('✓', '⚓'):
                continue
            if name:
                cats.append(name)
                cat_rows[name] = row_idx
        # Total = row après end CAT (pas de scan texte)
        total_row = (cat_end_row + 1) if cat_end_row else None

        # Postes budgétaires : col A (nom), col B (type Fixe/Variable)
        posts = []
        post_rows = {}
        post_types = {}
        posts_total_row = None
        scan_start = (postes_start_row + 1) if postes_start_row else 4
        scan_end = postes_end_row or start_row or 200
        for row_idx in range(scan_start, scan_end):
            val = ws.cell(row_idx, 1).value  # col A
            if not val:
                continue
            name_a = str(val).strip()
            if name_a in ('✓', '⚓'):
                continue
            if name_a:
                posts.append(name_a)
                post_rows[name_a] = row_idx
                type_val = ws.cell(row_idx, 2).value  # col B
                post_types[name_a] = str(type_val).strip() if type_val else ''
        # Total postes = row après end POSTES (pas de scan texte)
        posts_total_row = (postes_end_row + 1) if postes_end_row else None

        # Détecter la première colonne devise = cat_col + 1
        first_devise_col = cat_col + 1
        # Détecter la dernière devise (scan en-tête = 2 lignes au-dessus de start CAT)
        # Compter colonnes non-vides depuis cat_col, soustraire 6 structurelles
        # Structure : CATÉGORIES(1) + devises(N) + Total(1) + Alloc%(1) + Alloc(1) + Poste(1)
        header_row = (start_row - 2) if start_row else 27
        total_cols = 0
        for col_idx in range(cat_col, cat_col + 30):
            val = ws.cell(header_row, col_idx).value
            if not val:
                break
            total_cols += 1
        n_devises = total_cols - 5  # 5 = CATÉGORIES + Total + Alloc% + Alloc + Poste
        budget_last_devise = first_devise_col + max(0, n_devises - 1)

        # Collecter les codes devises depuis le header Budget
        budget_devises = set()
        for col_idx in range(first_devise_col, budget_last_devise + 1):
            val = ws.cell(header_row, col_idx).value
            if val and str(val).strip():
                budget_devises.add(str(val).strip())

        # Détecter la dernière colonne devise dans Contrôles
        # CTRL2drill pointe sur la colonne EUR (première devise).
        # Le header devises est 2 lignes au-dessus de CTRL2type START.
        ctrl_last_devise = None
        ctrl2_header_row = None
        if SHEET_CONTROLES in wb_values.sheetnames:
            ws_ctrl = wb_values[SHEET_CONTROLES]
            ctrl2_s, ctrl2_e = target.cr.rows('CTRL2type')
            if ctrl2_s:
                # v3.6 : drill CTRL2 en r1-1 (sentinelle ⚓ incluse dans NR).
                ctrl2_header_row = ctrl2_s - 1
            else:
                ctrl2_header_row = 61
            # Scanner depuis CTRL2drill (colonne EUR) pour la dernière devise
            eur_col = target.cr.col('CTRL2drill') if 'CTRL2drill' in target.cr._cols else None
            if eur_col:
                for col_idx in range(eur_col, eur_col + 30):
                    val = ws_ctrl.cell(ctrl2_header_row, col_idx).value
                    if not val or not str(val).strip():
                        ctrl_last_devise = col_idx - 1
                        break
            if ctrl_last_devise is None:
                ctrl_last_devise = 29  # fallback AC

        if close_wb:
            wb_values.close()
        target.budget_categories = cats
        target.budget_cat_rows = cat_rows
        target.budget_cat_col = cat_col
        # start_row = start CAT (model row). Les taux sont 1 ligne au-dessus.
        # Le header devises (EUR, USD...) est 2 lignes au-dessus.
        target.budget_start_row = (start_row - 1) if start_row else None  # ligne taux (START)
        target.budget_header_row = (start_row - 2) if start_row else 27   # ligne headers devises
        target.budget_total_row = total_row
        # Insertion devant ⚓ bot (dernière ligne du NR CATnom)
        target.budget_insert_row = cat_end_row
        target.budget_posts = posts
        target.budget_post_rows = post_rows
        target.budget_post_types = post_types
        target.budget_posts_total_row = posts_total_row
        target.budget_first_devise_col = first_devise_col
        target.budget_last_devise_col = budget_last_devise
        target.budget_devises = budget_devises
        target.ctrl_last_devise_col = ctrl_last_devise
        target.ctrl2_header_row = ctrl2_header_row
    except Exception:
        target.budget_categories = []
        target.budget_cat_rows = {}
        target.budget_cat_col = target.cr.col('CATnom')
        target.budget_start_row = None
        target.budget_header_row = 27
        target.budget_total_row = None
        target.budget_insert_row = None
        target.budget_posts = []
        target.budget_post_rows = {}
        target.budget_post_types = {}
        target.budget_posts_total_row = None
        target.budget_first_devise_col = target.cr.col('CATnom') + 1
        target.budget_last_devise_col = target.cr.col('CATtotal_euro') - 1
        target.budget_devises = {'EUR'}
        target.ctrl_last_devise_col = 29
        target.ctrl2_header_row = 62
