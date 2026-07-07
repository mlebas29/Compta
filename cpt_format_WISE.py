#!/usr/bin/env python3
"""
Format Wise statement files (ZIP or XLSX) to standardized CSV format.

Input: ZIP containing multiple XLSX statements, or individual XLSX files
Output: Temporary CSV file(s) with 9-field standard format + #Solde

Tier 2 script: Raw XLSX → Standardized CSV
"""

import sys
import csv
import openpyxl
import zipfile
import re
from pathlib import Path
from datetime import datetime, timedelta
import configparser
import inc_categorize
from inc_format import process_files, lines_to_tuples, log_csv_debug as _log_csv_debug, site_name_from_file, base_dir

SITE = site_name_from_file(__file__)

# Currency to account name mapping : chargé depuis config_accounts.json
import json
_ACCOUNTS_JSON = base_dir() / 'config_accounts.json'
with open(_ACCOUNTS_JSON, 'r', encoding='utf-8') as _f:
    _wise_config = json.load(_f).get(SITE, {})
CURRENCY_ACCOUNTS = {}
for _a in _wise_config.get('accounts', []):
    _name = _a['name']
    # Extraire la devise du nom (ex: "Compte Wise EUR" → "EUR")
    _parts = _name.rsplit(' ', 1)
    if len(_parts) == 2 and len(_parts[1]) == 3 and _parts[1].isupper():
        CURRENCY_ACCOUNTS[_parts[1]] = _name

EXPECTED_FILES = [
    # Deux chemins alternatifs (l'un OU l'autre) :
    #  - legacy : relevés multi-devises XLSX dans un ZIP (assistant Wise) ;
    #  - nouveau : export unique « all-transactions » (1 clic, cf. #131).
    ('statement_*.zip', 'glob', '0-1'),
    ('transaction-history*.csv', 'glob', '0-1'),
    ('all-transactions*.csv', 'glob', '0-1'),
    ('wise_balances.csv', 'exact', '0-1'),   # soldes par devise (#Solde, #131 choix b)
]

def log(message, verbose=False):
    """Print log message if verbose mode enabled."""
    if verbose:
        print(f"[WISE_FORMAT] {message}", file=sys.stderr)

def extract_zips(zip_path, dest_dir, verbose=False):
    """Extrait les fichiers XLSX d'un ZIP Wise.

    Pour les relevés Wise multi-devises, l'utilisateur télécharge un ZIP contenant
    4 fichiers XLSX (EUR, USD, SGD, SEK).

    Args:
        zip_path: Path du fichier ZIP
        dest_dir: Path du répertoire de destination
        verbose: Activer les logs

    Returns:
        list: Liste des fichiers XLSX extraits (Path objects), vide si erreur ou pas de XLSX
    """
    zip_path = Path(zip_path)
    dest_dir = Path(dest_dir)
    extracted_files = []

    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            # Lister les fichiers XLSX dans le ZIP
            all_files = zf.namelist()
            xlsx_files = [f for f in all_files if f.endswith('.xlsx')]

            # Alerte si le ZIP ne contient pas de XLSX
            if not xlsx_files:
                other_extensions = set(Path(f).suffix.lower() for f in all_files if Path(f).suffix)
                if other_extensions:
                    log(
                        f"ZIP '{zip_path.name}' ne contient pas de fichiers XLSX "
                        f"(trouvé: {', '.join(sorted(other_extensions))}). "
                        f"Retélécharger le relevé au format XLSX sur wise.com",
                        verbose=True
                    )
                return []

            # Extraire les XLSX dans le répertoire de destination
            for xlsx_file in xlsx_files:
                zf.extract(xlsx_file, dest_dir)
                extracted_path = dest_dir / xlsx_file
                extracted_files.append(extracted_path)
                log(f"Extrait: {xlsx_file}", verbose)

    except Exception as e:
        log(f"Erreur extraction {zip_path.name}: {e}", verbose=True)

    return extracted_files

def extract_currency_from_filename(filename):
    """
    Extract currency code from Wise statement filename.
    Pattern: statement_XXXXXXXX_<CURRENCY>_YYYY-MM-DD_YYYY-MM-DD.xlsx

    Args:
        filename: Name of the XLSX file

    Returns:
        Currency code (EUR, USD, etc.) or None if not found
    """
    match = re.search(r'statement_\d+_([A-Z]{3})_\d{4}-\d{2}-\d{2}_\d{4}-\d{2}-\d{2}\.xlsx', filename)
    return match.group(1) if match else None

def parse_wise_date(date_value):
    """
    Parse Wise date format to DD/MM/YYYY.

    Args:
        date_value: Date string in format YYYY-MM-DD HH:MM:SS or datetime object

    Returns:
        Date string in format DD/MM/YYYY
    """
    if isinstance(date_value, datetime):
        return date_value.strftime('%d/%m/%Y')
    elif isinstance(date_value, str):
        # Extract date part (YYYY-MM-DD) from datetime string
        date_str = date_value.split()[0]
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        return dt.strftime('%d/%m/%Y')
    return ''

def parse_wise_xlsx(xlsx_file, verbose=False):
    """
    Parse a Wise XLSX statement file.

    Note: Le filtrage par date est centralisé dans inc_format.process_files()

    Args:
        xlsx_file: Path to XLSX file
        verbose: Enable verbose logging

    Returns:
        Tuple: (operations list, final_balance, currency, account_name)
    """
    log(f"Parsing {xlsx_file.name}", verbose)

    # Extract currency from filename
    currency = extract_currency_from_filename(xlsx_file.name)
    if not currency:
        raise ValueError(f"Cannot extract currency from filename: {xlsx_file.name}")

    account_name = CURRENCY_ACCOUNTS.get(currency)
    if not account_name:
        raise ValueError(
            f"Devise {currency} non configurée dans cpt_format_WISE.py (CURRENCY_ACCOUNTS). "
            f"Ajouter '{currency}': 'Compte Wise {currency}' et créer le compte dans comptes.xlsx (feuille Avoirs)"
        )

    log(f"Currency: {currency}, Account: {account_name}", verbose)

    # Load workbook
    wb = openpyxl.load_workbook(xlsx_file, data_only=True)
    ws = wb['All transactions']

    operations = []
    final_balance = 0.0  # Default to 0 for empty files

    # Get final balance from row 2 (most recent transaction) regardless of filtering
    if ws.max_row >= 2:
        balance_val = ws.cell(2, 8).value  # Column H: Solde actuel
        if balance_val is not None:
            final_balance = float(balance_val)

    # Process rows (skip header row 1, start from row 2)
    for row_idx in range(2, ws.max_row + 1):
        # Column indices (1-based in openpyxl)
        date_val = ws.cell(row_idx, 2).value  # Column B: Date
        amount_val = ws.cell(row_idx, 4).value  # Column D: Montant
        currency_val = ws.cell(row_idx, 5).value  # Column E: Devise
        description_val = ws.cell(row_idx, 6).value  # Column F: Description
        balance_val = ws.cell(row_idx, 8).value  # Column H: Solde actuel

        # Skip empty rows
        if not date_val or amount_val is None:
            continue

        # Format operation
        formatted_date = parse_wise_date(date_val)
        amount = float(amount_val)
        description = str(description_val) if description_val else ''

        # Catégorisation automatique via patterns
        category, opts = inc_categorize.categorize_operation(description, SITE)
        ref = opts.get('ref', '')
        equiv = opts.get('equiv', '')

        # Build operation tuple: (Date, Libellé, Montant, Devise, Equiv, Réf, Catégorie, Compte, Commentaire)
        operation = (
            formatted_date,
            description,
            f"{amount:.2f}",
            currency,
            equiv,
            ref,
            category,
            account_name,
            ''   # Commentaire
        )
        operations.append(operation)

    wb.close()

    # Reverse operations (Wise exports newest first, we need oldest first)
    operations.reverse()

    log(f"Extracted {len(operations)} operations, final balance: {final_balance:.2f} {currency}", verbose)

    return operations, final_balance, currency, account_name

# ============================================================================
# EXPORT « ALL-TRANSACTIONS » (CSV unique, #131) — remplace l'assistant XLSX
# ============================================================================
#
# Colonnes (index 0-based, ordre stable ; les entêtes sont localisées donc on
# adresse par position) :
#   0 Identifiant (ACCRUAL_CHARGE-/TRANSFER-/BALANCE_TRANSACTION-…)  1 Statut
#   2 Direction (IN/OUT/NEUTRAL)  3 Créé le  4 Terminé le
#   5 Frais départ (montant)  6 Frais départ (devise)  7-8 Frais arrivée
#   9 Nom d'origine  10 Montant départ (après frais)  11 Devise départ
#   12 Nom cible  13 Montant arrivée (après frais)  14 Devise arrivée
#   15 Taux  16 Référence  17 Paiement de masse  18 Créé par  19 Catégorie  20 Note
#
# Modèle de jambes (vérifié en croisant CSV ↔ ancien XLSX « Solde actuel ») :
#   ACCRUAL_CHARGE (OUT)  → 1 jambe : débit Compte Wise <dev> du montant (frais)
#   TRANSFER IN           → 1 jambe : crédit Compte Wise <dev arrivée>
#   TRANSFER OUT          → 1 jambe : débit Compte Wise <dev départ> (montant + frais).
#                            La cible d'un OUT est TOUJOURS externe (Kraken, banque…),
#                            jamais une 2e jambe Wise — même en conversion.
#   BALANCE_TRANSACTION (NEUTRAL) → 2 jambes (conversion INTERNE entre soldes) :
#                            débit Compte Wise <dev départ> (montant + frais)
#                            + crédit Compte Wise <dev arrivée> (montant arrivée),
#                            appariées par ref='-' (comme un change YUH).
# Le débit inclut TOUJOURS les frais de départ (confirmé : 4986.04+13.96 = -5000 USD).

_TXN_TYPE_RE = re.compile(r'^"?(ACCRUAL_CHARGE|TRANSFER|BALANCE_TRANSACTION)')


def is_all_transactions_csv(csv_file):
    """Reconnaît un export « all-transactions » Wise par le motif d'ID de sa
    1re ligne de données (indépendant du nom de fichier et de la langue)."""
    try:
        with open(csv_file, newline='', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            next(reader, None)          # entête
            first = next(reader, None)  # 1re ligne data
        return bool(first and _TXN_TYPE_RE.match((first[0] or '').strip()))
    except Exception:
        return False


def _to_float(value):
    """Parse un montant Wise (point OU virgule décimale, espaces/guillemets)."""
    s = (value or '').strip().strip('"').replace('\xa0', '').replace(' ', '')
    if not s:
        return 0.0
    if ',' in s and '.' not in s:      # virgule décimale (locale FR)
        s = s.replace(',', '.')
    return float(s)


def _fmt_fr(x):
    """Formate un montant façon relevé Wise FR : « 5 000,00 » (espace milliers,
    virgule décimale). Aligne le libellé des conversions sur le rendu de l'ancien
    XLSX → le dédup à l'import reconnaît la même op à la bascule (#131). NB : le
    dédup normalise les espaces, seuls comptent le séparateur espace + la virgule."""
    return f'{x:,.2f}'.replace(',', ' ').replace('.', ',')


def parse_all_transactions_csv(csv_file, verbose=False):
    """Décompose l'export « all-transactions » en opérations par compte-devise.

    Retourne une liste de tuples 9 champs (Date, Libellé, Montant, Devise, Equiv,
    Réf, Catégorie, Compte, Commentaire), du plus ancien au plus récent. Le
    filtrage par date est appliqué en aval par inc_format.process_files().
    """
    with open(csv_file, newline='', encoding='utf-8-sig') as f:
        rows = list(csv.reader(f))

    data = [r for r in rows[1:] if r and _TXN_TYPE_RE.match((r[0] or '').strip())]
    data.reverse()  # Wise exporte du plus récent au plus ancien → on remet ancien→récent

    operations = []
    for row in data:
        if len(row) < 21:
            continue
        txn = row[0].strip().strip('"')
        typ = _TXN_TYPE_RE.match(txn).group(1)
        if row[1].strip() != 'COMPLETED':      # ignorer PENDING/CANCELLED/…
            continue

        direction = row[2].strip()
        date = parse_wise_date(row[4] or row[3])
        fee = _to_float(row[5])
        src_amt, src_cur = _to_float(row[10]), row[11].strip()
        tgt_amt, tgt_cur = _to_float(row[13]), row[14].strip()
        name_orig, name_cible = row[9].strip(), row[12].strip()

        # (montant signé, devise, libellé, ref forcée)
        legs = []
        if typ == 'ACCRUAL_CHARGE':
            legs.append((-src_amt, src_cur, 'Frais Wise Assets Europe', ''))
        elif typ == 'TRANSFER' and direction == 'IN':
            legs.append((tgt_amt, tgt_cur, f'Argent reçu de {name_orig}', ''))
        elif typ == 'TRANSFER' and direction == 'OUT':
            legs.append((-(src_amt + fee), src_cur, f'Argent envoyé à {name_cible}', ''))
        elif typ == 'BALANCE_TRANSACTION':      # NEUTRAL — conversion interne
            label = (f'{_fmt_fr(src_amt + fee)} {src_cur} convertis en '
                     f'{_fmt_fr(tgt_amt)} {tgt_cur}')
            legs.append((-(src_amt + fee), src_cur, label, '-'))
            legs.append((tgt_amt, tgt_cur, label, '-'))
        else:
            log(f"Type/direction non géré ({typ}/{direction}) : {txn} — ignoré", verbose=True)
            continue

        for amount, cur, label, forced_ref in legs:
            account = CURRENCY_ACCOUNTS.get(cur)
            if not account:
                log(f"Devise {cur} sans compte configuré ({txn}) — jambe ignorée", verbose=True)
                continue
            category, opts = inc_categorize.categorize_operation(label, SITE)
            ref = forced_ref or opts.get('ref', '')
            equiv = opts.get('equiv', '')
            operations.append(
                (date, label, f'{amount:.2f}', cur, equiv, ref, category, account, '')
            )

    log(f"all-transactions: {len(operations)} jambes depuis {csv_file.name}", verbose)
    return operations


def read_balances_soldes(site_dir, date_str, verbose=False):
    """Lit wise_balances.csv (déposé par le fetch, #131 choix b) → une ligne
    #Solde par devise configurée, datée `date_str`. Format du fichier :
    « DEVISE,solde » par ligne. Absent → [] (comptes Wise auto-calculés)."""
    bal = Path(site_dir) / 'wise_balances.csv'
    if not bal.exists():
        return []
    soldes = []
    with open(bal, newline='', encoding='utf-8-sig') as f:
        for row in csv.reader(f):
            if len(row) < 2:
                continue
            cur = row[0].strip().upper()
            account = CURRENCY_ACCOUNTS.get(cur)
            if not account:
                log(f"solde devise {cur} sans compte configuré — ignoré", verbose=True)
                continue
            amount = _to_float(row[1])
            soldes.append((date_str, f'Relevé {account}', f'{amount:.2f}',
                           cur, '', '', '#Solde', account, ''))
    log(f"wise_balances: {len(soldes)} #Solde depuis {bal.name}", verbose)
    return soldes


# ============================================================================
# API POUR UPDATE - NOUVELLE INTERFACE
# ============================================================================

def format_site(site_dir, verbose=False, logger=None):
    """API pour Update.

    Note: Le filtrage par date est centralisé dans inc_format.process_files()
    """
    if logger is None:
        from inc_logging import Logger
        logger = Logger(SITE, verbose=verbose)

    # Vérification fichiers dropbox
    from inc_format import verify_dropbox_files
    for w in verify_dropbox_files(site_dir, SITE):
        logger.warning(w)

    site_dir = Path(site_dir)

    # Dossier temporaire pour extraction des ZIPs
    temp_dir = site_dir / '.wise_temp'
    all_operations = []

    try:
        # EXCLUSIVITÉ DE SOURCE (#131). L'export « all-transactions » contient
        # TOUT l'historique → s'il est présent, il PRIME et rend les relevés
        # XLSX/ZIP redondants. Traiter les deux doublonnerait : entre XLSX et CSV
        # la date (fuseau, transferts proches de minuit) ET le libellé (formatage)
        # divergent → le dédup ne les reconnaît pas. On choisit donc UNE source.
        csv_all_tx = [f for f in site_dir.glob('*.csv') if is_all_transactions_csv(f)]

        if csv_all_tx:
            def _parse_csv(csv_file):
                if not is_all_transactions_csv(csv_file):
                    return []
                return parse_all_transactions_csv(csv_file, verbose)
            ops_csv, _ = process_files(site_dir, [('*.csv', _parse_csv, 'ops')],
                                       verbose, SITE, logger=logger)
            all_operations.extend(ops_csv)

            # #Solde par devise depuis wise_balances.csv (fetch, #131 choix b) —
            # daté de la dernière opération importée. Absent → auto-calculé.
            dates = [o[0] for o in ops_csv if o and o[0]]
            if dates:
                try:
                    last = max(dates, key=lambda d: datetime.strptime(d, '%d/%m/%Y'))
                except ValueError:
                    last = dates[-1]
                all_operations.extend(read_balances_soldes(site_dir, last, verbose))

            if list(site_dir.glob('*.zip')) or list(site_dir.glob('statement_*.xlsx')):
                logger.warning(
                    "Export all-transactions présent → relevés XLSX/ZIP ignorés "
                    "(source unique, évite les doublons)")
        else:
            # Chemin legacy : relevés multi-devises XLSX (ZIP de l'assistant Wise).
            for zip_path in site_dir.glob('*.zip'):
                temp_dir.mkdir(exist_ok=True)
                extract_zips(zip_path, temp_dir, verbose)

            def _parse_xlsx(xlsx_file):
                operations, final_balance, currency, account_name = parse_wise_xlsx(
                    xlsx_file, verbose
                )
                # Ne pas générer de #Solde si le statement est vide (solde inconnu)
                if not operations:
                    return []
                solde_date = operations[-1][0]
                solde_line = (solde_date, f'Relevé {account_name}', f'{final_balance:.2f}',
                              currency, '', '', '#Solde', account_name, '')
                return operations + [solde_line]

            handlers = [('statement_*.xlsx', _parse_xlsx, 'ops')]
            if temp_dir.exists():
                ops_temp, _ = process_files(temp_dir, handlers, verbose, SITE, logger=logger)
                all_operations.extend(ops_temp)
            ops_direct, _ = process_files(site_dir, handlers, verbose, SITE, logger=logger)
            all_operations.extend(ops_direct)

    finally:
        # Nettoyer le dossier temporaire
        if temp_dir.exists():
            import shutil
            shutil.rmtree(temp_dir)

    logger.verbose(f"format_site: {len(all_operations)} ops, 0 pos")
    return all_operations, []

def log_csv_debug(operations, positions, site_dir, logger=None):
    """Wrapper vers inc_format.log_csv_debug()"""
    _log_csv_debug(SITE, operations, positions, logger)

if __name__ == '__main__':
    from inc_format import cli_main
    cli_main(format_site)
