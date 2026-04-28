#!/usr/bin/env python3
"""
tool_audit_formats.py — Audit read-only des formats d'un classeur comptes.xlsm
contre la charte graphique v3.6.

Usage:
  python3 tool_audit_formats.py <fichier.xlsm> [--sheet NOM] [--quiet]

Principe :
  - Lecture openpyxl, aucun write.
  - Identifie les "tableaux" par regroupement des named ranges colonne sur mêmes lignes.
  - Délimite tête (fond D2C195 contigu au-dessus du NR) et pied (fond EEEBDB contigu en
    dessous), zonage symétrique par fond.
  - Vérifie fonds par rôle (tête / col ref / data / pied), grille hair D2C195, BORDURE_PIED.
  - Signale séparément les cellules jaune (annotations user — jamais en violation).

Exceptions tolérées :
  - JAUNE (#FFFF00) : partout, signalement user, non remonté en violation.
  - Gamme BEIGE (#D2C195, #EEEBDB) : en zone data uniquement, effet de section.

Voir ~/.claude/projects/-home-marc-Compta-Claude/memory/charte_v3_6.md.
"""

import argparse
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from inc_formats import (
    TETE_FILL, PIED_FILL, COL_REF_FILL, DATA_FILL, JAUNE as JAUNE_INT,
    HAIR_COLOR as HAIR_COLOR_INT, PIED_BORDER_COLOR,
    EXC_DATA as EXC_DATA_INT,
    EXC_HEAD as EXC_HEAD_INT, EXC_FOOT as EXC_FOOT_INT,
)


def _ff(n):
    """Convertit un int 0xRRGGBB vers le format 'FFRRGGBB' utilisé par openpyxl."""
    return f'FF{n:06X}'


# ─── Charte v3.6 (forme openpyxl) ──────────────────────────────────────────
TETE            = _ff(TETE_FILL)
PIED            = _ff(PIED_FILL)
COL_REF         = _ff(COL_REF_FILL)
DATA_BLANC      = _ff(DATA_FILL)
JAUNE           = _ff(JAUNE_INT)
EXC_DATA        = {_ff(n) for n in EXC_DATA_INT}
EXC_HEAD        = {_ff(n) for n in EXC_HEAD_INT}
EXC_FOOT        = {_ff(n) for n in EXC_FOOT_INT}

HAIR_COLOR      = _ff(HAIR_COLOR_INT)
PIED_BORDER     = ('thick', _ff(PIED_BORDER_COLOR))
HAIR            = ('hair',  HAIR_COLOR)


def _norm_color(rgb):
    """Normalise une couleur openpyxl en 'FFxxxxxx' uppercase, ou None."""
    if rgb is None:
        return None
    s = str(rgb).upper()
    if len(s) == 6:
        return 'FF' + s
    if len(s) == 8:
        return s
    return None


def _cell_fill(cell):
    """Renvoie la couleur de fond normalisée, ou None si pas de fond solide."""
    fill = cell.fill
    if fill is None or fill.patternType != 'solid':
        return None
    fg = fill.fgColor
    if fg is None:
        return None
    # openpyxl peut exposer rgb, indexed, theme. On ne gère que rgb (cas courant).
    if fg.type == 'rgb':
        return _norm_color(fg.rgb)
    return None


def _side(side):
    """Renvoie (style, color_norm) d'un côté de bordure, ou (None, None)."""
    if side is None or side.style is None:
        return (None, None)
    col = None
    if side.color is not None and side.color.type == 'rgb':
        col = _norm_color(side.color.rgb)
    return (side.style, col)


# ─── Identification des tableaux ────────────────────────────────────────────
def collect_tables(wb):
    """
    Regroupe les named ranges colonne en "tableaux" par (sheet, start_row, end_row).
    Retourne list[dict] : {sheet, min_col, max_col, first_row, last_row, names}.
    """
    groups = defaultdict(lambda: {'cols': set(), 'names': []})
    for name in wb.defined_names:
        dn = wb.defined_names[name]
        dests = list(dn.destinations)
        if len(dests) != 1:
            continue
        sheet_name, ref = dests[0]
        # ref = "$B$5:$B$32" ou "$A$1"
        ref = ref.replace('$', '')
        if ':' not in ref:
            continue
        a, b = ref.split(':')
        # parse coordinates
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
        try:
            col_a, row_a = coordinate_from_string(a)
            col_b, row_b = coordinate_from_string(b)
        except Exception:
            continue
        if col_a != col_b:
            continue  # on ne garde que les NR colonne
        ci = column_index_from_string(col_a)
        if row_b - row_a < 1:
            continue  # mono-cell (déjà filtré par le ':' check, garde-fou)
        key = (sheet_name, row_a, row_b)
        groups[key]['cols'].add(ci)
        groups[key]['names'].append(name)

    tables = []
    for (sheet, first_row, last_row), data in groups.items():
        cols = sorted(data['cols'])
        tables.append({
            'sheet': sheet,
            'min_col': cols[0],
            'max_col': cols[-1],
            'cols': cols,
            'first_row': first_row,
            'last_row': last_row,
            'names': sorted(data['names']),
        })
    return tables


# ─── Délimitation tête / pied ───────────────────────────────────────────────
def _row_any_content(ws, r, min_col, max_col):
    """True si au moins une cellule de la ligne a une valeur non vide."""
    for c in range(min_col, max_col + 1):
        if ws.cell(row=r, column=c).value not in (None, ''):
            return True
    return False


def find_head(ws, min_col, max_col, first_row):
    """Tête = ligne adjacente à la sentinelle ⚓ haut (first_row-1) toujours incluse,
    + lignes au-dessus tant qu'il y a du contenu."""
    if first_row <= 1:
        return None
    head_bot = first_row - 1
    head_top = head_bot
    r = head_top - 1
    while r >= 1 and _row_any_content(ws, r, min_col, max_col):
        head_top = r
        r -= 1
    return head_top


def find_foot(ws, min_col, max_col, last_row):
    """Pied = ligne adjacente à la sentinelle ⚓ bas (last_row+1) toujours incluse,
    + lignes en dessous tant qu'il y a du contenu. None si adjacente hors feuille."""
    first = last_row + 1
    if first > ws.max_row:
        return None
    foot_bot = first
    r = foot_bot + 1
    while r <= ws.max_row and _row_any_content(ws, r, min_col, max_col):
        foot_bot = r
        r += 1
    return foot_bot


# ─── Audit ──────────────────────────────────────────────────────────────────
def audit_cell_fill(cell, expected, exceptions, role):
    """Renvoie une violation (dict) ou None."""
    got = _cell_fill(cell)
    if got == expected:
        return None
    if got in exceptions:
        return None
    return {
        'kind': 'fill',
        'role': role,
        'expected': expected,
        'got': got,
    }


def audit_cell_borders(cell, is_first_foot_row):
    """Grille hair D2C195 partout dans le tableau (tête, col ref, pied, data).
    Visuellement invisible sur beige, visible sur blanc — mais techniquement
    posée partout pour écraser les bordures hors charte.

    Exceptions :
    - BORDURE_PIED (thick 6C2E24) sur top 1re ligne pied
    - jaune exempt (liberté user totale)."""
    violations = []
    if _cell_fill(cell) == JAUNE:
        return violations  # jaune = libre
    bd = cell.border

    if is_first_foot_row:
        got = _side(bd.top)
        if got != PIED_BORDER:
            violations.append({
                'kind': 'border',
                'side': 'top',
                'expected': PIED_BORDER,
                'got': got,
            })

    for side_name, side_obj in [
        ('top', bd.top) if not is_first_foot_row else (None, None),
        ('bottom', bd.bottom),
        ('left', bd.left),
        ('right', bd.right),
    ]:
        if side_name is None:
            continue
        got = _side(side_obj)
        if got != HAIR:
            violations.append({
                'kind': 'border',
                'side': side_name,
                'expected': HAIR,
                'got': got,
            })
    return violations


def _row_is_used(ws, r, min_col, max_col):
    """True si au moins une cellule de la ligne a une valeur non vide.
    Le fond seul n'est pas suffisant (certaines feuilles préformatent des milliers
    de lignes blanches dans leur NR)."""
    for c in range(min_col, max_col + 1):
        if ws.cell(row=r, column=c).value not in (None, ''):
            return True
    return False


def audit_table(ws, table):
    """Audit un tableau, renvoie (violations, jaunes)."""
    first_data = table['first_row']
    last_data_nr = table['last_row']
    min_col = table['min_col']
    max_col = table['max_col']

    # Borner la data au dernier rang effectivement utilisé (évite les NR géants
    # type OP 4..10000 dont 99% des lignes sont vides non-formatées).
    last_data = last_data_nr
    while last_data >= first_data and not _row_is_used(ws, last_data, min_col, max_col):
        last_data -= 1
    if last_data < first_data:
        last_data = first_data  # tableau vide — on garde au moins 1 ligne

    head_top = find_head(ws, min_col, max_col, first_data)
    foot_bot = find_foot(ws, min_col, max_col, last_data_nr)

    head_first = head_top if head_top is not None else first_data
    foot_last = foot_bot if foot_bot is not None else last_data
    foot_first = last_data + 1 if foot_bot is not None else None

    violations = []
    jaunes = []

    def add_cell(r, c, role, expected_fill, fill_exceptions, is_first_foot):
        cell = ws.cell(row=r, column=c)
        got = _cell_fill(cell)
        # Tolérance openpyxl : cellule totalement vide (value None) + fill None
        # → openpyxl ignore le fill/bordures même si UNO les a posés. On n'audite pas.
        if cell.value in (None, '') and got is None:
            return
        if got == JAUNE:
            jaunes.append({
                'cell': f"{get_column_letter(c)}{r}",
                'role': role,
            })
        else:
            v = audit_cell_fill(cell, expected_fill, fill_exceptions, role)
            if v is not None:
                v['cell'] = f"{get_column_letter(c)}{r}"
                violations.append(v)
        # bordures auditées uniquement si la cellule a un fond lisible par openpyxl
        for bv in audit_cell_borders(cell, is_first_foot):
            bv['cell'] = f"{get_column_letter(c)}{r}"
            bv['role'] = role
            violations.append(bv)

    # Tête
    if head_top is not None:
        for r in range(head_top, first_data):
            for c in range(min_col, max_col + 1):
                add_cell(r, c, 'tête', TETE, EXC_HEAD, False)

    # Data
    for r in range(first_data, last_data + 1):
        for c in range(min_col, max_col + 1):
            if c == min_col:
                add_cell(r, c, 'col_ref', COL_REF, EXC_DATA, False)
            else:
                add_cell(r, c, 'data', DATA_BLANC, EXC_DATA, False)

    # Pied
    if foot_bot is not None:
        for r in range(foot_first, foot_last + 1):
            is_first = (r == foot_first)
            for c in range(min_col, max_col + 1):
                add_cell(r, c, 'pied', PIED, EXC_FOOT, is_first)

    meta = {
        'head': (head_top, first_data - 1) if head_top else None,
        'data': (first_data, last_data),
        'foot': (foot_first, foot_last) if foot_bot else None,
    }
    return violations, jaunes, meta


# ─── Rapport ────────────────────────────────────────────────────────────────
def _read_alarm_sqrefs(path):
    """Pré-pass : retourne {sheet_name: set((col_letter, row))} des cells dans
    sqref de CF alarme (dxf fill ∈ {FFC7CE, FFEB9C})."""
    import zipfile, re as _re
    ALARM_FILLS = {'FFC7CE', 'FFEB9C'}

    with zipfile.ZipFile(path) as z:
        styles = z.read('xl/styles.xml').decode('utf-8')

    m = _re.search(r'<dxfs[^>]*>(.*?)</dxfs>', styles, _re.DOTALL)
    dxfs_xml = m.group(1) if m else ''
    dxf_blocks = _re.findall(r'<dxf>(.*?)</dxf>', dxfs_xml, _re.DOTALL)
    alarm_dxf_set = set()
    for i, blk in enumerate(dxf_blocks):
        fills = _re.findall(r'(?:fg|bg)Color[^/]*rgb="(?:FF)?([0-9A-Fa-f]{6})"', blk)
        if any(f.upper() in ALARM_FILLS for f in fills):
            alarm_dxf_set.add(i)
    return alarm_dxf_set


def _parse_sqref(s):
    import re as _re
    cells = []
    for part in str(s).split():
        m = _re.match(r'^([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?$', part)
        if not m:
            continue
        def _c2i(letter):
            n = 0
            for c in letter:
                n = n * 26 + (ord(c) - 64)
            return n
        c1 = _c2i(m.group(1))
        r1 = int(m.group(2))
        c2 = _c2i(m.group(3) or m.group(1))
        r2 = int(m.group(4) or m.group(2))
        for c in range(c1, c2 + 1):
            cl = ''
            n = c
            while n > 0:
                n, r = divmod(n - 1, 26)
                cl = chr(65 + r) + cl
            for r in range(r1, r2 + 1):
                cells.append((cl, r))
    return cells


def audit_alarm_bold(wb, path):
    """Vérifie que les cells de contrôle (sqref CF alarme) non-vides hors col
    drill sont en bold direct.

    Retourne {sheet_name: [(cell_addr, value_repr), ...]} des divergences.
    """
    import re as _re
    alarm_dxf_set = _read_alarm_sqrefs(path)
    if not alarm_dxf_set:
        return {}

    # Cols drill exclues (CATmontant Budget, CTRL2drill Contrôles)
    drill_cols = {}
    for sheet, nm in [('Budget', 'CATmontant'), ('Contrôles', 'CTRL2drill')]:
        if nm in wb.defined_names:
            v = wb.defined_names[nm].value
            m = _re.match(r"[^!]+!\$([A-Z]+)\$", v)
            if m:
                drill_cols[sheet] = m.group(1)

    violations = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        drill = drill_cols.get(sheet_name)
        seen = set()
        sheet_viol = []
        for cfr in ws.conditional_formatting:
            for rule in ws.conditional_formatting[cfr]:
                if rule.dxfId not in alarm_dxf_set:
                    continue
                for cl, r in _parse_sqref(cfr.sqref):
                    if (cl, r) in seen:
                        continue
                    seen.add((cl, r))
                    if drill and cl == drill:
                        continue
                    cell = ws[f"{cl}{r}"]
                    if cell.value is None:
                        continue
                    is_bold = (cell.font.b if cell.font else False) or False
                    if not is_bold:
                        v_str = repr(cell.value)
                        if len(v_str) > 50:
                            v_str = v_str[:47] + '...'
                        sheet_viol.append((f"{cl}{r}", v_str))
        if sheet_viol:
            violations[sheet_name] = sheet_viol
    return violations


def fmt_expected(exp):
    if isinstance(exp, tuple):
        style, col = exp
        return f"{style}/{col[-6:] if col else '?'}"
    return str(exp[-6:]) if isinstance(exp, str) and len(exp) == 8 else str(exp)


def fmt_got(got):
    if isinstance(got, tuple):
        style, col = got
        if style is None:
            return '(aucune)'
        return f"{style}/{col[-6:] if col else '?'}"
    if got is None:
        return '(aucun)'
    return got[-6:] if isinstance(got, str) and len(got) == 8 else str(got)


def _pattern_key(v):
    """Clé de regroupement pour mode synthèse."""
    if v['kind'] == 'fill':
        return ('fill', v['role'], fmt_got(v['expected']), fmt_got(v['got']))
    return ('border', v['role'], v['side'], fmt_expected(v['expected']), fmt_got(v['got']))


def _sample_cells(items, limit=5):
    cells = [x['cell'] for x in items[:limit]]
    suffix = f" +{len(items)-limit}" if len(items) > limit else ""
    return ', '.join(cells) + suffix


def report(results, args):
    total_v = 0
    total_j = 0
    for sheet, tables in results.items():
        if args.sheet and sheet != args.sheet:
            continue
        print(f"\n━━━ {sheet} ━━━")
        if not tables:
            print("  (aucun tableau NR détecté)")
            continue
        for t in tables:
            meta = t['meta']
            names = ', '.join(t['table']['names'][:4])
            if len(t['table']['names']) > 4:
                names += f" +{len(t['table']['names'])-4}"
            col_lbl = f"{get_column_letter(t['table']['min_col'])}→{get_column_letter(t['table']['max_col'])}"
            print(f"\n▸ Tableau {col_lbl} | {names}")
            if meta['head']:
                print(f"    tête : {meta['head'][0]}..{meta['head'][1]}")
            else:
                print(f"    tête : (absente)")
            print(f"    data : {meta['data'][0]}..{meta['data'][1]}")
            if meta['foot']:
                print(f"    pied : {meta['foot'][0]}..{meta['foot'][1]}")
            else:
                print(f"    pied : (absent)")

            v = t['violations']
            j = t['jaunes']
            total_v += len(v)
            total_j += len(j)

            if not v:
                print(f"    ✓ conforme" + (f" ({len(j)} jaune(s) user)" if j else ""))
                if not j:
                    continue

            if v:
                fills = [x for x in v if x['kind'] == 'fill']
                borders = [x for x in v if x['kind'] == 'border']
                if args.verbose:
                    if fills:
                        print(f"    ✗ {len(fills)} écart(s) fond :")
                        for x in fills:
                            print(f"        {x['cell']:<6} [{x['role']}] attendu={fmt_got(x['expected'])} obs={fmt_got(x['got'])}")
                    if borders:
                        print(f"    ✗ {len(borders)} écart(s) bordure :")
                        for x in borders:
                            print(f"        {x['cell']:<6} [{x['role']}] {x['side']:<6} attendu={fmt_expected(x['expected'])} obs={fmt_got(x['got'])}")
                else:
                    # Mode synthèse : regroupe par pattern
                    if fills:
                        print(f"    ✗ {len(fills)} écart(s) fond :")
                        groups = defaultdict(list)
                        for x in fills:
                            groups[_pattern_key(x)].append(x)
                        for key, items in sorted(groups.items(), key=lambda kv: -len(kv[1])):
                            _, role, exp, got = key
                            print(f"        {len(items):>5}× [{role}] attendu={exp} obs={got}   ex: {_sample_cells(items)}")
                    if borders:
                        print(f"    ✗ {len(borders)} écart(s) bordure :")
                        groups = defaultdict(list)
                        for x in borders:
                            groups[_pattern_key(x)].append(x)
                        for key, items in sorted(groups.items(), key=lambda kv: -len(kv[1])):
                            _, role, side, exp, got = key
                            print(f"        {len(items):>5}× [{role}] {side:<6} attendu={exp} obs={got}   ex: {_sample_cells(items)}")
            if j:
                print(f"    ℹ {len(j)} cellule(s) jaune (annotations user) :")
                for x in j[:10]:
                    print(f"        {x['cell']:<6} [{x['role']}]")
                if len(j) > 10:
                    print(f"        ... +{len(j)-10} autres")

    print(f"\n━━━ Bilan ━━━")
    print(f"  Violations : {total_v}")
    print(f"  Jaunes (info) : {total_j}")
    return total_v


# ─── Main ───────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('xlsm', help='fichier xlsm à auditer')
    parser.add_argument('--sheet', help='limiter à une feuille')
    parser.add_argument('--verbose', action='store_true', help='liste chaque écart (par défaut : synthèse par pattern)')
    args = parser.parse_args()

    path = Path(args.xlsm)
    if not path.exists():
        print(f"✗ fichier introuvable : {path}", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(path, keep_vba=True, data_only=False)
    tables = collect_tables(wb)

    # grouper par feuille, trier par first_row
    by_sheet = defaultdict(list)
    for t in tables:
        by_sheet[t['sheet']].append(t)
    for sheet in by_sheet:
        by_sheet[sheet].sort(key=lambda t: t['first_row'])

    results = {}
    for sheet_name in wb.sheetnames:
        if sheet_name not in by_sheet:
            results[sheet_name] = []
            continue
        ws = wb[sheet_name]
        sheet_results = []
        for t in by_sheet[sheet_name]:
            violations, jaunes, meta = audit_table(ws, t)
            sheet_results.append({
                'table': t,
                'violations': violations,
                'jaunes': jaunes,
                'meta': meta,
            })
        results[sheet_name] = sheet_results

    total_v = report(results, args)

    # === Audit bold cells contrôle (CF alarme) hors drill ===
    bold_viol = audit_alarm_bold(wb, path)
    if bold_viol:
        print(f"\n━━━ Bold cells contrôle (CF alarme) ━━━")
        n_bold = 0
        for sheet_name, items in bold_viol.items():
            if args.sheet and sheet_name != args.sheet:
                continue
            print(f"\n  {sheet_name} : {len(items)} cell(s) attendue(s) bold, observée(s) not-bold :")
            for addr, v_str in items:
                print(f"      {addr:<6} {v_str}")
            n_bold += len(items)
        if n_bold:
            print(f"\n  ✗ {n_bold} violation(s) bold")
            total_v += n_bold

    print(f"\n━━━ Total ━━━")
    print(f"  Violations totales : {total_v}")
    return 0 if total_v == 0 else 1


if __name__ == '__main__':
    sys.exit(main())
