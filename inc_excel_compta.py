#!/usr/bin/env python3
"""
Module de base Excel pour comptes.xlsm — infrastructure et pairing.

Classe ComptaExcel : gestion workbook, lectures pairing, écritures pairing.
Utilisée par composition dans ComptaPairer (cpt_pair.py) et ComptaRefsTool (tool_refs.py).
Étendue par ComptaExcelImport (inc_excel_import.py) pour l'import.
"""

import sys
import re
import shutil
from pathlib import Path
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
    import warnings
    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
except ImportError:
    print("❌ Erreur: openpyxl n'est pas installé")
    print("Installation: pip3 install openpyxl")
    sys.exit(1)

from inc_excel_schema import (
    OpCol, SHEET_OPERATIONS, SHEET_CONTROLES, SHEET_CONTROLES_LEGACY,
    PAIRING_COUNTER_CELL, Operation,
)


# ============================================================================
# CLASSES DE RÉFÉRENCES (déterminent le préfixe des Réf)
# ============================================================================
CLASSE_VIREMENT = ['Virement']
CLASSE_TITRES = ['Achat titres', 'Vente titres', 'Arbitrage titres']
CLASSE_CHANGE = ['Change']
CLASSE_METAUX = ['Achat métaux']

# ============================================================================
# CONFIGURATION DES OPÉRATIONS LIÉES ET SOLDE AUTO (chargées depuis config_pipeline.json)
# ============================================================================

def _load_pipeline_config():
    """Charge LINKED_OPERATIONS et SOLDE_AUTO_ACCOUNTS depuis config_pipeline.json.

    Clés requises : "linked_operations" et "solde_auto".
    """
    import json
    from inc_mode import get_base_dir
    config_path = get_base_dir() / 'config_pipeline.json'
    if not config_path.exists():
        raise FileNotFoundError(f'config_pipeline.json introuvable : {config_path}')

    with open(config_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if 'linked_operations' not in data:
        raise ValueError('config_pipeline.json : clé "linked_operations" manquante')
    if 'solde_auto' not in data:
        raise ValueError('config_pipeline.json : clé "solde_auto" manquante')

    # Normaliser les clés linked_operations en UPPER
    linked = {}
    for pattern, entry in data['linked_operations'].items():
        linked[pattern.upper()] = {
            'compte_cible': entry['compte_cible'],
            'description': entry['description'],
        }

    solde_auto = {}
    for compte, entry in data['solde_auto'].items():
        solde_auto[compte] = {
            'categorie_trigger': entry['categorie_trigger'],
            'devise': entry['devise'],
        }

    return linked, solde_auto

LINKED_OPERATIONS, SOLDE_AUTO_ACCOUNTS = _load_pipeline_config()

# ============================================================================
# FONCTIONS UTILITAIRES
# ============================================================================

def normalize_devise(devise):
    """Normalise une devise pour l'utiliser comme préfixe de référence.

    - Supprime les suffixes 'Pr' (Premium) et 'Jo' (Joaillerie) : OrPr → or, OrJo → or
    - Alias : SAT → btc (satoshis = bitcoin)
    - Convertit en minuscules : USD → usd, BTC → btc
    """
    if not devise:
        return 'eur'
    devise = devise.strip()
    # Alias crypto (avant autres transformations)
    if devise.upper() == 'SAT':
        return 'btc'
    # Supprimer suffixes Premium et Joaillerie
    if devise.endswith('Pr'):
        devise = devise[:-2]
    if devise.endswith('Jo'):
        devise = devise[:-2]
    return devise.lower()


def get_ref_prefix(categorie, devise_credit=None):
    """Détermine le préfixe de référence selon la classe et la devise crédit.

    Args:
        categorie: Catégorie de l'opération (Virement, Achat titres, Change, etc.)
        devise_credit: Devise du côté crédit (pour Change et Achat métaux)

    Returns:
        Préfixe en minuscules: 'v', 't', 'btc', 'usd', 'orjo', etc.
    """
    if categorie in CLASSE_VIREMENT:
        return 'v'
    elif categorie in CLASSE_TITRES:
        return 't'
    elif categorie in CLASSE_CHANGE:
        return normalize_devise(devise_credit)
    elif categorie in CLASSE_METAUX:
        return normalize_devise(devise_credit) + 'jo'
    else:
        # Fallback pour catégories non reconnues
        return 'v'


def parse_date(date_str):
    """Parse une date en gérant les formats DD/MM/YY et DD/MM/YYYY"""
    try:
        if isinstance(date_str, datetime):
            return date_str
        if len(date_str.split('/')[-1]) == 2:
            return datetime.strptime(date_str, '%d/%m/%y')
        else:
            return datetime.strptime(date_str, '%d/%m/%Y')
    except:
        return None


def parse_montant(montant_str):
    """Parse un montant string en float"""
    try:
        return float(str(montant_str).replace(' ', '').replace(',', '.'))
    except:
        return None


def classify_reference_pattern(ref_str):
    """Détermine le pattern d'une référence.

    Returns:
        str: 'vxx', 'btcxx', 'orxx', 'txx', 'orjoxx', etc.
    """
    ref_lower = ref_str.lower()

    # Typo: 0r → or
    if re.match(r'^0r\d+$', ref_str):
        return 'orxx (typo)'

    # Métaux physiques (bijoux) - avec suffixe 'jo'
    if re.match(r'^orjo\d+$', ref_lower):
        return 'orjoxx'
    if re.match(r'^agjo\d+$', ref_lower):
        return 'agjoxx'

    # Classes principales
    if re.match(r'^v\d+$', ref_lower):
        return 'vxx'
    if re.match(r'^t\d+$', ref_lower):
        return 'txx'

    # Devises et crypto
    if re.match(r'^btc\d+$', ref_lower):
        return 'btcxx'
    if re.match(r'^xmr\d+$', ref_lower):
        return 'xmrxx'
    if re.match(r'^or\d+$', ref_lower):
        return 'orxx'
    if re.match(r'^ag\d+$', ref_lower):
        return 'agxx'
    if re.match(r'^usd\d+$', ref_lower):
        return 'usdxx'
    if re.match(r'^eur\d+$', ref_lower):
        return 'eurxx'
    if re.match(r'^chf\d+$', ref_lower):
        return 'chfxx'
    if re.match(r'^sgd\d+$', ref_lower):
        return 'sgdxx'
    if re.match(r'^sek\d+$', ref_lower):
        return 'sekxx'
    if re.match(r'^jpy\d+$', ref_lower):
        return 'jpyxx'
    if re.match(r'^gbp\d+$', ref_lower):
        return 'gbpxx'

    return 'autres'


def normalize_reference_case(ref_str):
    """Normalise la casse d'une référence.

    Règles:
      - Tout en minuscules (v, t, btc, or, usd, etc.)
      - 0rxx (typo): corrige en orxx

    Returns:
        str: Référence normalisée
    """
    # Typo: 0r → or
    if re.match(r'^0r\d+$', ref_str):
        return 'or' + ref_str[2:]

    # Tout en minuscules
    if re.match(r'^[a-zA-Z]+\d+$', ref_str):
        return ref_str.lower()

    # Autres: ne pas modifier
    return ref_str


# ============================================================================
# CLASSE PRINCIPALE
# ============================================================================

class ComptaExcel:
    """Accès Excel pour comptes.xlsm — gestion workbook et pairing."""

    def __init__(self, comptes_file, verbose=False, all_soldes=False, logger=None):
        self.comptes_file = Path(comptes_file)
        self.verbose = verbose
        self.all_soldes = all_soldes
        self.logger = logger

        # État workbook
        self.wb = None
        self.ws_operations = None
        self.ws_controle = None
        self._pairing_counter = None  # initialisé paresseusement par _init_pairing_counter

    # ====================================================================
    # Gestion workbook
    # ====================================================================

    def open_workbook(self):
        """Ouvre le fichier comptes.xlsm"""
        try:
            self.wb = openpyxl.load_workbook(self.comptes_file, keep_vba=True)

            if SHEET_OPERATIONS not in self.wb.sheetnames:
                self.logger.error(f"Onglet '{SHEET_OPERATIONS}' introuvable")
                return False

            self.ws_operations = self.wb[SHEET_OPERATIONS]

            # Bornes tableaux via named ranges
            from inc_excel_schema import get_named_ranges, get_table_start, PV_FIRST_ROW
            named = get_named_ranges(self.wb)
            pvl_start = get_table_start(named, 'PVL')
            self._pvl_data_start = (pvl_start or PV_FIRST_ROW) + 1

            # Onglet Contrôles optionnel
            if SHEET_CONTROLES in self.wb.sheetnames:
                self.ws_controle = self.wb[SHEET_CONTROLES]
            elif SHEET_CONTROLES_LEGACY in self.wb.sheetnames:
                self.ws_controle = self.wb[SHEET_CONTROLES_LEGACY]

            self.logger.verbose(f"Fichier ouvert: {len(self.wb.sheetnames)} onglets")
            return True
        except Exception as e:
            self.logger.error(f"Erreur ouverture fichier: {e}")
            return False

    def close_workbook(self, save=True):
        """Ferme le fichier en sauvegardant ou non"""
        if self.wb:
            try:
                if save:
                    self.wb.save(self.comptes_file)
                    self.logger.verbose("Fichier sauvegardé")
                self.wb.close()
            except Exception as e:
                self.logger.error(f"Erreur fermeture fichier: {e}")
                return False
        return True

    # Rows 1-3 = en-têtes, row 4 = model row (template de style)
    OP_MODEL_ROW = 4

    def find_last_data_row(self):
        """Trouve la dernière ligne contenant des données réelles.

        Retourne au minimum OP_MODEL_ROW (row 4) pour que l'insertion
        commence en row 5 et copie le format de la model row, pas de l'en-tête.
        Ignore les model rows ✓ (coches START/END).
        """
        if not self.ws_operations:
            return self.OP_MODEL_ROW

        for row in range(self.ws_operations.max_row, self.OP_MODEL_ROW, -1):
            date = self.ws_operations.cell(row, OpCol.DATE).value
            if date and str(date).strip() == '✓':
                continue  # model row — ignorer
            label = self.ws_operations.cell(row, OpCol.LABEL).value
            montant = self.ws_operations.cell(row, OpCol.MONTANT).value

            if date or label or montant is not None:
                return row

        return self.OP_MODEL_ROW

    # ====================================================================
    # Lectures pairing
    # ====================================================================

    def load_unpaired_operations(self):
        """Charge toutes les opérations avec ref='-' depuis Excel

        Returns:
            list[Operation] avec champs enrichis (row, date_parsed, montant_parsed, equiv_parsed)
        """
        operations = []
        last_row = self.find_last_data_row()

        for row in range(3, last_row + 1):
            ref = self.ws_operations.cell(row, OpCol.REF).value
            if ref != '-':
                continue

            date_val = self.ws_operations.cell(row, OpCol.DATE).value
            label = self.ws_operations.cell(row, OpCol.LABEL).value or ''
            montant_raw = self.ws_operations.cell(row, OpCol.MONTANT).value
            devise = self.ws_operations.cell(row, OpCol.DEVISE).value or 'EUR'
            equiv_raw = self.ws_operations.cell(row, OpCol.EQUIV).value
            categorie = self.ws_operations.cell(row, OpCol.CATEGORIE).value or '-'
            compte = self.ws_operations.cell(row, OpCol.COMPTE).value or ''
            commentaire = self.ws_operations.cell(row, OpCol.COMMENTAIRE).value or ''

            date_parsed = parse_date(date_val)
            montant_parsed = parse_montant(montant_raw)
            equiv_parsed = parse_montant(equiv_raw)

            if date_parsed is None or montant_parsed is None:
                continue

            operations.append(Operation(
                date=date_val,
                label=label,
                montant=montant_raw,
                devise=devise,
                equiv=equiv_raw,
                ref=ref,
                categorie=categorie,
                compte=compte,
                commentaire=commentaire,
                row=row,
                date_parsed=date_parsed,
                montant_parsed=montant_parsed,
                equiv_parsed=equiv_parsed,
            ))

        self.logger.verbose(f"{len(operations)} opération(s) non appairée(s) chargée(s)")
        return operations

    # ====================================================================
    # Écritures pairing
    # ====================================================================

    def write_ref_to_excel(self, row, ref, categorie=None):
        """Écrit une ref (et optionnellement une catégorie) dans Excel.

        Ne remplace pas une ref existante (différente de '' et '-').
        """
        existing = self.ws_operations.cell(row, OpCol.REF).value
        if existing and str(existing).strip() not in ('', '-'):
            self.logger.verbose(f"Ref existante '{existing}' préservée (row {row}), "
                                f"ref '{ref}' ignorée")
            return
        self.ws_operations.cell(row, OpCol.REF).value = ref
        if categorie:
            self.ws_operations.cell(row, OpCol.CATEGORIE).value = categorie

    def write_equiv_to_excel(self, row, equiv_value):
        """Écrit une valeur Equiv dans Excel"""
        self.ws_operations.cell(row, OpCol.EQUIV).value = equiv_value

    def _init_pairing_counter(self):
        """Initialise le compteur d'appariement à MAX(partie numérique des refs) + 1.

        Scanne la colonne F pour trouver le plus grand suffixe numérique existant,
        évitant ainsi les collisions avec les refs pré-existantes.
        """
        import re
        max_num = 0
        col = PAIRING_COUNTER_CELL[1]  # colonne F
        for row_idx in range(3, self.ws_operations.max_row + 1):
            ref = self.ws_operations.cell(row_idx, col).value
            if ref is None:
                continue
            m = re.search(r'(\d+)$', str(ref))
            if m:
                num = int(m.group(1))
                if num > max_num:
                    max_num = num
        self._pairing_counter = (max_num + 1) % 1000
        self.logger.verbose(f"Compteur appariement initialisé à {self._pairing_counter} (max existant: {max_num})")

    def get_next_pairing_ref(self, categorie='Virement', devise_credit=None):
        """Génère la prochaine référence d'appariement selon la classe et la devise.

        Le préfixe dépend de la classe (dérivée de la catégorie) et de la devise crédit:
        - Virement → v{xxx}
        - Titres (Achat/Vente/Arbitrage titres) → t{xxx}
        - Change → {devise}{xxx} (ex: btc, usd, or, chf...)
        - Achat métaux → {devise}jo{xxx} (ex: orjo)

        Le compteur est initialisé au premier appel par scan des refs existantes (MAX+1),
        modulo 1000. La cellule F2 est mise à jour pour affichage.
        """
        if not self.ws_operations:
            self.logger.error("Onglet Opérations non ouvert pour compteur")
            return None

        if self._pairing_counter is None:
            self._init_pairing_counter()

        prefix = get_ref_prefix(categorie, devise_credit)
        ref = f"{prefix}{self._pairing_counter}"
        self._pairing_counter = (self._pairing_counter + 1) % 1000

        # Mise à jour F2 pour affichage
        row, col = PAIRING_COUNTER_CELL
        self.ws_operations.cell(row, col).value = self._pairing_counter

        self.logger.verbose(f"Appariement: {ref} (compteur → {self._pairing_counter})")
        return ref

    # ====================================================================
    # Chargement références (pour audit/maintenance)
    # ====================================================================

    def load_all_references(self, year_filter=None, ref_regex_filter=None, account_pair_filter=None):
        """Charge toutes les références depuis l'onglet Opérations avec filtres optionnels.

        Args:
            year_filter: int, filtrer par année (ex: 2023)
            ref_regex_filter: str, regex pour filtrer références (ex: '^v24', '^v2')
            account_pair_filter: tuple, paire de comptes (compte1, compte2) ou None

        Returns:
            dict: {ref_str: [(row, date_str, cat_str, amount, equiv, account_str, devise_str), ...]}
        """
        refs = defaultdict(list)

        for row in range(4, self.ws_operations.max_row + 1):
            ref_val = self.ws_operations.cell(row, OpCol.REF).value

            if ref_val and str(ref_val) not in ['-', '#Solde']:
                ref_str = str(ref_val)
                date_val = self.ws_operations.cell(row, OpCol.DATE).value
                cat_val = self.ws_operations.cell(row, OpCol.CATEGORIE).value
                amount_val = self.ws_operations.cell(row, OpCol.MONTANT).value
                devise_val = self.ws_operations.cell(row, OpCol.DEVISE).value
                equiv_val = self.ws_operations.cell(row, OpCol.EQUIV).value
                account_val = self.ws_operations.cell(row, OpCol.COMPTE).value

                # Filtrer par année
                if year_filter and date_val:
                    if hasattr(date_val, 'year'):
                        if date_val.year != year_filter:
                            continue

                # Filtrer par regex référence
                if ref_regex_filter:
                    if not re.match(ref_regex_filter, ref_str):
                        continue

                # Filtrer par paire de comptes
                if account_pair_filter:
                    account_str_temp = str(account_val) if account_val else ''
                    compte1, compte2 = account_pair_filter
                    if compte1.lower() not in account_str_temp.lower() and compte2.lower() not in account_str_temp.lower():
                        continue

                # Convertir date
                if hasattr(date_val, 'strftime'):
                    date_str = date_val.strftime('%d/%m/%Y')
                else:
                    date_str = str(date_val) if date_val else ''

                # Convertir montant
                try:
                    amount = float(str(amount_val).replace(',', '.')) if amount_val else 0.0
                except Exception:
                    amount = 0.0

                # Convertir equiv
                try:
                    equiv = float(str(equiv_val).replace(',', '.')) if equiv_val else None
                except Exception:
                    equiv = None

                cat_str = str(cat_val) if cat_val else ''
                account_str = str(account_val) if account_val else ''
                devise_str = str(devise_val) if devise_val else 'EUR'

                refs[ref_str].append((row, date_str, cat_str, amount, equiv, account_str, devise_str))

        return dict(refs)

    # ====================================================================
    # Backup
    # ====================================================================

    def create_backup(self, prefix="BACKUP"):
        """Crée un backup du fichier comptes.xlsm dans archives/.

        Args:
            prefix: préfixe du fichier backup (ex: BACKUP, BACKUP_FIX)

        Returns:
            Path: chemin du backup créé
        """
        backup_dir = self.comptes_file.parent / 'archives'
        backup_dir.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = backup_dir / f'comptes_{prefix}_{timestamp}.xlsx'

        shutil.copy2(self.comptes_file, backup_path)
        self.logger.verbose(f"Backup: {backup_path.name}")
        return backup_path
