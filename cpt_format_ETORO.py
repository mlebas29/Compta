#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
cpt_format_ETORO.py - Convertit les fichiers eToro en formats standardisés

Fichiers supportés:

  1. Money operations (TSV): eToroTransactions_*.tsv
     Colonnes: Name, Date, Amount, Currency
     → Sortie stdout: format 9 colonnes

  2. Reserve operations (XLSX): etoro-account-statement*.xlsx
     Onglet "Activité du compte", colonnes: Date, Type, Détails, Montant, Unités
     → Sortie stdout: format 9 colonnes

  3. PDF accueil eToro (nommage flexible): *.pdf contenant "USD disponible", "EUR disponible"
     Capture d'écran de la page d'accueil eToro
     → Génère: soldes_comptes_parsed.csv (Réserve USD + Money EUR)

  4. PDF portfolio eToro (nommage flexible): *.pdf contenant "Actifs", "Prix", "Unités"
     Capture d'écran de la page portfolio eToro
     → Génère: positions_titres_parsed.csv + mise à jour soldes (Titres USD)

  5. Positions (CSV): positions_titres_parsed.csv
     Colonnes: Date;Ligne;Montant
     → Sortie stdout: format 4 colonnes

  6. Soldes (CSV): soldes_comptes_parsed.csv
     Colonnes: Compte;Solde
     → Ignoré (lu par les autres processus)

Comptes Excel:
  - Compte eToro Money (EUR)
  - Portefeuille eToro USD Réserve (USD)
  - Portefeuille eToro USD Titres (USD)

Usage:
  ./cpt_format_ETORO.py <fichier>
"""

import sys
import csv
import re
import configparser
from pathlib import Path
from datetime import datetime
import openpyxl
try:
    import pdfplumber
except ImportError:
    pdfplumber = None
import inc_categorize
from inc_format import process_files, lines_to_tuples, log_csv_debug as _log_csv_debug, get_file_date, site_name_from_file, require_account

SITE = site_name_from_file(__file__)

# Noms de comptes : chargés depuis config_accounts.json
import json
_ACCOUNTS_JSON = Path(__file__).parent / 'config_accounts.json'
with open(_ACCOUNTS_JSON, 'r', encoding='utf-8') as _f:
    _etoro_config = json.load(_f).get(SITE, {})
_etoro_accounts = [a['name'] for a in _etoro_config.get('accounts', [])]
ACCOUNT_TITRES = require_account(_etoro_accounts, 'Titres', SITE)
ACCOUNT_RESERVE = require_account(_etoro_accounts, 'Réserve', SITE)
ACCOUNT_MONEY = require_account(_etoro_accounts, 'Money', SITE)


def get_temp_dir():
    """Retourne le répertoire temporaire logs/debug/ETORO/ pour les fichiers parsed"""
    script_dir = Path(__file__).parent
    config_file = script_dir / 'config.ini'
    config = configparser.ConfigParser()
    config.read(config_file)
    logs_dir = script_dir / config.get('paths', 'logs', fallback='./logs')
    temp_dir = logs_dir / 'debug' / 'ETORO'
    temp_dir.mkdir(parents=True, exist_ok=True)
    return temp_dir


def format_date(date_str):
    """
    Convertit diverses formats de date en DD/MM/YYYY

    Entrée: "24/09/2025 10:22:56" ou "27/02/2025" ou "24/09/2025"
    Sortie: "24/09/2025"
    """
    date_str = str(date_str).strip()

    # Si déjà au format DD/MM/YYYY
    if re.match(r'\d{2}/\d{2}/\d{4}', date_str):
        return date_str.split()[0]  # Retirer l'heure si présente

    # Si format YYYY-MM-DD
    if re.match(r'\d{4}-\d{2}-\d{2}', date_str):
        try:
            dt = datetime.strptime(date_str, '%Y-%m-%d')
            return dt.strftime('%d/%m/%Y')
        except:
            pass

    return date_str.split()[0]  # Fallback: retirer l'heure


def format_amount(amount_str, target_currency='EUR'):
    """
    Convertit le montant en format décimal avec virgule

    Entrée: "-900" ou "1050.06" ou "-4 500,00" ou "-7,92"
    Sortie: "-900,00" ou "1050,06" ou "-4500,00" ou "-7,92"
    """
    amount_str = str(amount_str).strip()

    # Retirer les espaces
    amount_str = amount_str.replace(' ', '')

    # Retirer les guillemets
    amount_str = amount_str.replace('"', '')

    # Si point décimal, convertir en virgule
    if '.' in amount_str and ',' not in amount_str:
        amount_str = amount_str.replace('.', ',')

    # S'assurer qu'on a 2 décimales
    if ',' in amount_str:
        parts = amount_str.split(',')
        if len(parts[1]) == 1:
            amount_str = amount_str + '0'
    else:
        # Pas de décimales
        amount_str = amount_str + ',00'

    return amount_str


def extract_eur_amount(label):
    """
    Extrait le montant EUR d'un libellé

    Entrée: "900.00 EUR eToroMoney" ou "Dépôt 900.00 EUR eToroMoney"
    Sortie: "900.00" ou None
    """
    # Pattern: chercher un nombre suivi de "EUR"
    match = re.search(r'([\d.,]+)\s*EUR', label, re.IGNORECASE)
    if match:
        return match.group(1).replace(',', '.')
    return None


# Note: Fonctions categorize_money_operation et categorize_reserve_operation supprimées
# Catégorisation maintenant centralisée dans inc_category_mappings.py (via cpt_categorize)


# =============================================================================
# Fonctions PDF - Extraction soldes et positions depuis captures d'écran eToro
# =============================================================================

def extract_pdf_text(pdf_path):
    """Extrait le texte de toutes les pages d'un PDF

    Returns:
        str: Texte complet du PDF ou None si erreur
    """
    if pdfplumber is None:
        print("❌ pdfplumber non installé. Installer avec: pip3 install pdfplumber", file=sys.stderr)
        return None

    try:
        full_text = ""
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"
        return full_text
    except Exception as e:
        print(f"❌ Erreur extraction PDF: {e}", file=sys.stderr)
        return None


def detect_pdf_type(pdf_text):
    """Détecte le type de PDF eToro par son contenu

    Returns:
        'pdf_home'      : Page d'accueil avec soldes (USD/EUR disponible)
        'pdf_portfolio' : Page portfolio avec positions (Actifs, Prix, Unités)
        None            : PDF non reconnu
    """
    if not pdf_text:
        return None

    # PDF accueil : contient les soldes disponibles
    if 'USD disponible' in pdf_text and 'EUR disponible' in pdf_text:
        return 'pdf_home'

    # PDF portfolio : contient la liste des actifs
    if 'Actifs' in pdf_text and ('Prix' in pdf_text or 'Unités' in pdf_text):
        return 'pdf_portfolio'

    return None


def parse_pdf_home(pdf_text, default_date=None):
    """Parse le PDF page d'accueil pour extraire les soldes

    Extrait:
    - USD disponible $XXX.XX
    - EUR disponible €XXX.XX

    Args:
        pdf_text: Texte extrait du PDF
        default_date: Date fallback DD/MM/YYYY (défaut: date du fichier via get_file_date)

    Returns:
        dict: {'usd_disponible': float, 'eur_disponible': float, 'date': str}
    """
    result = {
        'usd_disponible': None,
        'eur_disponible': None,
        'date': default_date or datetime.now().strftime('%d/%m/%Y')
    }

    # Extraire date du PDF (format DD/MM/YYYY HH:MM en haut)
    date_match = re.search(r'(\d{2}/\d{2}/\d{4})\s+\d{2}:\d{2}', pdf_text)
    if date_match:
        result['date'] = date_match.group(1)

    # USD disponible $XXX.XX ou $X,XXX.XX
    usd_match = re.search(r'USD disponible\s*\$?([\d,]+\.?\d*)', pdf_text)
    if usd_match:
        usd_str = usd_match.group(1).replace(',', '')
        try:
            result['usd_disponible'] = float(usd_str)
        except ValueError:
            pass

    # EUR disponible €XXX.XX ou €X,XXX.XX
    eur_match = re.search(r'EUR disponible\s*€?([\d,]+\.?\d*)', pdf_text)
    if eur_match:
        eur_str = eur_match.group(1).replace(',', '')
        try:
            result['eur_disponible'] = float(eur_str)
        except ValueError:
            pass

    return result


def parse_pdf_portfolio(pdf_text, default_date=None):
    """Parse le PDF portfolio pour extraire les positions

    Structure attendue (lignes):
    TICKER
    PRICE
    Name                    UNITS    COST   G/P   G/P%  VALUE   ...

    Exemple:
    BTC
    89897.77
    Bitcoin <0.01 112890.00-$203.67-20.37%$796.31 89902.64 89902.65

    Pattern valeur: après le pourcentage (G/P%), le premier $XXX.XX est la valeur

    Args:
        pdf_text: Texte extrait du PDF
        default_date: Date fallback DD/MM/YYYY (défaut: date du fichier via get_file_date)

    Returns:
        dict: {
            'positions': [{'ticker': str, 'name': str, 'value': float}, ...],
            'total_titres': float,
            'date': str
        }
    """
    result = {
        'positions': [],
        'total_titres': 0.0,
        'date': default_date or datetime.now().strftime('%d/%m/%Y')
    }

    # Extraire date du PDF
    date_match = re.search(r'(\d{2}/\d{2}/\d{4})\s+\d{2}:\d{2}', pdf_text)
    if date_match:
        result['date'] = date_match.group(1)

    lines = pdf_text.split('\n')

    # Chercher les lignes d'actifs
    # Pattern: TICKER suivi du prix, puis ligne avec détails incluant la valeur
    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # Détecter un ticker (lettres majuscules, peut contenir . pour AI.PA)
        ticker_match = re.match(r'^([A-Z][A-Z0-9.]{0,10})$', line)
        if ticker_match:
            ticker = ticker_match.group(1)

            # Ligne suivante devrait être le prix
            if i + 1 < len(lines):
                price_line = lines[i + 1].strip()

                # Ligne après devrait contenir les détails avec la valeur $XXX.XX
                if i + 2 < len(lines):
                    detail_line = lines[i + 2].strip()

                    # Extraire le nom (début de la ligne jusqu'au premier nombre ou <)
                    name_match = re.match(r'^([A-Za-z].*?)\s+[\d<]', detail_line)
                    name = name_match.group(1).strip() if name_match else ticker

                    # Pattern: G/P suivi de G/P% puis VALUE
                    # Exemple: -$203.67-20.37%$796.31 ou $8.20 4.10% $208.19
                    # Chercher le montant $XXX.XX qui suit un pourcentage XX.XX%
                    value_match = re.search(r'[\d.]+%\s*\$?([\d,]+\.?\d*)', detail_line)

                    if value_match:
                        try:
                            val = float(value_match.group(1).replace(',', ''))
                            if val > 0:
                                result['positions'].append({
                                    'ticker': ticker,
                                    'name': name,
                                    'value': val
                                })
                                result['total_titres'] += val
                        except ValueError:
                            pass

                    i += 3
                    continue

        i += 1

    return result


def process_pdf_to_csv_files(input_file):
    """Traite un PDF eToro et génère les fichiers CSV appropriés

    Pour PDF accueil (soldes):
        - Génère/met à jour soldes_comptes_parsed.csv

    Pour PDF portfolio (positions):
        - Génère/met à jour positions_titres_parsed.csv
        - Met à jour soldes_comptes_parsed.csv avec le total Titres

    Returns:
        tuple: (success: bool, message: str)
    """
    pdf_text = extract_pdf_text(input_file)
    if not pdf_text:
        return False, "Impossible d'extraire le texte du PDF"

    pdf_type = detect_pdf_type(pdf_text)
    if not pdf_type:
        return False, "Type de PDF non reconnu (ni accueil ni portfolio)"

    # Fichiers temporaires dans logs/debug/ETORO/ (pas dans dropbox/)
    temp_dir = get_temp_dir()
    soldes_file = temp_dir / 'soldes_comptes_parsed.csv'
    positions_file = temp_dir / 'positions_titres_parsed.csv'

    # Charger soldes existants si présent
    existing_soldes = {}
    if soldes_file.exists():
        try:
            with open(soldes_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f, delimiter=';')
                for row in reader:
                    existing_soldes[row['Compte']] = row['Solde']
        except:
            pass

    if pdf_type == 'pdf_home':
        # Parser les soldes
        data = parse_pdf_home(pdf_text, default_date=get_file_date(input_file))

        # Mettre à jour les soldes
        if data['usd_disponible'] is not None:
            existing_soldes[ACCOUNT_RESERVE] = f"{data['usd_disponible']:.2f}"
        if data['eur_disponible'] is not None:
            existing_soldes[ACCOUNT_MONEY] = f"{data['eur_disponible']:.2f}"

        # Écrire soldes_comptes_parsed.csv
        with open(soldes_file, 'w', encoding='utf-8', newline='') as f:
            f.write("Compte;Solde\n")
            for compte, solde in existing_soldes.items():
                f.write(f"{compte};{solde}\n")

        msg = f"Soldes extraits: USD={data['usd_disponible']}, EUR={data['eur_disponible']}"
        print(f"✓ {input_file.name} → soldes_comptes_parsed.csv ({msg})", file=sys.stderr)
        return True, msg

    elif pdf_type == 'pdf_portfolio':
        # Parser les positions
        data = parse_pdf_portfolio(pdf_text, default_date=get_file_date(input_file))

        if not data['positions']:
            return False, "Aucune position trouvée dans le PDF"

        # Écrire positions_titres_parsed.csv
        with open(positions_file, 'w', encoding='utf-8', newline='') as f:
            f.write("Date;Ligne;Montant\n")
            for pos in data['positions']:
                f.write(f"{data['date']};{pos['ticker']};{pos['value']:.2f}\n")

        # Mettre à jour solde Titres
        existing_soldes[ACCOUNT_TITRES] = f"{data['total_titres']:.2f}"

        # Écrire soldes_comptes_parsed.csv
        with open(soldes_file, 'w', encoding='utf-8', newline='') as f:
            f.write("Compte;Solde\n")
            for compte, solde in existing_soldes.items():
                f.write(f"{compte};{solde}\n")

        msg = f"{len(data['positions'])} positions, total=${data['total_titres']:.2f}"
        print(f"✓ {input_file.name} → positions_titres_parsed.csv ({msg})", file=sys.stderr)
        return True, msg

    return False, f"Type PDF non géré: {pdf_type}"


def process_pdf_etoro(input_file):
    """Traite un PDF eToro et retourne les lignes formatées pour stdout

    Pour PDF accueil (soldes):
        - Format 9 colonnes avec #Solde pour USD Réserve et EUR Money

    Pour PDF portfolio (positions):
        - Format 4 colonnes pour les positions
        - Ajoute #Solde Titres à la fin

    Returns:
        tuple: (output_lines: list, header: str, is_positions: bool)
    """
    pdf_text = extract_pdf_text(input_file)
    if not pdf_text:
        return None, None, False

    pdf_type = detect_pdf_type(pdf_text)
    if not pdf_type:
        return None, None, False

    if pdf_type == 'pdf_home':
        # Parser les soldes → format 9 colonnes
        data = parse_pdf_home(pdf_text, default_date=get_file_date(input_file))
        output_lines = []

        # #Solde Réserve USD
        if data['usd_disponible'] is not None:
            solde_str = f"{data['usd_disponible']:.2f}".replace('.', ',')
            output_lines.append(f"{data['date']};Relevé solde;{solde_str};USD;;;#Solde;{ACCOUNT_RESERVE};")

        # #Solde Money EUR
        if data['eur_disponible'] is not None:
            solde_str = f"{data['eur_disponible']:.2f}".replace('.', ',')
            output_lines.append(f"{data['date']};Relevé solde;{solde_str};EUR;;;#Solde;{ACCOUNT_MONEY};")

        header = "Date;Libellé;Montant;Devise;Equiv;Réf;Catégorie;Compte;Commentaire"
        print(f"📄 PDF soldes: USD={data['usd_disponible']}, EUR={data['eur_disponible']}", file=sys.stderr)
        return output_lines, header, False

    elif pdf_type == 'pdf_portfolio':
        # Parser les positions → format 4 colonnes
        data = parse_pdf_portfolio(pdf_text, default_date=get_file_date(input_file))
        output_lines = []

        for pos in data['positions']:
            output_lines.append(f"{data['date']};{pos['ticker']};{pos['value']:.2f};{ACCOUNT_TITRES}")

        # Ajouter #Solde Titres
        output_lines.append(f"{data['date']};#Solde Titres;{data['total_titres']:.2f};{ACCOUNT_TITRES}")

        header = "Date;Ligne;Montant;Compte"
        print(f"📄 PDF positions: {len(data['positions'])} titres, total=${data['total_titres']:.2f}", file=sys.stderr)
        return output_lines, header, True

    return None, None, False


def process_money_operations(input_file):
    """
    Traite les opérations Money (EUR) depuis le fichier TSV

    Note: Le filtrage par date est centralisé dans inc_format.process_files()

    Format: Name	Date	Amount	Currency

    Règles colonne Equiv:
    - DPT/WDL (Change EUR↔USD): Equiv = montant EUR (pour appariement avec Transfer)
    - Virements EUR↔EUR: Equiv vide
    """
    output_lines = []
    operations = []

    with open(input_file, 'r', encoding='utf-8') as f:
        # TSV avec tabulation
        reader = csv.DictReader(f, delimiter='\t')

        for row in reader:
            name = row['Name'].strip()
            date_str = row['Date'].strip()
            amount_str = row['Amount'].strip()
            currency = row['Currency'].strip()

            operations.append({
                'date_str': format_date(date_str),
                'name': name,
                'amount': format_amount(amount_str, currency)
            })

    # Générer les lignes formatées
    for op in operations:
        # Catégorisation automatique via patterns
        category, opts = inc_categorize.categorize_operation(op['name'], SITE)
        ref = opts.get('ref', '')

        # Format standardisé (equiv déterminé par cpt_pair lors de l'appariement)
        output_line = f"{op['date_str']};{op['name']};{op['amount']};EUR;;{ref};{category};{ACCOUNT_MONEY};"
        output_lines.append(output_line)

    # Note: Le #Solde Money vient du PDF eToro, pas du TSV transactions

    return output_lines


def process_reserve_operations(input_file):
    """
    Traite les opérations Réserve (USD) depuis le fichier XLSX

    Note: Le filtrage par date est centralisé dans inc_format.process_files()

    Format: onglet "Activité du compte"
    Date	Type	Détails	Montant	Unités

    Règles colonne Equiv:
    - Equiv uniquement si opération appariée (avec Réf) ET impliquant devise non-EUR
    - Opérations non appariées: Equiv vide
    """
    output_lines = []

    # Ouvrir le fichier XLSX
    wb = openpyxl.load_workbook(input_file, data_only=True)

    # Onglet "Activité du compte"
    sheet_name = None
    for name in wb.sheetnames:
        if 'activit' in name.lower() or 'account' in name.lower():
            sheet_name = name
            break

    if not sheet_name:
        print(f"⚠ Onglet 'Activité du compte' non trouvé dans {input_file.name}", file=sys.stderr)
        return output_lines

    ws = wb[sheet_name]

    operations = []

    # Lire les lignes (skip header row 1)
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for row in rows:
        if not row[0]:  # Skip empty rows
            continue

        date_val = row[0]
        type_op = str(row[1]).strip() if row[1] else ''
        details = str(row[2]).strip() if row[2] else ''
        montant = row[3]

        # Format date
        if isinstance(date_val, datetime):
            date_str = date_val.strftime('%d/%m/%Y')
        else:
            date_str = format_date(str(date_val))

        # Format montant
        if montant is None or montant == '':
            continue

        montant_value = float(str(montant).replace(',', '.'))

        operations.append({
            'date_str': date_str,
            'type_op': type_op,
            'details': details,
            'montant_value': montant_value
        })

    # Générer les lignes formatées
    for op in operations:
        # Construire le libellé
        label = f"{op['type_op']} {op['details']}".strip()

        # Catégorisation automatique via patterns
        category, opts = inc_categorize.categorize_operation(label, SITE)
        ref = opts.get('ref', '')

        # IMPORTANT: Pour les achats de titres (Position ouverte), inverser le signe
        # Le XLSX affiche la valeur de l'achat (+999.98), mais la Réserve est débitée (-999.98)
        montant_value = op['montant_value']
        if category == '@Achat titres' and montant_value > 0:
            montant_value = -montant_value

        amount = format_amount(str(montant_value), 'USD')

        # Colonne Equiv: vide (opérations non appariées)
        equiv = ''

        # Format standardisé
        output_line = f"{op['date_str']};{label};{amount};USD;{equiv};{ref};{category};{ACCOUNT_RESERVE};"
        output_lines.append(output_line)

    # Note: Le #Solde Réserve vient du PDF eToro, pas du xlsx
    return output_lines


def process_positions(input_file):
    """
    Traite les positions titres (format déjà standardisé 4 colonnes)

    Format d'entrée: Date;Ligne;Montant
    Format de sortie: Date;Ligne;Montant;Compte

    Ajoute la colonne Compte
    """
    output_lines = []

    with open(input_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=';')

        # Skip header
        next(reader, None)

        for row in reader:
            if len(row) < 3:
                continue

            date = row[0].strip()
            ticker = row[1].strip()
            value = row[2].strip()

            # Format standardisé 4 colonnes
            output_line = f"{date};{ticker};{value};{ACCOUNT_TITRES}"
            output_lines.append(output_line)

    # Ajouter #Solde Titres depuis soldes_comptes_parsed.csv (si disponible)
    # Note: #Solde Réserve vient du PDF eToro, pas d'ici
    solde_titres = None
    temp_dir = get_temp_dir()
    soldes_file = temp_dir / 'soldes_comptes_parsed.csv'

    if soldes_file.exists():
        try:
            with open(soldes_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f, delimiter=';')
                for row in reader:
                    compte = row['Compte'].strip()
                    if 'Titres' in compte:
                        solde_titres = row['Solde'].strip().replace(',', '.')
                        break
        except:
            pass

    # Ajouter #Solde Titres seulement si disponible
    if solde_titres:
        date_aujourdhui = get_file_date(input_file)
        solde_line = f"{date_aujourdhui};#Solde Titres;{solde_titres};{ACCOUNT_TITRES}"
        output_lines.append(solde_line)

    return output_lines


def detect_file_type(input_file):
    """
    Détecte le type de fichier eToro par patterns dans le nom + extension + contenu

    Mode manuel : noms de fichiers originaux (téléchargés depuis eToro)

    Returns:
        'money_operations'   : TSV (eToroTransactions_*.tsv)
        'reserve_operations' : XLSX (etoro-account-statement*.xlsx)
        'positions'          : CSV positions (positions_titres_parsed.csv)
        'soldes'             : CSV soldes (soldes_comptes_parsed.csv - ignoré)
        'pdf_etoro'          : PDF eToro (accueil ou portfolio, détecté par contenu)
        None                 : type non reconnu
    """
    filename = input_file.name.lower()

    # 1. Fichiers soldes (optionnel, ne pas formatter)
    if 'soldes' in filename and 'parsed' in filename:
        return 'soldes'

    # 2. Fichiers positions (générés par PDF ou mode auto)
    if 'positions' in filename and 'parsed' in filename:
        return 'positions'

    # 3. Money operations (TSV) - Nom original eToro
    if filename.endswith('.tsv'):
        if 'etorotransactions' in filename or 'transactions' in filename:
            return 'money_operations'

    # 4. Reserve operations (XLSX) - Nom original eToro
    if filename.endswith('.xlsx'):
        if 'account-statement' in filename or 'accountstatement' in filename:
            return 'reserve_operations'

    # 5. PDF eToro - Détection flexible par contenu
    if filename.endswith('.pdf'):
        # Vérifier le contenu pour confirmer que c'est un PDF eToro
        pdf_text = extract_pdf_text(input_file)
        if pdf_text and ('etoro' in pdf_text.lower() or 'eToro' in pdf_text):
            pdf_type = detect_pdf_type(pdf_text)
            if pdf_type:
                return 'pdf_etoro'

    return None


# ============================================================================
# API POUR UPDATE - NOUVELLE INTERFACE
# ============================================================================

def _wrap_lines_to_tuples(func, num_fields):
    """Wrapper pour convertir une fonction retournant des lignes CSV en tuples."""
    def wrapper(file_path):
        lines = func(file_path)
        result = []
        for line in lines:
            parts = line.split(';')
            if len(parts) >= num_fields:
                result.append(tuple(parts[:num_fields]))
        return result
    return wrapper


def _process_pdf_etoro_wrapper(file_path):
    """Wrapper pour traiter les PDFs eToro (détection par contenu).

    Returns:
        list: liste de tuples (ops ou pos selon type PDF)
    """
    pdf_text = extract_pdf_text(file_path)
    if not pdf_text:
        return []

    pdf_type = detect_pdf_type(pdf_text)
    if not pdf_type:
        return []

    results = []

    if pdf_type == 'pdf_home':
        # Extraire soldes → opérations #Solde
        data = parse_pdf_home(pdf_text, default_date=get_file_date(file_path))
        if data['usd_disponible'] is not None:
            solde_str = f"{data['usd_disponible']:.2f}".replace('.', ',')
            results.append((
                data['date'], 'Relevé solde', solde_str, 'USD', '', '',
                '#Solde', ACCOUNT_RESERVE, ''
            ))
        if data['eur_disponible'] is not None:
            solde_str = f"{data['eur_disponible']:.2f}".replace('.', ',')
            results.append((
                data['date'], 'Relevé solde', solde_str, 'EUR', '', '',
                '#Solde', ACCOUNT_MONEY, ''
            ))

    return results


def _process_pdf_portfolio_wrapper(file_path):
    """Wrapper pour extraire les positions du PDF portfolio eToro.

    Returns:
        list: liste de tuples positions (4 champs)
    """
    pdf_text = extract_pdf_text(file_path)
    if not pdf_text:
        return []

    pdf_type = detect_pdf_type(pdf_text)
    if pdf_type != 'pdf_portfolio':
        return []

    data = parse_pdf_portfolio(pdf_text, default_date=get_file_date(file_path))
    results = []

    for pos in data['positions']:
        results.append((
            data['date'],
            pos['ticker'],
            f"{pos['value']:.2f}",
            ACCOUNT_TITRES,
        ))

    # Ajouter #Solde Titres
    if data['total_titres'] > 0:
        results.append((
            data['date'],
            '#Solde Titres',
            f"{data['total_titres']:.2f}",
            ACCOUNT_TITRES,
        ))

    return results


def _deduplicate_cross_file_ops(ops_by_file, verbose=False):
    """Déduplique les opérations présentes dans plusieurs fichiers TSV.

    Les doublons intra-fichier sont conservés (légitimes).
    Seuls les doublons inter-fichiers sont éliminés.

    Même logique que cpt_format_BB._deduplicate_cross_file_ops.
    """
    if len(ops_by_file) < 2:
        result = []
        for lines in ops_by_file.values():
            result.extend(lines)
        return result

    # Compter occurrences par ligne par fichier
    line_per_file = {}  # line -> {fname: count}
    for fname, lines in ops_by_file.items():
        for line in lines:
            d = line_per_file.setdefault(line, {})
            d[fname] = d.get(fname, 0) + 1

    # Pour les lignes cross-fichier, choisir le fichier propriétaire (max occurrences)
    line_owner = {}
    for line, fmap in line_per_file.items():
        if len(fmap) > 1:
            line_owner[line] = max(fmap, key=lambda f: fmap[f])

    # Émettre les lignes, exclure les cross-fichier non propriétaires
    result = []
    n_dedup = 0
    for fname, lines in ops_by_file.items():
        for line in lines:
            if line in line_owner and line_owner[line] != fname:
                n_dedup += 1
            else:
                result.append(line)

    if n_dedup:
        print(f"⚠ [ETORO] Dédup inter-fichiers TSV: {n_dedup} doublon(s) éliminé(s)", file=sys.stderr)

    return result


def format_site(site_dir, verbose=False, logger=None):
    """API pour Update.

    Traite tous les fichiers eToro présents dans le répertoire :
    - TSV (eToroTransactions_*.tsv) → opérations Money EUR
    - XLSX (etoro-account-statement*.xlsx) → opérations Reserve USD
    - PDF → soldes (pdf_home) et/ou positions (pdf_portfolio)
    - CSV positions_titres_parsed.csv → positions titres
    """
    if logger is None:
        from inc_logging import Logger
        logger = Logger(SITE, verbose=verbose)

    # Vérification fichiers dropbox
    from inc_format import verify_dropbox_files
    for w in verify_dropbox_files(site_dir, SITE):
        logger.warning(w)

    site_dir = Path(site_dir)

    # Phase 1: TSV Money — traitement séparé si multiples fichiers (dédup inter-fichiers)
    tsv_files = sorted(site_dir.glob('*[Tt]ransactions*.tsv'))
    tsv_files = [f for f in tsv_files if '_formatted' not in f.name and '_temp' not in f.name]

    if len(tsv_files) >= 2:
        # Traiter chaque TSV séparément puis dédupliquer
        ops_by_file = {}
        money_wrapper = _wrap_lines_to_tuples(process_money_operations, 9)
        for f in tsv_files:
            if verbose:
                logger.verbose(f"TSV: {f.name}")
            ops_by_file[f.name] = money_wrapper(f)
        ops_money = _deduplicate_cross_file_ops(ops_by_file, verbose)
    else:
        ops_money = []  # sera traité par process_files ci-dessous

    skip_tsv = len(tsv_files) >= 2

    # Phase 2: Autres fichiers via process_files
    reserve_wrapper = _wrap_lines_to_tuples(process_reserve_operations, 9)
    positions_wrapper = _wrap_lines_to_tuples(process_positions, 4)

    def _money_wrapper_skip(file_path):
        """Skip TSV déjà traités en phase 1."""
        if skip_tsv:
            return []
        return _wrap_lines_to_tuples(process_money_operations, 9)(file_path)

    handlers = [
        # TSV opérations Money EUR (skip si déjà dédupliqués)
        ('*[Tt]ransactions*.tsv', _money_wrapper_skip, 'ops'),
        # XLSX opérations Reserve USD
        ('*account-statement*.xlsx', reserve_wrapper, 'ops'),
        ('*accountstatement*.xlsx', reserve_wrapper, 'ops'),
        # PDF soldes (détection par contenu)
        ('*.pdf', _process_pdf_etoro_wrapper, 'ops'),
        # PDF positions (détection par contenu)
        ('*.pdf', _process_pdf_portfolio_wrapper, 'pos'),
        # CSV positions déjà parsés
        ('positions_*_parsed.csv', positions_wrapper, 'pos'),
    ]

    ops_other, positions = process_files(site_dir, handlers, verbose, SITE, logger=logger)

    return ops_money + ops_other, positions


def log_csv_debug(operations, positions, site_dir, logger=None):
    """Wrapper vers inc_format.log_csv_debug()"""
    _log_csv_debug(SITE, operations, positions, logger)


if __name__ == '__main__':
    from inc_format import cli_main
    cli_main(format_site)
