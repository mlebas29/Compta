#!/usr/bin/env python3
"""
Module d'import Excel pour comptes.xlsm — extension de ComptaExcel.

Classe ComptaExcelImport(ComptaExcel) : logique d'import (append_to_comptes,
process_valorisations, generate_linked_operations, UNO...).
Utilisée par composition dans ComptaUpdater (cpt_update.py).
"""

import os
import sys
import subprocess
from pathlib import Path
from datetime import datetime
from copy import copy

from inc_exchange_rates import convert_to_eur

try:
    import openpyxl
    from openpyxl.styles import Font, Color, PatternFill
    import warnings
    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
except ImportError:
    print("❌ Erreur: openpyxl n'est pas installé")
    print("Installation: pip3 install openpyxl")
    sys.exit(1)

from inc_excel_compta import ComptaExcel, LINKED_OPERATIONS, SOLDE_AUTO_ACCOUNTS, parse_date, parse_montant
from inc_excel_schema import (
    SHEET_OPERATIONS, SHEET_AVOIRS, SHEET_PLUS_VALUE,
    SHEET_CONTROLES, Operation,
)


# ============================================================================
# FORMATS DE NOMBRE PAR DEVISE (pour opérations multi-devises)
# ============================================================================
# Appliqués à la colonne Montant — générés depuis config_cotations.json (style openpyxl/US)
from inc_formats import devise_format, _load_decimals, _DEFAULT_DECIMALS

def _build_currency_formats():
    """Construit les formats openpyxl depuis config_cotations.json."""
    decimals = _load_decimals()
    result = {'EUR': devise_format('EUR', _DEFAULT_DECIMALS, style='openpyxl')}
    for code, dec in decimals.items():
        result[code] = devise_format(code, dec, style='openpyxl')
    return result

CURRENCY_NUMBER_FORMATS = _build_currency_formats()
EQUIV_EUR_FORMAT = CURRENCY_NUMBER_FORMATS['EUR']

# Couleur de fond grise pour les devises non-EUR (colonnes Montant et Devise)
NON_EUR_FILL = PatternFill(start_color='FFDCDCDC', end_color='FFDCDCDC', fill_type='solid')  # Gris clair (0xDCDCDC)

# Couleur bleue pour les champs mis à jour par cpt_update (PVL date_solde, solde)
BLUE_COLOR = '000000FF'


# ============================================================================
# FONCTIONS UTILITAIRES IMPORT
# ============================================================================

def copy_cell_formatting(source_cell, target_cell):
    """Copie tous les attributs de formatage d'une cellule source vers une cellule cible"""
    if source_cell.font:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            underline=source_cell.font.underline,
            color=source_cell.font.color
        )
    if source_cell.border:
        target_cell.border = copy(source_cell.border)
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format


def copy_row_formatting(ws, source_row, target_row, num_cols=9):
    """Copie le formatage d'une ligne template vers une ligne cible"""
    for col in range(1, num_cols + 1):
        copy_cell_formatting(
            ws.cell(source_row, col),
            ws.cell(target_row, col)
        )


def extract_currency_from_account(compte):
    """Extrait la devise du nom de compte

    Exemples:
        "Compte Yuh CHF" → "CHF"
        "Compte Wise USD" → "USD"
        "Compte principal SG" → "EUR" (défaut)
    """
    if not compte:
        return 'EUR'

    currencies = ['CHF', 'USD', 'SGD', 'SEK', 'JPY', 'EUR']
    compte_upper = compte.upper()

    # Cas spécial : comptes BTC → devise SAT
    if compte_upper.endswith('BTC') or ' BTC' in compte_upper:
        return 'SAT'

    # Cas spécial : comptes XMR → devise XMR
    if compte_upper.endswith('XMR') or ' XMR' in compte_upper:
        return 'XMR'

    for currency in currencies:
        if compte_upper.endswith(currency):
            return currency
        if f' {currency}' in compte_upper:
            return currency

    return 'EUR'


def normalize_amount(amount_str):
    """Normalise un montant pour comparaison (string → float)"""
    if isinstance(amount_str, (int, float)):
        return float(amount_str)
    amount_str = str(amount_str).strip().replace(',', '.').replace(' ', '')
    try:
        return float(amount_str)
    except ValueError:
        return 0.0


def parse_csv_line(line, logger=None):
    """Parse une ligne CSV et retourne un dict (9 champs)"""
    parts = line.strip().split(';')

    if len(parts) < 8:
        if logger:
            logger.error(f"Ligne invalide (< 8 champs): {line[:80]}")
        return None

    return {
        'date': parts[0],
        'label': parts[1],
        'montant': parts[2],
        'devise': parts[3],
        'equiv': parts[4],
        'ref': parts[5],
        'categorie': parts[6],
        'compte': parts[7],
        'commentaire': parts[8] if len(parts) > 8 else None
    }


def get_valid_accounts(excel_file, verbose=False):
    """Lit la liste des comptes valides depuis la feuille Avoirs.

    Ne retient que les lignes où la colonne E (Devise) est renseignée.
    S'arrête au marqueur "Total".
    """
    excel_path = Path(excel_file)
    if not excel_path.exists():
        return []

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if SHEET_AVOIRS not in wb.sheetnames:
            return []

        sheet = wb[SHEET_AVOIRS]
        accounts = []

        from inc_excel_schema import ColResolver
        cr = ColResolver.from_openpyxl(wb)
        avr_start, _ = cr.rows('AVRintitulé')
        avr_start = avr_start or 4
        for row_idx in range(avr_start + 1, avr_start + 200):
            compte_name = sheet.cell(row_idx, cr.col('AVRintitulé')).value
            if compte_name and 'total' in str(compte_name).lower():
                break
            devise = sheet.cell(row_idx, cr.col('AVRdevise')).value
            if compte_name and devise:
                compte_clean = str(compte_name).strip()
                if compte_clean:
                    accounts.append(compte_clean)

        wb.close()
        return accounts
    except Exception:
        return []


# ============================================================================
# CLASSE IMPORT (extension de ComptaExcel)
# ============================================================================

class ComptaExcelImport(ComptaExcel):
    """Extension import de ComptaExcel — append, valorisations, linked ops, UNO."""

    def __init__(self, comptes_file, verbose=False, all_soldes=False, logger=None):
        super().__init__(comptes_file=comptes_file, verbose=verbose, all_soldes=all_soldes, logger=logger)

        # Stats (remplies par append_to_comptes, process_valorisations)
        self.stats = {
            'pdf_converted': 0,
            'csv_converted': 0,
            'lines_added': 0,
            'duplicates_skipped': 0,
            'operations_by_compte': {},
            'comptes_avec_solde': set(),
            'positions_by_compte': {},
            'errors': []
        }
        self.comptes_avec_operations = set()
        self.comptes_valorisations = {}

    # ====================================================================
    # Lectures
    # ====================================================================

    def get_account_balance(self, compte):
        """Récupère le dernier solde d'un compte depuis la feuille Opérations"""
        if not self.ws_operations:
            return 0.0

        last_solde = 0.0
        last_row = self.find_last_data_row()

        for row in range(2, last_row + 1):
            row_compte = self.ws_operations.cell(row, self.cr.col('OPcompte')).value
            categorie = self.ws_operations.cell(row, self.cr.col('OPcatégorie')).value
            montant = self.ws_operations.cell(row, self.cr.col('OPmontant')).value

            if row_compte == compte and categorie and str(categorie).lower() == '#solde':
                if montant is not None:
                    last_solde = normalize_amount(montant)

        return last_solde

    def find_last_solde_date(self, compte):
        """Trouve la date du dernier #Solde pour un compte donné

        Returns:
            datetime ou None si aucun #Solde trouvé
        """
        if not self.ws_operations:
            return None

        last_solde_date = None
        last_row = self.find_last_data_row()

        for row in range(2, last_row + 1):
            row_compte = self.ws_operations.cell(row, self.cr.col('OPcompte')).value
            categorie = self.ws_operations.cell(row, self.cr.col('OPcatégorie')).value
            date = self.ws_operations.cell(row, self.cr.col('OPdate')).value

            if row_compte == compte and categorie and str(categorie).lower() == '#solde':
                if date:
                    if hasattr(date, 'strftime'):
                        last_solde_date = date
                    else:
                        try:
                            last_solde_date = datetime.strptime(str(date), '%d/%m/%Y')
                        except:
                            pass

        return last_solde_date

    def load_existing_soldes(self):
        """Charge les #Solde existants d'Excel pour éviter les doublons

        Returns:
            dict: clé = (date_str, compte_norm, categorie_norm) → montant_norm
        """
        existing_soldes = {}

        if not self.ws_operations:
            return existing_soldes

        try:
            last_row = self.find_last_data_row()

            for row in range(2, last_row + 1):
                categorie = self.ws_operations.cell(row, self.cr.col('OPcatégorie')).value
                if not categorie or not str(categorie).startswith('#'):
                    continue

                date = self.ws_operations.cell(row, self.cr.col('OPdate')).value
                montant = self.ws_operations.cell(row, self.cr.col('OPmontant')).value
                compte = self.ws_operations.cell(row, self.cr.col('OPcompte')).value

                if not date or not compte:
                    continue

                if hasattr(date, 'strftime'):
                    date_str = date.strftime('%d/%m/%Y')
                else:
                    date_str = str(date)

                compte_norm = str(compte).lower()
                categorie_norm = str(categorie).lower()
                montant_norm = normalize_amount(montant) if montant else ''

                solde_key = (date_str, compte_norm, categorie_norm)
                existing_soldes[solde_key] = montant_norm

            self.logger.verbose(f"{len(existing_soldes)} #Solde existants chargés")
        except Exception as e:
            self.logger.error(f"Erreur lecture #Solde existants: {e}")

        return existing_soldes

    def load_existing_operations(self, compte_filtre=None, depuis_date=None):
        """Charge les opérations existantes pour détecter les doublons

        Args:
            compte_filtre: Si spécifié, ne charge que les opérations de ce compte
            depuis_date: Si spécifié, ne charge que les opérations à partir de cette date
        """
        existing_ops = set()

        if not self.ws_operations:
            return existing_ops

        try:
            last_row = self.find_last_data_row()
            if not compte_filtre:
                self.logger.verbose(f"Dernière ligne avec données: {last_row}")

            for row in range(2, last_row + 1):
                date = self.ws_operations.cell(row, self.cr.col('OPdate')).value
                label = self.ws_operations.cell(row, self.cr.col('OPlibellé')).value
                montant_cell = self.ws_operations.cell(row, self.cr.col('OPmontant'))
                montant = montant_cell.value
                row_compte = self.ws_operations.cell(row, self.cr.col('OPcompte')).value

                # Si montant est une formule, essayer de l'évaluer
                if isinstance(montant, str) and montant.startswith('='):
                    try:
                        montant = eval(montant[1:])
                    except:
                        self.logger.verbose(f"⚠️  Ligne {row}: formule non évaluable: {montant}")
                        continue

                if date and label and montant is not None and row_compte:
                    if compte_filtre and str(row_compte).lower() != str(compte_filtre).lower():
                        continue

                    date_obj = None
                    if hasattr(date, 'strftime'):
                        date_obj = date
                        date_str = date.strftime('%d/%m/%Y')
                    else:
                        date_str = str(date)
                        try:
                            date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                        except:
                            pass

                    if depuis_date and date_obj:
                        if date_obj < depuis_date:
                            continue

                    montant_norm = normalize_amount(montant)
                    label_norm = ' '.join(str(label).split())
                    compte_norm = str(row_compte).lower()

                    key = f"{date_str};{compte_norm};{montant_norm};{label_norm}"
                    existing_ops.add(key)

            if not compte_filtre:
                self.logger.verbose(f"{len(existing_ops)} opérations existantes chargées")
        except Exception as e:
            self.logger.error(f"Erreur lecture opérations existantes: {e}")

        return existing_ops

    # ====================================================================
    # Enrichissement et génération
    # ====================================================================

    def enrich_equiv_for_non_eur(self, operations):
        """Enrichit la colonne Equiv pour les opérations à appairer

        Règle : Equiv uniquement si opération appariée (avec Réf) ou Achat titres

        - EUR : equiv = montant (par définition, pas de conversion)
        - Non-EUR : equiv = conversion via taux de change ECB

        Exclusions :
        - Opérations non appariées (sans Réf ou Réf = "-") sauf Achat titres
        - Virements (devise constante, pas de plus-value)
        """
        enriched_count = 0

        for op in operations:
            devise = op.devise
            equiv = op.equiv.strip()
            categorie = op.categorie.lower()
            ref = op.ref.strip()

            if not ref or ref == '-':
                if devise == 'EUR':
                    # EUR : laisser passer les catégories cross-currency (equiv = montant)
                    if categorie not in ('achat titres', 'vente titres', 'change', 'achat métaux'):
                        continue
                else:
                    # Non-EUR : seulement titres (ECB interférerait avec _match_equiv)
                    if categorie not in ('achat titres', 'vente titres'):
                        continue

            if 'virement' in categorie:
                continue

            if equiv:
                continue

            montant_str = op.montant or '0'
            try:
                montant = float(str(montant_str).replace(',', '.').replace(' ', ''))
            except (ValueError, TypeError):
                continue

            if devise == 'EUR':
                # EUR : equiv = montant seulement pour les catégories cross-currency
                # (Change, Achat métaux) — pas pour Achat titres EUR interne
                if categorie in ('change', 'achat métaux'):
                    op.equiv = montant_str
                    enriched_count += 1
            else:
                date_str = op.date
                eur_amount = convert_to_eur(montant, devise, date_str)
                if eur_amount is not None:
                    op.equiv = f"{eur_amount:.2f}"
                    enriched_count += 1
                    self.logger.verbose(f"  Equiv enrichi: {montant:.2f} {devise} → {eur_amount:.2f} EUR ({op.label[:40]})")

        if enriched_count > 0:
            self.logger.info(f"Equiv enrichi pour {enriched_count} opération(s)")

        return operations

    def generate_linked_operations(self, operations):
        """Génère les opérations liées (Espèces, Créances, Titres)

        Pour chaque opération matchant un pattern LINKED_OPERATIONS :
        - Génère l'opération symétrique (signe inversé) dans le compte cible
        - Génère un nouveau #Solde pour le compte cible
        - Les deux opérations reçoivent ref='-' (appariement différé à cpt_pair)
        """
        if not operations:
            return operations

        # Trier par date pour que les soldes cumulés soient corrects
        def parse_date_key(op):
            try:
                return datetime.strptime(op.date, '%d/%m/%Y')
            except (ValueError, TypeError):
                return datetime.min
        operations.sort(key=parse_date_key)

        # Retirer les #Solde manuels des comptes auto-solde (seront recalculés)
        if SOLDE_AUTO_ACCOUNTS:
            filtered = []
            for op in operations:
                if op.compte in SOLDE_AUTO_ACCOUNTS and op.categorie == '#Solde':
                    self.logger.verbose(f"#Solde manuel retiré pour {op.compte} (sera auto-généré)")
                else:
                    filtered.append(op)
            operations = filtered

        # Charger les soldes initiaux des comptes auto-solde
        solde_auto_balances = {}
        solde_auto_last_date = {}
        for compte, config in SOLDE_AUTO_ACCOUNTS.items():
            solde_auto_balances[compte] = self.get_account_balance(compte)
            solde_auto_last_date[compte] = None
            self.logger.verbose(f"Solde auto initial {compte}: {solde_auto_balances[compte]} {config['devise']}")

        # Charger les soldes initiaux des comptes cibles
        balances = {}
        for pattern_config in LINKED_OPERATIONS.values():
            compte_cible = pattern_config['compte_cible']
            if compte_cible not in balances:
                balances[compte_cible] = self.get_account_balance(compte_cible)
                self.logger.verbose(f"Solde initial {compte_cible}: {balances[compte_cible]:.2f} €")

        enriched_ops = []

        for op in operations:
            label = op.label
            matched_pattern = None

            for pattern, config in LINKED_OPERATIONS.items():
                if pattern in label.upper():
                    matched_pattern = pattern
                    break

            if matched_pattern:
                config = LINKED_OPERATIONS[matched_pattern]
                compte_cible = config['compte_cible']

                if op.ref in ('', '-'):
                    op.ref = '-'
                op.categorie = 'Virement'

                montant_original = normalize_amount(op.montant)
                montant_cible = -montant_original

                balances[compte_cible] += montant_cible

                op_symetrique = Operation(
                    date=op.date,
                    label='',
                    montant=str(montant_cible).replace('.', ','),
                    devise=op.devise,
                    ref='-',
                    categorie='Virement',
                    compte=compte_cible,
                )

                op_solde = Operation(
                    date=op.date,
                    label='Relevé compte',
                    montant=str(balances[compte_cible]).replace('.', ','),
                    devise=op.devise,
                    categorie='#Solde',
                    compte=compte_cible,
                )

                enriched_ops.append(op)
                enriched_ops.append(op_symetrique)
                enriched_ops.append(op_solde)

                self.logger.verbose(f"Opération liée détectée: {matched_pattern}")
                self.logger.verbose(f"  → {compte_cible}: {montant_cible:+.2f} € (solde: {balances[compte_cible]:.2f} €)")

            else:
                compte = op.compte
                categorie = op.categorie
                ref = op.ref

                if ('Réserve' in compte and
                    categorie in ['Achat titres', 'Vente titres'] and
                    ref in ['', '-']):

                    op.ref = '-'

                    montant_reserve = normalize_amount(op.montant)
                    montant_titres = -montant_reserve

                    categorie_titres = categorie

                    equiv_reserve = op.equiv.strip()
                    equiv_titres = ''
                    if equiv_reserve:
                        try:
                            equiv_val = float(equiv_reserve.replace(',', '.').replace(' ', ''))
                            equiv_titres = str(-equiv_val).replace('.', ',')
                        except:
                            pass

                    compte_titres = compte.replace('Réserve', 'Titres')

                    op_titres = Operation(
                        date=op.date,
                        label=op.label,
                        montant=str(montant_titres).replace('.', ','),
                        devise=op.devise,
                        equiv=equiv_titres,
                        ref='-',
                        categorie=categorie_titres,
                        compte=compte_titres,
                        commentaire=op.commentaire,
                    )

                    enriched_ops.append(op)
                    enriched_ops.append(op_titres)

                    self.logger.verbose(f"Opération Titres générée: {compte} {categorie}")
                    self.logger.verbose(f"  Réserve: {montant_reserve:+.2f} € → Titres: {montant_titres:+.2f} €")

                elif (compte in SOLDE_AUTO_ACCOUNTS and
                      categorie == SOLDE_AUTO_ACCOUNTS[compte]['categorie_trigger']):
                    montant_val = normalize_amount(op.montant)
                    solde_auto_balances[compte] += montant_val
                    solde_auto_last_date[compte] = op.date
                    enriched_ops.append(op)

                else:
                    enriched_ops.append(op)

        # Émettre les #Solde auto-générés
        for compte, config in SOLDE_AUTO_ACCOUNTS.items():
            if solde_auto_last_date[compte] is not None:
                op_solde = Operation(
                    date=solde_auto_last_date[compte],
                    label='Relevé compte',
                    montant=str(solde_auto_balances[compte]).replace('.', ','),
                    devise=config['devise'],
                    categorie='#Solde',
                    compte=compte,
                )
                enriched_ops.append(op_solde)
                self.logger.verbose(f"#Solde auto-généré {compte}: {solde_auto_balances[compte]} {config['devise']}")

        return enriched_ops

    # ====================================================================
    # Gros traitements L+E
    # ====================================================================

    def append_to_comptes(self, operations):
        """Ajoute les nouvelles lignes à comptes.xlsm

        Args:
            operations: list[Operation] — opérations à importer

        Approche brutale (sans Td) : Compare TOUTES les opérations avec TOUTES
        les opérations Excel, sans hypothèse de rangement chronologique.
        Performance O(n) grâce aux sets Python.
        """
        if not self.ws_operations:
            self.logger.error("Onglet Opérations non ouvert")
            return False

        all_ops = [op for op in operations if op.compte]

        if not all_ops:
            self.logger.verbose("Aucune ligne valide à traiter")
            return True

        self.logger.verbose(f"Chargement de toutes les opérations Excel...")
        existing_ops = self.load_existing_operations()
        existing_soldes = self.load_existing_soldes()

        unique_lines = []
        soldes_par_compte = {}
        seen_keys = set()

        for op in all_ops:
            categorie = op.categorie
            if categorie and categorie.lower() == '#solde':
                date_obj = parse_date(op.date)

                compte_norm = str(op.compte).lower()
                # Identité métier d'un #Solde = (compte, date) : une observation à une date.
                # Plusieurs #Solde même compte à dates distinctes (ancrage + relevé) = légitimes.
                solde_key = (compte_norm, date_obj)
                montant_norm = normalize_amount(op.montant) if op.montant else ''

                if solde_key not in soldes_par_compte:
                    soldes_par_compte[solde_key] = []
                soldes_par_compte[solde_key].append((date_obj, montant_norm, op))
                continue

            montant_norm = normalize_amount(op.montant)
            label_norm = ' '.join(str(op.label).split())

            date_obj = parse_date(op.date)
            date_norm = date_obj.strftime('%d/%m/%Y') if date_obj else op.date

            compte_norm = str(op.compte).lower()

            key = f"{date_norm};{compte_norm};{montant_norm};{label_norm}"

            if key not in existing_ops and key not in seen_keys:
                unique_lines.append(op)
                seen_keys.add(key)
            elif key in seen_keys:
                self.logger.verbose(f"  INTRA-BATCH doublon: {key}")
            else:
                self.stats['duplicates_skipped'] += 1

        # Dédupliquer les #Solde collectés (même compte + même date)
        for solde_key, soldes_list in soldes_par_compte.items():
            compte_norm, _ = solde_key

            soldes_list.sort(key=lambda x: x[0] if x[0] else datetime.min, reverse=True)

            date_obj, montant_norm, best_op = soldes_list[0]
            date_str = date_obj.strftime('%d/%m/%Y') if date_obj else ''

            excel_key = (date_str, compte_norm, '#solde')
            if excel_key in existing_soldes:
                self.stats['duplicates_skipped'] += len(soldes_list)
                continue

            unique_lines.append(best_op)
            self.stats['duplicates_skipped'] += len(soldes_list) - 1

            if len(soldes_list) > 1:
                self.logger.verbose(f"  #Solde {compte_norm}: gardé {date_str}, ignoré {len(soldes_list)-1} autre(s)")

        if not unique_lines:
            self.logger.verbose("Aucune nouvelle ligne (tous des doublons)")
            return True

        unique_lines = self.enrich_equiv_for_non_eur(unique_lines)
        unique_lines = self.generate_linked_operations(unique_lines)

        # Filtrer les #Solde orphelins
        comptes_avec_operations = set()
        nb_soldes = 0
        for op in unique_lines:
            if op.categorie.lower() == '#solde':
                nb_soldes += 1
            else:
                comptes_avec_operations.add(op.compte)

        self.comptes_avec_operations.update(comptes_avec_operations)

        fichier_soldes_pur = (nb_soldes > 0 and len(comptes_avec_operations) == 0)

        if not self.all_soldes:
            filtered_lines = []
            for op in unique_lines:
                if op.categorie.lower() == '#solde':
                    if op.compte not in comptes_avec_operations:
                        self.stats['duplicates_skipped'] += 1
                        self.logger.verbose(f"  #Solde ignoré (pas de nouvelles opérations): {op.compte}")
                        continue
                filtered_lines.append(op)

            unique_lines = filtered_lines

        if not unique_lines:
            self.logger.verbose("Aucune nouvelle ligne (tous doublons ou soldes orphelins)")
            return True

        # Vérifier que tous les comptes existent dans Avoirs (détection précoce)
        valid_accounts = set(get_valid_accounts(self.comptes_file))
        unknown_accounts = set()
        for op in unique_lines:
            compte = op.compte
            if compte and compte not in valid_accounts:
                unknown_accounts.add(compte)
        if unknown_accounts:
            for compte in sorted(unknown_accounts):
                self.logger.error(f"Compte '{compte}' absent de la feuille Avoirs")
            self.logger.error("Ajouter le(s) compte(s) dans Avoirs avec une devise, puis relancer")
            return False

        last_data_row = self.find_last_data_row()
        next_row = last_data_row + 1
        self.logger.verbose(f"Ajout à partir de la ligne {next_row}")

        template_cache = {}

        for op in unique_lines:
            try:
                compte = op.compte
                if compte not in template_cache:
                    template_row = None
                    for row in range(last_data_row, 1, -1):
                        row_compte = self.ws_operations.cell(row, self.cr.col('OPcompte')).value
                        if row_compte == compte:
                            template_row = row
                            break

                    if template_row:
                        template_cache[compte] = [self.ws_operations.cell(template_row, col) for col in range(1, 11)]
                    else:
                        self.logger.verbose(f"Compte '{compte}' sans opération existante - template par devise")

                        devise = op.devise
                        for row in range(last_data_row, 1, -1):
                            row_devise = self.ws_operations.cell(row, self.cr.col('OPdevise')).value
                            if row_devise == devise:
                                template_row = row
                                break

                        if template_row:
                            template_cache[compte] = [self.ws_operations.cell(template_row, col) for col in range(1, 11)]
                            self.logger.verbose(f"   → Format cloné depuis une ligne {devise} (compte différent)")
                        else:
                            template_cache[compte] = [self.ws_operations.cell(last_data_row, col) for col in range(1, 11)]
                            self.logger.verbose(f"   → Format cloné depuis la dernière ligne (fallback)")

                template_cells = template_cache[compte]

                # Date
                cell = self.ws_operations.cell(next_row, self.cr.col('OPdate'))
                date_obj = parse_date(op.date)
                cell.value = date_obj if date_obj else op.date
                if template_cells:
                    copy_cell_formatting(template_cells[self.cr.col('OPdate') - 1], cell)

                # Libellé
                cell = self.ws_operations.cell(next_row, self.cr.col('OPlibellé'))
                cell.value = op.label
                if template_cells:
                    copy_cell_formatting(template_cells[self.cr.col('OPlibellé') - 1], cell)

                # Montant
                cell = self.ws_operations.cell(next_row, self.cr.col('OPmontant'))
                montant_str = op.montant.replace(',', '.')
                try:
                    cell.value = float(montant_str)
                except ValueError:
                    cell.value = montant_str
                if template_cells:
                    copy_cell_formatting(template_cells[self.cr.col('OPmontant') - 1], cell)
                devise = op.devise
                if devise in CURRENCY_NUMBER_FORMATS:
                    cell.number_format = CURRENCY_NUMBER_FORMATS[devise]
                if devise != 'EUR':
                    cell.fill = NON_EUR_FILL

                # Devise
                cell_devise = self.ws_operations.cell(next_row, self.cr.col('OPdevise'))
                cell_devise.value = op.devise
                if template_cells:
                    copy_cell_formatting(template_cells[self.cr.col('OPdevise') - 1], cell_devise)
                if devise != 'EUR':
                    cell_devise.fill = NON_EUR_FILL

                # Equiv
                cell = self.ws_operations.cell(next_row, self.cr.col('OPequiv_euro'))
                equiv_str = op.equiv
                if equiv_str:
                    try:
                        cell.value = float(str(equiv_str).replace(',', '.'))
                    except (ValueError, AttributeError):
                        cell.value = equiv_str
                else:
                    cell.value = None
                if template_cells:
                    copy_cell_formatting(template_cells[self.cr.col('OPequiv_euro') - 1], cell)
                cell.number_format = EQUIV_EUR_FORMAT

                # Ref
                cell = self.ws_operations.cell(next_row, self.cr.col('OPréf'))
                cell.value = op.ref if op.ref else None
                if template_cells:
                    copy_cell_formatting(template_cells[self.cr.col('OPréf') - 1], cell)

                # Catégorie
                cell = self.ws_operations.cell(next_row, self.cr.col('OPcatégorie'))
                cell.value = op.categorie
                if template_cells:
                    copy_cell_formatting(template_cells[self.cr.col('OPcatégorie') - 1], cell)

                # Compte
                cell = self.ws_operations.cell(next_row, self.cr.col('OPcompte'))
                cell.value = op.compte
                if template_cells:
                    copy_cell_formatting(template_cells[self.cr.col('OPcompte') - 1], cell)

                # Commentaire
                cell = self.ws_operations.cell(next_row, self.cr.col('OPcommentaire'))
                cell.value = op.commentaire or None
                if template_cells and len(template_cells) > self.cr.col('OPcommentaire') - 1:
                    copy_cell_formatting(template_cells[self.cr.col('OPcommentaire') - 1], cell)

                next_row += 1
                self.stats['lines_added'] += 1
                if compte not in self.stats['operations_by_compte']:
                    self.stats['operations_by_compte'][compte] = 0
                self.stats['operations_by_compte'][compte] += 1
                if op.categorie == '#Solde':
                    self.stats['comptes_avec_solde'].add(compte)
            except Exception as e:
                self.logger.error(f"Erreur ajout ligne {next_row}: {e}")
                self.stats['errors'].append(f"Ligne {next_row}: {e}")

        self.logger.verbose(f"{len(unique_lines)} nouvelles lignes ajoutées")
        return True

    def process_valorisations(self, valorisation_files):
        """Traite les fichiers positions/supports pour mettre à jour Plus_value.

        Lit les fichiers au format unifié 5 colonnes (Date;Ligne;Montant;Compte;Sous-compte)
        et met à jour la feuille Plus_value.
        """
        if not valorisation_files:
            return True

        self.logger.verbose(f"\n=== Traitement valorisations ({len(valorisation_files)} fichiers) ===")

        ws_pv = self.wb[SHEET_PLUS_VALUE]
        date_aujourdhui = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

        positions_not_found = []
        comptes_valorises = set()

        for file_path in valorisation_files:
            try:
                self.logger.verbose(f"\nTraitement: {file_path.name}")

                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        csv_content = f.read()
                except Exception as e:
                    self.logger.error(f"  ❌ Erreur lecture {file_path.name}: {e}")
                    continue

                first_line = csv_content.split('\n')[0] if csv_content else ''
                if not first_line.startswith('Date;Ligne;Montant;Compte;Sous-compte'):
                    self.logger.error(f"  ❌ Format invalide: {file_path.name}")
                    continue

                valorisation_compte = 0.0
                nb_updates = 0

                for line in csv_content.strip().split('\n')[1:]:
                    if not line.strip():
                        continue

                    fields = line.split(';')
                    if len(fields) < 4:
                        continue

                    date_str = fields[0].strip()
                    ligne = fields[1].strip()
                    montant_str = fields[2].strip()
                    compte = fields[3].strip()

                    try:
                        montant = float(montant_str)
                        valorisation_compte += montant
                    except ValueError:
                        self.logger.error(f"  ⚠ Montant invalide pour {ligne}: {montant_str}")
                        continue

                    # Chercher la ligne correspondante dans Plus_value
                    pv_row_idx = None

                    for row_idx in range(self._pvl_data_start, ws_pv.max_row + 1):
                        compte_pv = ws_pv.cell(row_idx, self.cr.col('PVLcompte')).value
                        ligne_pv = ws_pv.cell(row_idx, self.cr.col('PVLtitre')).value

                        if compte_pv != compte:
                            continue

                        if not ligne_pv:
                            continue

                        ligne_pv_str = str(ligne_pv)
                        if not ligne_pv_str.startswith('*'):
                            continue

                        ligne_normalized = ' '.join(ligne.upper().split())
                        ligne_pv_normalized = ' '.join(ligne_pv_str.replace('*', '').upper().split())

                        if ligne_normalized in ligne_pv_normalized or ligne_pv_normalized in ligne_normalized:
                            pv_row_idx = row_idx
                            break

                    # Traiter les lignes #Solde Titres
                    if ligne.startswith('#Solde') or ligne == 'Relevé compte':
                        ecrire_solde_operations = self.all_soldes or compte in self.comptes_avec_operations

                        if not ecrire_solde_operations:
                            self.logger.verbose(f"  ℹ #Solde Opérations ignoré (pas de nouvelles opérations): {compte}")

                        if ecrire_solde_operations:
                            ws_operations = self.wb[SHEET_OPERATIONS]
                            last_data_row = self.find_last_data_row()
                            target_row = last_data_row + 1

                            template_row = None
                            fallback_template_row = None

                            for row_idx in range(last_data_row, max(1, last_data_row - 200), -1):
                                if ws_operations.cell(row_idx, self.cr.col('OPcatégorie')).value == '#Solde':
                                    if fallback_template_row is None:
                                        fallback_template_row = row_idx

                                    compte_cell = ws_operations.cell(row_idx, self.cr.col('OPcompte'))
                                    if compte_cell.value == compte:
                                        template_row = row_idx
                                        break

                            if template_row is None and fallback_template_row is not None:
                                template_row = fallback_template_row

                            if template_row:
                                copy_row_formatting(ws_operations, template_row, target_row, num_cols=9)

                            try:
                                date_solde = datetime.strptime(date_str, '%d/%m/%Y')
                            except ValueError:
                                date_solde = date_aujourdhui

                            ws_operations.cell(target_row, self.cr.col('OPdate')).value = date_solde
                            ws_operations.cell(target_row, self.cr.col('OPlibellé')).value = 'Relevé compte'
                            ws_operations.cell(target_row, self.cr.col('OPmontant')).value = montant
                            ws_operations.cell(target_row, self.cr.col('OPdevise')).value = extract_currency_from_account(compte)
                            ws_operations.cell(target_row, self.cr.col('OPequiv_euro')).value = None
                            ws_operations.cell(target_row, self.cr.col('OPréf')).value = None
                            ws_operations.cell(target_row, self.cr.col('OPcatégorie')).value = '#Solde'
                            ws_operations.cell(target_row, self.cr.col('OPcompte')).value = compte
                            ws_operations.cell(target_row, self.cr.col('OPcommentaire')).value = None

                            if template_row:
                                template_date_format = ws_operations.cell(template_row, self.cr.col('OPdate')).number_format
                                ws_operations.cell(target_row, self.cr.col('OPdate')).number_format = template_date_format

                            devise = extract_currency_from_account(compte)
                            if devise in CURRENCY_NUMBER_FORMATS:
                                ws_operations.cell(target_row, self.cr.col('OPmontant')).number_format = CURRENCY_NUMBER_FORMATS[devise]
                            ws_operations.cell(target_row, self.cr.col('OPequiv_euro')).number_format = EQUIV_EUR_FORMAT
                            if devise != 'EUR':
                                ws_operations.cell(target_row, self.cr.col('OPmontant')).fill = NON_EUR_FILL
                                ws_operations.cell(target_row, self.cr.col('OPdevise')).fill = NON_EUR_FILL

                            self.stats['comptes_avec_solde'].add(compte)
                            self.logger.verbose(f"  ✓ #Solde Opérations ajouté (ligne {target_row}): {compte} = {montant:.2f}")

                    elif pv_row_idx:
                        cell_solde = ws_pv.cell(pv_row_idx, self.cr.col('PVLmontant'))
                        is_formula = isinstance(cell_solde.value, str) and cell_solde.value.startswith('=')

                        if not is_formula:
                            cell_date = ws_pv.cell(pv_row_idx, self.cr.col('PVLdate'))
                            cell_date.value = date_aujourdhui
                            cell_date.font = cell_date.font.copy(color=BLUE_COLOR)
                            cell_solde = ws_pv.cell(pv_row_idx, self.cr.col('PVLmontant'))
                            cell_solde.value = montant
                            cell_solde.font = cell_solde.font.copy(color=BLUE_COLOR)
                            nb_updates += 1
                            comptes_valorises.add(compte)
                            self.stats['positions_by_compte'][compte] = self.stats['positions_by_compte'].get(compte, 0) + 1
                            self.logger.verbose(f"  ✓ {ligne}: {montant:.2f} € (ligne {pv_row_idx})")
                        else:
                            self.logger.verbose(f"  ℹ {ligne} (ligne {pv_row_idx}): utilise une formule (non modifié)")
                    else:
                        if not ligne.startswith('#'):
                            # Ne signaler que si le compte a un sous-compte Titres dans Plus_value
                            # (sinon c'est un compte simple sans titres individuels)
                            has_titres = any(
                                ws_pv.cell(r, self.cr.col('PVLcompte')).value == compte
                                and ws_pv.cell(r, self.cr.col('PVLtitre')).value == 'Titres'
                                for r in range(self._pvl_data_start, ws_pv.max_row + 1)
                            )
                            if has_titres:
                                positions_not_found.append((ligne, compte, montant))
                                self.logger.verbose(f"  ⚠ {ligne}: non trouvé dans Plus_value")

                if valorisation_compte > 0:
                    self.logger.verbose(f"  {nb_updates} ligne(s) mise(s) à jour, valorisation: {valorisation_compte:.2f}€")

            except Exception as e:
                self.logger.error(f"  ❌ Erreur traitement {file_path.name}: {e}")
                import traceback
                traceback.print_exc()
                return False

        if positions_not_found:
            self.logger.warning(f"\n⚠ POSITIONS NON TROUVÉES dans Plus_value ({len(positions_not_found)}):")
            self.logger.warning("  Ces titres existent dans les fichiers source mais pas dans Excel.")
            self.logger.warning("  → Ajouter les lignes manquantes dans la feuille Plus_value")
            for ligne, compte, montant in positions_not_found:
                self.logger.warning(f"    - {ligne} ({compte}) = {montant:.2f}")

        if not hasattr(self, 'comptes_valorises'):
            self.comptes_valorises = set()
        self.comptes_valorises.update(comptes_valorises)

        return True

    def generate_missing_soldes(self):
        """Génère un #Solde calculé pour les comptes avec opérations mais sans #Solde fourni.

        Pour chaque compte concerné : dernier #Solde existant (ou 0 si compte nouveau)
        + somme des opérations postérieures = solde calculé.
        """
        if not self.wb:
            return

        comptes_sans_solde = set()
        for compte, count in self.stats['operations_by_compte'].items():
            if count > 0 and compte not in self.stats['comptes_avec_solde']:
                is_pf_titres = compte.startswith('Portefeuille') and 'Titres' in compte
                if not is_pf_titres and not compte.startswith('Espèces') and not compte.startswith('Créance'):
                    comptes_sans_solde.add(compte)

        if not comptes_sans_solde:
            return

        last_data_row = self.find_last_data_row()

        for compte in sorted(comptes_sans_solde):
            # Trouver le dernier #Solde existant (0 si compte nouveau) et la devise
            last_solde = 0.0
            last_solde_row = 0
            last_op_date = None
            devise = None

            for row in range(2, last_data_row + 1):
                row_compte = self.ws_operations.cell(row, self.cr.col('OPcompte')).value
                if row_compte != compte:
                    continue
                categorie = self.ws_operations.cell(row, self.cr.col('OPcatégorie')).value
                montant = self.ws_operations.cell(row, self.cr.col('OPmontant')).value
                if categorie and str(categorie).lower() == '#solde':
                    if montant is not None:
                        last_solde = normalize_amount(montant)
                        last_solde_row = row
                else:
                    if not devise:
                        devise = self.ws_operations.cell(row, self.cr.col('OPdevise')).value

            # Sommer les opérations après le dernier #Solde
            sum_ops = 0.0
            for row in range(max(last_solde_row + 1, 2), last_data_row + 1):
                row_compte = self.ws_operations.cell(row, self.cr.col('OPcompte')).value
                if row_compte != compte:
                    continue
                categorie = self.ws_operations.cell(row, self.cr.col('OPcatégorie')).value
                if categorie and str(categorie).lower() in ('#solde', '#info'):
                    continue
                montant = self.ws_operations.cell(row, self.cr.col('OPmontant')).value
                if montant is not None:
                    sum_ops += normalize_amount(montant)
                date_val = self.ws_operations.cell(row, self.cr.col('OPdate')).value
                if date_val:
                    last_op_date = date_val

            if last_op_date is None:
                continue

            solde_calc = last_solde + sum_ops
            if not devise:
                devise = 'EUR'

            # Écrire la ligne #Solde calculé
            next_row = last_data_row + 1

            # Template de formatage depuis une ligne existante du même compte
            template_cells = None
            for row in range(last_data_row, 1, -1):
                if self.ws_operations.cell(row, self.cr.col('OPcompte')).value == compte:
                    template_cells = [self.ws_operations.cell(row, col) for col in range(1, 11)]
                    break

            cell = self.ws_operations.cell(next_row, self.cr.col('OPdate'))
            cell.value = last_op_date
            if template_cells:
                copy_cell_formatting(template_cells[self.cr.col('OPdate') - 1], cell)

            cell = self.ws_operations.cell(next_row, self.cr.col('OPlibellé'))
            cell.value = 'Solde calculé'
            if template_cells:
                copy_cell_formatting(template_cells[self.cr.col('OPlibellé') - 1], cell)

            cell = self.ws_operations.cell(next_row, self.cr.col('OPmontant'))
            cell.value = round(solde_calc, 2)
            if template_cells:
                copy_cell_formatting(template_cells[self.cr.col('OPmontant') - 1], cell)
            if devise in CURRENCY_NUMBER_FORMATS:
                cell.number_format = CURRENCY_NUMBER_FORMATS[devise]
            if devise != 'EUR':
                cell.fill = NON_EUR_FILL

            cell = self.ws_operations.cell(next_row, self.cr.col('OPdevise'))
            cell.value = devise
            if template_cells:
                copy_cell_formatting(template_cells[self.cr.col('OPdevise') - 1], cell)
            if devise != 'EUR':
                cell.fill = NON_EUR_FILL

            cell = self.ws_operations.cell(next_row, self.cr.col('OPequiv_euro'))
            cell.value = None
            if template_cells:
                copy_cell_formatting(template_cells[self.cr.col('OPequiv_euro') - 1], cell)
            cell.number_format = EQUIV_EUR_FORMAT

            cell = self.ws_operations.cell(next_row, self.cr.col('OPréf'))
            cell.value = None
            if template_cells:
                copy_cell_formatting(template_cells[self.cr.col('OPréf') - 1], cell)

            cell = self.ws_operations.cell(next_row, self.cr.col('OPcatégorie'))
            cell.value = '#Solde'
            if template_cells:
                copy_cell_formatting(template_cells[self.cr.col('OPcatégorie') - 1], cell)

            cell = self.ws_operations.cell(next_row, self.cr.col('OPcompte'))
            cell.value = compte
            if template_cells:
                copy_cell_formatting(template_cells[self.cr.col('OPcompte') - 1], cell)

            for col in (self.cr.col('OPcommentaire'),):
                cell = self.ws_operations.cell(next_row, col)
                cell.value = None
                if template_cells and len(template_cells) > col - 1:
                    copy_cell_formatting(template_cells[col - 1], cell)

            self.stats['comptes_avec_solde'].add(compte)
            last_data_row = next_row

            self.logger.info(f"#Solde calculé {compte}: {solde_calc:.2f} {devise}")

    def verify_plus_value_dates(self):
        """Vérifie que les dates Plus_value ont été mises à jour pour les comptes valorisés."""
        if not self.wb:
            return

        if not hasattr(self, 'comptes_valorises') or not self.comptes_valorises:
            return

        ws_pv = self.wb[SHEET_PLUS_VALUE]
        date_aujourdhui = datetime.now().date()

        lignes_non_maj = []

        for row_idx in range(self._pvl_data_start, ws_pv.max_row + 1):
            compte_pv = ws_pv.cell(row_idx, self.cr.col('PVLcompte')).value
            ligne_pv = ws_pv.cell(row_idx, self.cr.col('PVLtitre')).value
            date_cell = ws_pv.cell(row_idx, self.cr.col('PVLdate'))
            solde_cell = ws_pv.cell(row_idx, self.cr.col('PVLmontant'))

            if compte_pv not in self.comptes_valorises:
                continue

            if date_cell.value is None:
                continue

            if isinstance(date_cell.value, str) and date_cell.value.startswith('='):
                continue

            solde_val = solde_cell.value
            if solde_val is None:
                continue
            if isinstance(solde_val, str) and solde_val.startswith('='):
                continue
            try:
                if float(solde_val) == 0:
                    continue
            except (ValueError, TypeError):
                pass

            date_pv = date_cell.value
            date_ok = False
            if hasattr(date_pv, 'date'):
                date_ok = (date_pv.date() == date_aujourdhui)
            elif isinstance(date_pv, str):
                try:
                    from datetime import datetime as dt
                    parsed = dt.strptime(date_pv, '%d/%m/%Y').date()
                    date_ok = (parsed == date_aujourdhui)
                except:
                    pass

            if not date_ok:
                date_str = date_pv.strftime('%d/%m/%Y') if hasattr(date_pv, 'strftime') else str(date_pv)
                label = ligne_pv if ligne_pv else '(sans libellé)'
                lignes_non_maj.append((label, compte_pv, date_str, row_idx))

        if lignes_non_maj:
            self.logger.warning(f"\n⚠ DATES PLUS_VALUE NON MISES À JOUR ({len(lignes_non_maj)}):")
            self.logger.warning("  Ces titres n'ont pas été trouvés dans les fichiers source.")
            self.logger.warning("  → Vérifier si le titre existe toujours dans le portefeuille")
            for ligne, compte, date_str, row_idx in lignes_non_maj:
                self.logger.warning(f"    - Ligne {row_idx}: {ligne} ({compte}) - date: {date_str}")

    # ====================================================================
    # UNO (LibreOffice)
    # ====================================================================

    def read_cell_with_uno(self, sheet_name, cell_address):
        """Lit la valeur calculée d'une cellule via LibreOffice UNO."""
        from inc_uno import UnoDocument, check_lock_file, HAS_UNO

        if not HAS_UNO:
            self.logger.verbose("Module UNO non disponible")
            return None

        if check_lock_file(self.comptes_file):
            self.logger.error("Fichier verrouillé (LibreOffice ouvert ?)")
            return None

        try:
            with UnoDocument(self.comptes_file, read_only=False) as doc:
                doc.calculate_all()
                sheet = doc.get_sheet(sheet_name)
                col = ord(cell_address[0].upper()) - ord('A')
                row = int(cell_address[1:]) - 1
                value = sheet.getCellByPosition(col, row).getString()
                doc.save()  # save() écrit automatiquement le miroir C1
                return value
        except Exception as e:
            self.logger.error(f"Erreur lecture UNO: {e}")
            return None

    def check_control_sheet(self):
        """Vérifie l'onglet Contrôles et affiche le statut"""
        if not self.ws_controle:
            self.logger.verbose("Onglet Contrôles non disponible, vérification ignorée")
            return True

        try:
            cell_value = self.read_cell_with_uno(SHEET_CONTROLES, "A1")

            if cell_value is None:
                self.logger.warning("Contrôles!A1 non lisible (LibreOffice ouvert ? UNO indisponible ?)")
                self.logger.warning("Vérification des contrôles impossible — ouvrir le fichier dans Excel pour vérifier")
                return True  # non bloquant mais warning visible

            cell_a1_str = str(cell_value).strip()

            if cell_a1_str in ('.', '✓'):
                self.logger.verbose("Contrôles OK")
                return True

            # Détecter les erreurs connues (nouveau format ✓/✗/⚠ + ancien format rétrocompat)
            errors = []
            if "COMPTES" in cell_a1_str or (len(cell_a1_str) >= 1 and cell_a1_str[0] == '✗'):
                errors.append("COMPTES (écarts de solde)")
            if "CATÉGORIES" in cell_a1_str or (len(cell_a1_str) >= 2 and cell_a1_str[1] == '✗'):
                errors.append("CATÉGORIES (catégories inconnues)")
            if "INCONNUS" in cell_a1_str or (len(cell_a1_str) >= 6 and cell_a1_str[5] == '✗'):
                errors.append("INCONNUS (comptes absents d'Avoirs)")

            if errors:
                self.logger.error(f"Contrôles!A1 = '{cell_a1_str}'")
                for err in errors:
                    self.logger.error(f"  ERREUR {err}")
                self.logger.error("Utiliser --fallback pour annuler l'import")
                try:
                    tool_controles = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tool_controles.py")
                    subprocess.run([sys.executable, tool_controles, "-f", str(self.comptes_file)], check=False)
                except Exception as e:
                    self.logger.error(f"Impossible de lancer tool_controles: {e}")
                return False

            # Flags non critiques (appariements en cours, etc.)
            self.logger.warning(f"Contrôles!A1 = '{cell_a1_str}'")
            self.logger.warning("Vérifier dans Excel (appariements, soldes, etc.)")
            return True

        except Exception as e:
            self.logger.error(f"Erreur lecture onglet Contrôles: {e}")
            return False
