#!/usr/bin/env python3-uno
"""Migration idempotente : enveloppe la ligne « Changes Eq € » du bloc BALANCES
(feuille Contrôles) dans ROUND(...,0).

Contexte (#176) : le compteur BALANCES `L = 3-COUNTIFS(...,0)` teste l'égalité
EXACTE à zéro. Ses trois lignes sont « Virements € », « Titres € », « Changes
Eq € ». Les deux premières sont déjà `ROUND(SUMIFS(...),0)` ; seule « Changes
Eq € » ne l'était pas → un résidu flottant (ex. 4.08e-14, affiché « 0.00 »)
comptait comme anomalie → K = ⚠ → A1 = ⚠, sans aucun vrai déséquilibre. Enrober
de ROUND absorbe le bruit, comme ses sœurs.

Localisation — DOCTRINE NAMED RANGES (jamais de cellule/plage en dur, jamais
`max_row`/zone-utilisée) : le bloc BALANCES se déplace verticalement (le tableau
CTRL1 au-dessus grossit avec le nombre de comptes). La ligne est trouvée en
scannant le named range des LIBELLÉS `CTRL2type` (col J) pour « Changes Eq », et
la valeur est prise dans le named range `CTRL2general` (col L) à la MÊME ligne —
les deux NR se recalent seuls aux insertions (LO). La formule existante est
réutilisée VERBATIM et simplement enveloppée → aucune hypothèse de séparateur
(UNO rend `;` en locale FR, le stockage xlsx est en `,`).

Idempotent : ne fait rien si la formule est déjà enveloppée de ROUND.

Usage :
    python3 tool_migrate_ctrl_changes_round.py ~/Compta/comptes.xlsm [--dry-run]

--dry-run : sonde openpyxl SANS LibreOffice (rc 3 = changerait, 0 = déjà à jour).
"""
import argparse
import sys
from pathlib import Path

SHEET = 'Contrôles'
LABEL = 'Changes Eq'      # sous-chaîne du libellé, cherchée DANS le NR CTRL2type
LABEL_NR = 'CTRL2type'    # named range des libellés (col J) — BORNE le scan
VALUE_NR = 'CTRL2general' # named range des valeurs € (col L) — la cible


def _nr_bounds_openpyxl(wb, name):
    """(col_1indexed, start_row, end_row) d'un named range colonne, ou None.

    Résout depuis les `defined_names` openpyxl (chemin dry-run, sans LibreOffice)
    — pendant openpyxl de `inc_uno.get_col_range_bounds` (chemin UNO)."""
    import re
    from openpyxl.utils import column_index_from_string
    try:
        dn = wb.defined_names.get(name)
    except AttributeError:
        dn = None
    if dn is None:
        return None
    m = re.search(r"\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)", str(dn.value))
    if not m:
        return None
    return column_index_from_string(m.group(1)), int(m.group(2)), int(m.group(4))


def _already_round(xlsm_path):
    """Sonde openpyxl (SANS LibreOffice) : la formule € de « Changes Eq » est-elle
    déjà enveloppée de ROUND ? True = migration inutile (NR/libellé absent inclus)."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsm_path, data_only=False,
                               keep_vba=str(xlsm_path).endswith('.xlsm'))
    try:
        if SHEET not in wb.sheetnames:
            return True                          # pas de feuille Contrôles → rien à faire
        ws = wb[SHEET]
        lab = _nr_bounds_openpyxl(wb, LABEL_NR)
        val = _nr_bounds_openpyxl(wb, VALUE_NR)
        if not lab or not val:
            return True                          # NR absents → classeur trop ancien
        lcol, start, end = lab
        vcol = val[0]
        r = None
        for rr in range(start, end + 1):         # borné par CTRL2type, jamais max_row
            v = ws.cell(rr, lcol).value
            if v and LABEL in str(v):
                r = rr
                break
        if r is None:
            return True                          # libellé absent → rien à faire
        formula = ws.cell(r, vcol).value or ''
        return str(formula).upper().lstrip().startswith('=ROUND(')
    finally:
        wb.close()


def _apply(xlsm_path):
    """Enveloppe la formule € de « Changes Eq » dans ROUND(...,0) via UNO. Idempotent."""
    from inc_uno import UnoDocument, check_lock_file, get_col_range_bounds

    busy = check_lock_file(xlsm_path)
    if busy:
        print(f"✗ Classeur ouvert ({busy}) — ferme LibreOffice puis relance.", file=sys.stderr)
        return 1

    with UnoDocument(str(xlsm_path)) as doc:
        xdoc = doc.document
        ws = doc.get_sheet(SHEET)
        # Doctrine NR : borner le scan par CTRL2type (libellés), valeur via CTRL2general
        # (col L) à la même ligne. Patron identique à tool_controles.read_ctrl_tokens.
        lab_b = get_col_range_bounds(xdoc, LABEL_NR)
        val_b = get_col_range_bounds(xdoc, VALUE_NR)
        if not lab_b or not val_b:
            print(f"✗ Named ranges {LABEL_NR}/{VALUE_NR} absents ({SHEET}) — "
                  f"classeur trop ancien ?", file=sys.stderr)
            return 1
        lab_col, start, end = lab_b[1], lab_b[2], lab_b[3]   # col 0-indexée, lignes 1-indexées
        val_col = val_b[1]
        r0 = None
        for r1 in range(start, end + 1):
            if LABEL in ws.getCellByPosition(lab_col, r1 - 1).getString():
                r0 = r1 - 1
                break
        if r0 is None:
            print(f"✗ Libellé « {LABEL} » introuvable dans {LABEL_NR}.", file=sys.stderr)
            return 1

        vcell = ws.getCellByPosition(val_col, r0)
        formula = vcell.getFormula()             # locale UNO (séparateur `;`)
        if formula.upper().lstrip().startswith('=ROUND('):
            print("✓ « Changes Eq € » déjà enveloppée de ROUND — rien à faire.")
            return 0
        if not formula.startswith('='):
            print(f"✗ Cellule € de « {LABEL} » sans formule ({formula!r}).", file=sys.stderr)
            return 1

        body = formula[1:]                       # retire le `=`
        new_formula = f"=ROUND({body};0)"        # enveloppe verbatim, séparateur UNO
        vcell.setFormula(new_formula)
        col_letter = chr(ord('A') + val_col)
        doc.save()
        print(f"✓ « Changes Eq € » enveloppée de ROUND ({col_letter}{r0 + 1}).")

    # Hors du `with` (fichier fermé) : recadrer la vue salie par le save UNO
    # (patch ZIP-XML pur, idempotent). Cf. doctrine « --reframe avant livraison ».
    from tool_fix_formats import frame_views
    frame_views(str(xlsm_path), verbose=False)
    return 0


def main():
    ap = argparse.ArgumentParser(
        description="Enveloppe la ligne « Changes Eq € » (BALANCES) dans ROUND(...,0).")
    ap.add_argument('xlsm', help='chemin du classeur comptes.xlsm')
    ap.add_argument('--dry-run', action='store_true',
                    help='sonde openpyxl (rc 3 = changerait, 0 = déjà à jour), sans LibreOffice')
    args = ap.parse_args()

    p = Path(args.xlsm).expanduser()
    if not p.exists():
        print(f"✗ Introuvable : {p}", file=sys.stderr)
        return 1

    if args.dry_run:
        return 0 if _already_round(p) else 3

    return _apply(p)


if __name__ == '__main__':
    sys.exit(main())
