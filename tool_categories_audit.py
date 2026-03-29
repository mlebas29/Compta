#!/usr/bin/env python3
"""
Audit des catégorisations : compare les patterns category_mappings vs Excel.

Usage:
    ./cpt_categories_audit.py           # Audit complet
    ./cpt_categories_audit.py --summary # Résumé uniquement
    ./cpt_categories_audit.py -v        # Détails par pattern

Exit codes:
    0: Aucune divergence
    1: Divergences trouvées (informatif)
    2: Erreur technique
"""

import sys
import re
import argparse
from pathlib import Path

# Configuration chemin
import os
if os.environ.get('COMPTA_MODE') == 'prod':
    BASE_DIR = Path.home() / 'Compta'
else:
    BASE_DIR = Path.home() / 'Compta' / 'Claude'

sys.path.insert(0, str(BASE_DIR))

import json
import openpyxl
from inc_category_mappings import _load_patterns

# Patterns par site : chargés dynamiquement depuis config_category_mappings.json
SITE_PATTERNS = _load_patterns()

# Mapping Compte Excel → Site : dérivé de config_accounts.json
_ACCOUNTS_PATH = BASE_DIR / 'config_accounts.json'
if _ACCOUNTS_PATH.exists():
    with open(_ACCOUNTS_PATH, 'r', encoding='utf-8') as _f:
        _accounts_data = json.load(_f)
    COMPTE_TO_SITE = {
        acct['name']: site
        for site, site_data in _accounts_data.items()
        for acct in site_data.get('accounts', [])
    }
else:
    COMPTE_TO_SITE = {}


def get_site_from_compte(compte):
    """Détermine le site à partir du nom de compte Excel."""
    if not compte:
        return None

    # Recherche exacte
    if compte in COMPTE_TO_SITE:
        return COMPTE_TO_SITE[compte]

    # Recherche partielle (ordre important : plus spécifique d'abord)
    compte_lower = compte.lower()

    # Sites avec noms pouvant être confondus (tester en premier)
    if 'wise' in compte_lower:
        return 'WISE'
    if 'etoro' in compte_lower:
        return 'ETORO'
    if 'kraken' in compte_lower:
        return 'KRAKEN'
    if 'degiro' in compte_lower:
        return 'DEGIRO'

    # Crypto (wallets)
    if 'btc' in compte_lower or 'bitcoin' in compte_lower or 'lightning' in compte_lower:
        return 'BTC'
    if 'xmr' in compte_lower or 'monero' in compte_lower:
        return 'XMR'

    # Banques traditionnelles (après les autres pour éviter collision SG/SGD)
    if 'bourso' in compte_lower or compte_lower.endswith(' bb'):
        return 'BB'
    if 'société générale' in compte_lower or compte_lower.endswith(' sg'):
        return 'SG'
    if 'ass vie' in compte_lower or 'compte chèque sg' in compte_lower or 'compte livret sg' in compte_lower or 'compte titre sg' in compte_lower:
        return 'SG'

    # Autres
    if 'pee' in compte_lower or 'syngenta' in compte_lower:
        return 'PEE'
    if 'sci' in compte_lower or 'yvelles' in compte_lower:
        return 'BG_GESTION'

    return None


def find_matching_pattern(libelle, site):
    """Trouve le premier pattern qui matche le libellé."""
    if not libelle:
        return None, None, None

    # Patterns du site + génériques
    site_patterns = SITE_PATTERNS.get(site, [])
    all_patterns = site_patterns + SITE_PATTERNS.get('GENERIC', [])

    for pattern, category, opts in all_patterns:
        if re.search(pattern, str(libelle)):
            return pattern, category, opts

    return None, None, None


def audit_categories(xlsx_path, verbose=False, last_lines=None):
    """Audit les catégorisations Excel vs patterns.

    Args:
        xlsx_path: Chemin vers comptes.xlsm
        verbose: Afficher les exemples
        last_lines: Si défini, auditer uniquement les N dernières lignes
    """

    # Note: read_only=False requis car read_only=True retourne max_row=1048576
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Opérations"]

    # Collecter les lignes avec données
    all_rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
    total_rows = len(all_rows)

    # Filtrer si last_lines
    if last_lines and last_lines > 0:
        start_idx = max(0, total_rows - last_lines)
        rows_with_nums = [(i + 2 + start_idx, all_rows[start_idx + i]) for i in range(min(last_lines, total_rows - start_idx))]
    else:
        rows_with_nums = [(i + 2, row) for i, row in enumerate(all_rows)]

    divergences = {}  # (pattern, cat_attendue) -> [(row_num, libelle, cat_excel, compte)]
    stats = {
        'total': 0,
        'matched': 0,
        'divergent': 0,
        'no_pattern': 0,
        'no_category': 0,
    }

    for row_num, row in rows_with_nums:
        if not row[0]:
            continue

        date, libelle, montant, devise, equiv, ref, categorie, compte, commentaire = row[:9]

        if not libelle:
            continue

        stats['total'] += 1

        # Ignorer les méta-opérations (#Solde, #Info, etc.)
        if categorie and str(categorie).startswith('#'):
            continue
        if libelle and str(libelle).startswith('#'):
            continue

        if not categorie:
            stats['no_category'] += 1
            continue

        # Déterminer le site
        site = get_site_from_compte(compte)

        # Chercher le pattern
        pattern, cat_attendue, opts = find_matching_pattern(libelle, site)

        if not pattern:
            stats['no_pattern'] += 1
            continue

        stats['matched'] += 1

        # Comparer
        if categorie != cat_attendue:
            stats['divergent'] += 1
            key = (pattern, cat_attendue)
            if key not in divergences:
                divergences[key] = []
            divergences[key].append((row_num, str(libelle)[:50], categorie, compte))

    wb.close()

    return divergences, stats


def print_report(divergences, stats, verbose=False, summary_only=False, last_lines=None):
    """Affiche le rapport d'audit."""

    total_div = sum(len(v) for v in divergences.values())

    # Résumé
    scope = f" (dernières {last_lines} lignes)" if last_lines else ""
    print(f"📊 Audit catégorisations{scope}")
    print(f"   Total opérations: {stats['total']}")
    print(f"   Avec pattern: {stats['matched']}")
    print(f"   Sans pattern: {stats['no_pattern']}")
    print(f"   Sans catégorie: {stats['no_category']}")
    print()

    if total_div == 0:
        print("✓ Aucune divergence")
        return 0

    print(f"⚠ {total_div} divergence(s) pour {len(divergences)} pattern(s)")

    if summary_only:
        return 1

    print()

    # Détails par pattern
    for (pattern, cat_attendue), cases in sorted(divergences.items(), key=lambda x: -len(x[1])):
        print(f"📌 {pattern} → {cat_attendue}")

        # Grouper par catégorie Excel avec numéros de lignes
        cats = {}  # cat -> [row_nums]
        for row_num, lib, cat, compte in cases:
            if cat not in cats:
                cats[cat] = []
            cats[cat].append(row_num)

        for cat, row_nums in sorted(cats.items(), key=lambda x: -len(x[1])):
            # Afficher les numéros de lignes (max 10)
            nums_str = ', '.join(str(n) for n in row_nums[:10])
            if len(row_nums) > 10:
                nums_str += f'... (+{len(row_nums)-10})'
            print(f"   • '{cat}': L{nums_str}")

        if verbose and cases:
            row_num, lib, cat, compte = cases[0]
            print(f"   Ex L{row_num}: '{lib}...' ({compte})")

        print()

    return 1


def main():
    parser = argparse.ArgumentParser(description='Audit catégorisations vs patterns')
    parser.add_argument('--summary', action='store_true', help='Résumé uniquement')
    parser.add_argument('--lines', type=int, metavar='N', help='Auditer les N dernières lignes uniquement')
    parser.add_argument('-v', '--verbose', action='store_true', help='Afficher exemples')
    args = parser.parse_args()

    xlsx_path = BASE_DIR / 'comptes.xlsm'

    if not xlsx_path.exists():
        print(f"❌ Fichier non trouvé: {xlsx_path}", file=sys.stderr)
        sys.exit(2)

    try:
        divergences, stats = audit_categories(xlsx_path, args.verbose, args.lines)
        exit_code = print_report(divergences, stats, args.verbose, args.summary, args.lines)
        sys.exit(exit_code)
    except Exception as e:
        print(f"❌ Erreur: {e}", file=sys.stderr)
        sys.exit(2)


if __name__ == '__main__':
    main()
