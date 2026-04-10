"""Mixin Comptes pour ConfigGUI."""

from datetime import datetime
from tkinter import messagebox
from tkinter import ttk
import openpyxl
import re
import shutil
import subprocess
import sys
import threading
import time
import tkinter as tk
import unicodedata

from inc_excel_schema import (
    ColResolver,
    SHEET_AVOIRS, SHEET_CONTROLES, SHEET_OPERATIONS, SHEET_PLUS_VALUE,
)


class AccountsMixin:
    """Onglet Comptes (Avoirs, titres, UNO ops comptes)."""

    def _build_tab_accounts(self, tab=None):
        if tab is None:
            tab = ttk.Frame(self.notebook)
            self.notebook.add(tab, text='Comptes')

        # --- Treeview groupé par Site ---
        tree_frame = ttk.LabelFrame(tab, text='Comptes et biens matériels', padding=5)
        tree_frame.pack(fill='both', expand=True, padx=5, pady=(5, 2))
        self._make_help_button(tree_frame)

        cols = ('devise', 'type')
        self.acct_tree = ttk.Treeview(tree_frame, columns=cols,
                                       show='tree headings', selectmode='browse')
        self.acct_tree.heading('#0', text='Site / Intitulé')
        self.acct_tree.column('#0', width=350, minwidth=200)
        self.acct_tree.heading('devise', text='Devise')
        self.acct_tree.column('devise', width=60, minwidth=40)
        self.acct_tree.heading('type', text='Type')
        self.acct_tree.column('type', width=120, minwidth=80)

        vsb = ttk.Scrollbar(tree_frame, orient='vertical',
                            command=self.acct_tree.yview)
        self.acct_tree.configure(yscrollcommand=vsb.set)
        self.acct_tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

        self._populate_accounts_tree()

        # --- Boutons ---
        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill='x', padx=5, pady=(2, 5))
        ttk.Button(btn_frame, text='\u2795 Ajouter compte',
                   command=self._acct_add).pack(side='left', padx=2)
        ttk.Button(btn_frame, text='\u2795 Ajouter bien',
                   command=self._bien_add).pack(side='left', padx=2)
        ttk.Separator(btn_frame, orient='vertical').pack(
            side='left', fill='y', padx=8)
        ttk.Button(btn_frame, text='\u2795 Nouvelle devise',
                   command=self._devise_add).pack(side='left', padx=2)
        ttk.Button(btn_frame, text='\u2716 Supprimer devise',
                   command=self._devise_delete_dialog).pack(side='left', padx=2)

        # --- Menu contextuel (clic droit) ---
        self._acct_context_menu = tk.Menu(self.acct_tree, tearoff=0)
        self._acct_context_menu.add_command(
            label='\u270f Modifier', command=self._acct_edit)
        self._acct_context_menu.add_command(
            label='\u2716 Supprimer', command=self._acct_delete)
        self._acct_context_menu.add_command(
            label='\u2795 Ajout titre', command=self._acct_add_title)
        self._acct_context_menu.add_separator()
        self._acct_context_menu.add_command(
            label='\u2672 Purger opérations', command=self._acct_purge)

        # Menu contextuel pour les titres (3e niveau)
        self._title_context_menu = tk.Menu(self.acct_tree, tearoff=0)
        self._title_context_menu.add_command(
            label='\u270f Renommer titre', command=self._pv_title_rename)
        self._title_context_menu.add_command(
            label='\u2716 Supprimer titre', command=self._pv_title_delete)

        self.acct_tree.bind('<Button-3>', self._acct_show_context_menu)
        self.acct_tree.bind('<Double-1>', lambda e: self._acct_edit())

    def _acct_show_context_menu(self, event):
        """Affiche le menu contextuel sur clic droit (compte ou titre)."""
        item = self.acct_tree.identify_row(event.y)
        if not item or not self.acct_tree.parent(item):
            return  # groupe Site ou vide → pas de menu
        self.acct_tree.selection_set(item)
        if self._is_title_node(item):
            # Titre (profondeur 2)
            self._title_context_menu.tk_popup(event.x_root, event.y_root)
        else:
            # Compte (profondeur 1)
            entry = self._find_display_by_tree_selection()
            can_add_title = (entry and entry.get('type') == 'Portefeuilles'
                            and 'Réserve' not in entry['intitule'])
            self._acct_context_menu.entryconfigure(2, state='normal' if can_add_title else 'disabled')
            self._acct_context_menu.tk_popup(event.x_root, event.y_root)

    def _populate_accounts_tree(self):
        """Remplit le Treeview avec les comptes groupés par Site."""
        self.acct_tree.delete(*self.acct_tree.get_children())
        self._acct_group_nodes = {}

        # Grouper les display_accounts par site
        groups = {}
        for entry in self.display_accounts:
            site = entry.get('site') or 'N/A'
            groups.setdefault(site, []).append(entry)

        # Ordre : sites activés (config.ini) puis les autres
        enabled_str = self.config.get('sites', 'enabled', fallback='')
        enabled_list = [s.strip() for s in enabled_str.split(',') if s.strip()]
        ordered = list(enabled_list)
        for g in groups:
            if g not in ordered:
                ordered.append(g)

        for site in ordered:
            if site not in groups:
                continue
            entries = groups[site]
            site_name = self.config.get(site, 'name', fallback=site)
            parent_id = self.acct_tree.insert(
                '', 'end', text=site_name,
                open=True, values=('', ''))
            self._acct_group_nodes[site] = parent_id
            for entry in entries:
                acct_id = self.acct_tree.insert(
                    parent_id, 'end', text=entry['intitule'],
                    values=(entry['devise'], entry.get('type', '')))
                # Titres portefeuille (3e niveau)
                if entry.get('type') == 'Portefeuilles' and 'Réserve' not in entry['intitule']:
                    titles = getattr(self, 'pv_titles', {}).get(entry['intitule'], [])
                    if titles:
                        self.acct_tree.item(acct_id, open=True)
                        for title, dev, pv_row in titles:
                            self.acct_tree.insert(
                                acct_id, 'end', text=title,
                                values=(dev, ''))

    def _is_title_node(self, item_id):
        """Vérifie si un noeud Treeview est un titre (profondeur 2 = a un grand-parent)."""
        parent = self.acct_tree.parent(item_id)
        if not parent:
            return False
        return bool(self.acct_tree.parent(parent))

    def _find_display_by_tree_selection(self):
        """Retourne le display_account dict correspondant à la sélection Treeview.

        Si un titre est sélectionné, retourne le compte parent.
        """
        sel = self.acct_tree.selection()
        if not sel:
            return None
        item_id = sel[0]
        # Si c'est un titre, remonter au compte parent
        if self._is_title_node(item_id):
            item_id = self.acct_tree.parent(item_id)
        item = self.acct_tree.item(item_id)
        intitule = item['text']
        if not intitule:
            return None
        for entry in self.display_accounts:
            if entry['intitule'] == intitule:
                return entry
        return None

    def _select_display_in_tree(self, intitule):
        """Sélectionne une entrée dans le Treeview par intitulé."""
        for parent_id in self.acct_tree.get_children():
            for child_id in self.acct_tree.get_children(parent_id):
                item = self.acct_tree.item(child_id)
                if item['text'] == intitule:
                    self.acct_tree.selection_set(child_id)
                    self.acct_tree.see(child_id)
                    return

    def _after_accounts_save(self, message):
        """Callback après sauvegarde UNO comptes : recharge + site_map + message."""
        self._load_accounts_data()
        self._load_pv_titles()
        self._populate_accounts_tree()
        self._save_site_map()
        swept = self._last_sweep_count
        if swept:
            message += f", {swept} ligne(s) balayée(s) dans « Compte clos »"
        self._set_status(message)

    # --- Champs conditionnels par site et gardes comptes ---

    # Limites max comptes par site
    SITE_MAX_ACCOUNTS = {
        'BOURSOBANK': 4, 'NATIXIS': 1, 'PAYPAL': 1, 'AMAZON': 1, 'ORCHESTRA': 1,
    }

    # Champs conditionnels par site : (label, key, widget, values)
    @staticmethod
    def _site_account_fields(site, type_sg=None):
        """Retourne les champs conditionnels pour un site donné."""
        if site == 'SOCGEN':
            fields = [('Type SG :', 'type_sg', 'combo',
                        ['principal', 'epargne', 'assurance_vie'])]
            if type_sg == 'principal':
                fields.append(('Numéro :', 'numero', 'entry', None))
            elif type_sg == 'epargne':
                fields.append(('Numéro :', 'numero', 'entry', None))
                fields.append(('ID technique :', 'id_technique', 'entry', None))
            elif type_sg == 'assurance_vie':
                fields.append(('ID technique :', 'id_technique', 'entry', None))
                fields.append(('Clé fichiers :', 'file_key', 'entry', None))
            return fields
        elif site == 'BOURSOBANK':
            return [('Numéro :', 'numero', 'entry', None)]
        elif site == 'BTC':
            return [
                ('Clé wallet :', 'wallet_key', 'entry', None),
                ('Adresses :', 'addresses', 'entry', None),
            ]
        elif site == 'XMR':
            return [
                ('Clé :', 'wallet_key', 'entry', None),
                ('Portefeuille :', 'wallet_name', 'entry', None),
            ]
        return []

    def _count_site_accounts(self, site, type_sg=None):
        """Compte les comptes existants d'un site (optionnel: par type_sg)."""
        count = 0
        for acct in self.accounts_data:
            if acct.get('site') == site:
                if type_sg is None:
                    count += 1
                else:
                    # Chercher type_sg dans le JSON
                    json_acct = self._find_json_account(acct['intitule'], site)
                    if json_acct and json_acct.get('type_sg') == type_sg:
                        count += 1
        return count

    def _find_json_account(self, name, site):
        """Trouve l'entrée JSON d'un compte par nom et site."""
        site_data = self.accounts_json_data.get(site, {})
        for acct in site_data.get('accounts', []):
            if acct.get('name') == name:
                return acct
        return None

    def _validate_site_fields(self, site, extra_fields, dlg, intitule=None):
        """Valide les champs conditionnels. Retourne dict ou None si erreur."""
        values = {}
        for key, var in extra_fields.items():
            val = var.get().strip()
            values[key] = val

        # Gardes SOCGEN
        if site == 'SOCGEN':
            type_sg = values.get('type_sg', '')
            if not type_sg:
                messagebox.showwarning('Champ requis', 'Type SG obligatoire.',
                                       parent=dlg)
                return None
            if type_sg == 'principal':
                if not values.get('numero'):
                    messagebox.showwarning('Champ requis',
                        'Numéro obligatoire pour un compte principal.', parent=dlg)
                    return None
                if self._count_site_accounts('SOCGEN', 'principal') > 0:
                    messagebox.showwarning('Limite',
                        'Un seul compte principal SG autorisé.', parent=dlg)
                    return None
            elif type_sg == 'epargne':
                if not values.get('numero') or not values.get('id_technique'):
                    messagebox.showwarning('Champ requis',
                        'Numéro et ID technique obligatoires pour un compte épargne.',
                        parent=dlg)
                    return None
            elif type_sg == 'assurance_vie':
                if not values.get('id_technique'):
                    messagebox.showwarning('Champ requis',
                        'ID technique obligatoire pour une assurance vie.', parent=dlg)
                    return None
                # Auto-générer file_key si vide
                if not values.get('file_key') and intitule:
                    import unicodedata
                    fk = unicodedata.normalize('NFD', intitule)
                    fk = ''.join(c for c in fk if unicodedata.category(c) != 'Mn')
                    fk = fk.replace(' ', '_')
                    values['file_key'] = fk
                # Unicité file_key
                fk = values.get('file_key', '')
                if fk:
                    for a in self.accounts_json_data.get(site, {}).get('accounts', []):
                        if a.get('file_key', '').lower() == fk.lower():
                            messagebox.showwarning('Doublon',
                                f"La clé fichiers '{fk}' est déjà utilisée.", parent=dlg)
                            return None

        # Gardes BTC/XMR : unicité wallet_key
        if site in ('BTC', 'XMR') and values.get('wallet_key'):
            wk = values['wallet_key']
            for a in self.accounts_json_data.get(site, {}).get('accounts', []):
                if a.get('wallet_key') == wk:
                    messagebox.showwarning('Doublon',
                        f"La clé wallet '{wk}' est déjà utilisée pour {site}.",
                        parent=dlg)
                    return None

        # Garde XMR : wallet_name obligatoire
        if site == 'XMR' and not values.get('wallet_name'):
            messagebox.showwarning('Champ requis',
                'Le nom de portefeuille MoneroGUI est obligatoire.', parent=dlg)
            return None

        # Convertir addresses en liste
        if 'addresses' in values and values['addresses']:
            values['addresses'] = [a.strip() for a in values['addresses'].split(',')
                                   if a.strip()]

        return values

    def _inject_json_fields(self, name, site, extra_values):
        """Injecte les champs techniques dans accounts_json_data pour un nouveau compte."""
        if site not in self.accounts_json_data:
            self.accounts_json_data[site] = {'accounts': []}
        site_data = self.accounts_json_data[site]
        # Trouver ou créer l'entrée
        json_acct = None
        for a in site_data['accounts']:
            if a.get('name') == name:
                json_acct = a
                break
        if json_acct is None:
            json_acct = {'name': name}
            site_data['accounts'].append(json_acct)
        # Injecter les champs non vides
        for key, val in extra_values.items():
            if val:
                json_acct[key] = val

    def _acct_add(self):
        """Dialog pour ajouter un nouveau compte."""
        dlg = tk.Toplevel(self.root)
        dlg.title('Ajouter un compte')
        dlg.resizable(True, False)
        dlg.transient(self.root)
        dlg.grab_set()

        site_values = ['N/A'] + [s for s in self.all_sites if s != 'MANUEL']

        fields = {}
        row = 0
        for label, key, widget_type, values, default in [
            ('Site :', 'site', 'combo', site_values, 'N/A'),
            ('Intitulé :', 'intitule', 'entry', None, ''),
            ('Devise :', 'devise', 'combo', self.ACCOUNT_DEVISES, 'EUR'),
            ('Type :', 'type', 'combo', self.ACCOUNT_TYPES, 'Euros'),
            ('Sous-type :', 'sous_type', 'combo', self.ACCOUNT_SOUS_TYPES, 'Euro'),
            ('Domiciliation :', 'domiciliation', 'entry', None, ''),
            ('Titulaire :', 'titulaire', 'entry', None, ''),
            ('Propriété :', 'propriete', 'entry', None, ''),
        ]:
            ttk.Label(dlg, text=label).grid(row=row, column=0,
                                             sticky='w', padx=10, pady=3)
            var = tk.StringVar(value=default)
            if widget_type == 'combo':
                w = ttk.Combobox(dlg, textvariable=var, values=values, width=25)
            else:
                w = ttk.Entry(dlg, textvariable=var, width=28)
            w.grid(row=row, column=1, padx=10, pady=3, sticky='w')
            fields[key] = var
            row += 1

        # Checkbox Contrôle de solde
        ctrl_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(dlg, text='Contrôle de solde',
                        variable=ctrl_var).grid(
            row=row, column=0, columnspan=2, sticky='w', padx=10, pady=3)
        row += 1

        # Frame dynamique pour champs conditionnels
        extra_frame = ttk.Frame(dlg)
        extra_frame.grid(row=row, column=0, columnspan=2, sticky='ew')
        row += 1
        extra_fields = {}
        extra_widgets = []

        # Hint file_key (affiché quand file_key a une valeur)
        hint_label = None

        _current_type_sg = tk.StringVar()

        def update_extra_fields(*_args):
            nonlocal hint_label
            # Sauvegarder type_sg avant nettoyage
            if 'type_sg' in extra_fields:
                _current_type_sg.set(extra_fields['type_sg'].get())

            # Nettoyer
            for w in extra_widgets:
                w.destroy()
            extra_widgets.clear()
            extra_fields.clear()

            site = fields['site'].get().strip()
            type_sg = _current_type_sg.get() or None
            field_defs = self._site_account_fields(site, type_sg)

            for i, (label, key, wtype, values) in enumerate(field_defs):
                lbl = ttk.Label(extra_frame, text=label)
                lbl.grid(row=i, column=0, sticky='w', padx=10, pady=2)
                extra_widgets.append(lbl)
                var = tk.StringVar()
                if key == 'type_sg' and _current_type_sg.get():
                    var.set(_current_type_sg.get())
                if wtype == 'combo':
                    w = ttk.Combobox(extra_frame, textvariable=var, values=values, width=25)
                    if key == 'type_sg':
                        w.bind('<<ComboboxSelected>>', update_extra_fields)
                else:
                    w = ttk.Entry(extra_frame, textvariable=var, width=28)
                w.grid(row=i, column=1, padx=10, pady=2, sticky='w')
                extra_widgets.append(w)
                extra_fields[key] = var

                # Hint pour wallet_key
                if key == 'wallet_key' and site in ('BTC', 'XMR'):
                    prefix = 'xmr' if site == 'XMR' else 'btc'
                    def make_hint(v=var, p=prefix):
                        val = v.get().strip()
                        return f'(→ {p}_{val}_operations.csv)' if val else ''
                    hint_label = ttk.Label(extra_frame, text='', style='Hint.TLabel')
                    hint_label.grid(row=i, column=2, padx=2, pady=2, sticky='w')
                    extra_widgets.append(hint_label)
                    var.trace_add('write', lambda *_, h=hint_label, mk=make_hint: h.configure(text=mk()))

            # Ajuster la taille du dialogue
            dlg.update_idletasks()

        # Lier le changement de site
        fields['site'].trace_add('write', update_extra_fields)

        # Solde initial (optionnel)
        ttk.Label(dlg, text='Solde initial :').grid(row=row, column=0,
                                                     sticky='w', padx=10, pady=3)
        solde_var = tk.StringVar()
        ttk.Entry(dlg, textvariable=solde_var, width=28).grid(
            row=row, column=1, padx=10, pady=3, sticky='w')
        row += 1

        row += 1

        def on_ok():
            intitule = fields['intitule'].get().strip()
            if not intitule:
                messagebox.showwarning('Champ requis', 'Intitulé obligatoire.',
                                       parent=dlg)
                return
            devise = fields['devise'].get().strip()
            if not devise:
                messagebox.showwarning('Champ requis', 'Devise obligatoire.',
                                       parent=dlg)
                return
            for a in self.accounts_data:
                if a['intitule'] == intitule:
                    messagebox.showwarning('Doublon',
                                           f"Le compte '{intitule}' existe déjà.",
                                           parent=dlg)
                    return

            site = fields['site'].get().strip()
            if not site or site == 'N/A':
                site = 'N/A'

            # Garde max comptes par site
            if site in self.SITE_MAX_ACCOUNTS:
                current = self._count_site_accounts(site)
                if current >= self.SITE_MAX_ACCOUNTS[site]:
                    messagebox.showwarning('Limite',
                        f'{site} : maximum {self.SITE_MAX_ACCOUNTS[site]} compte(s).',
                        parent=dlg)
                    return

            # Valider champs conditionnels
            extra_values = None
            if extra_fields:
                extra_values = self._validate_site_fields(
                    site, extra_fields, dlg, intitule)
                if extra_values is None:
                    return

            # Parser le solde initial
            solde_init = None
            solde_str = solde_var.get().strip()
            if solde_str:
                try:
                    solde_init = float(solde_str.replace(' ', '').replace(',', '.'))
                except ValueError:
                    messagebox.showwarning('Solde invalide',
                                           f"'{solde_str}' n'est pas un nombre valide.",
                                           parent=dlg)
                    return

            acct_type = fields['type'].get().strip() or 'Euros'
            new_acct = {
                'row': None,
                '_is_new': True,
                'intitule': intitule,
                'type': acct_type,
                'sous_type': fields['sous_type'].get().strip() or 'Euro',
                'domiciliation': fields['domiciliation'].get().strip(),
                'devise': devise,
                'titulaire': fields['titulaire'].get().strip(),
                'propriete': fields['propriete'].get().strip(),
                'date_anter': None,
                'montant_anter': None,
                'montant_debut': solde_init,
                'site': site,
            }
            self.accounts_data.append(new_acct)
            self.account_site_map[intitule] = site

            # Injecter les champs techniques dans le JSON
            if extra_values:
                self._inject_json_fields(intitule, site, extra_values)

            new_display = {
                'ctrl_row': None,
                'intitule': intitule,
                'devise': devise,
                'controle': ctrl_var.get(),
                'site': site,
                'type': new_acct['type'],
                'avoirs_ref': new_acct,
            }
            self.display_accounts.append(new_display)
            dlg.destroy()

            self._run_uno_operation(
                'Enregistrement compte',
                self._save_accounts,
                lambda: (self._after_accounts_save(f"Compte '{intitule}' ajouté"),
                         self._select_display_in_tree(intitule))
            )

        btn_frame = ttk.Frame(dlg)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text='OK', command=on_ok).pack(side='left', padx=5)
        ttk.Button(btn_frame, text='Annuler',
                   command=dlg.destroy).pack(side='left', padx=5)

    def _bien_add(self):
        """Dialog pour ajouter un bien matériel (foncier ou mobilier)."""
        dlg = tk.Toplevel(self.root)
        dlg.title('Ajouter un bien matériel')
        dlg.resizable(True, False)
        dlg.transient(self.root)
        dlg.grab_set()

        # Domiciliations suggérées (liste ouverte)
        doms = ['Maison', 'Voiture']

        fields = {}
        row = 0
        for label, key, widget_type, values, default in [
            ('Intitulé :', 'intitule', 'entry', None, ''),
            ('Nature :', 'sous_type', 'combo', ['Foncier', 'Mobilier'], 'Foncier'),
            ('Devise :', 'devise', 'combo', self.ACCOUNT_DEVISES, ''),
            ('Domiciliation :', 'domiciliation', 'combo', doms, ''),
            ('Titulaire :', 'titulaire', 'entry', None, ''),
            ('Propriété :', 'propriete', 'entry', None, ''),
            ('Montant :', 'montant', 'entry', None, ''),
        ]:
            ttk.Label(dlg, text=label).grid(row=row, column=0,
                                             sticky='w', padx=10, pady=3)
            var = tk.StringVar(value=default)
            if widget_type == 'combo':
                w = ttk.Combobox(dlg, textvariable=var, values=values, width=25)
            else:
                w = ttk.Entry(dlg, textvariable=var, width=28)
            w.grid(row=row, column=1, padx=10, pady=3, sticky='w')
            fields[key] = var
            row += 1

        # Hint devise
        ttk.Label(dlg, text='Devise vide pour immobilier/véhicules, devise cotée pour métaux',
                  style='Hint.TLabel').grid(
            row=row, column=0, columnspan=2, sticky='w', padx=10)
        row += 1

        def on_ok():
            intitule = fields['intitule'].get().strip()
            if not intitule:
                messagebox.showwarning('Champ requis', 'Intitulé obligatoire.',
                                       parent=dlg)
                return
            for a in self.accounts_data:
                if a['intitule'] == intitule:
                    messagebox.showwarning('Doublon',
                                           f"Le bien '{intitule}' existe déjà.",
                                           parent=dlg)
                    return

            devise = fields['devise'].get().strip()
            sous_type = fields['sous_type'].get().strip()
            domiciliation = fields['domiciliation'].get().strip()
            titulaire = fields['titulaire'].get().strip()
            propriete = fields['propriete'].get().strip()
            montant_str = fields['montant'].get().strip()

            # Parser le montant
            montant = None
            if montant_str:
                try:
                    montant = float(montant_str.replace(' ', '').replace(',', '.'))
                except ValueError:
                    messagebox.showwarning('Montant invalide',
                                           f"'{montant_str}' n'est pas un nombre valide.",
                                           parent=dlg)
                    return

            dlg.destroy()

            # Ajouter via add_account (controle=False, pas de #Solde)
            new_acct = {
                'row': None,
                '_is_new': True,
                'intitule': intitule,
                'type': 'Biens matériels',
                'sous_type': sous_type,
                'domiciliation': domiciliation,
                'devise': devise,
                'titulaire': titulaire,
                'propriete': propriete,
                'date_anter': None,
                'montant_anter': None,
                'date_debut': None,
                'montant_debut': montant,
                'date_solde': None,
                'site': 'N/A',
            }
            self.accounts_data.append(new_acct)
            self.display_accounts.append({
                'intitule': intitule,
                'devise': devise,
                'controle': False,
                'ctrl_row': None,
                'avoirs_ref': new_acct,
            })
            dlg.destroy()

            self._run_uno_operation(
                'Enregistrement bien',
                self._save_accounts,
                lambda: (self._after_accounts_save(f"Bien ajouté : {intitule}"),
                         self._select_display_in_tree(intitule))
            )

        btn_frame = ttk.Frame(dlg)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text='OK', command=on_ok).pack(side='left', padx=5)
        ttk.Button(btn_frame, text='Annuler',
                   command=dlg.destroy).pack(side='left', padx=5)

    def _acct_edit(self):
        """Dialog pour modifier le compte sélectionné."""
        entry = self._find_display_by_tree_selection()
        if not entry:
            messagebox.showwarning('Sélection', 'Aucun compte sélectionné.',
                                   parent=self.root)
            return

        avoirs_acct = entry['avoirs_ref']
        old_intitule = entry['intitule']
        old_site = avoirs_acct.get('site', '') or entry.get('site', '') or 'N/A'
        json_acct = self._find_json_account(old_intitule, old_site) or {}

        dlg = tk.Toplevel(self.root)
        dlg.title('Modifier un compte')
        dlg.resizable(True, False)
        dlg.transient(self.root)
        dlg.grab_set()

        site_values = ['N/A'] + [s for s in self.all_sites if s != 'MANUEL']

        fields = {}
        row = 0
        for label, key, widget_type, values in [
            ('Site :', 'site', 'combo', site_values),
            ('Intitulé :', 'intitule', 'entry', None),
            ('Devise :', 'devise', 'combo', self.ACCOUNT_DEVISES),
            ('Type :', 'type', 'combo', self.ACCOUNT_TYPES),
            ('Sous-type :', 'sous_type', 'combo', self.ACCOUNT_SOUS_TYPES),
            ('Domiciliation :', 'domiciliation', 'entry', None),
            ('Titulaire :', 'titulaire', 'entry', None),
            ('Propriété :', 'propriete', 'entry', None),
        ]:
            ttk.Label(dlg, text=label).grid(row=row, column=0,
                                             sticky='w', padx=10, pady=3)
            # Pré-remplir avec la valeur existante
            if key == 'site':
                current_val = old_site
            else:
                current_val = avoirs_acct.get(key, '')
            var = tk.StringVar(value=current_val)
            # Devise et Type en lecture seule
            readonly = key in ('devise', 'type')
            if widget_type == 'combo':
                state = 'disabled' if readonly else 'normal'
                w = ttk.Combobox(dlg, textvariable=var, values=values,
                                 width=25, state=state)
            else:
                state = 'readonly' if readonly else 'normal'
                w = ttk.Entry(dlg, textvariable=var, width=28, state=state)
            w.grid(row=row, column=1, padx=10, pady=3, sticky='w')
            fields[key] = var
            row += 1

        # Checkbox Contrôle de solde
        ctrl_var = tk.BooleanVar(value=entry['controle'])
        ttk.Checkbutton(dlg, text='Contrôle de solde',
                        variable=ctrl_var).grid(
            row=row, column=0, columnspan=2, sticky='w', padx=10, pady=3)
        row += 1

        # Frame dynamique pour champs conditionnels
        extra_frame = ttk.Frame(dlg)
        extra_frame.grid(row=row, column=0, columnspan=2, sticky='ew')
        row += 1
        extra_fields = {}
        extra_widgets = []

        # Champs figés en édition
        frozen_keys = {'type_sg', 'file_key'}

        _current_type_sg = tk.StringVar(value=json_acct.get('type_sg', ''))

        def build_extra_fields():
            # Sauvegarder type_sg avant nettoyage
            if 'type_sg' in extra_fields:
                _current_type_sg.set(extra_fields['type_sg'].get())

            for w in extra_widgets:
                w.destroy()
            extra_widgets.clear()
            extra_fields.clear()

            site = fields['site'].get().strip()
            type_sg = _current_type_sg.get() or None
            field_defs = self._site_account_fields(site, type_sg)

            for i, (label, key, wtype, values) in enumerate(field_defs):
                lbl = ttk.Label(extra_frame, text=label)
                lbl.grid(row=i, column=0, sticky='w', padx=10, pady=2)
                extra_widgets.append(lbl)
                # Pré-remplir : d'abord le combobox courant, sinon le JSON
                if key == 'type_sg' and _current_type_sg.get():
                    current = _current_type_sg.get()
                else:
                    current = json_acct.get(key, '')
                if key == 'addresses' and isinstance(current, list):
                    current = ', '.join(current)
                var = tk.StringVar(value=current)
                is_frozen = key in frozen_keys and json_acct.get(key, '')
                if wtype == 'combo':
                    state = 'disabled' if is_frozen else 'normal'
                    w = ttk.Combobox(extra_frame, textvariable=var, values=values,
                                     width=25, state=state)
                    if key == 'type_sg':
                        w.bind('<<ComboboxSelected>>', lambda *_: build_extra_fields())
                else:
                    state = 'readonly' if is_frozen else 'normal'
                    w = ttk.Entry(extra_frame, textvariable=var, width=28, state=state)
                w.grid(row=i, column=1, padx=10, pady=2, sticky='w')
                extra_widgets.append(w)
                extra_fields[key] = var

                # Hint pour wallet_key
                if key == 'wallet_key' and site in ('BTC', 'XMR'):
                    prefix = 'xmr' if site == 'XMR' else 'btc'
                    def make_hint(v=var, p=prefix):
                        val = v.get().strip()
                        return f'(→ {p}_{val}_operations.csv)' if val else ''
                    hint = ttk.Label(extra_frame, text=make_hint(), style='Hint.TLabel')
                    hint.grid(row=i, column=2, padx=2, pady=2, sticky='w')
                    extra_widgets.append(hint)
                    var.trace_add('write', lambda *_, h=hint, mk=make_hint: h.configure(text=mk()))

            dlg.update_idletasks()

        build_extra_fields()
        fields['site'].trace_add('write', lambda *_: build_extra_fields())
        row += 1

        def on_ok():
            new_intitule = fields['intitule'].get().strip()
            if not new_intitule:
                messagebox.showwarning('Erreur', 'Le nom du compte ne peut pas être vide.',
                                       parent=dlg)
                return
            # Vérifier doublon si renommage
            if new_intitule != old_intitule:
                existing = {e['intitule'] for e in self.display_accounts}
                if new_intitule in existing:
                    messagebox.showwarning('Doublon',
                        f'Le compte "{new_intitule}" existe déjà.',
                        parent=dlg)
                    return

            site = fields['site'].get().strip()
            if not site or site == 'N/A':
                site = 'N/A'

            # Valider et sauvegarder champs conditionnels
            if extra_fields:
                extra_values = self._validate_site_fields(
                    site, extra_fields, dlg, new_intitule)
                if extra_values is None:
                    return
                self._inject_json_fields(old_intitule, old_site, extra_values)

            # Mettre à jour avoirs_ref
            avoirs_acct['site'] = site
            avoirs_acct['sous_type'] = fields['sous_type'].get().strip()
            avoirs_acct['domiciliation'] = fields['domiciliation'].get().strip()
            avoirs_acct['titulaire'] = fields['titulaire'].get().strip()
            avoirs_acct['propriete'] = fields['propriete'].get().strip()
            # Mettre à jour le display_account
            entry['controle'] = ctrl_var.get()
            entry['type'] = avoirs_acct['type']
            entry['site'] = site

            # Renommage du compte
            renamed = new_intitule != old_intitule
            if renamed:
                avoirs_acct['intitule'] = new_intitule
                entry['intitule'] = new_intitule
                # Mettre à jour site_map
                self.account_site_map.pop(old_intitule, None)
            self.account_site_map[new_intitule] = site

            dlg.destroy()

            if renamed:
                self._run_uno_operation(
                    'Renommage du compte',
                    lambda: self._rename_account(old_intitule, new_intitule),
                    lambda: (self._after_accounts_save(
                        f"Compte '{old_intitule}' renommé en '{new_intitule}'"),
                             self._select_display_in_tree(new_intitule))
                )
            else:
                self._run_uno_operation(
                    'Enregistrement compte',
                    self._save_accounts,
                    lambda: (self._after_accounts_save(f"Compte '{new_intitule}' modifié"),
                             self._select_display_in_tree(new_intitule))
                )

        btn_frame = ttk.Frame(dlg)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text='OK', command=on_ok).pack(side='left', padx=5)
        ttk.Button(btn_frame, text='Annuler',
                   command=dlg.destroy).pack(side='left', padx=5)

    def _rename_account(self, old_name, new_name):
        """Worker UNO : renomme un compte dans Avoirs + Opérations + Plus_value.

        Avoirs col A : nom direct.
        Contrôles col A : texte brut renommé, formules =Avoirs.A{row} auto-propagées.
        Opérations col H : texte brut, scanné et remplacé.
        Plus_value col A : texte brut, scanné et remplacé.
        config_accounts.json : mis à jour (fait dans on_ok avant appel).
        """
        from inc_uno import UnoDocument
        from inc_excel_schema import uno_row, uno_col, ColResolver

        bak_path = self.xlsx_path.with_suffix('.xlsm.bak')
        shutil.copy2(self.xlsx_path, bak_path)

        with UnoDocument(self.xlsx_path) as doc:
            cr = doc.cr
            # 1. Avoirs col A — trouver et renommer
            ws_av = doc.get_sheet(SHEET_AVOIRS)
            col_a_av = cr.col('AVRintitulé')
            avr_data = self._start_avr + 1
            for r in range(uno_row(avr_data), uno_row(self._end_avr + 1)):
                cell = ws_av.getCellByPosition(col_a_av, r)
                if cell.getString() == old_name:
                    cell.setString(new_name)
                    break

            # 2. Avoirs — sauvegarder aussi les autres champs éditables
            for acct in self.accounts_data:
                r = acct.get('row')
                if r is None or acct.get('_is_new'):
                    continue
                r0 = uno_row(r)
                ws_av.getCellByPosition(cr.col('AVRdomiciliation'), r0).setString(acct.get('domiciliation') or '')
                ws_av.getCellByPosition(cr.col('AVRtitulaire'), r0).setString(acct.get('titulaire') or '')
                ws_av.getCellByPosition(cr.col('AVRpropriete'), r0).setString(acct.get('propriete') or '')

            # 3. Opérations col H — renommer toutes les occurrences
            ws_ops = doc.get_sheet(SHEET_OPERATIONS)
            col_h = cr.col('OPcompte')
            cursor = ws_ops.createCursor()
            cursor.gotoStartOfUsedArea(False)
            cursor.gotoEndOfUsedArea(True)
            last_row_0 = cursor.getRangeAddress().EndRow
            count_ops = 0
            for r in range(2, last_row_0 + 1):
                cell = ws_ops.getCellByPosition(col_h, r)
                if cell.getString() == old_name:
                    cell.setString(new_name)
                    count_ops += 1

            # 4. Plus_value col A — renommer toutes les occurrences
            ws_pv = doc.get_sheet(SHEET_PLUS_VALUE)
            col_a = cr.col('PVLcompte')
            cursor_pv = ws_pv.createCursor()
            cursor_pv.gotoStartOfUsedArea(False)
            cursor_pv.gotoEndOfUsedArea(True)
            last_row_pv = cursor_pv.getRangeAddress().EndRow
            count_pv = 0
            for r in range(0, last_row_pv + 1):
                cell = ws_pv.getCellByPosition(col_a, r)
                if cell.getString() == old_name:
                    cell.setString(new_name)
                    count_pv += 1

            # 5. Contrôles col A — texte brut uniquement (formules =Avoirs.A{row} auto-propagées)
            ws_ctrl = doc.get_sheet(SHEET_CONTROLES)
            col_a_ctrl = cr.col('CTRL1compte')
            ctrl_data = self._start_ctrl1 + 1
            for r in range(uno_row(ctrl_data), uno_row(self._end_ctrl1 + 1)):
                cell = ws_ctrl.getCellByPosition(col_a_ctrl, r)
                if cell.getType().value == 0:  # EMPTY
                    continue
                # getFormula() retourne la formule si présente, sinon le texte
                if cell.getFormula().startswith('='):
                    continue  # formule → auto-propagée par calculateAll
                if cell.getString() == old_name:
                    cell.setString(new_name)

            self._uno_finalize(doc)

        print(f'Compte renommé : Opérations={count_ops}, Plus_value={count_pv}', file=sys.stderr)

    def _acct_count_ops(self, account_name):
        """Compte les opérations appariées et non appariées d'un compte."""
        paired = 0
        unpaired = 0
        try:
            wb = openpyxl.load_workbook(self.xlsx_path, read_only=True, data_only=True)
            cr_xl = ColResolver.from_openpyxl(wb)
            ws = wb[SHEET_OPERATIONS]
            for row in ws.iter_rows(min_row=3, min_col=1, max_col=max(cr_xl.col('OPcompte'), cr_xl.col('OPréf'))):
                compte = row[cr_xl.col('OPcompte') - 1].value
                if not compte or str(compte).strip() != account_name:
                    continue
                ref = str(row[cr_xl.col('OPréf') - 1].value or '').strip()
                cat = str(row[cr_xl.col('OPcatégorie') - 1].value or '').strip()
                # Appariée = ref non vide, non "-", et catégorie ne commence pas par "#"
                if ref and ref != '-' and not cat.startswith('#'):
                    paired += 1
                else:
                    unpaired += 1
            wb.close()
        except Exception:
            pass
        return paired, unpaired

    def _acct_delete(self):
        """Supprime le compte sélectionné après confirmation."""
        entry = self._find_display_by_tree_selection()
        if not entry:
            messagebox.showwarning('Sélection', 'Aucun compte sélectionné.',
                                   parent=self.root)
            return
        label = entry['intitule']
        # Garder au moins un compte (template de style pour les ajouts futurs)
        if len(self.accounts_data) <= 1:
            messagebox.showwarning(
                'Suppression impossible',
                'Impossible de supprimer le dernier compte.\n'
                'Il sert de modèle pour les ajouts futurs.',
                parent=self.root)
            return

        # Compter les opérations appariées et non appariées
        paired, unpaired = self._acct_count_ops(entry['intitule'])

        if paired > 0:
            msg = (f"Le compte '{label}' a {paired} opération(s) appariée(s) "
                   f"et {unpaired} non appariée(s).\n\n"
                   f"• {unpaired} opération(s) non appariées seront supprimées\n"
                   f"• {paired} opération(s) appariées → transférées dans « Compte clos »\n"
                   f"• Le compte sera entièrement retiré\n\n"
                   "Continuer ?")
        else:
            msg = (f"Supprimer '{label}' ?\n\n"
                   f"{unpaired} opération(s) seront supprimées.\n"
                   "Le compte sera entièrement retiré.")

        if not messagebox.askyesno('Confirmer la suppression', msg,
                                   parent=self.root):
            return

        self.display_accounts.remove(entry)
        avoirs_ref = entry.get('avoirs_ref')
        if avoirs_ref and avoirs_ref in self.accounts_data:
            still_referenced = any(
                e.get('avoirs_ref') is avoirs_ref
                for e in self.display_accounts if e is not entry
            )
            if not still_referenced:
                self.accounts_data.remove(avoirs_ref)
        self._deleted_accounts.append(entry['intitule'])
        if paired > 0:
            self._soft_deleted_accounts.append(entry['intitule'])

        self.account_site_map.pop(entry['intitule'], None)
        if entry.get('ctrl_row') is not None:
            self._deleted_ctrl_rows.append(entry['ctrl_row'])

        # Dernier compte du site → désactiver le site
        site = entry.get('site', '')
        if site and site != 'N/A':
            remaining = sum(1 for a in self.accounts_data if a.get('site') == site)
            if remaining == 0:
                if site in self.site_vars:
                    self.site_vars[site].set(False)
                    self._refresh_site_list()
                messagebox.showinfo('Site désactivé',
                    f"Plus aucun compte pour {site}.\nLe site a été désactivé.",
                    parent=self.root)

        self._run_uno_operation(
            'Suppression compte',
            self._save_accounts,
            lambda: self._after_accounts_save(f"'{label}' supprimé")
        )

    def _acct_purge(self):
        """Purge un compte : supprime ses opérations et titres, garde la structure."""
        entry = self._find_display_by_tree_selection()
        if not entry:
            messagebox.showwarning('Sélection', 'Aucun compte sélectionné.',
                                   parent=self.root)
            return
        label = entry['intitule']
        paired, unpaired = self._acct_count_ops(label)
        total = paired + unpaired

        if total == 0:
            messagebox.showinfo('Purge', f"'{label}' n'a aucune opération.",
                                parent=self.root)
            return

        # Titres portefeuille
        titles = getattr(self, 'pv_titles', {}).get(label, [])
        if paired:
            parts = [f"• {unpaired} opération(s) supprimée(s)",
                     f"• {paired} opération(s) appariée(s) → « Compte clos »"]
        else:
            parts = [f"• {total} opération(s) supprimée(s)"]
        if titles:
            parts.append(f"• {len(titles)} titre(s) portefeuille supprimé(s)")
        parts.append("• Le compte reste dans la liste active (sans #Solde)")

        msg = f"Purger '{label}' ?\n\n" + "\n".join(parts) + "\n\nContinuer ?"

        if not messagebox.askyesno('Confirmer la purge', msg, parent=self.root):
            return

        self._run_uno_operation(
            f'Purge {label}',
            lambda: self._purge_account_uno(label, titles),
            lambda: (self._load_pv_titles(),
                     self._populate_accounts_tree(),
                     self._refresh_status_bar(),
                     self._set_status(self._purge_status_msg(label, total)))
        )

    def _purge_status_msg(self, label, total):
        """Construit le message status après purge, incluant le balai."""
        msg = f"'{label}' purgé — {total} opération(s) supprimée(s)"
        swept = self._last_sweep_count
        if swept:
            msg += f", {swept} ligne(s) balayée(s) dans « Compte clos »"
        return msg

    def _purge_account_uno(self, account_name, titles):
        """Worker UNO : purge les opérations et titres d'un compte."""
        from inc_uno import UnoDocument
        from inc_excel_schema import (uno_row, uno_col, SHEET_PLUS_VALUE)

        with UnoDocument(self.xlsx_path) as doc:
            cr = doc.cr
            # 1. Opérations : supprimer non appariées, reloger appariées → "Compte clos"
            COMPTE_CLOS = 'Compte clos'
            ws_ops = doc.get_sheet(SHEET_OPERATIONS)
            cursor = ws_ops.createCursor()
            cursor.gotoStartOfUsedArea(False)
            cursor.gotoEndOfUsedArea(True)
            last_row_0 = cursor.getRangeAddress().EndRow
            rows_to_delete = []
            rehoused = 0
            for row_0 in range(2, last_row_0 + 1):
                compte = ws_ops.getCellByPosition(cr.col('OPcompte'), row_0).getString()
                if not compte or compte.strip() != account_name:
                    continue
                ref = ws_ops.getCellByPosition(cr.col('OPréf'), row_0).getString().strip()
                cat = ws_ops.getCellByPosition(cr.col('OPcatégorie'), row_0).getString().strip()
                if ref and ref != '-' and not cat.startswith('#'):
                    ws_ops.getCellByPosition(cr.col('OPcompte'), row_0).setString(COMPTE_CLOS)
                    rehoused += 1
                else:
                    rows_to_delete.append(row_0)
            for row_0 in reversed(rows_to_delete):
                ws_ops.Rows.removeByIndex(row_0, 1)

            # 1a. Balai : supprimer les paires entièrement dans "Compte clos"
            swept = 0
            if rehoused > 0:
                swept = self._sweep_compte_clos_uno(ws_ops, cr=cr)
            self._last_sweep_count = swept

            # 1b. Ajouter 2 lignes #Solde (hier + aujourd'hui)
            devise = ''
            for acct in self.accounts_data:
                if acct['intitule'] == account_name:
                    devise = acct.get('devise', '')
                    break
            self._append_solde_lines(ws_ops, account_name, devise, doc=doc)

            # 2. Plus_value : supprimer uniquement les titres, garder la structure
            if titles:
                ws_pv = doc.get_sheet(SHEET_PLUS_VALUE)
                col_b = cr.col('PVLcompte')
                col_c = cr.col('PVLtitre')
                # Identifier les lignes titres (*nom*) et la ligne Total
                title_rows = []
                total_row_pv = None
                pvl_data = (self._start_pvl or 5) + 1
                for row_idx in range(pvl_data, self._end_pvl + 1):
                    val_b = ws_pv.getCellByPosition(col_b, uno_row(row_idx)).getString().strip()
                    val_c = ws_pv.getCellByPosition(col_c, uno_row(row_idx)).getString().strip()
                    if val_b != account_name:
                        if title_rows or total_row_pv:
                            break  # Sorti du bloc
                        continue
                    if val_c.startswith('*') and val_c.endswith('*') and len(val_c) > 2:
                        title_rows.append(row_idx)
                    elif val_c == 'Total':
                        total_row_pv = row_idx

                # Supprimer les titres en ordre inverse
                for row_idx in reversed(title_rows):
                    ws_pv.Rows.removeByIndex(uno_row(row_idx), 1)

                # Mettre les formules Total à 0 (plus de titres)
                if total_row_pv is not None:
                    adjusted = total_row_pv - len(title_rows)
                    t0 = uno_row(adjusted)
                    for pv_c in ('PVLdate_init', 'PVLmontant_init',
                                 'PVLsigma', 'PVLdate', 'PVLmontant'):
                        ws_pv.getCellByPosition(cr.col(pv_c), t0).setValue(0)

            # 3. Avoirs : ne pas toucher (les formules J/K se recalculent automatiquement)

            # 4. Créer "Compte clos" dans Avoirs si des ops ont été relogées
            ws_av = doc.get_sheet(SHEET_AVOIRS)
            if rehoused > 0:
                total_row = None
                total_row = (self._end_avr + 1) if self._end_avr else None
                self._ensure_compte_clos(ws_av, total_row, cr=cr)

            self._uno_finalize(doc)

    def _acct_add_title(self):
        """Dialog pour ajouter un titre à un portefeuille."""
        entry = self._find_display_by_tree_selection()
        if not entry or entry.get('type') != 'Portefeuilles':
            return
        if not self.xlsx_path:
            return
        # Le compte doit être sauvegardé (bloc Plus_value créé) avant d'ajouter un titre
        if entry.get('avoirs_ref', {}).get('_is_new'):
            messagebox.showinfo(
                'Sauvegarde requise',
                "Le compte doit être sauvegardé avant d'ajouter un titre.\n\n"
                "Attendez la fin de l'enregistrement puis réessayez.",
                parent=self.root)
            return

        account_name = entry['intitule']
        account_devise = entry.get('devise') or 'EUR'

        dlg = tk.Toplevel(self.root)
        dlg.title('Ajout titre')
        dlg.geometry('400x210')
        dlg.transient(self.root)
        dlg.wait_visibility()
        dlg.grab_set()

        ttk.Label(dlg, text='Nom du titre :').grid(
            row=0, column=0, sticky='w', padx=10, pady=5)
        name_var = tk.StringVar()
        name_entry = ttk.Entry(dlg, textvariable=name_var, width=30)
        name_entry.grid(row=0, column=1, padx=10, pady=5, sticky='w')

        ttk.Label(dlg, text='Devise :').grid(
            row=1, column=0, sticky='w', padx=10, pady=5)
        devise_var = tk.StringVar(value=account_devise)
        ttk.Combobox(dlg, textvariable=devise_var,
                      values=self.ACCOUNT_DEVISES,
                      width=12, state='readonly').grid(
            row=1, column=1, padx=10, pady=5, sticky='w')

        ttk.Label(dlg, text='Date initiale :').grid(
            row=2, column=0, sticky='w', padx=10, pady=5)
        date_var = tk.StringVar()
        ttk.Entry(dlg, textvariable=date_var, width=14).grid(
            row=2, column=1, padx=10, pady=5, sticky='w')
        ttk.Label(dlg, text='(JJ/MM/AAAA, optionnel)',
                  foreground='grey').grid(
            row=2, column=1, padx=130, sticky='w')

        status_label = ttk.Label(dlg, text='', foreground='red')
        status_label.grid(row=3, column=0, columnspan=2, padx=10)

        def on_ok():
            title = name_var.get().strip()
            if not title:
                status_label.config(text='Le nom du titre est obligatoire.')
                return
            devise = devise_var.get().strip() or account_devise
            date_str = date_var.get().strip()
            date_init = None
            if date_str:
                try:
                    date_init = datetime.strptime(date_str, '%d/%m/%Y')
                except ValueError:
                    status_label.config(text='Format date invalide (JJ/MM/AAAA).')
                    return
            dlg.destroy()
            self._run_title_save(account_name, title, devise, date_init)

        btn_frame = ttk.Frame(dlg)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text='OK', command=on_ok).pack(
            side='left', padx=5)
        ttk.Button(btn_frame, text='Annuler',
                   command=dlg.destroy).pack(side='left', padx=5)
        name_entry.focus()

    def _get_selected_title_info(self):
        """Retourne (account_name, title_name, devise, pv_row) du titre sélectionné, ou None."""
        sel = self.acct_tree.selection()
        if not sel:
            return None
        item_id = sel[0]
        if not self._is_title_node(item_id):
            return None
        title_item = self.acct_tree.item(item_id)
        title_name = title_item['text'].strip()
        vals = title_item['values']
        devise = str(vals[0]) if vals else ''
        # Remonter au compte parent pour l'account_name
        parent_id = self.acct_tree.parent(item_id)
        account_name = self.acct_tree.item(parent_id)['text']
        # Trouver le pv_row dans self.pv_titles
        pv_row = None
        for t_name, t_dev, t_row in self.pv_titles.get(account_name, []):
            if t_name == title_name and t_dev == devise:
                pv_row = t_row
                break
        return account_name, title_name, devise, pv_row

    def _pv_title_rename(self):
        """Dialog pour renommer un titre de portefeuille."""
        info = self._get_selected_title_info()
        if not info:
            return
        account_name, old_name, devise, pv_row = info
        if not pv_row or not self.xlsx_path:
            return

        dlg = tk.Toplevel(self.root)
        dlg.title('Renommer titre')
        dlg.geometry('400x150')
        dlg.transient(self.root)
        dlg.wait_visibility()
        dlg.grab_set()

        ttk.Label(dlg, text=f'Titre actuel : {old_name}').grid(
            row=0, column=0, columnspan=2, sticky='w', padx=10, pady=5)
        ttk.Label(dlg, text='Nouveau nom :').grid(
            row=1, column=0, sticky='w', padx=10, pady=5)
        name_var = tk.StringVar(value=old_name)
        name_entry = ttk.Entry(dlg, textvariable=name_var, width=30)
        name_entry.grid(row=1, column=1, padx=10, pady=5, sticky='w')

        ttk.Label(dlg, text='Les opérations existantes ne seront pas renommées.',
                  foreground='grey').grid(
            row=2, column=0, columnspan=2, padx=10, sticky='w')

        def on_ok():
            new_name = name_var.get().strip()
            if not new_name or new_name == old_name:
                dlg.destroy()
                return
            dlg.destroy()
            self._run_uno_operation(
                'Renommage en cours',
                lambda: self._rename_pv_title(account_name, old_name, new_name, pv_row),
                lambda: self._after_pv_title_change(
                    account_name, f'Titre "{old_name}" renommé en "{new_name}".')
            )

        btn_frame = ttk.Frame(dlg)
        btn_frame.grid(row=3, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text='OK', command=on_ok).pack(side='left', padx=5)
        ttk.Button(btn_frame, text='Annuler',
                   command=dlg.destroy).pack(side='left', padx=5)
        name_entry.focus()
        name_entry.select_range(0, 'end')

    def _pv_title_delete(self):
        """Supprime un titre de portefeuille après confirmation.

        Garde : le solde du titre (PVLmontant) doit être à zéro.
        """
        info = self._get_selected_title_info()
        if not info:
            return
        account_name, title_name, devise, pv_row = info
        if not pv_row or not self.xlsx_path:
            return

        # Garde : vérifier que le solde est à zéro
        wb = openpyxl.load_workbook(self.xlsx_path, data_only=True)
        try:
            from inc_excel_schema import ColResolver as _CR
            solde = wb[SHEET_PLUS_VALUE].cell(pv_row, _CR.from_openpyxl(wb).col('PVLmontant')).value
        finally:
            wb.close()
        solde_f = float(solde) if solde else 0.0
        if abs(solde_f) > 0.01:
            messagebox.showwarning(
                'Suppression impossible',
                f'Le titre "{title_name}" a un solde non nul '
                f'({solde_f:,.2f} {devise}).\n\n'
                'Seuls les titres soldés (vendus ou arbitrés) '
                'peuvent être retirés du portefeuille.',
                parent=self.root)
            return

        if not messagebox.askyesno(
                'Confirmer la suppression',
                f'Retirer le titre "{title_name}" du portefeuille {account_name} ?\n\n'
                'Le titre est soldé — il sera retiré du suivi.\n'
                'Les opérations associées restent dans Opérations.',
                parent=self.root):
            return

        self._run_uno_operation(
            'Suppression en cours',
            lambda: self._delete_pv_title(account_name, title_name, pv_row),
            lambda: self._after_pv_title_change(
                account_name, f'Titre "{title_name}" supprimé.')
        )

    def _after_pv_title_change(self, account_name, msg):
        """Callback après renommage/suppression de titre : recharge + refresh."""
        self._load_pv_titles()
        self._populate_accounts_tree()
        self._set_status(msg)

    def _rename_pv_title(self, account_name, old_name, new_name, pv_row):
        """Worker UNO : renomme un titre dans Plus_value (col B)."""
        from inc_uno import UnoDocument
        from inc_excel_schema import uno_row

        bak_path = self.xlsx_path.with_suffix('.xlsm.bak')
        shutil.copy2(self.xlsx_path, bak_path)

        with UnoDocument(self.xlsx_path) as doc:
            ws_pv = doc.get_sheet(SHEET_PLUS_VALUE)
            ws_pv.getCellByPosition(cr.col('PVLtitre'), uno_row(pv_row)).setString(f'*{new_name}*')
            self._uno_finalize(doc)

    def _delete_pv_title(self, account_name, title_name, pv_row):
        """Worker UNO : supprime un titre dans Plus_value.

        - removeByIndex sur la ligne titre
        - Si c'était le dernier titre, met les formules Total à 0
        - Sinon removeByIndex ajuste automatiquement les SUM/MIN du Total
        """
        from inc_uno import UnoDocument
        from inc_excel_schema import uno_row, uno_col
        cr = doc.cr
        import re

        bak_path = self.xlsx_path.with_suffix('.xlsm.bak')
        shutil.copy2(self.xlsx_path, bak_path)

        with UnoDocument(self.xlsx_path) as doc:
            ws_pv = doc.get_sheet(SHEET_PLUS_VALUE)

            # Compter les titres restants pour ce compte
            titles_for_account = self.pv_titles.get(account_name, [])
            is_last = len(titles_for_account) <= 1

            if is_last:
                # Dernier titre : trouver le Total et mettre ses formules à 0
                total_row = None
                col_c = cr.col('PVLtitre')
                for scan in range(pv_row + 1, pv_row + 10):
                    val_c = ws_pv.getCellByPosition(col_c, uno_row(scan)).getString().strip()
                    if val_c == 'Total':
                        total_row = scan
                        break
                ws_pv.Rows.removeByIndex(uno_row(pv_row), 1)
                if total_row:
                    # Total a décalé de -1
                    total_r0 = uno_row(total_row - 1)
                    for pv_c in ('PVLdate_init', 'PVLmontant_init',
                                 'PVLsigma', 'PVLdate', 'PVLmontant'):
                        ws_pv.getCellByPosition(uno_col(pv_c), total_r0).setValue(0)
            else:
                # Pas le dernier : removeByIndex ajuste auto les SUM
                ws_pv.Rows.removeByIndex(uno_row(pv_row), 1)

            self._uno_finalize(doc)

    @staticmethod
    def _append_solde_lines(ws_ops, account_name, devise,
                            date_debut=None, date_solde=None,
                            montant_debut=None, doc=None, **_kwargs):
        """Ajoute 0 ou 1 ligne #Solde initial après la dernière opération.

        Refonte 0..N #Solde : si aucune valeur n'est fournie en GUI
        (montant_debut=None), aucune ligne n'est créée. Si une valeur est
        donnée, une seule ligne #Solde est ajoutée à date_debut (par défaut
        aujourd'hui) avec ce montant.
        """
        if montant_debut is None:
            return  # Pas de valeur initiale → aucun #Solde

        from inc_excel_schema import uno_col, uno_row
        from inc_uno import copy_row_style
        from datetime import datetime

        cr = doc.cr

        cursor = ws_ops.createCursor()
        cursor.gotoEndOfUsedArea(True)
        ops_last_0 = cursor.getRangeAddress().EndRow
        min_data_0 = cr.rows('OPdate')[0] - 1  # model row START (0-indexed)
        while ops_last_0 > min_data_0 and not ws_ops.getCellByPosition(0, ops_last_0).getString():
            ops_last_0 -= 1
        ops_next_0 = ops_last_0 + 1
        template_ops_0 = max(ops_last_0, min_data_0)
        epoch = datetime(1899, 12, 30)
        if date_debut is None:
            date_debut = datetime.today()
        from inc_formats import FORMATS_DEVISE, FORMAT_EUR, GRIS
        is_non_eur = devise and devise != 'EUR'
        if doc is not None:
            fmt_devise = doc.register_number_format(FORMATS_DEVISE.get(devise, FORMAT_EUR)) if devise else None
            fmt_eur = doc.register_number_format(FORMAT_EUR)
        else:
            fmt_devise = None
            fmt_eur = None

        serial = (date_debut - epoch).days
        copy_row_style(ws_ops, template_ops_0, ops_next_0, col_start=0, col_end=9)
        ws_ops.getCellByPosition(cr.col('OPdate'), ops_next_0).setValue(serial)
        ws_ops.getCellByPosition(cr.col('OPlibellé'), ops_next_0).setString('Relevé compte')
        c_cell = ws_ops.getCellByPosition(cr.col('OPmontant'), ops_next_0)
        c_cell.setValue(montant_debut)
        if fmt_devise is not None:
            c_cell.NumberFormat = fmt_devise
        if fmt_eur is not None:
            ws_ops.getCellByPosition(cr.col('OPequiv_euro'), ops_next_0).NumberFormat = fmt_eur
        ws_ops.getCellByPosition(cr.col('OPdevise'), ops_next_0).setString(devise or '')
        ws_ops.getCellByPosition(cr.col('OPcatégorie'), ops_next_0).setString('#Solde')
        ws_ops.getCellByPosition(cr.col('OPcompte'), ops_next_0).setString(account_name)
        if is_non_eur:
            c_cell.CellBackColor = GRIS
            ws_ops.getCellByPosition(cr.col('OPdevise'), ops_next_0).CellBackColor = GRIS

    @staticmethod
    def _sweep_compte_clos_uno(ws_ops, cr=None):
        """Supprime les paires où les deux côtés sont dans 'Compte clos'.

        Scan toutes les ops : pour chaque ref, si AUCUNE op n'est dans un
        compte actif, toutes les lignes de cette ref dans 'Compte clos' sont
        supprimées (la paire n'a plus de valeur).

        Returns:
            int: nombre de lignes supprimées
        """
        from inc_excel_schema import uno_col
        COMPTE_CLOS = 'Compte clos'

        cursor = ws_ops.createCursor()
        cursor.gotoStartOfUsedArea(False)
        cursor.gotoEndOfUsedArea(True)
        last_row_0 = cursor.getRangeAddress().EndRow

        # Collecter les refs par localisation (clos vs actif)
        clos_rows = {}    # ref -> [row_0, ...]
        active_refs = set()  # refs présentes dans un compte actif

        for row_0 in range(2, last_row_0 + 1):
            ref = ws_ops.getCellByPosition(cr.col('OPréf'), row_0).getString().strip()
            if not ref or ref == '-':
                continue
            cat = ws_ops.getCellByPosition(cr.col('OPcatégorie'), row_0).getString().strip()
            if cat.startswith('#'):
                continue
            compte = ws_ops.getCellByPosition(cr.col('OPcompte'), row_0).getString().strip()
            if compte == COMPTE_CLOS:
                clos_rows.setdefault(ref, []).append(row_0)
            else:
                active_refs.add(ref)

        # Refs entièrement dans "Compte clos" → suppression
        rows_to_delete = []
        for ref, rows in clos_rows.items():
            if ref not in active_refs:
                rows_to_delete.extend(rows)

        for row_0 in sorted(rows_to_delete, reverse=True):
            ws_ops.Rows.removeByIndex(row_0, 1)

        return len(rows_to_delete)

    def _ensure_compte_clos(self, ws_avoirs, total_row, cr=None):
        """Crée 'Compte clos' dans Avoirs s'il n'existe pas. Retourne True si créé."""
        from inc_excel_schema import uno_row, uno_col
        COMPTE_CLOS = 'Compte clos'
        avr_data = (self._start_avr or 4) + 1
        end_avr = self._end_avr or (total_row and total_row - 1)
        if not end_avr:
            return False
        for row_idx in range(avr_data, end_avr + 1):
            val = ws_avoirs.getCellByPosition(cr.col('AVRintitulé'), uno_row(row_idx)).getString()
            if val and val.strip() == COMPTE_CLOS:
                return False
        # Insérer avant end AVR (dans la zone données)
        insert_0 = uno_row(end_avr)
        ws_avoirs.Rows.insertByIndex(insert_0, 1)
        ws_avoirs.getCellByPosition(cr.col('AVRintitulé'), insert_0).setString(COMPTE_CLOS)
        self._end_avr += 1
        return True

    def _uno_finalize(self, doc):
        """calculateAll + lecture Contrôles A1 + save. Appelé par les workers UNO."""
        doc.calculate_all()
        doc.save()  # save() écrit automatiquement les miroirs C1 et L1

    def _run_uno_operation(self, label, worker_fn, on_success=None):
        """Lance une opération UNO dans un thread avec fenêtre d'attente animée.

        Args:
            label: message affiché dans la modale (ex: 'Écriture en cours')
            worker_fn: callable exécuté dans le thread (pas d'arguments)
            on_success: callable exécuté dans le thread principal après succès
        """
        # Garde 1 : lock file (LibreOffice a le fichier ouvert)
        if self.xlsx_path:
            lock = self.xlsx_path.parent / f".~lock.{self.xlsx_path.name}#"
            if lock.exists():
                # Vérifier qu'un processus soffice tourne réellement
                import subprocess as _sp
                lo_running = _sp.run(
                    ['pgrep', '-x', 'soffice.bin'],
                    capture_output=True).returncode == 0
                if lo_running:
                    messagebox.showwarning('Fichier verrouillé',
                        'Le fichier est ouvert dans LibreOffice.\n'
                        'Fermer LibreOffice avant de continuer.',
                        parent=self.root)
                    return

        # Garde 2 : soffice déjà actif sur le port 2002
        # Tolérance : soffice peut mettre quelques secondes à libérer le port
        import socket
        import time
        port_busy = False
        for attempt in range(4):  # 3 retries, ~3s max
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                if s.connect_ex(('localhost', 2002)) != 0:
                    port_busy = False
                    break
                port_busy = True
            if attempt < 3:
                time.sleep(1)
        if port_busy:
            messagebox.showwarning('Port UNO occupé',
                'Un processus soffice est déjà actif (port 2002).\n'
                'Fermer LibreOffice ou tuer le processus soffice.',
                parent=self.root)
            return

        wait = tk.Toplevel(self.root)
        wait.title('')
        wait.geometry('320x80')
        wait.transient(self.root)
        wait.resizable(False, False)
        wait.protocol('WM_DELETE_WINDOW', lambda: None)
        wait.wait_visibility()
        wait.grab_set()

        msg_var = tk.StringVar(value=label)
        ttk.Label(wait, textvariable=msg_var, font=('', 11)).pack(
            expand=True, pady=15)

        dots = [0]

        def animate():
            dots[0] = (dots[0] % 3) + 1
            msg_var.set(label + '.' * dots[0])
            wait.after(400, animate)

        animate()

        result = {}

        def worker():
            try:
                worker_fn()
            except Exception as e:
                result['error'] = e

        t = threading.Thread(target=worker, daemon=True)
        t.start()

        def check_done():
            if t.is_alive():
                wait.after(100, check_done)
                return
            wait.grab_release()
            wait.destroy()
            if 'error' in result:
                messagebox.showerror('Erreur UNO',
                                     f'Erreur :\n{result["error"]}',
                                     parent=self.root)
                self._load_accounts_data()
                self._populate_accounts_tree()
                self._load_budget_categories()
                self._set_status('Erreur UNO — état restauré depuis le fichier.', 'error')
            else:
                if on_success:
                    on_success()
                # Rafraîchir statut (Contrôles + Total Avoirs)
                self._refresh_status_bar()

        check_done()

    def _run_title_save(self, account_name, title_name, devise, date_init):
        """Lance _insert_pv_title dans un thread avec fenêtre d'attente animée."""
        self._run_uno_operation(
            'Écriture en cours',
            lambda: self._insert_pv_title(account_name, title_name, devise, date_init),
            lambda: self._set_status(f'Titre "{title_name}" ajouté à {account_name}.')
        )

