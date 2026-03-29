#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
cpt_format_SG.py - Convertit les fichiers SG en format standardisé

Pour opérations CSV: Parse directement avec cpt_categorize (remplace AWK)
Pour positions (supports assurance vie): Parse Excel et génère format 4 colonnes
Pour PDF imprimés: Parse texte avec pdfplumber (fallback collecte manuelle)

Usage:
  ./cpt_format_SG.py <fichier.csv|.xlsx|.pdf>
"""

import sys
import csv
import re
import json
from pathlib import Path
from datetime import datetime
import inc_categorize
from inc_format import process_files, lines_to_tuples, log_csv_debug as _log_csv_debug, get_file_date, site_name_from_file

SITE = site_name_from_file(__file__)

try:
    import openpyxl as xl
except ImportError:
    xl = None

try:
    import pdfplumber
except ImportError:
    pdfplumber = None


# Comptes SG : chargés depuis config_accounts.json
_ACCOUNTS_JSON = Path(__file__).parent / 'config_accounts.json'
with open(_ACCOUNTS_JSON, 'r', encoding='utf-8') as _f:
    _sg_config = json.load(_f).get('SG', {})
_sg_accounts = {a['name']: a for a in _sg_config.get('accounts', [])}

# Mapping nom fichier → nom compte (détection assurances vie)
COMPTE_MAPPING = {}
for _name in _sg_accounts:
    if 'ébène' in _name.lower() or 'ebene' in _name.lower():
        # "Ass vie ébène 2 Cécile" → clé "ass-vie-ebene-2-cecile"
        import unicodedata
        _key = unicodedata.normalize('NFD', _name.lower())
        _key = ''.join(c for c in _key if unicodedata.category(c) != 'Mn')
        _key = _key.replace(' ', '-')
        COMPTE_MAPPING[_key] = _name

# Compte chèque (opérations CSV)
ACCOUNT_CHEQUE = next((n for n in _sg_accounts if 'chèque' in n.lower()), 'Compte chèque SG')
COMPTE_NUMERO_MAPPING = {
    a['numero']: a['name']
    for a in _sg_config.get('accounts', [])
    if 'numero' in a
}
# Retirer le compte principal (pas dans les Export_*.csv)
_compte_principal = _sg_config.get('compte_principal', '')
COMPTE_NUMERO_MAPPING.pop(_compte_principal, None)

# Mapping des noms de supports (renommages SG → noms Plus_value)
SUPPORT_NAME_MAPPING = {
    # Renommages historiques (SRI AGIR)
    'OPPORTUNITES SRI AGIR': 'DNCA INVEST - EVOLUTIF SRI',
    'OBLIGATIONS SRI AGIR': 'DNCA INVEST - OBLIGATIONS SRI',
    'MONETAIRE SRI AGIR': 'DNCA INVEST - FLEX FUTUR',

    # Renommages SG IS FUND → MOOREA FUND (ébène 1)
    'SG IS FUND - CREDIT MILLESIME 2030 RE': 'MOOREA FUND - HIGH YIELD',
    'SG IS FUND - SG CREDIT MILLESIME 2028 - PART RE': 'MOOREA FUND - SG CREDIT',
    'SG IS FUND - US EQUITY RE': 'MOOREA FUND - US EQUITY',

    # Troncatures (Plus_value a des noms tronqués)
    'SCI PIERRE PATRIMOINE': 'SCI PIERRE PATRIM',
}


def detect_compte(file_path):
    """Détecte le nom du compte depuis le nom de fichier"""
    filename = file_path.stem.lower()
    # Trier par longueur décroissante pour éviter les matchs partiels
    # Ex: "ass-vie-ebene-2-cecile" doit matcher avant "ass-vie-ebene-cecile"
    for key, compte_name in sorted(COMPTE_MAPPING.items(), key=lambda x: -len(x[0])):
        if key in filename:
            return compte_name
    return None


def process_operations(file_path):
    """
    Parse le CSV SG et génère le format standardisé 9 colonnes

    Format CSV SG (5 champs):
      Ligne 1: ;;;;date_solde;montant_solde EUR
      Lignes 2-3: Headers (skip)
      Lignes 4+: Date;Court;Long;Montant;Devise

    Remplace l'appel à cpt_format_SG.awk
    Note: Le filtrage par date est centralisé dans inc_format.process_files()
    """
    ACCOUNT_NAME = ACCOUNT_CHEQUE
    COMMENTAIRE = ""

    output_lines = []

    # Lire le fichier avec encoding latin-1 (format SG)
    with open(file_path, 'r', encoding='latin-1') as f:
        lines = f.readlines()

    if len(lines) < 4:
        print("⚠ Fichier CSV trop court (< 4 lignes)", file=sys.stderr)
        return ""

    # Ligne 1: Récupérer le solde
    line1 = lines[0].strip()
    fields1 = line1.split(';')

    if len(fields1) < 6:
        print(f"⚠ Ligne 1 invalide (attendu >= 6 champs, reçu {len(fields1)})", file=sys.stderr)
        date_solde = ""
        solde_montant = "0,00"
    else:
        date_solde = fields1[4].strip()  # Colonne 5 (index 4)
        solde_avec_devise = fields1[5].strip()  # Colonne 6 (index 5)
        # Split par espace pour retirer "EUR"
        solde_parts = solde_avec_devise.split()
        solde_montant = solde_parts[0] if solde_parts else "0,00"

    # Parser toutes les opérations d'abord (pour filtrage)
    operations = []
    for line in lines[3:]:  # Skip lignes 1, 2, 3 (index 0, 1, 2)
        line = line.strip()
        if not line:
            continue

        fields = line.split(';')
        if len(fields) < 5:
            continue

        date_str = fields[0].strip()
        label_long = fields[2].strip()  # Utilisé pour catégorisation
        montant = fields[3].strip()
        devise = fields[4].strip() if len(fields) > 4 else "EUR"

        operations.append({
            'date_str': date_str,
            'label_long': label_long,
            'montant': montant,
            'devise': devise
        })

    # Générer les lignes formatées
    for op in operations:
        # Catégorisation automatique via patterns
        cat, opts = inc_categorize.categorize_operation(op['label_long'], SITE)
        ref = opts.get('ref', '')

        # Format standardisé: Date;Libellé;Montant;Devise;Equiv;Réf;Catégorie;Compte;Commentaire
        output_line = f"{op['date_str']};{op['label_long']};{op['montant']};{op['devise']};;{ref};{cat};{ACCOUNT_NAME};{COMMENTAIRE}"
        output_lines.append(output_line)

    # Ligne #Solde finale
    if date_solde:
        solde_line = f"{date_solde};Relevé compte;{solde_montant};EUR;;;#Solde;{ACCOUNT_NAME};{COMMENTAIRE}"
        output_lines.append(solde_line)

    # Construire la sortie complète (header + lignes)
    result = "Date;Libellé;Montant;Devise;Equiv;Réf;Catégorie;Compte;Commentaire\n"
    result += "\n".join(output_lines)

    return result


def process_epargne_csv(file_path):
    """
    Traite les CSV épargne SG (Export_*.csv) - format direct depuis le site

    Format CSV:
      Ligne 1: ="0006100030472944";date_debut;date_fin;
      Ligne 2: date_comptabilisation;libellé_complet_operation;montant_operation;devise;
      Lignes 3+: DD/MM/YYYY;Libellé;Montant;EUR;

    Output: Format standardisé 9 colonnes avec catégorisation
    """
    COMMENTAIRE = ""
    output_lines = []
    operations = []

    # Lire le fichier avec encoding latin-1 (format SG)
    with open(file_path, 'r', encoding='latin-1') as f:
        lines = f.readlines()

    if len(lines) < 3:
        print("⚠ Fichier CSV trop court (< 3 lignes)", file=sys.stderr)
        return "Date;Libellé;Montant;Devise;Equiv;Réf;Catégorie;Compte;Commentaire\n"

    # Ligne 1: Extraire le numéro de compte
    # Format: ="0006100030472944";29/07/2025;28/01/2026;
    line1 = lines[0].strip()
    fields1 = line1.split(';')

    compte = None
    date_fin = None

    if len(fields1) >= 3:
        # Extraire le numéro de compte (retirer ="0006" et garder les 11 derniers chiffres)
        numero_raw = fields1[0].strip().strip('="')
        # Le numéro peut être "0006100030472944" → extraire "00030472944"
        if len(numero_raw) >= 11:
            numero = numero_raw[-11:]
            compte = COMPTE_NUMERO_MAPPING.get(numero)
        date_fin = fields1[2].strip()  # Date de fin pour le solde

    if not compte:
        # Essayer d'extraire depuis le nom de fichier
        filename = Path(file_path).stem
        for num, nom in COMPTE_NUMERO_MAPPING.items():
            if num in filename:
                compte = nom
                break

    if not compte:
        print(f"⚠ Compte non identifié dans {file_path}", file=sys.stderr)
        compte = "Compte SG"

    # Ligne 2: En-têtes (skip)
    # Lignes 3+: Opérations
    for line in lines[2:]:
        line = line.strip()
        if not line:
            continue

        fields = line.split(';')
        if len(fields) < 4:
            continue

        date_str = fields[0].strip()
        libelle = fields[1].strip()
        montant = fields[2].strip()
        devise = fields[3].strip() if fields[3].strip() else "EUR"

        # Ignorer si date vide
        if not date_str or not re.match(r'\d{2}/\d{2}/\d{4}', date_str):
            continue

        operations.append({
            'date_str': date_str,
            'libelle': libelle,
            'montant': montant,
            'devise': devise
        })

    # Générer les lignes formatées
    for op in operations:
        cat, opts = inc_categorize.categorize_operation(op['libelle'], SITE)
        ref = opts.get('ref', '')

        output_line = f"{op['date_str']};{op['libelle']};{op['montant']};{op['devise']};;{ref};{cat};{compte};{COMMENTAIRE}"
        output_lines.append(output_line)

    # Note: Pas de #Solde dans ce format - le solde vient du PDF "Mes comptes en ligne _ SG.pdf"

    result = "Date;Libellé;Montant;Devise;Equiv;Réf;Catégorie;Compte;Commentaire\n"
    result += "\n".join(output_lines)

    return result


def process_operations_parsed(file_path):
    """
    Traite les fichiers CSV RAW parsed (assurances vie / épargne) - format 4 colonnes

    Format RAW:
      Header: Date;Opération;Montant;Compte (ou Date;Libellé;Montant;Compte)
      Lignes: Date;Libellé;Montant;Compte
      #Solde: Date;#Solde;Montant;Compte

    Output: Format standardisé 9 colonnes avec catégorisation

    Note: Le filtrage par date est centralisé dans inc_format.process_files()
    """
    COMMENTAIRE = ""

    output_lines = []
    soldes_par_compte = {}  # {nom_compte: {'date_str': ..., 'montant': ...}}

    # Lire le fichier avec encoding UTF-8 (format parsed)
    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=';')

        operations = []
        for row in reader:
            date_str_raw = row['Date'].strip()
            # Header peut être "Opération" ou "Libellé"
            libelle = row.get('Opération', row.get('Libellé', '')).strip()
            montant = row['Montant'].strip()
            compte = row['Compte'].strip()

            # Convertir date DD/MM/YY → DD/MM/YYYY pour parse_french_date
            # Exemple: "31/12/25" → "31/12/2025"
            if '/' in date_str_raw:
                parts = date_str_raw.split('/')
                if len(parts) == 3 and len(parts[2]) == 2:
                    # Année sur 2 chiffres → convertir en 4 chiffres
                    # 00-49 → 2000-2049, 50-99 → 1950-1999
                    year_2digit = int(parts[2])
                    year_4digit = 2000 + year_2digit if year_2digit < 50 else 1900 + year_2digit
                    date_str = f"{parts[0]}/{parts[1]}/{year_4digit}"
                else:
                    date_str = date_str_raw
            else:
                date_str = date_str_raw

            # Gérer le #Solde (dans la colonne Libellé/Opération)
            if libelle == '#Solde':
                # Stocker le solde par compte (peut y avoir plusieurs comptes)
                soldes_par_compte[compte] = {
                    'date_str': date_str,
                    'montant': montant
                }
                continue

            operations.append({
                'date_str': date_str,
                'libelle': libelle,
                'montant': montant,
                'compte': compte
            })

    # Identifier les comptes qui ont des opérations
    comptes_avec_operations = set(op['compte'] for op in operations)

    # Générer les lignes formatées
    for op in operations:
        # Catégorisation automatique via patterns
        cat, opts = inc_categorize.categorize_operation(op['libelle'], SITE)
        ref = opts.get('ref', '')

        # Format standardisé: Date;Libellé;Montant;Devise;Equiv;Réf;Catégorie;Compte;Commentaire
        output_line = f"{op['date_str']};{op['libelle']};{op['montant']};EUR;;{ref};{cat};{op['compte']};{COMMENTAIRE}"
        output_lines.append(output_line)

    # Ajouter #Solde pour chaque compte qui a des opérations conservées
    for compte, solde_info in soldes_par_compte.items():
        if compte in comptes_avec_operations:
            solde_line = f"{solde_info['date_str']};Relevé compte;{solde_info['montant']};EUR;;;#Solde;{compte};{COMMENTAIRE}"
            output_lines.append(solde_line)

    # Construire la sortie complète (header + lignes)
    result = "Date;Libellé;Montant;Devise;Equiv;Réf;Catégorie;Compte;Commentaire\n"
    result += "\n".join(output_lines)

    return result


def process_positions(file_path):
    """
    Parse fichier Excel supports et génère format 4 colonnes
    Calcule SYNOE pour Ass vie ébène 2

    Détection du compte par nom de fichier :
    - SG_Ebene2_supports*.xlsx → ébène 2 (avec calcul SYNOE)
    - SG_Ebene_supports*.xlsx → ébène 1
    """
    if xl is None:
        print("❌ Module openpyxl requis: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    # Parser Excel d'abord (pour détecter le compte par contenu si nécessaire)
    import re
    supports_data = {}
    total_valorisation = 0.0

    wb = xl.load_workbook(file_path, data_only=True)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        support_name = ws.cell(row, 2).value  # Colonne B : Support
        valorisation_str = ws.cell(row, 5).value  # Colonne E : Valorisation

        if not support_name:
            continue

        support_name = str(support_name).strip()

        # Cas spécial : "SUPPORT EURO" → "SÉCURITÉ EUROS"
        if support_name == 'SUPPORT EURO':
            support_name = 'SÉCURITÉ EUROS'

        # Appliquer le mapping des noms (correspondance exacte prioritaire, puis startswith)
        if support_name in SUPPORT_NAME_MAPPING:
            support_name = SUPPORT_NAME_MAPPING[support_name]
        else:
            # Fallback: correspondance par préfixe
            for old_name, new_name in SUPPORT_NAME_MAPPING.items():
                if support_name.startswith(old_name):
                    support_name = new_name
                    break

        # Parser la valorisation
        if valorisation_str:
            valorisation_str = str(valorisation_str).strip()
            valorisation_str = re.sub(r'[\s€]', '', valorisation_str).replace(',', '.')
            try:
                valorisation = float(valorisation_str)
                # Additionner si le support existe déjà (fusion après mapping)
                if support_name in supports_data:
                    supports_data[support_name] += valorisation
                else:
                    supports_data[support_name] = valorisation
                total_valorisation += valorisation
            except ValueError:
                pass

    wb.close()

    # Détecter le compte : 1) par nom de fichier (critère principal)
    #                      2) par nombre de supports (fallback si nom ambigu)
    filename = Path(file_path).name.lower()
    if re.search(r'ebene[_\- ]?2', filename):
        compte = 'Ass vie ébène 2 Cécile'
    elif re.search(r'ebene', filename) and not re.search(r'2', filename):
        compte = 'Ass vie ébène Cécile'
    elif len(supports_data) > 10:
        # Fallback : Ébène 2 a 25+ supports, Ébène 1 en a ~5
        compte = 'Ass vie ébène 2 Cécile'
    else:
        compte = 'Ass vie ébène Cécile'

    # Calculer SYNOE si applicable (ébène 2 uniquement)
    # SYNOE = tous les supports sous mandat Synoe
    # Hors mandat = SÉCURITÉ EUROS + MOOREA FUND - HIGH YIELD (si existe)
    synoe = None
    supports_hors_mandat_names = {'SÉCURITÉ EUROS', 'MOOREA FUND - HIGH YIELD'}

    if compte == 'Ass vie ébène 2 Cécile':
        # Soustraire du total uniquement les supports présents hors mandat
        total_hors_mandat = sum(
            supports_data[name] for name in supports_hors_mandat_names
            if name in supports_data
        )
        synoe = total_valorisation - total_hors_mandat

    # Générer CSV format 4 colonnes
    output_lines = []
    date_aujourdhui = get_file_date(file_path)

    # En-tête
    output_lines.append('Date;Ligne;Montant;Compte')

    # Supports à exporter selon le compte
    if synoe is not None:
        # Ébène 2: exporter UNIQUEMENT les supports hors mandat (présents) + SYNOE
        for support_name in sorted(supports_hors_mandat_names):
            if support_name in supports_data:
                valorisation = supports_data[support_name]
                output_lines.append(f'{date_aujourdhui};{support_name};{valorisation:.2f};{compte}')
        # Ajouter SYNOE (regroupe tous les supports sous mandat)
        output_lines.append(f'{date_aujourdhui};SYNOE;{synoe:.2f};{compte}')
    else:
        # Ébène 1 (ou autres): exporter tous les supports
        for support_name in sorted(supports_data.keys()):
            valorisation = supports_data[support_name]
            output_lines.append(f'{date_aujourdhui};{support_name};{valorisation:.2f};{compte}')

    return '\n'.join(output_lines)


def process_pdf_synthese(file_path):
    """
    Parse le PDF synthèse SG "Mes comptes en ligne _ SG.pdf" pour extraire les soldes

    Ce PDF contient les soldes de tous les comptes SG.
    Génère uniquement des lignes #Solde (format 9 colonnes).

    Format extrait:
      Livret A Marc •••• 2944 20 220,40 €
      Ebene Cécile •••• 7423 111 575,31 €
    """
    if pdfplumber is None:
        print("❌ pdfplumber non installé. Exécutez: pip3 install pdfplumber", file=sys.stderr)
        sys.exit(1)

    output_lines = []
    date_aujourdhui = get_file_date(file_path)

    # Mapping des noms extraits → noms Excel
    compte_mapping = {
        'livret a marc': 'Livret A Marc',
        'livret a cécile': 'Livret A Cécile',
        'ldd marc': 'LDD Marc',
        'ldd cécile': 'LDD Cécile',
        'compte livret': 'Compte livret SG',
        'ebene cécile': 'Ass vie ébène Cécile',
        'ebene 2 cecile': 'Ass vie ébène 2 Cécile',
    }

    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            lines = text.split('\n')

            for line in lines:
                line_stripped = line.strip()

                # Pattern: "Nom compte •••• XXXX Montant €"
                # Ex: "Livret A Marc •••• 2944 20 220,40 €"
                match = re.match(r'^(.+?)\s+••••\s+\d{4}\s+([\d\s]+,\d{2})\s*€', line_stripped)
                if match:
                    nom_brut = match.group(1).strip().lower()
                    montant_str = match.group(2).replace(' ', '').replace(',', '.')

                    # Chercher le compte correspondant (clés plus longues d'abord)
                    compte = None
                    for key, nom_compte in sorted(compte_mapping.items(), key=lambda x: -len(x[0])):
                        if key in nom_brut:
                            compte = nom_compte
                            break

                    if compte:
                        try:
                            montant = float(montant_str)
                            output_lines.append(
                                f"{date_aujourdhui};Relevé compte;{montant:.2f};EUR;;;#Solde;{compte};".replace('.', ',')
                            )
                        except ValueError:
                            pass

    result = "Date;Libellé;Montant;Devise;Equiv;Réf;Catégorie;Compte;Commentaire\n"
    result += "\n".join(output_lines)

    return result


def process_pdf_assurance_vie(file_path):
    """
    Parse un PDF d'opérations assurance vie SG "Gestion des contrats - Assurance Vie.pdf"

    Format tableau fragmenté :
      - Lignes avec montant : "YYYY au cours de + 0,23 EUR Réalisé" ou "DD/MM/YYYY texte + X,XX EUR Réalisé"
      - Lignes libellé avant/après : "Intérêts crédités", "l'année 2025", etc.

    Stratégie : chercher les lignes avec montant, puis reconstituer le contexte.
    """
    if pdfplumber is None:
        print("❌ pdfplumber non installé. Exécutez: pip3 install pdfplumber", file=sys.stderr)
        sys.exit(1)

    COMMENTAIRE = ""
    output_lines = []
    operations = []

    # Détecter le compte depuis le nom de fichier
    # Patterns: "(1)" (nom original) ou "ebene2" / "ebene_2" (renommé)
    filename = Path(file_path).name.lower()
    if '(1)' in filename or 'ebene2' in filename or 'ebene_2' in filename:
        compte = 'Ass vie ébène 2 Cécile'
    else:
        compte = 'Ass vie ébène Cécile'

    with pdfplumber.open(file_path) as pdf:
        all_lines = []
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                all_lines.extend(text.split('\n'))

        if not all_lines:
            return "Date;Libellé;Montant;Devise;Equiv;Réf;Catégorie;Compte;Commentaire\n"

        # D'abord, identifier les lignes pertinentes (entre "Suivi des opérations" et footer)
        in_operations = False
        filtered_lines = []
        for line in all_lines:
            line = line.strip()
            if 'Suivi des opérations' in line:
                in_operations = True
                continue
            if 'Questions fréquentes' in line or 'https://' in line:
                in_operations = False
                continue
            if in_operations and line:
                # Ignorer header du tableau
                if line == 'Date Opération Montant Statut Origine':
                    continue
                filtered_lines.append(line)

        # Parser les opérations : chercher les lignes avec montant
        # Pattern: "+ X,XX EUR" ou "- X,XX EUR"
        i = 0
        while i < len(filtered_lines):
            line = filtered_lines[i]

            # Chercher un montant dans la ligne
            montant_match = re.search(r'([+-])\s*([\d\s]+,\d{2})\s*EUR', line)
            if montant_match:
                sign = montant_match.group(1)
                montant_str = montant_match.group(2).replace(' ', '').replace(',', '.')
                try:
                    montant = float(montant_str)
                    if sign == '-':
                        montant = -montant
                except ValueError:
                    i += 1
                    continue

                # Extraire la date et le libellé
                # La date peut être :
                # 1. En début de ligne: "31/12/2025 texte + X,XX EUR"
                # 2. Dans la partie texte: "2025 au cours de + 0,23 EUR"
                # 3. Dans une ligne précédente

                before_montant = line[:montant_match.start()].strip()

                # Chercher une date DD/MM/YYYY dans before_montant
                date_match = re.search(r'(\d{2}/\d{2}/\d{4})', before_montant)
                if date_match:
                    date_str = date_match.group(1)
                    # Libellé = partie entre date et montant
                    libelle_part = before_montant[date_match.end():].strip()
                else:
                    # Chercher une année seule (2025, 2024...)
                    year_match = re.match(r'^(\d{4})\s+(.+)', before_montant)
                    if year_match:
                        year = year_match.group(1)
                        date_str = f"31/12/{year}"  # Année → 31/12/année
                        libelle_part = year_match.group(2).strip()
                    else:
                        # Date inconnue → utiliser date du fichier (≈ date du fetch)
                        date_str = get_file_date(file_path)
                        libelle_part = before_montant

                # Récupérer le libellé complet (ligne précédente + partie courante + ligne suivante)
                libelle_parts = []

                # Ligne précédente si elle ne contient pas de montant et pas de date
                if i > 0:
                    prev_line = filtered_lines[i-1]
                    if not re.search(r'[+-]\s*[\d\s]+,\d{2}\s*EUR', prev_line):
                        if not re.match(r'^\d{2}/\d{2}/\d{4}', prev_line) and prev_line not in ('1/2', '2/2'):
                            # Vérifier que ce n'est pas un libellé d'une autre opération (pas de Réalisé)
                            if 'Réalisé' not in prev_line:
                                libelle_parts.append(prev_line)

                if libelle_part:
                    libelle_parts.append(libelle_part)

                # Ligne suivante si c'est une continuation (pas de montant, pas de date)
                if i + 1 < len(filtered_lines):
                    next_line = filtered_lines[i+1]
                    if not re.search(r'[+-]\s*[\d\s]+,\d{2}\s*EUR', next_line):
                        if not re.match(r'^\d{2}/\d{2}/\d{4}', next_line) and not re.match(r'^\d{4}$', next_line):
                            if next_line not in ('Réalisé', 'Sogecap', '1/2', '2/2') and 'Réalisé' not in next_line:
                                libelle_parts.append(next_line)

                # Construire le libellé final
                libelle = ' '.join(libelle_parts).strip()
                # Nettoyer
                libelle = re.sub(r'\s+', ' ', libelle)
                libelle = re.sub(r'\s*Réalisé.*$', '', libelle).strip()

                # Mapper vers des libellés connus
                # Aligner les libellés sur ceux du HTML (pour éviter doublons)
                if 'Intérêts crédités' in libelle or 'au cours de' in libelle:
                    year_in_libelle = re.search(r"l'année (\d{4})", libelle)
                    if year_in_libelle:
                        libelle = f"Intérêts crédités au cours de l'année {year_in_libelle.group(1)}"
                    else:
                        libelle = "Intérêts crédités"
                elif 'Participation aux bénéfices' in libelle or 'bénéfices sur le' in libelle:
                    libelle = "Participation aux bénéfices sur le(s) support(..."
                elif 'Prélèvements sociaux' in libelle or libelle == 'sociaux':
                    libelle = "Prélèvements sociaux"

                if libelle:
                    operations.append({
                        'date_str': date_str,
                        'libelle': libelle,
                        'montant': f"{montant:.2f}".replace('.', ',')
                    })

            i += 1

    # Générer les lignes formatées
    for op in operations:
        cat, opts = inc_categorize.categorize_operation(op['libelle'], SITE)
        ref = opts.get('ref', '')

        output_line = f"{op['date_str']};{op['libelle']};{op['montant']};EUR;;{ref};{cat};{compte};{COMMENTAIRE}"
        output_lines.append(output_line)

    result = "Date;Libellé;Montant;Devise;Equiv;Réf;Catégorie;Compte;Commentaire\n"
    result += "\n".join(output_lines)

    return result


def process_pdf_printed(file_path):
    """
    Parse un PDF imprimé depuis l'interface web SG (fallback collecte manuelle)

    Le PDF imprimé contient les opérations d'un ou plusieurs comptes épargne/assurance-vie.
    Format texte extrait typique:
      - Lignes avec date DD/MM/YYYY, libellé et montant
      - Soldes identifiables par "Solde" ou montant en fin de section

    Retourne le format standardisé 9 colonnes.
    """
    if pdfplumber is None:
        print("❌ pdfplumber non installé. Exécutez: pip3 install pdfplumber", file=sys.stderr)
        sys.exit(1)

    COMMENTAIRE = ""
    output_lines = []
    operations = []
    soldes_par_compte = {}

    # Mapping des noms de comptes depuis le PDF vers les noms Excel
    compte_mapping = {
        'livret a': 'Livret A',
        'ldd': 'LDD',
        'ldds': 'LDD',
        'pel': 'PEL',
        'cel': 'CEL',
        'csl': 'Compte livret SG',
        'compte sur livret': 'Compte livret SG',
    }

    current_compte = None

    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            lines = text.split('\n')

            for line in lines:
                line = line.strip()
                if not line:
                    continue

                # Détecter le nom du compte
                line_lower = line.lower()
                for key, compte_name in compte_mapping.items():
                    if key in line_lower and 'solde' not in line_lower:
                        current_compte = compte_name
                        break

                # Chercher les opérations : ligne commençant par une date DD/MM/YYYY
                date_match = re.match(r'(\d{2}/\d{2}/\d{4})\s+(.+)', line)
                if date_match and current_compte:
                    date_str = date_match.group(1)
                    rest = date_match.group(2)

                    # Extraire le montant (dernier nombre de la ligne, avec ou sans signe)
                    # Format: "1 234,56" ou "-1 234,56" ou "1234,56"
                    amount_match = re.search(r'(-?\d[\d\s]*,\d{2})\s*€?\s*$', rest)
                    if amount_match:
                        libelle = rest[:amount_match.start()].strip()
                        montant_str = amount_match.group(1).replace(' ', '').replace(',', '.')

                        try:
                            montant = float(montant_str)
                            operations.append({
                                'date_str': date_str,
                                'libelle': libelle,
                                'montant': f"{montant:.2f}".replace('.', ','),
                                'compte': current_compte
                            })
                        except ValueError:
                            pass

                # Détecter les soldes : ligne contenant "solde" et un montant
                if 'solde' in line_lower and current_compte:
                    solde_match = re.search(r'(-?\d[\d\s]*,\d{2})\s*€?\s*$', line)
                    if solde_match:
                        montant_str = solde_match.group(1).replace(' ', '').replace(',', '.')
                        try:
                            solde = float(montant_str)
                            date_aujourdhui = get_file_date(file_path)
                            soldes_par_compte[current_compte] = {
                                'date_str': date_aujourdhui,
                                'montant': f"{solde:.2f}".replace('.', ',')
                            }
                        except ValueError:
                            pass

    # Identifier les comptes qui ont des opérations
    comptes_avec_operations = set(op['compte'] for op in operations)

    # Générer les lignes formatées
    for op in operations:
        cat, opts = inc_categorize.categorize_operation(op['libelle'], SITE)
        ref = opts.get('ref', '')

        output_line = f"{op['date_str']};{op['libelle']};{op['montant']};EUR;;{ref};{cat};{op['compte']};{COMMENTAIRE}"
        output_lines.append(output_line)

    # Ajouter #Solde pour chaque compte qui a des opérations conservées
    for compte, solde_info in soldes_par_compte.items():
        if compte in comptes_avec_operations:
            solde_line = f"{solde_info['date_str']};Relevé compte;{solde_info['montant']};EUR;;;#Solde;{compte};{COMMENTAIRE}"
            output_lines.append(solde_line)

    # Construire la sortie
    result = "Date;Libellé;Montant;Devise;Equiv;Réf;Catégorie;Compte;Commentaire\n"
    result += "\n".join(output_lines)

    return result


# ============================================================================
# API POUR UPDATE - NOUVELLE INTERFACE
# ============================================================================

def _wrap_csv_to_tuples(func, num_fields):
    """Wrapper pour convertir une fonction retournant du CSV en liste de tuples."""
    def wrapper(file_path):
        output = func(file_path)
        return _parse_csv_output(output, num_fields)
    return wrapper


def format_site(site_dir, verbose=False, logger=None):
    """API pour Update.

    Traite tous les fichiers du répertoire SG:
    - CSV opérations (00*.csv, Export_*.csv, operations_*.csv)
    - XLSX supports assurance vie (Supports*.xlsx, *_supports*.xlsx, SG_Ebene*.xlsx)
    - PDF relevés (Mes comptes en ligne*.pdf, Gestion des contrats*.pdf, SG_Ebene*.pdf)
    """
    if logger is None:
        from inc_logging import Logger
        logger = Logger(SITE, verbose=verbose)

    # Vérification fichiers dropbox
    from inc_format import verify_dropbox_files
    for w in verify_dropbox_files(site_dir, SITE):
        logger.warning(w)

    # Wrappers pour convertir CSV text → tuples
    ops_wrapper = _wrap_csv_to_tuples(process_operations, 9)
    epargne_wrapper = _wrap_csv_to_tuples(process_epargne_csv, 9)
    parsed_wrapper = _wrap_csv_to_tuples(process_operations_parsed, 9)
    positions_wrapper = _wrap_csv_to_tuples(process_positions, 4)
    pdf_synthese_wrapper = _wrap_csv_to_tuples(process_pdf_synthese, 9)
    pdf_assurance_wrapper = _wrap_csv_to_tuples(process_pdf_assurance_vie, 9)

    handlers = [
        # CSV opérations
        ('00*.csv', ops_wrapper, 'ops'),
        ('Export_*.csv', epargne_wrapper, 'ops'),
        ('operations_*.csv', parsed_wrapper, 'ops'),
        # XLSX positions
        ('Supports*.xlsx', positions_wrapper, 'pos'),
        ('*_supports*.xlsx', positions_wrapper, 'pos'),
        ('SG_Ebene*.xlsx', positions_wrapper, 'pos'),
        # PDF soldes et opérations
        ('*[Mm]es comptes en ligne*.pdf', pdf_synthese_wrapper, 'ops'),
        ('*[Gg]estion des contrats*.pdf', pdf_assurance_wrapper, 'ops'),
        ('SG_[Ee]bene*.pdf', pdf_assurance_wrapper, 'ops'),
    ]

    return process_files(site_dir, handlers, verbose, SITE, logger=logger)


def _parse_csv_output(csv_output, num_fields):
    """Parse la sortie CSV texte et retourne une liste de tuples.

    Args:
        csv_output: Sortie CSV (string avec header + lignes)
        num_fields: Nombre de champs attendus (9 pour opérations, 4 pour positions)

    Returns:
        Liste de tuples (sans le header)
    """
    if not csv_output:
        return []

    lines = csv_output.strip().split('\n')
    if len(lines) < 2:
        return []

    result = []
    # Skip header (première ligne)
    for line in lines[1:]:
        if not line.strip():
            continue
        fields = line.split(';')
        # Compléter avec des chaînes vides si nécessaire
        while len(fields) < num_fields:
            fields.append('')
        result.append(tuple(fields[:num_fields]))

    return result


def log_csv_debug(operations, positions, site_dir, logger=None):
    """Wrapper vers inc_format.log_csv_debug()"""
    _log_csv_debug(SITE, operations, positions, logger)


if __name__ == '__main__':
    from inc_format import cli_main
    cli_main(format_site)
