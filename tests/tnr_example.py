#!/usr/bin/env python3
"""
tnr_example.py — TNR example (template → devises + comptes + titres + import ops)

Usage:
  python3 tests/tnr_example.py              # mode batch (défaut, 1 session UNO)
  python3 tests/tnr_example.py --legacy     # mode legacy (build-example séquentiel, ~5min)
  python3 tests/tnr_example.py --keep       # debug (pas de restore)

Part du template expected, construit l'exemple complet (devises, comptes, titres),
importe les opérations via cpt_update, compare vs expected.
"""

import argparse
import os
import shutil
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

_name = Path(__file__).stem.removeprefix('tnr_')
SCENARIO_DIR = Path(__file__).parent / 'tnr' / _name
EXPECTED = SCENARIO_DIR / 'expected.xlsm'
RESULT = SCENARIO_DIR / 'result.xlsm'

sys.path.insert(0, str(Path(__file__).parent))
import tnr_lib
from tnr_lib import (
    find_code_root, timestamp, check_libreoffice_running,
    compare_result, save_result, check_integrity_fast,
    setup_sandbox,
)

CODE_ROOT = find_code_root(__file__)
sys.path.insert(0, str(CODE_ROOT))
from inc_excel_schema import SHEET_OPERATIONS, SHEET_AVOIRS, ColResolver, uno_row
from tool_gui_cli import HeadlessGUI, TEST_DEVISES, TEST_COTATIONS_SPOT, TEST_ACCOUNTS, TEST_TITLES

_D = datetime  # raccourci pour les dates


def _load_biens_materiels():
    """Charge les biens matériels depuis la ref (Avoirs, type='Biens matériels')."""
    import openpyxl as _xl
    wb = _xl.load_workbook(EXPECTED, data_only=True)
    cr = ColResolver.from_openpyxl(wb)
    ws = wb['Avoirs']
    biens = []
    for r in range(4, 40):
        acct_type = ws.cell(r, cr.col('AVRtype')).value
        if not acct_type or str(acct_type).strip() != 'Biens matériels':
            continue
        intitule = str(ws.cell(r, cr.col('AVRintitulé')).value or '').strip()
        if intitule == 'Pièces or':
            continue  # déjà dans TEST_ACCOUNTS (géré par GUI)
        biens.append({
            'intitule': intitule,
            'domiciliation': str(ws.cell(r, cr.col('AVRdomiciliation')).value or '').strip(),
            'sous_type': str(ws.cell(r, cr.col('AVRsous_type')).value or '').strip(),
            'titulaire': str(ws.cell(r, cr.col('AVRtitulaire')).value or '').strip(),
            'propriete': str(ws.cell(r, cr.col('AVRpropriete')).value or '').strip(),
            'date_anter': ws.cell(r, cr.col('AVRdate_anter')).value,
            'montant_anter': ws.cell(r, cr.col('AVRmontant_anter')).value,
            'date_solde': ws.cell(r, cr.col('AVRdate_solde')).value,
            'montant_solde': ws.cell(r, cr.col('AVRmontant_solde')).value,
        })
    wb.close()
    return biens


def _load_cours_ref():
    """Charge les cours spot depuis la ref (Cotations)."""
    import openpyxl as _xl
    wb = _xl.load_workbook(EXPECTED, data_only=True)
    cr = ColResolver.from_openpyxl(wb)
    ws = wb['Cotations']
    cot_start, _ = cr.rows('COTcode')
    cot_start = cot_start or 3
    cours = {}
    for r in range(cot_start + 1, cot_start + 30):
        code = ws.cell(r, cr.col('COTcode')).value
        if not code:
            continue
        val = ws.cell(r, cr.col('COTcours')).value
        if isinstance(val, (int, float)):
            cours[str(code).strip()] = val
    wb.close()
    return cours


def _load_pvl_titres():
    """Charge date_solde + solde des lignes titre PVL depuis la ref."""
    import openpyxl as _xl
    wb = _xl.load_workbook(EXPECTED, data_only=True)
    cr = ColResolver.from_openpyxl(wb)
    ws = wb['Plus_value']
    titres = {}
    for r in range(4, 50):
        c = ws.cell(r, cr.col('PVLtitre')).value
        if c and str(c).startswith('*'):
            titre = str(c).strip()
            titres[titre] = {
                'date_init': ws.cell(r, cr.col('PVLdate_init')).value,
                'montant_init': ws.cell(r, cr.col('PVLmontant_init')).value,
                'date_solde': ws.cell(r, cr.col('PVLdate')).value,
                'solde': ws.cell(r, cr.col('PVLmontant')).value,
            }
    wb.close()
    return titres


def _check_totals_vs_ref():
    """Compare les totaux Patrimoine et PVL du result vs expected.

    Lit la ref via UNO (recalcul des formules) pour avoir les bonnes valeurs.
    """
    import openpyxl as _xl
    from inc_uno import UnoDocument

    # Ref via UNO (formules recalculées)
    from inc_uno import col_of
    ref_totals = {}
    with UnoDocument(EXPECTED, read_only=True) as doc:
        doc.calculate_all()
        xdoc = doc.document
        for sheet, label, label_nr, val_nr in [
            ('Patrimoine', 'TOTAL', 'PATlabel', 'PATvaleur'),
            ('Plus_value', 'GRAND TOTAL', 'PVLsection', 'PVLpvl'),
        ]:
            ws = doc.get_sheet(sheet)
            lc = col_of(xdoc, label_nr)
            vc = col_of(xdoc, val_nr)
            for r0 in range(1, 60):
                b = ws.getCellByPosition(lc, r0).getString()
                if label in b.upper():
                    ref_totals[(sheet, label)] = ws.getCellByPosition(vc, r0).getValue()
                    break

    # Result via openpyxl (valeurs cachées OK car recalculé par le batch)
    wb_res = _xl.load_workbook(tnr_lib.COMPTES_XLSX, data_only=True)
    cr_res = ColResolver.from_openpyxl(wb_res)

    ok = True
    for sheet, label, label_nr, val_nr in [
        ('Patrimoine', 'TOTAL', 'PATlabel', 'PATvaleur'),
        ('Plus_value', 'GRAND TOTAL', 'PVLsection', 'PVLpvl'),
    ]:
        ref_val = ref_totals.get((sheet, label))
        res_val = None
        ws = wb_res[sheet]
        for r in range(2, 100):
            b = ws.cell(r, cr_res.col(label_nr)).value
            if b and label in str(b).upper():
                res_val = ws.cell(r, cr_res.col(val_nr)).value
                break

        if ref_val is None or res_val is None:
            print(f"  ❌ {sheet} {label}: introuvable (ref={ref_val}, result={res_val})")
            ok = False
        elif abs(ref_val - res_val) > 0.01:
            print(f"  ❌ {sheet} {label}: ref={ref_val:.2f}, result={res_val:.2f}, delta={res_val - ref_val:+.2f}")
            ok = False
        else:
            print(f"  ✓ {sheet} {label}: {res_val:.2f}")

    wb_res.close()
    return ok


def setup_legacy():
    """Mode legacy : build-example via subprocess (séquentiel, ~5min)."""
    SCRIPT_DIR = tnr_lib.SCRIPT_DIR
    COMPTES_XLSX = tnr_lib.COMPTES_XLSX
    print("🔧 Build-example (legacy)...")
    build_example = SCRIPT_DIR / 'tool_gui_cli.py'
    env = {**os.environ, 'COMPTA_BASE_DIR': str(SCRIPT_DIR)}
    result = subprocess.run(
        [sys.executable, str(build_example), str(COMPTES_XLSX), 'build-example',
         '--source', str(COMPTES_XLSX)],
        cwd=SCRIPT_DIR, env=env, capture_output=True, text=True)
    if result.stdout:
        print(result.stdout)
    if result.returncode != 0:
        print(f"❌ build-example a échoué")
        if result.stderr:
            print(result.stderr)
        return False
    return True


def setup_batch():
    """Mode batch : template expected → devises + comptes + titres en 1 session UNO."""
    SCRIPT_DIR = tnr_lib.SCRIPT_DIR
    COMPTES_XLSX = tnr_lib.COMPTES_XLSX
    TEMPLATE = SCRIPT_DIR / 'comptes_template.xlsm'

    if not TEMPLATE.exists():
        print(f"❌ Template introuvable : {TEMPLATE}")
        print(f"   Fichier attendu : {TEMPLATE}")
        return False

    # Copier le template comme base
    shutil.copy2(TEMPLATE, COMPTES_XLSX)
    print(f"   {TEMPLATE.name} → comptes.xlsm")

    # Batch UNO : toutes les opérations en une session
    gui = HeadlessGUI(COMPTES_XLSX)
    errors = 0

    # Catégories et postes nécessaires à l'exemple (retirés du template)
    EXAMPLE_POSTES = ['Produits financiers', 'Loisirs et restaurant', 'Santé', 'Maison']
    EXAMPLE_CATEGORIES = [
        ('Ajustement', 'Divers'),
        ('Coupon', 'Produits financiers'),
        ('Frais bancaires', 'Divers'),
        ('Intérêts', 'Produits financiers'),
        ('Loisirs', 'Loisirs et restaurant'),
        ('Maison', 'Maison'),
        ('Restaurant', 'Loisirs et restaurant'),
        ('Santé', 'Santé'),
    ]

    with gui.batch() as doc:
        # Postes et catégories d'abord (les catégories référencent les postes)
        print(f"🔧 Ajout de {len(EXAMPLE_POSTES)} postes...")
        for poste in EXAMPLE_POSTES:
            if not gui.add_poste(poste, doc=doc):
                errors += 1

        print(f"🔧 Ajout de {len(EXAMPLE_CATEGORIES)} catégories...")
        for cat_name, cat_poste in EXAMPLE_CATEGORIES:
            if not gui.add_category(cat_name, poste=cat_poste, doc=doc):
                errors += 1

        # Cotations spot d'abord (XAU, BTC) — référence pour les dérivées
        print(f"🔧 Ajout de {len(TEST_COTATIONS_SPOT)} cotations spot...")
        for code, famille, nom, df, fm in TEST_COTATIONS_SPOT:
            if not gui.add_devise(code, famille, nom=nom, derived_from=df, formula=fm, doc=doc):
                errors += 1

        print(f"🔧 Ajout de {len(TEST_DEVISES)} devises...")
        for code, famille, nom, df, fm in TEST_DEVISES:
            if not gui.add_devise(code, famille, nom=nom, derived_from=df, formula=fm, doc=doc):
                errors += 1

        print(f"🔧 Ajout de {len(TEST_ACCOUNTS)} comptes...")
        for intitule, acct_type, devise, dom, st, tit, prop, da, ma in TEST_ACCOUNTS:
            ctrl = not (acct_type == 'Portefeuilles' and st == 'Titres')
            if not gui.add_account(intitule, acct_type, devise=devise,
                                   sous_type=st, domiciliation=dom,
                                   titulaire=tit, propriete=prop,
                                   date_anter=da, montant_anter=ma,
                                   controle=ctrl, doc=doc):
                errors += 1

        # Biens matériels via API add_account (avant _save_accounts, accumulés en batch)
        biens_materiels = _load_biens_materiels()
        print(f"🔧 Ajout de {len(biens_materiels)} biens matériels...")
        for b in biens_materiels:
            if not gui.add_account(
                b['intitule'], 'Biens matériels',
                devise='', sous_type=b['sous_type'],
                domiciliation=b['domiciliation'],
                titulaire=b['titulaire'], propriete=b['propriete'],
                date_anter=b['date_anter'], montant_anter=b['montant_anter'],
                date_solde=b['date_solde'],
                montant_debut=b['montant_solde'],
                controle=False, doc=doc):
                errors += 1

        # Écrire tous les comptes (normaux + biens) en UNO avant les titres
        gui._save_accounts(doc=doc)

        print(f"🔧 Ajout de {len(TEST_TITLES)} titres...")
        for compte, titre, devise in TEST_TITLES:
            if not gui.add_title(compte, titre, devise=devise, doc=doc):
                errors += 1

        # Écrire les cours depuis expected (pas d'API en TNR_MODE)
        from inc_excel_schema import SHEET_COTATIONS
        from inc_uno import get_col_range_bounds
        cot_bounds = get_col_range_bounds(doc.document, 'COTcode')
        cot_data = (cot_bounds[2] if cot_bounds else 3) + 1
        cr = doc.cr
        cours_ref = _load_cours_ref()
        ws_cot = doc.get_sheet(SHEET_COTATIONS)
        n_cours = 0
        for r0 in range(uno_row(cot_data), uno_row(cot_data) + 20):
            code = ws_cot.getCellByPosition(cr.col('COTcode'), r0).getString().strip()
            if not code:
                continue
            if code in cours_ref and cours_ref[code] is not None:
                cell = ws_cot.getCellByPosition(cr.col('COTcours'), r0)
                # Préserver les formules de dérivation
                formula = cell.getFormula()
                if formula and str(formula).startswith('='):
                    continue
                cell.setValue(cours_ref[code])
                n_cours += 1
        print(f"   {n_cours} cours écrits depuis expected")

        # Écrire les soldes PVL titres depuis expected (date_solde + solde)
        pvl_ref = _load_pvl_titres()
        if pvl_ref:
            from inc_excel_schema import SHEET_PLUS_VALUE
            ws_pvl = doc.get_sheet(SHEET_PLUS_VALUE)
            n_pvl = 0
            epoch = _D(1899, 12, 30)
            for r0 in range(3, 50):
                titre = ws_pvl.getCellByPosition(cr.col('PVLtitre'), r0).getString().strip()
                if titre and titre in pvl_ref:
                    d = pvl_ref[titre]
                    for col_name, val in [('PVLdate_init', d['date_init']),
                                          ('PVLmontant_init', d['montant_init']),
                                          ('PVLdate', d['date_solde']),
                                          ('PVLmontant', d['solde'])]:
                        if val is None:
                            continue
                        cell = ws_pvl.getCellByPosition(cr.col(col_name), r0)
                        if hasattr(val, 'toordinal'):
                            cell.setValue((val - epoch).days)
                        elif isinstance(val, (int, float)):
                            cell.setValue(val)
                    n_pvl += 1
            print(f"   {n_pvl} soldes PVL titres écrits depuis expected")

        # Supprimer les #Solde placeholder (bloc 1 à 2020) et injecter les
        # opérations de expected — dans le batch pour le bon format UNO.
        ws_ops = doc.get_sheet(SHEET_OPERATIONS)

        # 1. Supprimer les #Solde placeholder en ordre inverse
        deleted = 0
        for r0 in range(100, 2, -1):
            cat = ws_ops.getCellByPosition(cr.col('OPcatégorie'), r0).getString()
            if cat == '#Solde':
                ws_ops.Rows.removeByIndex(r0, 1)
                deleted += 1
        print(f"   {deleted} #Solde placeholder supprimés")

        # 2. Charger expected (data_only pour ops, formules pour Patrimoine/Cotations)
        import openpyxl as _xl
        wb_ref = _xl.load_workbook(EXPECTED, data_only=True)
        ws_ref = wb_ref[SHEET_OPERATIONS]
        ops_data = []
        for r in range(4, ws_ref.max_row + 1):
            row = [ws_ref.cell(r, c).value for c in range(1, 10)]
            if not row[0]:
                continue  # ligne vide
            if str(row[0]).strip() in ('✓', '⚓'):
                continue  # model row / ancre
            ops_data.append(row)
        wb_ref.close()

        # Append après la dernière donnée (pas de end OP)
        from inc_uno import copy_row_style
        cursor = ws_ops.createCursor()
        cursor.gotoEndOfUsedArea(True)
        last_0 = cursor.getRangeAddress().EndRow
        # Remonter jusqu'à une ligne non-vide (ignorer ✓ et vides)
        while last_0 > 3 and not ws_ops.getCellByPosition(cr.col('OPlibellé'), last_0).getString():
            last_0 -= 1
        next_0 = last_0 + 1
        template_0 = max(last_0, 3)  # model row pour le style

        # Mapping index dans row_data → named range
        OP_COLS = ['OPdate', 'OPlibellé', 'OPmontant', 'OPdevise',
                   'OPequiv_euro', 'OPréf', 'OPcatégorie', 'OPcompte', 'OPcommentaire']

        from datetime import datetime as _dt
        epoch = _dt(1899, 12, 30)
        for row_data in ops_data:
            copy_row_style(ws_ops, template_0, next_0, col_start=0, col_end=9)
            for i, col_name in enumerate(OP_COLS):
                val = row_data[i]
                if val is None:
                    continue
                cell = ws_ops.getCellByPosition(cr.col(col_name), next_0)
                if i == 0:  # Date
                    if hasattr(val, 'toordinal'):
                        cell.setValue((val - epoch).days)
                    else:
                        cell.setValue(float(val))
                elif i in (2, 4):  # Montant, Equiv
                    try:
                        cell.setValue(float(val))
                    except (ValueError, TypeError):
                        pass
                else:  # String columns
                    cell.setString(str(val or ''))
            next_0 += 1

        # Appliquer formats devise + fond gris pour les lignes non-EUR
        from inc_formats import formats_devise_uno, FORMAT_EUR, GRIS_BLANC
        fmt_cache = {}
        for devise_code, fmt_str in formats_devise_uno(doc.document).items():
            fmt_cache[devise_code] = doc.register_number_format(fmt_str)
        fmt_eur = doc.register_number_format(FORMAT_EUR)

        op_start_0 = next_0 - len(ops_data)  # première ligne injectée (0-indexed)
        for r0 in range(op_start_0, next_0):
            devise = ws_ops.getCellByPosition(cr.col('OPdevise'), r0).getString().strip()
            if not devise:
                continue
            expected_fmt = fmt_cache.get(devise, fmt_eur)
            ws_ops.getCellByPosition(cr.col('OPmontant'), r0).NumberFormat = expected_fmt
            ws_ops.getCellByPosition(cr.col('OPequiv_euro'), r0).NumberFormat = fmt_eur
            if devise != 'EUR':
                ws_ops.getCellByPosition(cr.col('OPmontant'), r0).CellBackColor = GRIS_BLANC

        print(f"   {len(ops_data)} opérations injectées depuis {EXPECTED.name}")
        wb_ref.close()

    if errors:
        print(f"❌ {errors} erreur(s) pendant le build")
        return False
    print("✓ Build-example batch terminé")
    return True




def main():
    parser = argparse.ArgumentParser(description='TNR example')
    parser.add_argument('--batch', action='store_true', default=True,
                        help='Mode batch (1 session UNO, plus rapide) — défaut')
    parser.add_argument('--legacy', action='store_true',
                        help='Mode legacy (build-example séquentiel, ~5min)')
    args = parser.parse_args()

    mode = 'legacy' if args.legacy else 'batch'
    success = True
    t_start = time.time()

    if not check_libreoffice_running():
        return 1

    print(f"\n{timestamp()} Setup sandbox")
    sandbox = setup_sandbox(SCENARIO_DIR)
    tnr_lib.set_base_dir(sandbox)
    SCRIPT_DIR = tnr_lib.SCRIPT_DIR
    COMPTES_XLSX = tnr_lib.COMPTES_XLSX
    print(f"  sandbox : {sandbox}")

    # Mode Classeur : vider config_pipeline.json dans la sandbox
    import json
    pipeline_json = sandbox / 'config_pipeline.json'
    if pipeline_json.exists():
        with open(pipeline_json, 'w') as f:
            json.dump({"linked_operations": {}}, f, indent=2)

    # tnr_example simule un démarrage from scratch : pas de config_cotations
    # utilisateur (sinon add_devise refuse les codes déjà présents, cf. 16baddd)
    (sandbox / 'config_cotations.json').unlink(missing_ok=True)

    print(f"\n{timestamp()} Build example ({mode})")
    if mode == 'batch':
        if not setup_batch():
            success = False
            return 1
    else:
        if not setup_legacy():
            success = False
            return 1

    if success:
        print(f"\n{timestamp()} Vérification intégrité")
        if not check_integrity_fast(COMPTES_XLSX):
            success = False

    if success:
        print(f"\n{timestamp()} Normalisation formats")
        # L'exemple est un livrable (install.sh) : on applique les corrections
        # de format pour garantir un rendu propre et cohérent.
        from tool_fix_formats import fix_formats
        from io import StringIO
        import contextlib
        buf = StringIO()
        with contextlib.redirect_stdout(buf):
            fix_formats(COMPTES_XLSX, apply=True)
        output = buf.getvalue()
        import re
        m = re.search(r'(\d+) correction', output)
        n_fixes = int(m.group(1)) if m else 0
        if n_fixes == 0:
            print("  ✓ Formats OK")
        else:
            print(f"  ✓ {n_fixes} correction(s) de format appliquée(s)")

    if success:
        print(f"\n{timestamp()} Contrôles métier (tool_controles)")
        ctrl_script = SCRIPT_DIR / 'tool_controles.py'
        env = {**os.environ, 'COMPTA_BASE_DIR': str(SCRIPT_DIR)}
        r = subprocess.run(
            [sys.executable, str(ctrl_script), '-f', str(COMPTES_XLSX)],
            cwd=SCRIPT_DIR, env=env, capture_output=True, text=True)
        if r.stdout:
            for line in r.stdout.strip().splitlines():
                print(f"  {line}")
        if r.returncode != 0:
            print(f"  ❌ tool_controles a retourné {r.returncode}")
            if r.stderr:
                print(r.stderr, file=sys.stderr)
            success = False

    if success and EXPECTED.exists():
        print(f"\n{timestamp()} Comparaison vs expected")
        cmp_result = compare_result(EXPECTED, tuples=True, brutal=False)
        if cmp_result is False:
            success = False
    elif success:
        print(f"\n{timestamp()} ⚠ Pas d'expected — valider puis:")
        print(f"   cp {RESULT} {EXPECTED}")

    if success:
        print(f"\n{timestamp()} Comparaison totaux vs expected")
        if not _check_totals_vs_ref():
            success = False

    print(f"\n{timestamp()} Sauvegarde résultat (hors sandbox pour archivage)")
    save_result(RESULT)
    # Plus de restore_context : la sandbox isole intrinsèquement le test.

    elapsed = time.time() - t_start
    print()
    if success:
        print(f"✅ TEST RÉUSSI ({elapsed:.0f}s)")
        return 0
    else:
        print(f"❌ TEST ÉCHOUÉ ({elapsed:.0f}s)")
        return 1


if __name__ == '__main__':
    sys.exit(main())
