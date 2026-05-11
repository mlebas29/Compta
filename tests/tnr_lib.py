"""
tnr_lib.py — Fonctions partagées pour les TNR (tests de non-régression)

Chaque scénario (tnr_pipe.py, tnr_gui.py, tnr_fast.py) importe ce module
pour les opérations communes : backup/restore, setup, exécution, comparaison.
"""

import configparser
import filecmp
import os
import shutil
import subprocess
import sys
import time
from pathlib import Path

# SCRIPT_DIR = racine du code applicatif. Convention :
#   tests/ vit soit à la racine projet, soit sous custom/ (overlay privé).
#   Le code public est toujours à la racine projet.
SCRIPT_DIR = Path(__file__).parent.parent
if SCRIPT_DIR.name == 'custom':
    SCRIPT_DIR = SCRIPT_DIR.parent
sys.path.insert(0, str(SCRIPT_DIR))
from inc_excel_schema import SHEET_OPERATIONS, ColResolver
from inc_compare_xlsx import compare_xlsx as do_compare_xlsx

# Répertoires projet
COMPTES_XLSX = SCRIPT_DIR / 'comptes.xlsm'
DROPBOX_DIR = SCRIPT_DIR / 'dropbox'
ARCHIVES_DIR = SCRIPT_DIR / 'archives'
JOURNAL_LOG = SCRIPT_DIR / 'logs' / 'journal.log'
LOCK_FILE = SCRIPT_DIR / '.~lock.comptes.xlsm#'

# Configs projet
CATEGORY_MAPPINGS_PY = SCRIPT_DIR / 'inc_category_mappings.py'
CATEGORY_MAPPINGS_JSON = SCRIPT_DIR / 'config_category_mappings.json'
COTATIONS_JSON = SCRIPT_DIR / 'config_cotations.json'
ACCOUNTS_JSON = SCRIPT_DIR / 'config_accounts.json'
PIPELINE_JSON = SCRIPT_DIR / 'config_pipeline.json'
CONFIG_INI = SCRIPT_DIR / 'config.ini'

# Config.ini (pour résoudre dossier par site)
_config = configparser.ConfigParser()
_config.read(SCRIPT_DIR / 'config.ini')


def timestamp():
    """Retourne l'heure courante formatée HH:MM:SS."""
    return time.strftime('%H:%M:%S')


def find_code_root(test_file):
    """Retourne la racine du code applicatif depuis le chemin d'un script TNR.

    Convention : tests/ vit soit à la racine projet, soit sous custom/ (overlay).
    Le code applicatif (cpt_*, gui_*, inc_*) est toujours à la racine projet.

    Exemples :
      ~/Compta/dev/tests/tnr_X.py        → ~/Compta/dev/        (DEV public)
      ~/Compta/dev/custom/tests/tnr_X.py → ~/Compta/dev/        (DEV privé)
      ~/Compta/tests/tnr_X.py            → ~/Compta/            (PROD public)
      ~/Compta/custom/tests/tnr_X.py     → ~/Compta/            (PROD privé)
    """
    p = Path(test_file).parent.parent  # tests/X.py → tests/.. = candidat code_root
    if p.name == 'custom':
        p = p.parent
    return p


def set_base_dir(new_base):
    """Bascule SCRIPT_DIR et toutes les variables dérivées sur une nouvelle racine.

    Permet aux TNR de pointer tnr_lib sur leur sandbox au lieu de DEV.
    À appeler juste après setup_sandbox() : aucun backup/restore nécessaire,
    la sandbox isole le test du DEV.

    Args:
        new_base : Path vers la nouvelle racine (typiquement la sandbox).
    """
    global SCRIPT_DIR, COMPTES_XLSX, DROPBOX_DIR, ARCHIVES_DIR, JOURNAL_LOG
    global LOCK_FILE, CATEGORY_MAPPINGS_PY, CATEGORY_MAPPINGS_JSON
    global COTATIONS_JSON, ACCOUNTS_JSON, PIPELINE_JSON, CONFIG_INI
    SCRIPT_DIR = Path(new_base)
    COMPTES_XLSX = SCRIPT_DIR / 'comptes.xlsm'
    DROPBOX_DIR = SCRIPT_DIR / 'dropbox'
    ARCHIVES_DIR = SCRIPT_DIR / 'archives'
    JOURNAL_LOG = SCRIPT_DIR / 'logs' / 'journal.log'
    LOCK_FILE = SCRIPT_DIR / '.~lock.comptes.xlsm#'
    CATEGORY_MAPPINGS_PY = SCRIPT_DIR / 'inc_category_mappings.py'
    CATEGORY_MAPPINGS_JSON = SCRIPT_DIR / 'config_category_mappings.json'
    COTATIONS_JSON = SCRIPT_DIR / 'config_cotations.json'
    ACCOUNTS_JSON = SCRIPT_DIR / 'config_accounts.json'
    PIPELINE_JSON = SCRIPT_DIR / 'config_pipeline.json'
    CONFIG_INI = SCRIPT_DIR / 'config.ini'


def setup_sandbox(scenario_dir, extra_copies=None, code_root=None):
    """Crée et peuple la sandbox d'un scénario TNR.

    Modèle : code apporté par symlink vers DEV (modifs code immédiatement testables),
    données et configs copiées (DEV jamais touché).

    Symlinks créés :
      sandbox/<*.py>     → DEV/<*.py>     (tous les modules Python racine)
      sandbox/custom/    → DEV/custom/    (overlay privé)

    Copies créées (depuis DEV) :
      sandbox/config*.json    : configs runtime
      sandbox/config.ini      : config principale
      sandbox/comptes_template.xlsm : template du classeur

    Args:
        scenario_dir : Path du dossier scénario (ex: tests/tnr/roundtrip).
        extra_copies : dict {Path source -> nom dans sandbox} pour fichiers
                       additionnels (ex: comptes.xlsm initial).
        code_root    : racine du code applicatif (par défaut : find_code_root
                       depuis scenario_dir, supporte DEV/PROD × public/custom).

    Returns:
        Path du dossier sandbox (à pointer via COMPTA_BASE_DIR si nécessaire).
    """
    scenario_dir = Path(scenario_dir)
    sandbox = scenario_dir / 'sandbox'

    # Reset complet pour idempotence
    if sandbox.exists():
        shutil.rmtree(sandbox)
    sandbox.mkdir(parents=True)

    if code_root is None:
        # scenario_dir = <root>/tests/tnr/<X>/  ou  <root>/custom/tests/tnr/<X>/
        # 3 niveaux pour atteindre <root> (ou <root>/custom) puis dépolluer custom.
        code_root = scenario_dir.parent.parent.parent
        if code_root.name == 'custom':
            code_root = code_root.parent

    # Symlinks code (tous les .py racine + dossier custom)
    for py_file in code_root.glob('*.py'):
        (sandbox / py_file.name).symlink_to(py_file)
    custom_dir = code_root / 'custom'
    if custom_dir.exists():
        (sandbox / 'custom').symlink_to(custom_dir)

    # Copies configs (sources de vérité = code_root)
    for cfg in code_root.glob('config*.json'):
        shutil.copy2(cfg, sandbox / cfg.name)
    config_ini = code_root / 'config.ini'
    if config_ini.exists():
        shutil.copy2(config_ini, sandbox / 'config.ini')

    # Copie template
    template = code_root / 'comptes_template.xlsm'
    if template.exists():
        shutil.copy2(template, sandbox / 'comptes_template.xlsm')

    # Copies additionnelles spécifiques au scénario
    if extra_copies:
        for src, dst in extra_copies.items():
            src = Path(src)
            if src.exists():
                if src.is_dir():
                    shutil.copytree(src, sandbox / dst)
                else:
                    shutil.copy2(src, sandbox / dst)

    return sandbox


def check_libreoffice_running():
    """Vérifie si LibreOffice est en cours d'exécution (bloquant pour le TNR)."""
    try:
        result = subprocess.run(
            ['pgrep', '-x', 'soffice.bin'],
            capture_output=True, text=True
        )
        if result.returncode == 0:
            print("❌ LibreOffice est en cours d'exécution !")
            print("   Ferme LibreOffice avant de lancer le TNR.")
            return False
    except FileNotFoundError:
        pass
    return True


def setup_dropbox(dropbox_src):
    """Copie le contenu de dropbox_src/ vers dropbox/ (projet).

    Chaque sous-dossier de dropbox_src est copié vers dropbox/{dossier},
    où {dossier} est résolu via config.ini (fallback = nom du sous-dossier).
    """
    dropbox_src = Path(dropbox_src)
    if not dropbox_src.exists():
        print(f"❌ Répertoire introuvable: {dropbox_src}")
        return False

    file_count = 0
    for site_dir in dropbox_src.iterdir():
        if not site_dir.is_dir():
            continue

        dossier = _config.get(site_dir.name, 'dossier', fallback=site_dir.name)
        dest_dir = DROPBOX_DIR / dossier
        # Vider le répertoire destination avant copie
        if dest_dir.exists():
            for old_file in dest_dir.iterdir():
                if old_file.is_file():
                    old_file.unlink()
        dest_dir.mkdir(parents=True, exist_ok=True)

        for file in site_dir.iterdir():
            if file.is_file():
                shutil.copy2(file, dest_dir / file.name)
                file_count += 1

    print(f"   {file_count} fichiers copiés vers dropbox/")
    return True


def setup_input_xlsm(input_xlsm):
    """Copie le fichier xlsm d'entrée vers comptes.xlsm."""
    input_path = Path(input_xlsm)
    if not input_path.exists():
        print(f"❌ Fichier introuvable: {input_path}")
        return False
    shutil.copy2(input_path, COMPTES_XLSX)
    print(f"   {input_path.name} → comptes.xlsm")
    return True


def run_cpt_update(flags):
    """Exécute cpt_update.py avec les flags donnés.

    Args:
        flags: liste de flags, ex: ['--no-pair', '--TNR', '--all-soldes']
    Returns:
        True si succès
    """
    cpt_update = SCRIPT_DIR / 'cpt_update.py'
    cmd = [sys.executable, str(cpt_update)] + flags
    env = {**os.environ, 'COMPTA_BASE_DIR': str(SCRIPT_DIR)}
    print(f"🚀 Exécution de cpt_update.py {' '.join(flags)}...")
    result = subprocess.run(cmd, cwd=SCRIPT_DIR, env=env, capture_output=True, text=True)
    if result.stdout:
        print(result.stdout)
    if result.stderr:
        print(result.stderr, file=sys.stderr)
    if result.returncode != 0:
        print(f"❌ cpt_update.py a retourné {result.returncode}")
        return False
    print("✓ Import terminé")
    return True


def run_cpt_pair():
    """Exécute cpt_pair.py (appariement)."""
    print("🚀 Exécution de cpt_pair.py...")
    cpt_pair = SCRIPT_DIR / 'cpt_pair.py'
    env = {**os.environ, 'COMPTA_BASE_DIR': str(SCRIPT_DIR)}
    result = subprocess.run(
        [sys.executable, str(cpt_pair)], cwd=SCRIPT_DIR, env=env,
        capture_output=True, text=True)
    if result.stdout:
        print(result.stdout)
    if result.stderr:
        print(result.stderr, file=sys.stderr)
    if result.returncode != 0:
        print(f"❌ cpt_pair.py a retourné {result.returncode}")
        return False
    return True


def apply_patches(patches_dir=None):
    """Applique les patches depuis la feuille Patch d'un fichier xlsx.

    Chaque ligne identifie une opération par (date, libellé regex, montant, devise, compte)
    et modifie sa Réf dans comptes.xlsm. Pas de Patch → rien à faire.

    Args:
        patches_dir: dossier contenant le xlsx avec la feuille Patch.
                     Si None, cherche dans dropbox/MANUEL/.
    """
    import re
    import openpyxl
    from datetime import datetime

    search_dir = Path(patches_dir) if patches_dir else DROPBOX_DIR / 'MANUEL'
    if not search_dir.exists():
        return
    xlsx_files = list(search_dir.glob('*.xlsx'))
    if not xlsx_files:
        return
    manuel_path = xlsx_files[0]

    wb_manuel = openpyxl.load_workbook(manuel_path, data_only=True)
    if 'Patch' not in wb_manuel.sheetnames:
        wb_manuel.close()
        return

    ws_patch = wb_manuel['Patch']
    patches = []
    for row in ws_patch.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        date_val, libelle_pattern, montant, devise, compte, new_ref = row[:6]
        new_equiv = row[6] if len(row) > 6 else None
        patches.append({
            'date': date_val,
            'libelle_re': re.compile(libelle_pattern, re.IGNORECASE) if libelle_pattern else None,
            'montant': float(montant) if montant is not None else None,
            'devise': str(devise).strip() if devise else None,
            'compte': str(compte).strip() if compte else None,
            'new_ref': str(new_ref).strip() if new_ref else '',
            'new_equiv': float(new_equiv) if new_equiv is not None else None,
        })
    wb_manuel.close()

    if not patches:
        return

    wb = openpyxl.load_workbook(COMPTES_XLSX, keep_vba=True)
    cr = ColResolver.from_openpyxl(wb)
    ws = wb[SHEET_OPERATIONS]
    applied = 0

    for row_idx in range(4, ws.max_row + 1):
        cell_date = ws.cell(row_idx, cr.col('OPdate')).value
        cell_libelle = ws.cell(row_idx, cr.col('OPlibellé')).value or ''
        cell_montant = ws.cell(row_idx, cr.col('OPmontant')).value
        cell_devise = ws.cell(row_idx, cr.col('OPdevise')).value or ''
        cell_compte = ws.cell(row_idx, cr.col('OPcompte')).value or ''

        for p in patches:
            if p.get('_applied'):
                continue
            if p['date'] and isinstance(cell_date, datetime):
                if cell_date.date() != p['date'].date():
                    continue
            if p['montant'] is not None and cell_montant is not None:
                if abs(float(cell_montant) - p['montant']) > 0.01:
                    continue
            if p['devise'] and str(cell_devise).strip() != p['devise']:
                continue
            if p['compte'] and str(cell_compte).strip() != p['compte']:
                continue
            if p['libelle_re'] and not p['libelle_re'].search(str(cell_libelle)):
                continue

            ws.cell(row_idx, cr.col('OPréf')).value = p['new_ref'] if p['new_ref'] else None
            if p['new_equiv'] is not None:
                ws.cell(row_idx, cr.col('OPequiv_euro')).value = p['new_equiv']
            p['_applied'] = True
            applied += 1
            print(f"  Patch L{row_idx}: Réf → {p['new_ref']!r} ({cell_libelle[:50]})")
            break

    wb.save(COMPTES_XLSX)
    wb.close()

    not_applied = [p for p in patches if not p.get('_applied')]
    if not_applied:
        for p in not_applied:
            print(f"  ⚠ Patch non appliqué: {p['date']} {p['montant']} {p['devise']} {p['compte']}")

    print(f"🔧 {applied}/{len(patches)} patch(es) appliqué(s)")


def compare_result(expected_path, tuples=True, brutal=True, skip_sheets=None):
    """Compare comptes.xlsm vs expected via tool_compare_xlsx."""
    expected_path = Path(expected_path)
    if not expected_path.exists():
        print(f"⚠ Pas de référence — valider manuellement puis:")
        print(f"   cp <result> {expected_path}")
        return None  # pas d'expected → pas d'échec, pas de succès

    import openpyxl as _xl
    _wb = _xl.load_workbook(COMPTES_XLSX, data_only=True)
    _cr = ColResolver.from_openpyxl(_wb)
    ignore = {_cr.col('OPréf'), _cr.col('OPcatégorie'), _cr.col('OPcommentaire')}
    _wb.close()
    result = do_compare_xlsx(
        str(COMPTES_XLSX), str(expected_path),
        override_ignore_cols=ignore,
        approx_tolerance=0.03,
        compare_tuples_flag=tuples,
        brutal=brutal,
        warn_threshold_override=0.001,
        labels=('RESULT', 'EXPECTED'),
        skip_sheets=skip_sheets,
    )
    # Filet named ranges : ajout/suppression/modif + mono-cell dégénérés
    nr_ok = compare_named_ranges(expected_path)
    if nr_ok is False and result is not False:
        result = False
    return result


def compare_named_ranges(expected_path, result_path=None):
    """Compare les defined_names de result vs expected + signale les mono-cells.

    Retourne True si pas de diff et pas de mono-cell, False sinon.
    """
    import openpyxl as _xl
    import re as _re

    if result_path is None:
        result_path = COMPTES_XLSX
    result_path = Path(result_path)
    expected_path = Path(expected_path)
    if not expected_path.exists():
        return None

    wb_r = _xl.load_workbook(result_path, keep_vba=True, data_only=False)
    wb_e = _xl.load_workbook(expected_path, keep_vba=True, data_only=False)

    names_r = {n: wb_r.defined_names[n].value for n in wb_r.defined_names}
    names_e = {n: wb_e.defined_names[n].value for n in wb_e.defined_names}

    added = set(names_r) - set(names_e)
    removed = set(names_e) - set(names_r)
    changed = {n for n in set(names_r) & set(names_e) if names_r[n] != names_e[n]}

    # Tolérance : OP* raccourci (borne end row réduite) = warning, pas erreur.
    # Cas typique : purge_account supprime N lignes -> UNO réduit OP* de N.
    def _parse_range(ref):
        m = _re.match(r'\$?[^.!]+[.!]\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)', ref or '')
        if not m:
            return None
        return (m.group(1), int(m.group(2)), m.group(3), int(m.group(4)))

    softened = set()
    for n in list(changed):
        if not n.startswith('OP'):
            continue
        pe = _parse_range(names_e[n])
        pr = _parse_range(names_r[n])
        if not (pe and pr):
            continue
        # Même col start, même col end, même row start, seule row end <= expected
        if pe[0] == pr[0] and pe[2] == pr[2] and pe[1] == pr[1] and pr[3] <= pe[3]:
            softened.add(n)
    changed -= softened

    # Détection range dégénéré : $Feuille.$Col$r1:$Col$r2 avec r1==r2 et col identique
    # Les refs sans ':' (constantes "Retenu", pointeurs $A$1) ne sont PAS des ranges dégénérés
    def _is_monocell(ref):
        if not ref or not isinstance(ref, str) or ':' not in ref:
            return False
        m = _re.match(r'\$?[^.!]+[.!]\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)', ref)
        if not m:
            return False
        return m.group(1) == m.group(3) and m.group(2) == m.group(4)

    monocells_r = {n: v for n, v in names_r.items() if _is_monocell(v)}
    refs_r = {n: v for n, v in names_r.items()
              if isinstance(v, str) and '#REF!' in v}

    ok = not (added or removed or changed or monocells_r or refs_r)

    print('🔖 Named ranges :')
    if added:
        print(f'  ⚠ ajoutés dans RESULT ({len(added)}):')
        for n in sorted(added):
            print(f'      {n} = {names_r[n]}')
    if removed:
        print(f'  ⚠ supprimés dans RESULT ({len(removed)}):')
        for n in sorted(removed):
            print(f'      {n} (était {names_e[n]})')
    if changed:
        print(f'  ⚠ modifiés ({len(changed)}):')
        for n in sorted(changed):
            print(f'      {n}: {names_e[n]} → {names_r[n]}')
    if softened:
        print(f'  ℹ OP* raccourcis ({len(softened)}) — toléré (artefact UNO post-purge):')
        for n in sorted(softened):
            print(f'      {n}: {names_e[n]} → {names_r[n]}')
    if monocells_r:
        print(f'  ⚠ mono-cell dans RESULT ({len(monocells_r)}) — risque de dégénérescence UNO:')
        for n in sorted(monocells_r):
            print(f'      {n} = {monocells_r[n]}')
    if refs_r:
        print(f'  ❌ #REF! dans RESULT ({len(refs_r)}) — named ranges cassés:')
        for n in sorted(refs_r):
            print(f'      {n} = {refs_r[n]}')
    if ok:
        print(f'  ✓ {len(names_r)} named ranges, aucun ajout/suppression/modification ni mono-cell')
    return ok


def save_result(result_path):
    """Sauvegarde le résultat du test avant restauration."""
    if COMPTES_XLSX.exists():
        result_path = Path(result_path)
        result_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(COMPTES_XLSX, result_path)
        print(f"📝 Résultat sauvé: {result_path}")


# Named ranges colonnes attendus dans un xlsm valide (un par tableau)
EXPECTED_NAMED_RANGES = {
    'AVRintitulé', 'CTRL1compte', 'PVLcompte',
    'COTcode', 'CATnom', 'POSTESnom', 'OPdate',
}


def check_integrity_fast(xlsm_path):
    """Vérification intégrité rapide via openpyxl (pas de UNO).

    Vérifie les noms définis et scanne les formules #REF!.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsm_path, data_only=False)
    cr = ColResolver.from_openpyxl(wb)
    errors = []
    checked = 0

    # 1. Noms définis
    missing = EXPECTED_NAMED_RANGES - set(cr._cols.keys())
    if missing:
        errors.append(f"Noms définis manquants: {', '.join(sorted(missing))}")

    # 2. Sentinelles ⚓ aux rows r1 et r2 du NR (règle NR ↔ sentinelles).
    from inc_excel_schema import ANCHOR_TABLES
    import re as _re
    SENT = ('✓', '⚓')
    for sheet_name, ref_nr, _target_end, only_start in ANCHOR_TABLES:
        dn = wb.defined_names.get(ref_nr)
        if not dn:
            continue  # NR absent : pas la responsabilité de ce check
        m = _re.match(r"'?([^'!]+)'?!\$([A-Z]+)\$(\d+):\$[A-Z]+\$(\d+)", dn.attr_text)
        if not m:
            continue
        col_letter = m.group(2)
        col_1 = 0
        for ch in col_letter:
            col_1 = col_1 * 26 + (ord(ch) - 64)
        r1, r2 = int(m.group(3)), int(m.group(4))
        ws = wb[sheet_name]
        bouts = [('top', r1)] if only_start else [('top', r1), ('bot', r2)]
        for label, r in bouts:
            val = ws.cell(r, col_1).value
            v = str(val).strip() if val else ''
            if v not in SENT:
                errors.append(
                    f"{sheet_name} {col_letter}{r} ({ref_nr} {label}): "
                    f"sentinelle attendue (✓/⚓), trouvé '{v}'")

    # 3. #REF! dans les formules
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    checked += 1
                    if '#REF!' in cell.value:
                        errors.append(f"{ws.title} {cell.coordinate}: #REF! — {cell.value[:60]}")

    wb.close()

    # 4. #NAME? dans les valeurs calculées (named ranges manquants)
    wb_val = openpyxl.load_workbook(xlsm_path, data_only=True)
    for ws in wb_val.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and '#NAME?' in cell.value:
                    errors.append(f"{ws.title} {cell.coordinate}: #NAME? (nom défini manquant)")
    wb_val.close()

    print(f"  {checked} formules vérifiées")
    if errors:
        for e in errors:
            print(f"  ✗ {e}")
        print(f"  ❌ {len(errors)} erreur(s)")
        return False
    print("  ✓ Aucune anomalie")
    return True


