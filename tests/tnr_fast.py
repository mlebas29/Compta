#!/usr/bin/env python3-uno
"""
tnr_fast.py — TNR fast (template → 1 devise + 1 compte via GUI)

Usage:
  python3 tests/tnr_fast.py

Part de comptes_template.xlsm, ajoute 1 devise
et 1 compte via HeadlessGUI (2 appels UNO), vérifie l'intégrité (openpyxl),
compare vs expected. Le DEV n'est jamais modifié.
"""

import argparse
import shutil
import sys
import time
from pathlib import Path

_name = Path(__file__).stem.removeprefix('tnr_')
SCENARIO_DIR = Path(__file__).parent / 'tnr' / _name
EXPECTED = SCENARIO_DIR / 'expected.xlsm'

sys.path.insert(0, str(Path(__file__).parent))
from tnr_lib import find_code_root, timestamp, check_libreoffice_running, check_integrity_fast, compare_named_ranges, setup_sandbox

CODE_ROOT = find_code_root(__file__)
sys.path.insert(0, str(CODE_ROOT))
from tool_gui_cli import HeadlessGUI
from gui_daemon import DaemonGUI


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--daemon', action='store_true',
                        help='Router via daemon JSON RPC (DaemonGUI) au lieu '
                             'du chemin in-process (HeadlessGUI)')
    args = parser.parse_args()
    gui_cls = DaemonGUI if args.daemon else HeadlessGUI
    backend = 'DaemonGUI (JSON RPC)' if args.daemon else 'HeadlessGUI (in-process)'

    t_start = time.time()
    success = True

    if not check_libreoffice_running():
        return 1

    print(f"  backend : {backend}")

    # 0. Sandbox : isole le test (DEV jamais touché).
    print(f"\n{timestamp()} Setup sandbox")
    sandbox = setup_sandbox(SCENARIO_DIR)
    TEMPLATE = sandbox / 'comptes_template.xlsm'
    RESULT = sandbox / 'result.xlsm'
    print(f"  sandbox : {sandbox}")

    # setup_sandbox copie config_cotations.json depuis code_root. En PROD ce
    # fichier contient les devises utilisateur (USD, CHF...) — incompatibles
    # avec un test qui démarre vierge et ajoute USD lui-même. On reset à {}.
    (sandbox / 'config_cotations.json').write_text('{}\n', encoding='utf-8')

    if not TEMPLATE.exists():
        print(f"❌ Template introuvable : {TEMPLATE}")
        return 1

    # 1. Copier template → result
    print(f"\n{timestamp()} Copie template → {RESULT.name}")
    shutil.copy2(TEMPLATE, RESULT)

    # 2-3. Ajout 1 devise + 1 compte. Une seule instance gui pour partager
    # le batch côté daemon (sinon 2 daemons spawnés). HeadlessGUI : pas de
    # batch maintenu entre appels en mode non-context, mais le test passe
    # quand même (chaque add_X ouvre/ferme sa propre session).
    with gui_cls(RESULT) as gui:
        print(f"\n{timestamp()} Ajout devise USD")
        if not gui.add_devise(code='USD', famille='fiat'):
            print("  ✗ Échec add-devise USD")
            success = False

        if success:
            print(f"\n{timestamp()} Ajout compte")
            if not gui.add_account(intitule='Compte Wise USD',
                                   acct_type='Devises étrangères',
                                   devise='USD', sous_type='Dollar US',
                                   domiciliation='Wise', titulaire='Barnabé',
                                   propriete='non'):
                print("  ✗ Échec add-account")
                success = False

    # 4. Vérification placement données entre START/END
    if success:
        print(f"\n{timestamp()} Vérification placement données")
        import openpyxl
        from inc_excel_schema import ColResolver
        wb_check = openpyxl.load_workbook(RESULT, data_only=True)
        cr = ColResolver.from_openpyxl(wb_check)
        # Vérification coches ✓ aux bornes des named ranges colonnes
        for table, ref_range, sheet, data_col in [
            ('AVR', 'AVRintitulé', 'Avoirs', 1),
            ('CTRL1', 'CTRL1compte', 'Contrôles', 1),
            ('OP', 'OPdate', 'Opérations', 1),
        ]:
            s, e = cr.rows(ref_range)
            if not s:
                print(f"  ❌ {table}: START introuvable")
                success = False
                continue
            if table != 'OP' and not e:
                print(f"  ❌ {table}: END introuvable")
                success = False
                continue
            ws = wb_check[sheet]
            # START doit contenir ✓ ; END seulement pour les tables avec model rows
            checks = [('START', s)]
            if table != 'OP' and e:
                checks.append(('END', e))
            for label, row in checks:
                val = str(ws.cell(row, data_col).value or '').strip()
                if val not in ('✓', '⚓'):
                    print(f"  ❌ {table}: {label} row {row} = '{val}' (attendu '⚓' ou '✓')")
                    success = False
            if table == 'OP' or not e:
                continue
            # Vérifier qu'il n'y a pas de données après END (avant le pied)
            for r in range(e + 1, e + 5):
                val = str(ws.cell(r, data_col).value or '').strip()
                if val and val not in ('Total', 'Erreurs', ''):
                    print(f"  ❌ {table}: donnée hors bornes row {r} = '{val}' (après END)")
                    success = False
                if val in ('Total', 'Erreurs', ''):
                    break
        # OP spécifique : vérifier que les #Solde sont entre START et END
        s_op, e_op = cr.rows('OPdate')
        if s_op and e_op:
            ws_op = wb_check['Opérations']
            for r in range(1, ws_op.max_row + 1):
                cat = ws_op.cell(r, cr.col('OPcatégorie')).value
                if cat and str(cat).strip() == '#Solde':
                    if r <= s_op or r >= e_op:
                        compte = ws_op.cell(r, cr.col('OPcompte')).value or ''
                        print(f"  ❌ OP: #Solde row {r} ({compte}) hors bornes START={s_op}/END={e_op}")
                        success = False
        if success:
            print("  ✓ Données entre START/END pour AVR, CTRL1 et OP")
        wb_check.close()

    # 5. Vérification intégrité (openpyxl, quasi instantané)
    if success:
        print(f"\n{timestamp()} Vérification intégrité")
        if not check_integrity_fast(RESULT):
            success = False

    # 6. Contrôles métier
    if success:
        import subprocess
        print(f"\n{timestamp()} Contrôles métier (tool_controles)")
        ctrl_script = CODE_ROOT / 'tool_controles.py'
        r = subprocess.run(
            [sys.executable, str(ctrl_script), '-f', str(RESULT)],
            cwd=CODE_ROOT, capture_output=True, text=True)
        if r.stdout:
            for line in r.stdout.strip().splitlines():
                print(f"  {line}")
        if r.returncode != 0:
            print(f"  ❌ tool_controles a retourné {r.returncode}")
            success = False

    # 7. Comparer vs expected
    if success and EXPECTED.exists():
        print(f"\n{timestamp()} Comparaison vs expected")
        from inc_compare_xlsx import compare_xlsx
        if not compare_xlsx(str(RESULT), str(EXPECTED)):
            success = False
        if compare_named_ranges(EXPECTED, result_path=RESULT) is False:
            success = False
    elif success:
        print(f"\n{timestamp()} ⚠ Pas d'expected — valider puis:")
        print(f"   cp {RESULT} {EXPECTED}")

    elapsed = time.time() - t_start
    print()
    if success:
        print(f"✅ TEST RÉUSSI ({elapsed:.0f}s)")
        return 0
    else:
        print(f"❌ TEST ÉCHOUÉ ({elapsed:.0f}s)")
        return 1


if __name__ == '__main__':
    sys.exit(main())
