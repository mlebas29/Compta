"""Mixin Devises/PV pour ConfigGUI."""

from datetime import datetime
from tkinter import messagebox
from tkinter import ttk
import json
import openpyxl
import re
import shutil
import threading
import time
import tkinter as tk

from inc_excel_schema import (
    ColResolver,
    DEVISE_SOURCES,
    SHEET_AVOIRS, SHEET_BUDGET, SHEET_CONTROLES,
    SHEET_OPERATIONS, SHEET_PLUS_VALUE,
)


class DevisesMixin:
    """Devises, Plus-values, sauvegarde comptes."""

    def _devise_add(self):
        """Ouvre le dialog d'ajout d'une nouvelle devise."""
        if not self.xlsx_path:
            return
        self._devise_add_dialog()

    def _devise_add_dialog(self):
        dlg = tk.Toplevel(self.root)
        dlg.title('Nouvelle devise / cotation')
        dlg.geometry('560x290')
        dlg.transient(self.root)
        dlg.wait_visibility()
        dlg.grab_set()

        row = 0
        ttk.Label(dlg, text='Code :').grid(
            row=row, column=0, sticky='w', padx=10, pady=4)
        code_var = tk.StringVar()
        code_entry = ttk.Entry(dlg, textvariable=code_var, width=12)
        code_entry.grid(row=row, column=1, padx=10, pady=4, sticky='w')

        row += 1
        ttk.Label(dlg, text='Nom complet :').grid(
            row=row, column=0, sticky='w', padx=10, pady=4)
        nom_var = tk.StringVar()
        ttk.Entry(dlg, textvariable=nom_var, width=35).grid(
            row=row, column=1, padx=10, pady=4, sticky='w')

        row += 1
        ttk.Label(dlg, text='Famille :').grid(
            row=row, column=0, sticky='w', padx=10, pady=4)
        famille_var = tk.StringVar()
        famille_combo = ttk.Combobox(dlg, textvariable=famille_var,
                                     values=list(DEVISE_SOURCES.keys()),
                                     width=12, state='readonly')
        famille_combo.grid(row=row, column=1, padx=10, pady=4, sticky='w')

        row += 1
        ttk.Label(dlg, text='Décimales :').grid(
            row=row, column=0, sticky='w', padx=10, pady=4)
        dec_var = tk.IntVar(value=2)
        ttk.Spinbox(dlg, textvariable=dec_var, from_=0, to=8, width=5).grid(
            row=row, column=1, padx=10, pady=4, sticky='w')

        row += 1
        ttk.Label(dlg, text='Dérivée de :').grid(
            row=row, column=0, sticky='w', padx=10, pady=4)
        # Codes spot uniquement (ceux avec une source API = pas dérivés)
        existing_codes = sorted(
            code for code, meta in self.cotations_meta.items()
            if meta.get('source1'))
        derived_var = tk.StringVar()
        ttk.Combobox(dlg, textvariable=derived_var,
                      values=[''] + existing_codes,
                      width=12).grid(row=row, column=1, padx=10, pady=4, sticky='w')

        row += 1
        ttk.Label(dlg, text='Formule :').grid(
            row=row, column=0, sticky='w', padx=10, pady=4)
        formula_var = tk.StringVar()
        ttk.Entry(dlg, textvariable=formula_var, width=20).grid(
            row=row, column=1, padx=10, pady=4, sticky='w')
        ttk.Label(dlg, text='ex: *1.043, /100000000, /2',
                  foreground='gray').grid(row=row, column=1, padx=160, sticky='w')

        row += 1
        status_label = ttk.Label(dlg, text='', foreground='red')
        status_label.grid(row=row, column=0, columnspan=2, padx=10)

        def on_ok():
            code = code_var.get().strip()
            famille = famille_var.get().strip()
            nom = nom_var.get().strip() or None
            decimals = dec_var.get()
            derived_from = derived_var.get().strip() or None
            formula = formula_var.get().strip() or None

            if not code:
                status_label.config(text='Le code est obligatoire.')
                return
            if not famille:
                status_label.config(text='La famille est obligatoire.')
                return
            if derived_from and not formula:
                status_label.config(text='Formule requise pour une dérivée.')
                return

            # Vérifier que le code n'existe pas déjà
            try:
                wb = openpyxl.load_workbook(self.xlsx_path, data_only=True)
                ws = wb[SHEET_BUDGET]
                header_row = self.budget_header_row
                for col_idx in range(self.budget_first_devise_col, self.budget_first_devise_col + 30):
                    val = ws.cell(header_row, col_idx).value
                    if not val:
                        break
                    if str(val).strip() == code:
                        wb.close()
                        status_label.config(text=f'La devise "{code}" existe déjà.')
                        return
                wb.close()
            except Exception as e:
                status_label.config(text=f'Erreur lecture : {e}')
                return

            # Sauvegarder dans config_cotations.json (pas de derived_from/formula, c'est dans Excel)
            meta = {'famille': famille, 'decimals': decimals}
            source1, source2 = DEVISE_SOURCES.get(famille, ('', ''))
            if source1 and not derived_from:
                meta['source1'] = source1
            if source2 and not derived_from:
                meta['source2'] = source2
            self.cotations_meta[code] = meta
            import json
            with open(self.cotations_json_path, 'w', encoding='utf-8') as f:
                json.dump(self.cotations_meta, f, ensure_ascii=False, indent=2)

            dlg.destroy()
            self._run_devise_save(code, famille, nom=nom,
                                   derived_from=derived_from, formula=formula)

        row += 1
        btn_frame = ttk.Frame(dlg)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text='OK', command=on_ok).pack(
            side='left', padx=5)
        ttk.Button(btn_frame, text='Annuler',
                   command=dlg.destroy).pack(side='left', padx=5)

        code_entry.focus()

    def _run_devise_save(self, code, famille, nom=None,
                          derived_from=None, formula=None):
        """Lance _save_devise dans un thread avec fenêtre d'attente animée."""
        def on_success():
            if code not in self.ACCOUNT_DEVISES:
                self.ACCOUNT_DEVISES.append(code)
            self._set_status(f'Devise "{code}" ajoutée.')

        self._run_uno_operation(
            'Écriture en cours',
            lambda: self._save_devise(code, famille, nom=nom,
                                       derived_from=derived_from, formula=formula),
            on_success
        )

        check_done()

    def _save_devise(self, code, famille, nom=None,
                      derived_from=None, formula=None, doc=None):
        """Insère une nouvelle devise dans Cotations, Budget et Contrôles via UNO.

        Args:
            doc: UnoDocument ouvert (mode batch). Si None, ouvre/ferme automatiquement.
        """
        from contextlib import nullcontext
        from inc_uno import UnoDocument, copy_col_style
        from inc_excel_schema import (
            uno_col, uno_row, col_letter, CotCol,
            SHEET_COTATIONS,
        )

        bak_path = self.xlsx_path.with_suffix('.xlsm.bak')
        shutil.copy2(self.xlsx_path, bak_path)

        source1, source2 = DEVISE_SOURCES.get(famille, ('', ''))
        decimals = self.cotations_meta.get(code, {}).get('decimals', 2)
        last_bud = self.budget_last_devise_col  # ex: 24 (X=SEK)
        last_ctrl = self.ctrl_last_devise_col   # ex: 29 (AC)

        owned = doc is None
        ctx = UnoDocument(self.xlsx_path) if owned else nullcontext(doc)
        with ctx as doc:
            cr = ColResolver.from_uno(doc.document)
            # ==== ÉTAPE A — Cotations : insérer une ligne dans le groupe ====
            ws_cot = doc.get_sheet(SHEET_COTATIONS)
            cot_data_start = self._start_cot + 1

            # Insérer les colonnes Nature/Famille/Décimales si absentes
            header_0 = uno_row(cot_data_start - 2)  # header = START - 1
            if ws_cot.getCellByPosition(cr.col('COTnature'), header_0).getString().strip() != 'Nature':
                ws_cot.Columns.insertByIndex(cr.col('COTnature'), 1)
                ws_cot.getCellByPosition(cr.col('COTnature'), header_0).setString('Nature')
            if ws_cot.getCellByPosition(cr.col('COTfamille'), header_0).getString().strip() != 'Famille':
                ws_cot.Columns.insertByIndex(cr.col('COTfamille'), 1)
                ws_cot.getCellByPosition(cr.col('COTfamille'), header_0).setString('Famille')
            if ws_cot.getCellByPosition(cr.col('COTdecimales'), header_0).getString().strip() != 'Décimales':
                ws_cot.Columns.insertByIndex(cr.col('COTdecimales'), 1)
                ws_cot.getCellByPosition(cr.col('COTdecimales'), header_0).setString('Décimales')

            # Scanner col A (code) + lookup famille dans cotations_meta
            cot_insert_pos = None
            cot_last_data = cot_data_start  # fallback
            for row_idx in range(cot_data_start, self._end_cot + 1):
                cell_code = ws_cot.getCellByPosition(cr.col('COTcode'), uno_row(row_idx))
                row_code = cell_code.getString().strip()
                cell_label = ws_cot.getCellByPosition(cr.col('COTlabel'), uno_row(row_idx))
                row_label = cell_label.getString().strip()
                if not row_code and not row_label:
                    continue
                if row_label == '✓' or row_code == '✓':
                    continue  # model row — ne pas inclure dans la zone de données
                cot_last_data = row_idx
                # Lookup famille via JSON meta (par code ou par label)
                row_famille = self.cotations_meta.get(row_code, {}).get('famille', '')
                if not row_famille:
                    row_famille = self.cotations_meta.get(row_label, {}).get('famille', '')
                if row_famille == famille:
                    cot_insert_pos = row_idx  # on continue pour trouver le dernier du groupe
                elif cot_insert_pos is not None:
                    # On a quitté le groupe → insérer ici
                    break

            if cot_insert_pos is None:
                # Famille non trouvée → insérer après la dernière ligne de données
                cot_insert_pos = cot_last_data

            # cot_insert_pos pointe vers le dernier row du groupe → insérer après
            cot_new_row = cot_insert_pos + 1
            ws_cot.Rows.insertByIndex(uno_row(cot_new_row), 1)

            # Mettre à jour _end_cot (l'insertion repousse END)
            if self._end_cot and cot_new_row <= self._end_cot:
                self._end_cot += 1

            # Style propagé automatiquement depuis la model row par insertByIndex

            # Remplir la ligne : A=label, B=nature, C=famille, D=décimales, E=code
            r0_cot = uno_row(cot_new_row)
            ws_cot.getCellByPosition(cr.col('COTlabel'), r0_cot).setString(nom or code)
            ws_cot.getCellByPosition(cr.col('COTcode'), r0_cot).setString(code)
            ws_cot.getCellByPosition(cr.col('COTfamille'), r0_cot).setString(famille)
            ws_cot.getCellByPosition(cr.col('COTdecimales'), r0_cot).setValue(decimals)

            # Nature + cours
            ws_cot.getCellByPosition(cr.col('COTnature'), r0_cot).setString(
                'dérivée' if derived_from else 'primaire')
            if derived_from and formula:
                # Formule dérivée : trouver le row du spot (col CODE)
                spot_row = None
                for ri in range(cot_data_start, cot_new_row + 5):
                    if ws_cot.getCellByPosition(cr.col('COTcode'), uno_row(ri)).getString().strip() == derived_from:
                        spot_row = ri
                        break
                if spot_row:
                    cot_cours_letter = cr.letter('COTcours')
                    cell = ws_cot.getCellByPosition(
                        cr.col('COTcours'), r0_cot)
                    cell.setFormula(
                        f'={cot_cours_letter}${spot_row}{formula}')
                    cell.CharColor = 0x000000  # noir = formule
            elif source1:
                try:
                    from cpt_fetch_quotes import API_FETCHERS
                    fetcher = API_FETCHERS.get(source1)
                    if fetcher:
                        result = fetcher([code])
                        if code in result:
                            ws_cot.getCellByPosition(
                                cr.col('COTcours'), r0_cot).setValue(result[code])
                            ws_cot.getCellByPosition(
                                cr.col('COTdate'), r0_cot).setString(
                                datetime.now().strftime('%d/%m/%Y'))
                except Exception:
                    pass  # pas bloquant — le cours sera renseigné au prochain fetch

            # Col H : cours de l'Euro = 1/cours_EUR
            cot_cours_letter = cr.letter('COTcours')
            ws_cot.getCellByPosition(cr.col('COTdate') + 1, r0_cot).setFormula(
                f'=1/{cot_cours_letter}{cot_new_row}')

            # Créer le named range cours_XXX → cellule cours de la nouvelle cotation
            cours_name = self.cours_name(code)
            if cours_name:
                from com.sun.star.table import CellAddress
                cot_cours_col_letter = cr.letter('COTcours')
                nr = doc.document.NamedRanges
                if nr.hasByName(cours_name):
                    nr.removeByName(cours_name)
                nr.addNewByName(cours_name,
                                f'$Cotations.${cot_cours_col_letter}${cot_new_row}',
                                CellAddress(), 0)

            # Déterminer la dernière ligne de données Cotations (pour SUMIF range)
            cot_last_row = cot_new_row
            for row_idx in range(cot_new_row + 1, cot_new_row + 30):
                cell_e = ws_cot.getCellByPosition(cr.col('COTcode'), uno_row(row_idx))
                if cell_e.getString().strip():
                    cot_last_row = row_idx
                else:
                    break

            # ==== ÉTAPE B — Budget : insérer 1 colonne avant Equiv EUR ====
            has_budget = self.budget_start_row is not None
            ws_bud = doc.get_sheet(SHEET_BUDGET)
            equiv_col = last_bud + 1  # position actuelle Equiv EUR (1-indexed)
            # Insérer avant Equiv EUR → la nouvelle colonne prend sa place
            ws_bud.Columns.insertByIndex(uno_col(equiv_col), 1)
            # Maintenant : new_col = equiv_col, Equiv EUR décalé à equiv_col+1

            new_col = equiv_col           # la nouvelle colonne devise
            new_col_letter = ColResolver._idx_to_letter(new_col)
            new_equiv_col = equiv_col + 1
            new_equiv_letter = ColResolver._idx_to_letter(new_equiv_col)
            new_alloc_pct_col = new_equiv_col + 1
            new_alloc_montant_col = new_equiv_col + 2
            new_poste_col = new_equiv_col + 3

            # Copier le style : EUR (col M) pour fiat, voisin pour les autres
            style_src_bud = self.budget_first_devise_col if famille == 'fiat' else last_bud
            bud_end = (self.budget_total_row or self.budget_header_row + 20) + 5
            # Identifier les model rows à ne pas toucher
            from inc_uno import get_named_range_pos
            _model_rows_0 = set()
            for _nr in ('START_CAT', 'END_CAT', 'START_POSTES', 'END_POSTES'):
                _p = get_named_range_pos(doc.document, _nr)
                if _p:
                    _model_rows_0.add(_p[2])
            copy_col_style(ws_bud, uno_col(style_src_bud), uno_col(new_col),
                           row_start=uno_row(self.budget_header_row), row_end=uno_row(bud_end),
                           skip_rows=_model_rows_0)

            # Appliquer le format devise + gris sur la nouvelle colonne Budget
            # Exclure "Montant Euros" (total_row+4) qui est en EUR
            from inc_formats import FORMATS_DEVISE, FORMAT_EUR, GRIS, devise_format
            fmt_devise = FORMATS_DEVISE.get(code, devise_format(code))
            montant_eur_row = (self.budget_total_row + 4) if self.budget_total_row else None
            if fmt_devise:
                fmt_id = doc.register_number_format(fmt_devise)
                fmt_eur_id = doc.register_number_format(FORMAT_EUR)
                for r in range(self.budget_header_row + 1, bud_end):
                    if uno_row(r) in _model_rows_0:
                        continue
                    cell = ws_bud.getCellByPosition(uno_col(new_col), uno_row(r))
                    if montant_eur_row and r == montant_eur_row:
                        cell.NumberFormat = fmt_eur_id
                    else:
                        cell.NumberFormat = fmt_id
                        cell.CellBackColor = GRIS

            # ==== ÉTAPE C — Budget : remplir la nouvelle colonne ====
            hr = self.budget_header_row
            nc0 = uno_col(new_col)
            nec0 = uno_col(new_equiv_col)  # colonne Equiv EUR (décalée)

            # Row header : code devise
            ws_bud.getCellByPosition(nc0, uno_row(hr)).setString(code)

            # Row START : taux = SUMIF(Cotations)
            cot_code_letter = cr.letter('COTcode')
            cot_cours_letter = cr.letter('COTcours')
            if has_budget:
                ws_bud.getCellByPosition(nc0, uno_row(self.budget_start_row)).setFormula(
                    f'=SUMIF(Cotations.${cot_code_letter}${cot_data_start}:${cot_code_letter}${cot_last_row};{new_col_letter}${hr};'
                    f'Cotations.${cot_cours_letter}${cot_data_start}:${cot_cours_letter}${cot_last_row})')

            # Rows catégories : SUMIFS par devise
            first_cat_row = min(self.budget_cat_rows.values()) if self.budget_cat_rows else (self.budget_start_row or hr) + 1
            sep_row = self.budget_insert_row or first_cat_row
            for r in range(first_cat_row, sep_row + 1):  # inclut le séparateur "-"
                cell_l = ws_bud.getCellByPosition(uno_col(self.budget_cat_col), uno_row(r))
                val_l = cell_l.getString().strip()
                if val_l and val_l != '✓':
                    cat_letter = ColResolver._idx_to_letter(self.budget_cat_col)
                    ws_bud.getCellByPosition(nc0, uno_row(r)).setFormula(
                        f'=SUMIFS(OPmontant;OPdevise;{new_col_letter}${hr};OPcatégorie;${cat_letter}{r};OPdate;">"&$C$2-365)')

            # Row Total et formules résumé (skip si pas de catégories Budget)
            total_row = self.budget_total_row
            if total_row:
                ws_bud.getCellByPosition(nc0, uno_row(total_row)).setFormula(
                    f'=SUM({new_col_letter}{first_cat_row}:{new_col_letter}{sep_row})')

                # Row Total+1 : SUMIFS hors Spéciale
                ws_bud.getCellByPosition(nc0, uno_row(total_row + 1)).setFormula(
                    f'=SUMIFS(OPmontant;OPdevise;{new_col_letter}${hr};OPdate;">"&$C$2-365;OPcatégorie;"<>"&Spéciale)')

                # Row Total+2 (95) : Écart = Total - Somme opérations
                ws_bud.getCellByPosition(nc0, uno_row(total_row + 2)).setFormula(
                    f'={new_col_letter}${total_row}-{new_col_letter}${total_row + 1}')

                # Row Total+3 (96) : Total hors Changes = SUM(col{hc_start}:col{sep})
                # Trouver la première catégorie avec Equiv EUR (= première catégorie budgétaire réelle)
                # Les catégories spéciales (Change, Virement, Achat...) n'ont pas de formule Equiv EUR
                hc_start = first_cat_row + 6  # fallback
                for r in range(first_cat_row, sep_row):
                    cell_eq = ws_bud.getCellByPosition(nec0, uno_row(r))
                    if cell_eq.getFormula():
                        hc_start = r
                        break
                changes_row = hc_start
                # Réécrire Total+3 pour TOUTES les colonnes devise (y compris EUR)
                # Les formules template single-cell ne s'étendent pas avec les inserts
                for dc in range(self.budget_first_devise_col, new_col + 1):
                    dcl = ColResolver._idx_to_letter(dc)
                    ws_bud.getCellByPosition(uno_col(dc), uno_row(total_row + 3)).setFormula(
                        f'=SUM({dcl}{changes_row}:{dcl}{sep_row})')

                # Row Total+4 : Montant Euros = col*taux
                sr = self.budget_start_row  # ligne des cours
                ws_bud.getCellByPosition(nc0, uno_row(total_row + 4)).setFormula(
                    f'={new_col_letter}${total_row + 3}*{new_col_letter}${sr}')

            # ==== ÉTAPE D — Budget : mettre à jour Equiv EUR ====
            # Après insertion, Equiv EUR est à new_equiv_col
            # Réécrire les formules en SUMPRODUCT première_devise → nouvelle devise incluse
            # Ne toucher que les lignes qui ont déjà une formule Equiv EUR
            sumproduct_last = new_col_letter  # ex: Y si PLN=col25 → inclut PLN
            fdl = ColResolver._idx_to_letter(self.budget_first_devise_col)

            for r in range(first_cat_row, sep_row + 1):  # inclut le séparateur
                cell_equiv = ws_bud.getCellByPosition(nec0, uno_row(r))
                existing = cell_equiv.getFormula()
                if existing:  # ne réécrire que si une formule existait
                    ws_bud.getCellByPosition(nec0, uno_row(r)).setFormula(
                        f'=SUMPRODUCT({fdl}{r}:{sumproduct_last}{r};{fdl}${sr}:{sumproduct_last}${sr})')

            # Pieds Equiv EUR : réécrire Total et Total+3 avec les bons ranges
            # (le template a des refs single-cell qui ne s'étendent pas)
            if total_row:
                # Total Equiv EUR : =SUM(equiv_col first_cat : equiv_col sep)
                ws_bud.getCellByPosition(nec0, uno_row(total_row)).setFormula(
                    f'=SUM({new_equiv_letter}{first_cat_row}:{new_equiv_letter}{sep_row})')
                # Total+3 Equiv EUR : =SUM(equiv_col changes : equiv_col sep)
                ws_bud.getCellByPosition(nec0, uno_row(total_row + 3)).setFormula(
                    f'=SUM({new_equiv_letter}{changes_row}:{new_equiv_letter}{sep_row})')

                # Total+3 Alloc montant
                alloc_m_col = new_equiv_col + 2
                alloc_m_letter = ColResolver._idx_to_letter(alloc_m_col)
                ws_bud.getCellByPosition(uno_col(alloc_m_col), uno_row(total_row + 3)).setFormula(
                    f'=SUM({alloc_m_letter}{changes_row}:{alloc_m_letter}{sep_row})')

            # Row Total+4 Equiv EUR : =SUM(first_dev:new_col)
            if total_row:
                ws_bud.getCellByPosition(nec0, uno_row(total_row + 4)).setFormula(
                    f'=SUM({fdl}${total_row + 4}:{new_col_letter}${total_row + 4})')

            # ==== ÉTAPE E — Contrôles tableau 2 : écrire la nouvelle colonne ====
            # Lignes relatives à ctrl2_header_row (header devises)
            ws_ctrl = doc.get_sheet(SHEET_CONTROLES)
            h = self.ctrl2_header_row  # row des codes devises (1-indexed)
            new_ctrl_col = last_ctrl + 1
            cc0 = uno_col(new_ctrl_col)
            ctrl_letter = ColResolver._idx_to_letter(new_ctrl_col)

            # Copier le style depuis la colonne EUR CTRL2 (= première devise CTRL2)
            # last_ctrl pointe sur la dernière colonne existante avant insertion
            first_ctrl_devise = last_ctrl  # pour la 1ère devise = EUR, ensuite = voisin
            copy_col_style(ws_ctrl, uno_col(first_ctrl_devise), cc0,
                           row_start=uno_row(h), row_end=uno_row(h + 15))

            # Appliquer les formats nombre spécifiques par ligne + gris
            from inc_formats import FORMATS_DEVISE, FORMAT_EUR, FORMAT_EUR_RED, GRIS, devise_format
            fmt_devise = FORMATS_DEVISE.get(code, devise_format(code))
            fmt_devise_id = doc.register_number_format(fmt_devise)
            fmt_eur_id = doc.register_number_format(FORMAT_EUR)
            fmt_eur_red_id = doc.register_number_format(FORMAT_EUR_RED)
            fmt_int_id = doc.register_number_format('#\xa0##0')
            # Rouge négatif pour Virements
            fmt_red_id = doc.register_number_format(f'{fmt_devise};[RED]\\-{fmt_devise}')

            # h+1: taux EUR, h+2/4/5/6: entier, h+3/9: devise+gris,
            # h+7: devise+gris+rouge, h+8/10/11/12: EUR
            for offset, fmt_id in ((1, fmt_eur_id), (2, fmt_int_id), (4, fmt_int_id),
                                    (5, fmt_int_id), (6, fmt_int_id),
                                    (8, fmt_eur_id), (10, fmt_eur_id),
                                    (11, fmt_eur_id), (12, fmt_eur_id)):
                ws_ctrl.getCellByPosition(cc0, uno_row(h + offset)).NumberFormat = fmt_id
            # Lignes devise + gris
            for offset in (3, 9):
                cell = ws_ctrl.getCellByPosition(cc0, uno_row(h + offset))
                cell.NumberFormat = fmt_devise_id
                cell.CellBackColor = GRIS
            # Virements : devise rouge + gris
            cell = ws_ctrl.getCellByPosition(cc0, uno_row(h + 7))
            cell.NumberFormat = fmt_red_id
            cell.CellBackColor = GRIS

            # Budget col correspondant à cette devise = new_col
            bud_letter = new_col_letter

            # h+0 : header devise
            ws_ctrl.getCellByPosition(cc0, uno_row(h)).setString(code)

            # h+1 : taux = Cotations.$C${cot_new_row}
            ws_ctrl.getCellByPosition(cc0, uno_row(h + 1)).setFormula(
                f'=Cotations.${cot_cours_letter}${cot_new_row}')

            # h+2 : COMPTES — COUNTIFS écarts par devise (CTRL1 ranges)
            # B=devise, K="Oui"(écart détecté), J=valeur écart (0=OK)
            ws_ctrl.getCellByPosition(cc0, uno_row(h + 2)).setFormula(
                f'=COUNTIFS($B4:$B58;{ctrl_letter}${h};$I4:$I58;"Oui")'
                f'-COUNTIFS($B4:$B58;{ctrl_letter}${h};$I4:$I58;"Oui";$H4:$H58;0)')

            # h+3 : CATÉGORIES = Budget.{bud_letter}${total_row+2}
            if total_row:
                ws_ctrl.getCellByPosition(cc0, uno_row(h + 3)).setFormula(
                    f'=Budget.{bud_letter}${total_row + 2}')

            # h+5 : Appariements
            ws_ctrl.getCellByPosition(cc0, uno_row(h + 5)).setFormula(
                f'=COUNTIFS(OPréf;"-";OPdevise;{ctrl_letter}${h})')

            # h+6 : Balances = COUNTA labels - COUNTIFS zéros
            ws_ctrl.getCellByPosition(cc0, uno_row(h + 6)).setFormula(
                f'=COUNTA($M{h+7}:$M{h+9})-COUNTIFS({ctrl_letter}{h+7}:{ctrl_letter}{h+9};0)')

            # h+7 : Virements = ROUND(SUMIFS par catégorie "Virement*")
            ws_ctrl.getCellByPosition(cc0, uno_row(h + 7)).setFormula(
                f'=ROUND(SUMIFS(OPmontant;OPdevise;{ctrl_letter}${h};OPcatégorie;"Virement*");0)')

            # h+8 : Virements € = h+7 * taux
            ws_ctrl.getCellByPosition(cc0, uno_row(h + 8)).setFormula(
                f'=ROUND(SUMIFS(OPmontant;OPdevise;{ctrl_letter}${h};OPcatégorie;"Virement*");0)*{ctrl_letter}${h+1}')

            # h+9 : Titres = ROUND(SUMIFS par catégorie "*titres")
            ws_ctrl.getCellByPosition(cc0, uno_row(h + 9)).setFormula(
                f'=ROUND(SUMIFS(OPmontant;OPdevise;{ctrl_letter}${h};OPcatégorie;"*titres");0)')

            # h+10 : Titres € = h+9 * taux
            ws_ctrl.getCellByPosition(cc0, uno_row(h + 10)).setFormula(
                f'=ROUND(SUMIFS(OPmontant;OPdevise;{ctrl_letter}${h};OPcatégorie;"*titres");0)*{ctrl_letter}${h+1}')

            # h+11 : Changes Eq € = SUMIFS equiv_euro par devise (Change + Achat métaux)
            ws_ctrl.getCellByPosition(cc0, uno_row(h + 11)).setFormula(
                f'=SUMIFS(OPequiv_euro;OPdevise;{ctrl_letter}${h};OPcatégorie;"Change")'
                f'+SUMIFS(OPequiv_euro;OPdevise;{ctrl_letter}${h};OPcatégorie;"Achat métaux")')

            # h+12 : Total € = Virements€ + Titres€ + Changes€
            ws_ctrl.getCellByPosition(cc0, uno_row(h + 12)).setFormula(
                f'={ctrl_letter}{h+8}+{ctrl_letter}{h+10}+{ctrl_letter}{h+11}')

            # ==== ÉTAPE F — Contrôles : étendre 4 ranges SUM ====
            new_end = ColResolver._idx_to_letter(new_ctrl_col)
            old_end = ColResolver._idx_to_letter(last_ctrl)

            # P(h+2) : IF(SUM(…)=0…) → étendre le range
            cell_p = ws_ctrl.getCellByPosition(uno_col(16), uno_row(h + 2))  # P=16
            old_formula = cell_p.getFormula()
            if old_formula:
                cell_p.setFormula(old_formula.replace(
                    f'${old_end}{h+2}', f'${new_end}{h+2}').replace(
                    f'${old_end}${h+2}', f'${new_end}${h+2}'))

            # P(h+3) : IF(SUM(…)=0…) → étendre le range
            cell_p3 = ws_ctrl.getCellByPosition(uno_col(16), uno_row(h + 3))
            old_formula = cell_p3.getFormula()
            if old_formula:
                cell_p3.setFormula(old_formula.replace(
                    f'${old_end}{h+3}', f'${new_end}{h+3}').replace(
                    f'${old_end}${h+3}', f'${new_end}${h+3}'))

            # Q(h+8) : SUM → étendre le range
            cell_q8 = ws_ctrl.getCellByPosition(uno_col(17), uno_row(h + 8))  # Q=17
            old_formula = cell_q8.getFormula()
            if old_formula:
                cell_q8.setFormula(old_formula.replace(
                    f'{old_end}{h+8}', f'{new_end}{h+8}'))

            # Q(h+10) : SUM → étendre le range
            cell_q10 = ws_ctrl.getCellByPosition(uno_col(17), uno_row(h + 10))
            old_formula = cell_q10.getFormula()
            if old_formula:
                cell_q10.setFormula(old_formula.replace(
                    f'{old_end}{h+10}', f'{new_end}{h+10}'))

            # ==== Finaliser ====
            if owned:
                self._uno_finalize(doc)

        # ==== ÉTAPE G — Mise à jour état mémoire + JSON ====
        self.budget_last_devise_col = new_col       # ex: 24→25
        self.ctrl_last_devise_col = new_ctrl_col    # ex: 29→30

        # Persister les métadonnées dans config_cotations.json
        self.cotations_meta[code] = {
            'famille': famille,
            'source1': source1,
            'source2': source2,
        }
        from cpt_gui import write_cotations_json
        write_cotations_json(self.cotations_json_path, self.cotations_meta)

    # ----------------------------------------------------------------
    # SUPPRESSION DEVISE
    # ----------------------------------------------------------------

    def _devise_delete_dialog(self):
        """Dialog de sélection et confirmation pour supprimer une devise."""
        if not self.xlsx_path:
            return

        # Lister les devises existantes depuis Budget row 27 (hors EUR)
        try:
            wb = openpyxl.load_workbook(self.xlsx_path, data_only=True)
            ws = wb[SHEET_BUDGET]
            devises = []
            header_row = self.budget_header_row or 27
            for col_idx in range(self.budget_first_devise_col + 1, self.budget_first_devise_col + 30):
                val = ws.cell(header_row, col_idx).value
                if not val:
                    break
                s = str(val).strip()
                if s.startswith('Équivalent') or s.startswith('Equivalent'):
                    break
                devises.append(s)
            wb.close()
        except Exception as e:
            messagebox.showerror('Erreur', f'Erreur lecture : {e}', parent=self.root)
            return

        if not devises:
            messagebox.showinfo('Info', 'Aucune devise supprimable (seul EUR reste).',
                                parent=self.root)
            return

        dlg = tk.Toplevel(self.root)
        dlg.title('Supprimer une devise')
        dlg.geometry('380x170')
        dlg.transient(self.root)
        dlg.wait_visibility()
        dlg.grab_set()

        ttk.Label(dlg, text='Devise à supprimer :').grid(
            row=0, column=0, sticky='w', padx=10, pady=5)
        devise_var = tk.StringVar()
        devise_combo = ttk.Combobox(dlg, textvariable=devise_var,
                                    values=devises, width=12, state='readonly')
        devise_combo.grid(row=0, column=1, padx=10, pady=5, sticky='w')

        status_label = ttk.Label(dlg, text='', foreground='red')
        status_label.grid(row=1, column=0, columnspan=2, padx=10)

        def on_ok():
            code = devise_var.get().strip()
            if not code:
                status_label.config(text='Sélectionner une devise.')
                return

            # Garde-fou : vérifier qu'aucun compte Avoirs n'utilise cette devise
            try:
                wb2 = openpyxl.load_workbook(self.xlsx_path, data_only=True)
                ws_av = wb2[SHEET_AVOIRS]
                avr_data = self._start_avr + 1
                for row in range(avr_data, ws_av.max_row + 1):
                    dev = ws_av.cell(row, self.cr.col('AVRdevise')).value
                    if dev and str(dev).strip() == code:
                        intitule = ws_av.cell(row, self.cr.col('AVRintitulé')).value or ''
                        wb2.close()
                        status_label.config(
                            text=f'Compte "{intitule.strip()}" utilise {code}.')
                        return
                wb2.close()
            except Exception as e:
                status_label.config(text=f'Erreur : {e}')
                return

            if not messagebox.askyesno(
                    'Confirmer',
                    f'Supprimer la devise "{code}" ?\n\n'
                    f'Cotations (ligne), Budget (colonne),\n'
                    f'Contrôles (colonne) et JSON seront nettoyés.',
                    parent=dlg):
                return

            dlg.destroy()
            self._run_devise_delete(code)

        btn_frame = ttk.Frame(dlg)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text='Supprimer', command=on_ok).pack(
            side='left', padx=5)
        ttk.Button(btn_frame, text='Annuler',
                   command=dlg.destroy).pack(side='left', padx=5)

    def _run_devise_delete(self, code):
        """Lance _delete_devise dans un thread avec fenêtre d'attente."""
        wait = tk.Toplevel(self.root)
        wait.title('')
        wait.geometry('320x80')
        wait.transient(self.root)
        wait.resizable(False, False)
        wait.protocol('WM_DELETE_WINDOW', lambda: None)
        wait.wait_visibility()
        wait.grab_set()

        msg_var = tk.StringVar(value='Suppression en cours')
        ttk.Label(wait, textvariable=msg_var, font=('', 11)).pack(
            expand=True, pady=15)

        dots = [0]

        def animate():
            dots[0] = (dots[0] % 3) + 1
            msg_var.set('Suppression en cours' + '.' * dots[0])
            wait.after(400, animate)

        animate()

        result = {}

        def worker():
            try:
                self._delete_devise(code)
            except Exception as e:
                result['error'] = e

        t = threading.Thread(target=worker, daemon=True)
        t.start()

        def check_done():
            if t.is_alive():
                wait.after(100, check_done)
                return
            wait.grab_release()
            wait.destroy()
            if 'error' in result:
                messagebox.showerror('Erreur UNO',
                                     f'Erreur lors de la suppression :\n{result["error"]}',
                                     parent=self.root)
            else:
                if code in self.ACCOUNT_DEVISES:
                    self.ACCOUNT_DEVISES.remove(code)
                self._set_status(f'Devise "{code}" supprimée.')

        check_done()

    def _delete_devise(self, code):
        """Supprime une devise de Cotations, Budget, Contrôles et JSON."""
        from inc_uno import UnoDocument
        from inc_excel_schema import (
            uno_col, uno_row, CotCol,
            SHEET_COTATIONS,
        )

        bak_path = self.xlsx_path.with_suffix('.xlsm.bak')
        shutil.copy2(self.xlsx_path, bak_path)

        with UnoDocument(self.xlsx_path) as doc:
            cr = ColResolver.from_uno(doc.document)
            # ==== Cotations : supprimer la ligne du code ====
            ws_cot = doc.get_sheet(SHEET_COTATIONS)
            cot_data_start = self._start_cot + 1
            cot_row = None
            for row_idx in range(cot_data_start, self._end_cot + 1):
                cell_code = ws_cot.getCellByPosition(
                    cr.col('COTcode'), uno_row(row_idx))
                if cell_code.getString().strip() == code:
                    cot_row = row_idx
                    break
            if cot_row is not None:
                ws_cot.Rows.removeByIndex(uno_row(cot_row), 1)

            # ==== Budget : supprimer la colonne de la devise ====
            ws_bud = doc.get_sheet(SHEET_BUDGET)
            bud_col = None
            bud_header_row = self.budget_header_row
            for col_idx in range(self.budget_first_devise_col, self.budget_first_devise_col + 30):
                val = ws_bud.getCellByPosition(
                    uno_col(col_idx), uno_row(bud_header_row)).getString().strip()
                if not val:
                    break
                if val == code:
                    bud_col = col_idx
                    break
            if bud_col is not None:
                ws_bud.Columns.removeByIndex(uno_col(bud_col), 1)

            # ==== Contrôles : supprimer la colonne de la devise ====
            ws_ctrl = doc.get_sheet(SHEET_CONTROLES)
            # Coche START_CTRL2 = première ligne data (h+2) et première col devise.
            # Header devises = 2 lignes au-dessus.
            from inc_uno import get_named_range_pos
            ctrl2_pos = get_named_range_pos(doc.document, 'START_CTRL2')
            first_devise_col_0 = ctrl2_pos[1]   # 0-indexed col du EUR
            header_row_0 = ctrl2_pos[2] - 2      # 0-indexed row du header devises
            ctrl_col_0 = None
            for col_0 in range(first_devise_col_0, first_devise_col_0 + 32):
                val = ws_ctrl.getCellByPosition(col_0, header_row_0).getString().strip()
                if not val:
                    break
                if val == code:
                    ctrl_col_0 = col_0
                    break
            if ctrl_col_0 is not None:
                ws_ctrl.Columns.removeByIndex(ctrl_col_0, 1)

            self._uno_finalize(doc)

        # ==== Mise à jour état mémoire ====
        if bud_col is not None:
            self.budget_last_devise_col -= 1
        if ctrl_col_0 is not None:
            self.ctrl_last_devise_col -= 1

        # ==== JSON : supprimer l'entrée ====
        self.cotations_meta.pop(code, None)
        from cpt_gui import write_cotations_json
        write_cotations_json(self.cotations_json_path, self.cotations_meta)

    # ----------------------------------------------------------------
    # HELPERS PLUS_VALUE
    # ----------------------------------------------------------------

    def _get_pv_section_for_account(self, acct_type, devise):
        """Détermine si/où créer une entrée Plus_value pour un nouveau compte.

        Returns: ('portfolio', None) | ('line', 'TOTAL métaux') | (None, None)
        """
        if acct_type == 'Portefeuilles':
            return ('portfolio', None)
        if not devise or devise == 'EUR':
            return (None, None)
        # Chercher la famille de la devise dans config_cotations.json
        meta = self.cotations_meta.get(devise, {})
        famille = meta.get('famille', 'fiat')  # défaut fiat pour devises inconnues
        total_label = self.PV_SECTION_TOTALS.get(famille)
        if total_label:
            return ('line', total_label)
        return (None, None)

    def _find_pv_row_by_label(self, ws_pv, label):
        """Trouve la ligne (1-indexed) d'un label dans la colonne B (Compte) de Plus_value."""
        from inc_excel_schema import uno_row, uno_col, ColResolver
        cr = ColResolver.from_uno(doc.document)
        col_b = cr.col('PVLcompte')
        for row_idx in range(1, 200):
            val = ws_pv.getCellByPosition(col_b, uno_row(row_idx)).getString().strip()
            if val == label:
                return row_idx
        return None

    def _create_pv_portfolio_block(self, ws_pv, doc, acct):
        """Insère un bloc Portefeuille vide (5 lignes) dans Plus_value.

        Structure :
          r   : en-tête (section + nom + "Portefeuille")
          r+1 : Total (formules SUM vides → valeurs 0)
          r+2 : #Solde Opérations (formules SUMIFS/MAXIFS)
          r+3 : Retenu (formules copie #Solde + IF)
          r+4 : ligne vide
        """
        from inc_uno import copy_row_style
        from inc_excel_schema import uno_col, uno_row, ColResolver
        cr = ColResolver.from_uno(doc.document)

        # Lettres de colonnes via PvCol
        cB = cr.letter('PVLcompte')
        cD = cr.letter('PVLdevise')
        cE = cr.letter('PVLpvl')
        cG = cr.letter('PVLdate_init')
        cH = cr.letter('PVLmontant_init')
        cI = cr.letter('PVLsigma')
        cJ = cr.letter('PVLdate')
        cK = cr.letter('PVLmontant')

        # Trouver l'emplacement — TOTAL portefeuilles est maintenant dans les TOTALs en pied
        # Les nouveaux blocs sont insérés avant la dernière ligne vide de la section portefeuilles
        # Chercher la dernière ligne de données portefeuilles (section = "portefeuilles")
        last_pf_data_row = None
        for scan in range(200, 0, -1):
            val_a = ws_pv.getCellByPosition(cr.col('PVLsection'), uno_row(scan)).getString().strip()
            if val_a == 'portefeuilles' or val_a == self.PV_SECTION_LABELS['portefeuilles']:
                last_pf_data_row = scan
                break

        if not last_pf_data_row:
            # Template vierge : pas de données, insérer avant GRAND TOTAL (début footer)
            footer_row = (self._end_pvl + 1) if self._end_pvl else None
            if not footer_row:
                return
            insert_row = footer_row
        else:
            # Insérer après le dernier bloc portefeuilles (+1 ligne vide de séparation)
            insert_row = last_pf_data_row + 2


        nom = acct['intitule']
        devise = acct.get('devise') or 'EUR'

        # Insérer 5 lignes
        insert_row = insert_row  # 1-indexed
        insert_0 = uno_row(insert_row)
        ws_pv.Rows.insertByIndex(insert_0, 5)

        r = insert_row   # 1-indexed, première ligne insérée
        r0 = uno_row(r)

        # Template de style : la dernière ligne Retenu avant notre insertion
        template_row = None
        for scan in range(r - 1, 0, -1):
            val_c = ws_pv.getCellByPosition(cr.col('PVLtitre'), uno_row(scan)).getString().strip()
            if val_c == 'Retenu':
                template_row = uno_row(scan)
                break
        if template_row is None:
            template_row = uno_row(r - 2)  # fallback

        # Copier le style pour les 4 lignes de données (pas la ligne vide)
        for offset in range(4):
            copy_row_style(ws_pv, template_row, r0 + offset, col_start=0, col_end=12)

        # --- Ligne r : en-tête ---
        ws_pv.getCellByPosition(cr.col('PVLsection'), r0).setString('portefeuilles')
        ws_pv.getCellByPosition(cr.col('PVLcompte'), r0).setString(nom)
        ws_pv.getCellByPosition(cr.col('PVLtitre'), r0).setString('Portefeuille')

        # --- Ligne r+1 : Total (vide, valeurs 0) ---
        ws_pv.getCellByPosition(cr.col('PVLsection'), r0 + 1).setString('portefeuilles')
        ws_pv.getCellByPosition(cr.col('PVLcompte'), r0 + 1).setString(nom)
        ws_pv.getCellByPosition(cr.col('PVLtitre'), r0 + 1).setString('Total')
        ws_pv.getCellByPosition(cr.col('PVLdevise'), r0 + 1).setString(devise)
        ws_pv.getCellByPosition(cr.col('PVLpvl'), r0 + 1).setFormula(
            f'={cK}{r+1}-({cH}{r+1}+{cI}{r+1})')
        ws_pv.getCellByPosition(cr.col('PVLdate_init'), r0 + 1).setValue(0)
        ws_pv.getCellByPosition(cr.col('PVLmontant_init'), r0 + 1).setValue(0)
        ws_pv.getCellByPosition(cr.col('PVLsigma'), r0 + 1).setValue(0)
        ws_pv.getCellByPosition(cr.col('PVLmontant'), r0 + 1).setValue(0)

        # --- Ligne r+2 : #Solde Opérations ---
        ws_pv.getCellByPosition(cr.col('PVLsection'), r0 + 2).setString('portefeuilles')
        ws_pv.getCellByPosition(cr.col('PVLcompte'), r0 + 2).setString(nom)
        ws_pv.getCellByPosition(cr.col('PVLtitre'), r0 + 2).setString('#Solde Opérations')
        ws_pv.getCellByPosition(cr.col('PVLdevise'), r0 + 2).setString(devise)
        ws_pv.getCellByPosition(cr.col('PVLpvl'), r0 + 2).setFormula(
            f'={cK}{r+2}-({cH}{r+2}+{cI}{r+2})')
        ws_pv.getCellByPosition(cr.col('PVLdate_init'), r0 + 2).setFormula(
            f'=MINIFS(OPdate;OPcompte;${cB}{r+2};OPdevise;{cD}{r+2};OPcatégorie;Solde)')
        ws_pv.getCellByPosition(cr.col('PVLmontant_init'), r0 + 2).setFormula(
            f'=SUMIFS(OPmontant;OPcompte;${cB}{r+2};OPdevise;${cD}{r+2};OPdate;${cG}{r+2};OPcatégorie;Solde)')
        ws_pv.getCellByPosition(cr.col('PVLsigma'), r0 + 2).setFormula(f'={cI}{r+1}')
        ws_pv.getCellByPosition(cr.col('PVLdate'), r0 + 2).setFormula(
            f'=MAXIFS(OPdate;OPcompte;${cB}{r+2};OPdevise;${cD}{r+2};OPcatégorie;Solde)')
        ws_pv.getCellByPosition(cr.col('PVLmontant'), r0 + 2).setFormula(
            f'=SUMIFS(OPmontant;OPcompte;${cB}{r+2};OPdevise;${cD}{r+2};OPcatégorie;Solde;OPdate;{cJ}{r+2})')

        # --- Ligne r+3 : Retenu ---
        ws_pv.getCellByPosition(cr.col('PVLsection'), r0 + 3).setString('portefeuilles')
        ws_pv.getCellByPosition(cr.col('PVLcompte'), r0 + 3).setString(nom)
        ws_pv.getCellByPosition(cr.col('PVLtitre'), r0 + 3).setString('Retenu')
        ws_pv.getCellByPosition(cr.col('PVLdevise'), r0 + 3).setString(devise)
        ws_pv.getCellByPosition(cr.col('PVLpvl'), r0 + 3).setFormula(
            f'={cK}{r+3}-({cH}{r+3}+{cI}{r+3})')
        ws_pv.getCellByPosition(cr.col('PVLdate_init'), r0 + 3).setFormula(f'={cG}{r+2}')
        ws_pv.getCellByPosition(cr.col('PVLmontant_init'), r0 + 3).setFormula(f'={cH}{r+2}')
        ws_pv.getCellByPosition(cr.col('PVLsigma'), r0 + 3).setFormula(f'={cI}{r+2}')
        ws_pv.getCellByPosition(cr.col('PVLdate'), r0 + 3).setFormula(
            f'=IF({cJ}{r+1}>{cJ}{r+2};{cJ}{r+1};{cJ}{r+2})')
        ws_pv.getCellByPosition(cr.col('PVLmontant'), r0 + 3).setFormula(
            f'=IF({cJ}{r+1}>{cJ}{r+2};{cK}{r+1};{cK}{r+2})')

        # --- Ligne r+4 : vide (déjà vide par insertByIndex) ---

        # Étendre TOTAL portefeuilles pour devise non-EUR
        # Trouver la ligne TOTAL portefeuilles dans le footer (col A = SECTION)
        total_pf_row = None
        for scan in range(r + 4, r + 30):
            val_a = ws_pv.getCellByPosition(cr.col('PVLsection'), uno_row(scan)).getString().strip()
            if 'TOTAL portefeuilles' in val_a:
                total_pf_row = scan
                break
        if total_pf_row:
            self._update_pv_total_portefeuilles(ws_pv, total_pf_row, r + 3, devise, doc=doc)

        # --- Formats nombre sur les 4 lignes de données ---
        self._apply_pv_formats(ws_pv, doc, r, devise, section='portefeuilles', count=3)

    def _apply_pv_formats(self, ws_pv, doc, first_row, devise, section, count):
        """Applique les formats nombre sur les lignes PVL insérées.

        Args:
            first_row: première ligne (1-indexed), Total ou ligne unique
            devise: code devise du compte
            section: 'portefeuilles' ou autre
            count: nombre de lignes à formater (3 pour bloc, 0 pour simple)
        """
        from inc_excel_schema import uno_col, uno_row, ColResolver
        cr = ColResolver.from_uno(doc.document)
        from inc_formats import FORMATS_DEVISE, FORMAT_EUR, FORMAT_EUR_RED, GRIS, BLANC

        fmt_date = doc.register_number_format('DD/MM/YY')
        is_portefeuille = section == 'portefeuilles'
        is_non_eur = devise != 'EUR'

        # E/K : format devise de la ligne (rouge négatif)
        devise_fmt_str = FORMATS_DEVISE.get(devise, FORMAT_EUR)
        fmt_ek = doc.register_number_format(
            f'{devise_fmt_str};[RED]\\-{devise_fmt_str}' if devise != 'EUR'
            else FORMAT_EUR_RED)

        # H/I : devise native pour portefeuilles, EUR pour les autres
        if is_portefeuille:
            fmt_hi = doc.register_number_format(FORMATS_DEVISE.get(devise, FORMAT_EUR))
        else:
            fmt_hi = doc.register_number_format(FORMAT_EUR)

        # Colonnes D-K pour fond blanc (lignes données, pas pieds portefeuille)
        ALL_DK = [cr.col('PVLdevise'), cr.col('PVLpvl'), cr.col('PVLpct'),
                  cr.col('PVLdate_init'), cr.col('PVLmontant_init'), cr.col('PVLsigma'),
                  cr.col('PVLdate'), cr.col('PVLmontant')]
        # Colonnes qui seront grisées (non-EUR) — pas de blanc dessus
        if is_non_eur:
            gris_set = {cr.col('PVLdevise'), cr.col('PVLpvl'), cr.col('PVLmontant')}
            if is_portefeuille:
                gris_set |= {cr.col('PVLmontant_init'), cr.col('PVLsigma')}
        else:
            gris_set = set()
        # count=0 → ligne donnée unique
        # count=3 → bloc portefeuille : offset 0 = en-tête, offsets 1-3 = pieds
        is_bloc = count > 0

        for offset in range(count + 1):
            r0 = uno_row(first_row + offset)
            is_header = is_bloc and offset == 0
            is_pied = is_bloc and offset > 0
            # Dates G et J
            ws_pv.getCellByPosition(cr.col('PVLdate_init'), r0).NumberFormat = fmt_date
            ws_pv.getCellByPosition(cr.col('PVLdate'), r0).NumberFormat = fmt_date
            # PVL E et Solde K : devise de la ligne (rouge négatif)
            ws_pv.getCellByPosition(cr.col('PVLpvl'), r0).NumberFormat = fmt_ek
            ws_pv.getCellByPosition(cr.col('PVLmontant'), r0).NumberFormat = fmt_ek
            # H et I
            ws_pv.getCellByPosition(cr.col('PVLmontant_init'), r0).NumberFormat = fmt_hi
            ws_pv.getCellByPosition(cr.col('PVLsigma'), r0).NumberFormat = fmt_hi
            # Fond blanc D-K pour lignes données (hors pieds et en-tête)
            if not is_pied and not is_header:
                for c0 in ALL_DK:
                    if c0 not in gris_set:
                        ws_pv.getCellByPosition(c0, r0).CellBackColor = BLANC
            # Gris non-EUR (se superpose au blanc) — pas sur l'en-tête (cellules vides)
            if is_non_eur and not is_header:
                ws_pv.getCellByPosition(cr.col('PVLdevise'), r0).CellBackColor = GRIS
                for c0 in (cr.col('PVLpvl'), cr.col('PVLmontant')):
                    ws_pv.getCellByPosition(c0, r0).CellBackColor = GRIS
                if is_portefeuille:
                    ws_pv.getCellByPosition(cr.col('PVLmontant_init'), r0).CellBackColor = GRIS
                    ws_pv.getCellByPosition(cr.col('PVLsigma'), r0).CellBackColor = GRIS

    def _update_pv_total_portefeuilles(self, ws_pv, total_row, retenu_row=None, devise=None, doc=None):
        """Reconstruit les SUMIFS du TOTAL portefeuilles pour toutes les devises présentes.

        Scanne les lignes Retenu pour trouver les devises, puis génère :
          SUMIFS(...;"EUR") + SUMIFS(...;"USD")*cours_USD + ...
        Appelé à chaque ajout de portefeuille.
        """
        from inc_excel_schema import uno_col, uno_row, ColResolver
        cr = ColResolver.from_uno(doc.document)

        # Scanner les devises des lignes Retenu portefeuilles
        # Utiliser total_row comme borne (pas _end_pvl qui peut être stale en batch)
        devises = set()
        pvl_data = (self._start_pvl or 5) + 1
        pvl_scan_end = total_row
        for scan in range(pvl_data, pvl_scan_end):
            a = ws_pv.getCellByPosition(cr.col('PVLsection'), uno_row(scan)).getString().strip()
            c = ws_pv.getCellByPosition(cr.col('PVLtitre'), uno_row(scan)).getString().strip()
            d = ws_pv.getCellByPosition(cr.col('PVLdevise'), uno_row(scan)).getString().strip()
            if a == 'portefeuilles' and c == 'Retenu' and d:
                devises.add(d)

        # Construire la formule pour chaque colonne (H, I, K)
        # Pas de devises ou EUR seul → formule générique (sans filtre devise)
        # Multi-devise ou non-EUR → formule pondérée par devise
        needs_weighted = devises and (len(devises) > 1 or devises != {'EUR'})

        for nr_name in ('PVLmontant_init', 'PVLsigma', 'PVLmontant'):
            cl = cr.letter(nr_name)
            if needs_weighted:
                terms = []
                for dev in sorted(devises):
                    term = f'SUMIFS({cl}:{cl};A:A;"portefeuilles";C:C;"Retenu";D:D;"{dev}")'
                    if dev != 'EUR':
                        cours = self.cours_name(dev)
                        if cours:
                            term = f'{term}*{cours}'
                    terms.append(term)
                formula = '=' + '+'.join(terms)
            else:
                formula = f'=SUMIFS({cl}:{cl};A:A;"portefeuilles";C:C;"Retenu")'
            ws_pv.getCellByPosition(cr.col(nr_name), uno_row(total_row)).setFormula(formula)

    def _update_pv_bloc_total(self, ws_pv, account_name, total_row, doc=None):
        """Reconstruit les formules Total d'un bloc portefeuille si multi-devise.

        Scanne les titres (*...*) du bloc pour trouver les devises.
        Mono-devise : SUM(range) — pas de changement.
        Multi-devise : SUMIFS par devise avec conversion cours.
        """
        from inc_excel_schema import uno_col, uno_row
        cr = ColResolver.from_uno(doc.document)

        col_b = cr.col('PVLcompte')
        col_c = cr.col('PVLtitre')
        col_d = cr.col('PVLdevise')
        total_r0 = uno_row(total_row)

        # Scanner les titres du bloc (lignes *...* avant Total)
        devises = {}  # devise → [row_1indexed, ...]
        for scan in range(total_row - 1, 0, -1):
            b = ws_pv.getCellByPosition(col_b, uno_row(scan)).getString().strip()
            if b != account_name:
                break
            c = ws_pv.getCellByPosition(col_c, uno_row(scan)).getString().strip()
            if c.startswith('*') and c.endswith('*'):
                d = ws_pv.getCellByPosition(col_d, uno_row(scan)).getString().strip()
                if d:
                    devises.setdefault(d, []).append(scan)

        if len(devises) <= 1:
            return  # Mono-devise : SUM(range) suffit

        # Multi-devise : reconstruire avec SUMIFS par devise * cours
        first_titre = min(r for rows in devises.values() for r in rows)
        last_titre = max(r for rows in devises.values() for r in rows)

        for nr_name in ('PVLmontant_init', 'PVLsigma', 'PVLmontant'):
            cl = cr.letter(nr_name)
            terms = []
            for dev in sorted(devises.keys()):
                term = f'SUMIFS({cl}{first_titre}:{cl}{last_titre};D{first_titre}:D{last_titre};"{dev}")'
                if dev != 'EUR':
                    cours = self.cours_name(dev)
                    if cours:
                        term = f'{term}*{cours}'
                    else:
                        term = f'SUMIFS({cl}{first_titre}:{cl}{last_titre};D{first_titre}:D{last_titre};"{dev}")'
                terms.append(term)
            ws_pv.getCellByPosition(cr.col(nr_name), total_r0).setFormula('=' + '+'.join(terms))

        # DATE_INIT : MIN de toutes les dates titres
        cg = cr.letter('PVLdate_init')
        ws_pv.getCellByPosition(
            cr.col('PVLdate_init'), total_r0).setFormula(
            f'=MIN({cg}{first_titre}:{cg}{last_titre})')

    def _create_pv_simple_line(self, ws_pv, doc, acct, total_label):
        """Insère une ligne simple dans Plus_value (métaux, crypto, devises).

        Les TOTALs en pied utilisent SUMIFS → pas besoin d'étendre les formules.
        """
        from inc_uno import copy_row_style
        from inc_excel_schema import uno_col, uno_row
        cr = ColResolver.from_uno(doc.document)

        # Déduire la section depuis le total_label
        section_map = {
            'TOTAL métaux': 'métaux',
            'TOTAL crypto-monnaies': 'crypto',
            'TOTAL devises': 'devises',
        }
        section = section_map.get(total_label)
        if not section:
            return

        # Trouver la dernière ligne de cette section
        section_label = self.PV_SECTION_LABELS.get(section)
        last_section_row = None
        for scan in range(200, 0, -1):
            val_a = ws_pv.getCellByPosition(cr.col('PVLsection'), uno_row(scan)).getString().strip()
            if val_a == section or val_a == section_label:
                last_section_row = scan
                break
        if not last_section_row:
            # Section vide → trouver la section précédente et insérer après
            section_order = ['portefeuilles', 'métaux', 'crypto', 'devises']
            idx = section_order.index(section)
            for prev_idx in range(idx - 1, -1, -1):
                prev_section = section_order[prev_idx]
                prev_label = self.PV_SECTION_LABELS.get(prev_section)
                for scan in range(200, 0, -1):
                    val_a = ws_pv.getCellByPosition(cr.col('PVLsection'), uno_row(scan)).getString().strip()
                    if val_a == prev_section or val_a == prev_label:
                        last_section_row = scan
                        break
                if last_section_row:
                    break
            if not last_section_row:
                # Template vierge : insérer avant la ligne TOTAL correspondante
                col_b = cr.col('PVLcompte')
                pvl_data = (self._start_pvl or 5) + 1
            for scan in range(pvl_data, pvl_data + 200):
                val_b = ws_pv.getCellByPosition(col_b, uno_row(scan)).getString().strip()
                if val_b == total_label:
                    last_section_row = scan - 1
                    break
            if not last_section_row:
                    return

        nom = acct['intitule']
        devise = acct.get('devise') or ''

        # Lettres de colonnes via PvCol
        cB = cr.letter('PVLcompte')
        cD = cr.letter('PVLdevise')
        cG = cr.letter('PVLdate_init')
        cH = cr.letter('PVLmontant_init')
        cI = cr.letter('PVLsigma')
        cK = cr.letter('PVLmontant')

        # Insérer 1 ligne après la dernière ligne de la section
        insert_row = last_section_row + 1
        insert_0 = uno_row(insert_row)
        ws_pv.Rows.insertByIndex(insert_0, 1)

        r = insert_row  # 1-indexed
        r0 = uno_row(r)

        # Template de style : ligne au-dessus
        template_0 = r0 - 1 if r0 > 0 else r0
        copy_row_style(ws_pv, template_0, r0, col_start=0, col_end=12)

        # Remplir
        ws_pv.getCellByPosition(cr.col('PVLsection'), r0).setString(section)
        ws_pv.getCellByPosition(cr.col('PVLcompte'), r0).setString(nom)
        ws_pv.getCellByPosition(cr.col('PVLdevise'), r0).setString(devise)
        ws_pv.getCellByPosition(cr.col('PVLpvl'), r0).setFormula(
            f'={cK}{r}-({cH}{r}+{cI}{r})')
        ws_pv.getCellByPosition(cr.col('PVLdate_init'), r0).setFormula(
            f'=MINIFS(OPdate;OPcompte;{cB}{r};OPdevise;{cD}{r};OPcatégorie;Solde)')
        ws_pv.getCellByPosition(cr.col('PVLmontant_init'), r0).setValue(0)
        ws_pv.getCellByPosition(cr.col('PVLsigma'), r0).setFormula(
            f'=SUMIFS(OPequiv_euro;OPcompte;{cB}{r};OPdevise;{cD}{r};OPcatégorie;"<>"&Spéciale;OPdate;">="&{cG}{r})')
        ws_pv.getCellByPosition(cr.col('PVLdate'), r0).setFormula(
            f'=SUMIF(AVRintitulé;${cB}{r};AVRdate_solde)')
        ws_pv.getCellByPosition(cr.col('PVLmontant'), r0).setFormula(
            f'=SUMIF(AVRintitulé;${cB}{r};AVRmontant_solde_euro)')

        # Formats nombre
        self._apply_pv_formats(ws_pv, doc, r, devise, section=section, count=0)

    def _insert_pv_title(self, account_name, title_name, devise, date_init, doc=None):
        """Insère un nouveau titre dans un bloc Portefeuille de Plus_value via UNO.

        Args:
            doc: UnoDocument ouvert (mode batch). Si None, ouvre/ferme automatiquement.
        """
        from contextlib import nullcontext
        from inc_uno import UnoDocument, copy_row_style
        from inc_excel_schema import uno_row, uno_col
        import re

        bak_path = self.xlsx_path.with_suffix('.xlsm.bak')
        shutil.copy2(self.xlsx_path, bak_path)

        owned = doc is None
        ctx = UnoDocument(self.xlsx_path) if owned else nullcontext(doc)
        with ctx as doc:
            cr = ColResolver.from_uno(doc.document)
            cB = cr.letter('PVLcompte')
            cC = cr.letter('PVLtitre')
            cD = cr.letter('PVLdevise')
            cG = cr.letter('PVLdate_init')
            cH = cr.letter('PVLmontant_init')
            cI = cr.letter('PVLsigma')
            cK = cr.letter('PVLmontant')
            ws_pv = doc.get_sheet(SHEET_PLUS_VALUE)

            # --- Trouver le bloc : header_row et total_row ---
            header_row = None
            total_row = None
            col_b = cr.col('PVLcompte')
            col_c = cr.col('PVLtitre')
            for scan in range(1, 200):
                val_b = ws_pv.getCellByPosition(col_b, uno_row(scan)).getString().strip()
                val_c = ws_pv.getCellByPosition(col_c, uno_row(scan)).getString().strip()
                if val_b != account_name:
                    continue
                if header_row is None and val_c != 'Total':
                    header_row = scan
                if val_c == 'Total':
                    total_row = scan
                    break

            if not total_row:
                raise ValueError(
                    f"Ligne 'Total' introuvable pour « {account_name} » "
                    f"dans Plus_value")

            # --- Trouver les titres existants (C = *...*) entre header et Total ---
            first_title_row = None
            for scan in range(header_row + 1, total_row):
                val_c = ws_pv.getCellByPosition(col_c, uno_row(scan)).getString().strip()
                if val_c.startswith('*') and val_c.endswith('*'):
                    if first_title_row is None:
                        first_title_row = scan

            has_existing = first_title_row is not None

            # --- Insérer 1 ligne à total_row (avant Total) ---
            ws_pv.Rows.insertByIndex(uno_row(total_row), 1)
            new_total_row = total_row + 1  # Total décalé de 1

            r = total_row   # 1-indexed, la nouvelle ligne titre
            r0 = uno_row(r)

            # --- Copier le style ---
            if has_existing:
                template_0 = uno_row(r - 1)
            else:
                template_0 = None
                for scan in range(new_total_row + 1, new_total_row + 4):
                    val_c = ws_pv.getCellByPosition(col_c, uno_row(scan)).getString().strip()
                    if val_c == '#Solde Opérations':
                        template_0 = uno_row(scan)
                        break
                if template_0 is None:
                    pvl_data2 = (self._start_pvl or 5) + 1
                    col_d = cr.col('PVLdevise')
                    for scan in range(pvl_data2, self._end_pvl + 1):
                        val_c = ws_pv.getCellByPosition(col_c, uno_row(scan)).getString().strip()
                        val_d = ws_pv.getCellByPosition(col_d, uno_row(scan)).getString().strip()
                        if val_c.startswith('*') and val_c.endswith('*') and val_d == devise:
                            template_0 = uno_row(scan)
                            break
                if template_0 is None:
                    template_0 = uno_row(header_row) if header_row else r0 - 1
            copy_row_style(ws_pv, template_0, r0, col_start=0, col_end=12)

            # --- Remplir la ligne titre ---
            ws_pv.getCellByPosition(cr.col('PVLsection'), r0).setString('portefeuilles')
            ws_pv.getCellByPosition(cr.col('PVLcompte'), r0).setString(account_name)
            ws_pv.getCellByPosition(cr.col('PVLtitre'), r0).setString(f'*{title_name}*')
            ws_pv.getCellByPosition(cr.col('PVLdevise'), r0).setString(devise)
            ws_pv.getCellByPosition(cr.col('PVLpvl'), r0).setFormula(
                f'={cK}{r}-({cH}{r}+{cI}{r})')

            # G = date initiale (serial date) ou 0
            if date_init:
                serial = (date_init - datetime(1899, 12, 30)).days
                ws_pv.getCellByPosition(cr.col('PVLdate_init'), r0).setValue(serial)
            else:
                ws_pv.getCellByPosition(cr.col('PVLdate_init'), r0).setValue(0)

            ws_pv.getCellByPosition(cr.col('PVLmontant_init'), r0).setValue(0)

            # I = formule SIGMA
            if has_existing:
                h_src = ws_pv.getCellByPosition(
                    cr.col('PVLsigma'), uno_row(first_title_row)).getFormula()
                h_formula = re.sub(
                    r'(\$?[A-Z]+)' + str(first_title_row) + r'(?!\d)',
                    lambda m: m.group(1) + str(r),
                    h_src)
                ws_pv.getCellByPosition(cr.col('PVLsigma'), r0).setFormula(h_formula)
            else:
                ws_pv.getCellByPosition(cr.col('PVLsigma'), r0).setFormula(
                    f'=SUMIFS(OPmontant;OPcompte;${cB}{r};OPdevise;${cD}{r}'
                    f';OPcatégorie;"<>"&Spéciale'
                    f';OPlibellé;${cC}{r};OPdate;">="&${cG}{r})')

            ws_pv.getCellByPosition(cr.col('PVLdate'), r0).setValue(0)
            ws_pv.getCellByPosition(cr.col('PVLmontant'), r0).setValue(0)

            # --- Mettre à jour les formules Total ---
            total_r0 = uno_row(new_total_row)

            if has_existing:
                for pv_col, func in [
                    ('PVLdate_init', 'MIN'),
                    ('PVLmontant_init', 'SUM'),
                    ('PVLsigma', 'SUM'),
                    ('PVLmontant', 'SUM'),
                ]:
                    ci = cr.col(pv_col)
                    formula = ws_pv.getCellByPosition(ci, total_r0).getFormula()
                    # Pattern 1 : range
                    m = re.match(
                        rf'(={func}\([A-Z]+\d+:[A-Z]+)(\d+)(\))', formula)
                    if m:
                        new_f = f'{m.group(1)}{int(m.group(2)) + 1}{m.group(3)}'
                        ws_pv.getCellByPosition(ci, total_r0).setFormula(new_f)
                        continue
                    # Pattern 2 : cellule unique
                    m2 = re.match(
                        rf'(={func}\()([A-Z]+)(\d+)(\))', formula)
                    if m2:
                        col_l = m2.group(2)
                        new_f = (f'{m2.group(1)}{col_l}{m2.group(3)}'
                                 f':{col_l}{r}{m2.group(4)}')
                        ws_pv.getCellByPosition(ci, total_r0).setFormula(new_f)
            else:
                # Premier titre : créer les formules du Total
                ws_pv.getCellByPosition(cr.col('PVLdate_init'), total_r0).setFormula(
                    f'=MIN({cG}{r}:{cG}{r})')
                ws_pv.getCellByPosition(cr.col('PVLmontant_init'), total_r0).setFormula(
                    f'=SUM({cH}{r}:{cH}{r})')
                ws_pv.getCellByPosition(cr.col('PVLsigma'), total_r0).setFormula(
                    f'=SUM({cI}{r}:{cI}{r})')
                ws_pv.getCellByPosition(cr.col('PVLdate'), total_r0).setFormula(
                    f'={cr.letter("PVLdate")}{r}')
                ws_pv.getCellByPosition(cr.col('PVLmontant'), total_r0).setFormula(
                    f'=SUM({cK}{r}:{cK}{r})')

            # Reconstruire les formules Total du bloc si multi-devise
            self._update_pv_bloc_total(ws_pv, account_name, new_total_row, doc=doc)

            # Formats nombre sur la ligne titre
            self._apply_pv_formats(ws_pv, doc, r, devise, section='portefeuilles', count=0)

            if owned:
                self._uno_finalize(doc)

    def _delete_pv_entries(self, ws_pv, deleted_names):
        """Supprime les entrées Plus_value des comptes supprimés.

        Détecte automatiquement bloc multi-lignes (Portefeuille, Assurance-vie...)
        vs ligne simple (métaux/crypto/devises).
        Les TOTALs en pied utilisent SUMIFS → pas besoin de les mettre à jour.
        """
        from inc_excel_schema import uno_row, uno_col, ColResolver
        cr = ColResolver.from_uno(doc.document)

        deleted_set = set(deleted_names)
        col_b = cr.col('PVLcompte')
        col_c = cr.col('PVLtitre')

        for name in list(deleted_set):
            # Scanner toute la feuille pour trouver la première occurrence col B (Compte)
            first_row = None
            for scan in range(1, 200):
                val = ws_pv.getCellByPosition(col_b, uno_row(scan)).getString().strip()
                if val == name:
                    first_row = scan
                    break
            if first_row is None:
                continue

            # Scanner toutes les lignes consécutives du compte (bloc ou ligne unique)
            block_rows = [first_row]
            for scan in range(first_row + 1, first_row + 50):
                val_b = ws_pv.getCellByPosition(col_b, uno_row(scan)).getString().strip()
                val_c = ws_pv.getCellByPosition(col_c, uno_row(scan)).getString().strip()
                if val_b == name:
                    block_rows.append(scan)
                elif not val_b and not val_c:
                    # Vérifier si c'est un spacer structurel (avant section header,
                    # TOTAL, ou END marker ✓ qui est sur col A)
                    next_a = ws_pv.getCellByPosition(
                        cr.col('PVLsection'), uno_row(scan + 1)).getString().strip()
                    next_b = ws_pv.getCellByPosition(
                        col_b, uno_row(scan + 1)).getString().strip()
                    if (next_a.startswith('Les ') or 'TOTAL' in next_b.upper()
                            or next_a == '✓'):
                        break  # Spacer structurel → préserver
                    # Ligne vide après le bloc → inclure puis stop
                    block_rows.append(scan)
                    break
                else:
                    break

            # Supprimer tout le bloc en une opération
            # Les TOTALs en pied avec SUMIFS se recalculent automatiquement
            count = len(block_rows)
            ws_pv.Rows.removeByIndex(uno_row(first_row), count)

        # Nettoyer les doubles spacers (2 lignes vides consécutives après un header section)
        col_a = cr.col('PVLsection')
        for scan in range(200, 5, -1):
            a2 = ws_pv.getCellByPosition(col_a, uno_row(scan - 2)).getString().strip()
            if not a2.startswith('Les '):
                continue
            a1 = ws_pv.getCellByPosition(col_a, uno_row(scan - 1)).getString().strip()
            b1 = ws_pv.getCellByPosition(col_b, uno_row(scan - 1)).getString().strip()
            a0 = ws_pv.getCellByPosition(col_a, uno_row(scan)).getString().strip()
            b0 = ws_pv.getCellByPosition(col_b, uno_row(scan)).getString().strip()
            if not a1 and not b1 and not a0 and not b0:
                ws_pv.Rows.removeByIndex(uno_row(scan), 1)

    @staticmethod
    def _cleanup_model_rows(ws, first_row, end_row, check_cols):
        """Supprime les model rows vides entre first_row et end_row (1-indexed).

        Une model row est une ligne où TOUTES les colonnes check_cols sont vides.
        Utilisé pour nettoyer les lignes de template après insertion de vrais comptes.
        """
        from inc_excel_schema import uno_row
        rows_to_del = []
        for r in range(first_row, end_row):
            all_empty = all(
                not ws.getCellByPosition(c, uno_row(r)).getString().strip()
                for c in check_cols
            )
            if all_empty:
                rows_to_del.append(r)
        for r in reversed(rows_to_del):
            ws.Rows.removeByIndex(uno_row(r), 1)
        return len(rows_to_del)

    @staticmethod
    def _cleanup_model_rows_ops(ws_ops, cr=None):
        """Supprime les model rows vides dans Opérations."""
        from inc_excel_schema import uno_row, uno_col
        cursor = ws_ops.createCursor()
        cursor.gotoEndOfUsedArea(True)
        last_0 = cursor.getRangeAddress().EndRow
        rows_to_del = []
        for r0 in range(4, last_0 + 1):  # row 5 = 0-indexed 4, skip model_head
            date_val = ws_ops.getCellByPosition(cr.col('OPdate'), r0).getString().strip()
            compte_val = ws_ops.getCellByPosition(cr.col('OPcompte'), r0).getString().strip()
            if not date_val and not compte_val:
                rows_to_del.append(r0)
        for r0 in reversed(rows_to_del):
            ws_ops.Rows.removeByIndex(r0, 1)

    def _save_accounts(self, doc=None):
        """Sauvegarde les comptes dans Avoirs + contrôle dans Contrôles (via UNO).

        Args:
            doc: UnoDocument ouvert (mode batch). Si None, ouvre/ferme automatiquement.
        """
        from contextlib import nullcontext
        from inc_uno import UnoDocument, copy_row_style, get_named_range_pos
        from inc_excel_schema import uno_col, uno_row

        # Backup
        bak_path = self.xlsx_path.with_suffix('.xlsm.bak')
        shutil.copy2(self.xlsx_path, bak_path)

        owned = doc is None
        ctx = UnoDocument(self.xlsx_path) if owned else nullcontext(doc)
        with ctx as doc:
            cr = ColResolver.from_uno(doc.document)
            ws = doc.get_sheet(SHEET_AVOIRS)
            ws_ctrl = doc.get_sheet(SHEET_CONTROLES)
            ws_ops = doc.get_sheet(SHEET_OPERATIONS)

            # --- Trouver la ligne Total actuelle (1-indexed) ---
            # Scanner au-delà de _end_avr : Total est après la dernière donnée
            avr_data = self._start_avr + 1
            total_row = (self._end_avr + 1) if self._end_avr else None

            # --- Avoirs : supprimer les lignes des comptes supprimés ---
            if self._deleted_accounts:
                deleted_set = set(self._deleted_accounts)
                av_rows_to_delete = []
                for row_idx in range(avr_data, total_row or self._end_avr + 1):
                    val = ws.getCellByPosition(cr.col('AVRintitulé'), uno_row(row_idx)).getString()
                    if val and val.strip() in deleted_set:
                        av_rows_to_delete.append(row_idx)
                # Supprimer en ordre inverse (indices hauts restent valides)
                for row_idx in reversed(av_rows_to_delete):
                    ws.Rows.removeByIndex(uno_row(row_idx), 1)
                # Ajuster total_row
                total_row -= len(av_rows_to_delete)
                # Ajuster acct['row'] pour les comptes existants restants
                for a in self.accounts_data:
                    if a.get('row') is not None:
                        shift = sum(1 for d in av_rows_to_delete if d < a['row'])
                        a['row'] -= shift

            # --- Avoirs : mettre à jour les champs éditables des comptes existants ---
            for acct in self.accounts_data:
                r = acct.get('row')
                if r is None or acct.get('_is_new'):
                    continue
                r0 = uno_row(r)
                ws.getCellByPosition(cr.col('AVRdomiciliation'), r0).setString(acct.get('domiciliation') or '')
                ws.getCellByPosition(cr.col('AVRsous_type'), r0).setString(acct.get('sous_type') or '')
                ws.getCellByPosition(cr.col('AVRtitulaire'), r0).setString(acct.get('titulaire') or '')
                ws.getCellByPosition(cr.col('AVRpropriete'), r0).setString(acct.get('propriete') or '')

            # --- Avoirs : insérer les nouveaux comptes dans le bloc Type correspondant ---
            new_accounts = [a for a in self.accounts_data if a.get('row') is None]
            if new_accounts and total_row:
                template_row_0 = uno_row(avr_data)

                for acct in new_accounts:
                    target_type = acct['type']

                    # Scanner pour trouver la dernière ligne du bloc Type
                    # (ou la première ligne vide/Clos après le bloc)
                    block_last = None
                    first_empty_or_clos = None
                    for row_idx in range(avr_data, total_row):
                        val_a = ws.getCellByPosition(cr.col('AVRintitulé'), uno_row(row_idx)).getString().strip()
                        val_b = ws.getCellByPosition(cr.col('AVRtype'), uno_row(row_idx)).getString().strip()
                        if val_b == target_type:
                            block_last = row_idx
                        if (not val_a or val_b == 'Clos') and first_empty_or_clos is None:
                            first_empty_or_clos = row_idx

                    # Point d'insertion : après la dernière ligne du bloc, avant END (✓)
                    avr_end_model = self._end_avr or total_row
                    if block_last is not None:
                        insert_row = min(block_last + 1, avr_end_model)
                    elif first_empty_or_clos is not None:
                        insert_row = min(first_empty_or_clos, avr_end_model)
                    else:
                        insert_row = avr_end_model

                    # Insérer 1 ligne
                    ws.Rows.insertByIndex(uno_row(insert_row), 1)
                    # Décaler les row en mémoire pour les comptes existants >= insert_row
                    for a in self.accounts_data:
                        if a.get('row') is not None and a['row'] >= insert_row:
                            a['row'] += 1
                    total_row += 1
                    if self._end_avr and insert_row <= self._end_avr:
                        self._end_avr += 1
                    # Décaler template_row_0 si l'insertion l'a poussé vers le bas
                    if uno_row(insert_row) <= template_row_0:
                        template_row_0 += 1

                    r = insert_row   # 1-indexed
                    r0 = uno_row(r)  # 0-indexed
                    acct['row'] = r

                    # Copier style du template (ligne 4)
                    copy_row_style(ws, template_row_0, r0, col_start=0, col_end=12)

                    # Écrire les données
                    ws.getCellByPosition(cr.col('AVRintitulé'), r0).setString(acct['intitule'])
                    ws.getCellByPosition(cr.col('AVRtype'), r0).setString(acct['type'])
                    ws.getCellByPosition(cr.col('AVRdomiciliation'), r0).setString(acct.get('domiciliation') or '')
                    ws.getCellByPosition(cr.col('AVRsous_type'), r0).setString(acct.get('sous_type') or '')
                    ws.getCellByPosition(cr.col('AVRdevise'), r0).setString(acct.get('devise') or '')
                    ws.getCellByPosition(cr.col('AVRtitulaire'), r0).setString(acct.get('titulaire') or '')
                    ws.getCellByPosition(cr.col('AVRpropriete'), r0).setString(acct.get('propriete') or '')
                    # DATE_ANTER et MONTANT_ANTER
                    if acct.get('date_anter'):
                        from datetime import datetime
                        epoch = datetime(1899, 12, 30)
                        serial = (acct['date_anter'] - epoch).days
                        cell_h = ws.getCellByPosition(cr.col('AVRdate_anter'), r0)
                        cell_h.setValue(serial)
                        cell_h.NumberFormat = doc.register_number_format('DD/MM/YYYY')
                    if acct.get('montant_anter') is not None:
                        from inc_formats import FORMAT_EUR
                        cell_i = ws.getCellByPosition(cr.col('AVRmontant_anter'), r0)
                        cell_i.setValue(acct['montant_anter'])
                        cell_i.NumberFormat = doc.register_number_format(FORMAT_EUR)

                    # Formules J/K/L (UNO : pas de préfixe _xlfn.)
                    devise = acct['devise']
                    acct_type = acct['type']
                    if acct_type == 'Portefeuilles':
                        ws.getCellByPosition(cr.col('AVRdate_solde'), r0).setFormula(
                            f'=SUMIFS(PVLdate;PVLcompte;$A{r};PVLtitre;Retenu)')
                        ws.getCellByPosition(cr.col('AVRmontant_solde'), r0).setFormula(
                            f'=SUMIFS(PVLmontant;PVLcompte;$A{r};PVLtitre;Retenu)')
                    elif acct_type == 'Biens matériels' and not devise:
                        # Bien matériel sans devise (immobilier, véhicules) : montant statique
                        montant = acct.get('montant_debut')
                        if montant is not None:
                            ws.getCellByPosition(cr.col('AVRmontant_solde'), r0).setValue(float(montant))
                    elif devise:
                        ws.getCellByPosition(cr.col('AVRdate_solde'), r0).setFormula(
                            f'=MAXIFS(OPdate;OPcompte;$A{r};OPdevise;$E{r};OPcatégorie;Solde)')
                        ws.getCellByPosition(cr.col('AVRmontant_solde'), r0).setFormula(
                            f'=SUMIFS(OPmontant;OPcompte;$A{r};OPdevise;$E{r};OPcatégorie;Solde;OPdate;$J{r})')

                    cours = self.cours_name(devise)
                    if cours:
                        lK = cr.letter('AVRmontant_solde')
                        ws.getCellByPosition(cr.col('AVRmontant_solde_euro'), r0).setFormula(f'={lK}{r}*{cours}')
                    elif devise in ('EUR', None, ''):
                        lK = cr.letter('AVRmontant_solde')
                        ws.getCellByPosition(cr.col('AVRmontant_solde_euro'), r0).setFormula(f'={lK}{r}')

                    # Format date sur J
                    j_cell = ws.getCellByPosition(cr.col('AVRdate_solde'), r0)
                    j_cell.NumberFormat = doc.register_number_format('DD/MM/YY')

                    # Format nombre sur K si devise spécifique
                    k_fmt_str = self.AVOIRS_K_FORMATS.get(devise)
                    if k_fmt_str:
                        k_cell = ws.getCellByPosition(cr.col('AVRmontant_solde'), r0)
                        k_cell.NumberFormat = doc.register_number_format(k_fmt_str)

                    # Format EUR sur L (Equiv EUR)
                    from inc_formats import FORMAT_EUR
                    l_cell = ws.getCellByPosition(cr.col('AVRmontant_solde_euro'), r0)
                    l_cell.NumberFormat = doc.register_number_format(FORMAT_EUR)
                    # Fond gris pour non-EUR (devise E + montant K)
                    if devise and devise not in ('EUR', ''):
                        ws.getCellByPosition(cr.col('AVRdevise'), r0).CellBackColor = 0xDCDCDC
                        ws.getCellByPosition(cr.col('AVRmontant_solde'), r0).CellBackColor = 0xDCDCDC

            # (recalibration AVR* + START/END_AVR déplacée après cleanup model rows)

            # --- Contrôles : supprimer les lignes des comptes supprimés ---
            ctrl_rows_to_delete = sorted(self._deleted_ctrl_rows)
            # Supprimer en ordre inverse (préserve le pied de page)
            for row_idx in reversed(ctrl_rows_to_delete):
                ws_ctrl.Rows.removeByIndex(uno_row(row_idx), 1)
            # Ajuster ctrl_row pour les display_accounts restants
            for entry in self.display_accounts:
                crow = entry.get('ctrl_row')
                if crow is not None:
                    shift = sum(1 for d in ctrl_rows_to_delete if d < crow)
                    entry['ctrl_row'] = crow - shift

            # --- Contrôles : mettre à jour formules + flags contrôle ---
            for entry in self.display_accounts:
                ctrl_row = entry.get('ctrl_row')
                if not ctrl_row:
                    continue
                avoirs_acct = entry['avoirs_ref']
                new_avoirs_row = avoirs_acct.get('row')
                if new_avoirs_row:
                    ws_ctrl.getCellByPosition(
                        cr.col('CTRL1compte'), uno_row(ctrl_row)
                    ).setFormula(f'=Avoirs.{cr.letter("AVRintitulé")}{new_avoirs_row}')
                ws_ctrl.getCellByPosition(
                    cr.col('CTRL1controle'), uno_row(ctrl_row)
                ).setString('Oui' if entry['controle'] else 'Non')

            # --- Contrôles : créer les lignes manquantes pour nouveaux comptes ---
            ctrl_last_data = max(
                (e['ctrl_row'] for e in self.display_accounts if e.get('ctrl_row')),
                default=self._start_ctrl1 + 1)
            # Borner par la model row END (✓) — ne jamais insérer après
            ctrl_end_model = self._end_ctrl1 or (ctrl_last_data + 1)
            ctrl_next_row = min(ctrl_last_data + 1, ctrl_end_model)

            for entry in self.display_accounts:
                if entry.get('ctrl_row') is not None:
                    continue
                # Biens matériels : exclus de CTRL1 (pas d'opérations, valeurs manuelles)
                if entry.get('type') == 'Biens matériels':
                    continue
                avoirs_acct = entry['avoirs_ref']
                avoirs_row = avoirs_acct.get('row')
                if not avoirs_row:
                    continue

                r = ctrl_next_row
                r0 = uno_row(r)
                devise = entry['devise']

                # Insérer une ligne avant END (✓)
                # Le style est propagé automatiquement depuis la model row
                ws_ctrl.Rows.insertByIndex(r0, 1)
                if self._end_ctrl1 and r <= self._end_ctrl1:
                    self._end_ctrl1 += 1

                # Formules nouveau modèle : ancrage min + relevé max via XLOOKUP
                # Tolère 0..N #Solde, doublons même date résolus déterministiquement
                ws_ctrl.getCellByPosition(cr.col('CTRL1compte'), r0).setFormula(
                    f'=Avoirs.A{avoirs_row}')
                if devise:
                    ws_ctrl.getCellByPosition(cr.col('CTRL1devise'), r0).setString(devise)
                # Col C = date ancrage : MINIFS si >=2 #Solde, sinon 0 (epoch)
                ws_ctrl.getCellByPosition(cr.col('CTRL1date_ancrage'), r0).setFormula(
                    f'=IF(COUNTIFS(OPcompte;$A{r};OPdevise;$B{r};OPcatégorie;Solde)>=2;'
                    f'MINIFS(OPdate;OPcompte;$A{r};OPdevise;$B{r};OPcatégorie;Solde);0)')
                # Col D = date relevé : MAXIFS sur les #Solde
                ws_ctrl.getCellByPosition(cr.col('CTRL1date_releve'), r0).setFormula(
                    f'=MAXIFS(OPdate;OPcompte;$A{r};OPdevise;$B{r};OPcatégorie;Solde)')
                # Col E = montant ancrage : XLOOKUP first occurrence à date C, 0 si C=0
                ws_ctrl.getCellByPosition(cr.col('CTRL1montant_ancrage'), r0).setFormula(
                    f'=IF($C{r}=0;0;'
                    f'XLOOKUP(1;(OPcompte=$A{r})*(OPdevise=$B{r})*(OPcatégorie=Solde)*(OPdate=$C{r});'
                    f'OPmontant;0;0;1))')
                # Col F = solde calculé : montant_ancrage + flux entre C (excl) et D (incl)
                ws_ctrl.getCellByPosition(cr.col('CTRL1solde_calc'), r0).setFormula(
                    f'=$E{r}+SUMIFS(OPmontant;OPcompte;$A{r};OPdevise;$B{r};'
                    f'OPdate;">"&$C{r};OPdate;"<="&$D{r};'
                    f'OPcatégorie;"<>Solde";OPcatégorie;"<>"&Spéciale)')
                # Col G = montant relevé : XLOOKUP last occurrence à date D
                ws_ctrl.getCellByPosition(cr.col('CTRL1montant_releve'), r0).setFormula(
                    f'=XLOOKUP(1;(OPcompte=$A{r})*(OPdevise=$B{r})*(OPcatégorie=Solde)*(OPdate=$D{r});'
                    f'OPmontant;0;0;-1)')
                # Col H = écart : relevé - calculé, tolérance 1 centime
                ws_ctrl.getCellByPosition(cr.col('CTRL1ecart'), r0).setFormula(
                    f'=IF(ABS($G{r}-$F{r})<0.015;0;ROUND($G{r}-$F{r};2))')
                # Col I = Oui/Non
                ws_ctrl.getCellByPosition(cr.col('CTRL1controle'), r0).setString(
                    'Oui' if entry['controle'] else 'Non')

                # Format nombre sur E/F/G/H si devise spécifique
                k_fmt_str = self.AVOIRS_K_FORMATS.get(devise)
                if k_fmt_str:
                    fmt_key = doc.register_number_format(k_fmt_str)
                    for c in (cr.col('CTRL1montant_ancrage'), cr.col('CTRL1solde_calc'),
                              cr.col('CTRL1montant_releve'), cr.col('CTRL1ecart')):
                        ws_ctrl.getCellByPosition(c, r0).NumberFormat = fmt_key
                # GRIS pour non-EUR ; blanc explicite pour EUR
                # (insertByIndex peut propager le GRIS d'une ligne voisine non-EUR ;
                # transparent -1 prendrait la couleur du thème = noir en mode sombre)
                bg_color = 0xDCDCDC if (devise and devise not in ('EUR', '')) else 0xFFFFFF
                for c in (cr.col('CTRL1montant_ancrage'), cr.col('CTRL1solde_calc'),
                          cr.col('CTRL1montant_releve'), cr.col('CTRL1ecart')):
                    ws_ctrl.getCellByPosition(c, r0).CellBackColor = bg_color

                entry['ctrl_row'] = r
                ctrl_next_row += 1

            # Model rows CTRL1 : préservées (pas de cleanup)
            # Elles servent d'ancrage aux named ranges et de modèle de format.

            # --- Recalibrer formules CTRL2 sur les bornes CTRL1 (model rows incluses) ---
            # On couvre START_CTRL1..END_CTRL1 pour matcher le template (B3:B4 vide)
            # plutôt que de pointer le data range qui peut être vide.
            from inc_uno import get_table_bounds_uno
            ctrl_start_now, ctrl_end_now = get_table_bounds_uno(doc.document, 'CTRL1')
            ctrl_end_now = ctrl_end_now or self._end_ctrl1 or ctrl_next_row
            f = ctrl_start_now or self._start_ctrl1
            l = ctrl_end_now

            # -- CTRL2 h+2 COMPTES : COUNTIFS par devise --
            ctrl2_pos = get_named_range_pos(doc.document, 'START_CTRL2')
            if ctrl2_pos:
                h2_row_0 = ctrl2_pos[2]       # 0-indexed row of h+2
                h0_row_0 = h2_row_0 - 2       # 0-indexed row of h+0 (header devises)
                first_col_0 = ctrl2_pos[1]     # 0-indexed col of first devise
                h1 = h0_row_0 + 1              # 1-indexed header row
                for col_0 in range(first_col_0, first_col_0 + 30):
                    code = ws_ctrl.getCellByPosition(col_0, h0_row_0).getString().strip()
                    if not code:
                        break
                    cl = ColResolver._idx_to_letter(col_0 + 1)  # 1-indexed → lettre
                    # h+2 : COMPTES
                    ws_ctrl.getCellByPosition(col_0, h2_row_0).setFormula(
                        f'=COUNTIFS($B{f}:$B{l};{cl}${h1};$I{f}:$I{l};"Oui")'
                        f'-COUNTIFS($B{f}:$B{l};{cl}${h1};$I{f}:$I{l};"Oui";$H{f}:$H{l};0)')
                    # h+4 : Dates — vider (seul O général est pertinent)
                    ws_ctrl.getCellByPosition(col_0, h2_row_0 + 2).setString('')

            # --- Plus_value : supprimer / créer ---
            new_pv_accounts = [a for a in self.accounts_data if a.get('_is_new')]
            need_pv = self._deleted_accounts or new_pv_accounts
            ws_pv = doc.get_sheet(SHEET_PLUS_VALUE) if need_pv else None

            if self._deleted_accounts and ws_pv:
                pass  # PvCol removed — using cr.col() instead
                self._delete_pv_entries(ws_pv, self._deleted_accounts)
                # Reconstruire la formule TOTAL portefeuilles (peut redevenir générique)
                pvl_data = (self._start_pvl or 5) + 1
                for scan in range(pvl_data, pvl_data + 300):
                    val_a = ws_pv.getCellByPosition(
                        cr.col('PVLsection'), uno_row(scan)).getString().strip()
                    if 'TOTAL portefeuilles' in val_a:
                        self._update_pv_total_portefeuilles(ws_pv, scan, doc=doc)
                        break

            if new_pv_accounts and ws_pv:
                for acct in new_pv_accounts:
                    pv_type, pv_total = self._get_pv_section_for_account(
                        acct['type'], acct.get('devise'))

                    if pv_type == 'portfolio':
                        self._create_pv_portfolio_block(ws_pv, doc, acct)
                    elif pv_type == 'line':
                        self._create_pv_simple_line(ws_pv, doc, acct, pv_total)

            # --- Opérations : créer 2 #Solde (début + fin) pour les nouveaux comptes ---
            # Exclure les biens matériels (pas d'opérations, valeurs manuelles)
            new_ops_accounts = [e for e in self.display_accounts
                                if 'formula_j' not in e.get('avoirs_ref', {})
                                and e.get('avoirs_ref', {}).get('type') != 'Biens matériels']
            for entry in new_ops_accounts:
                avoirs_ref = entry.get('avoirs_ref', {})
                self._append_solde_lines(
                    ws_ops, entry['intitule'], entry['devise'],
                    date_debut=avoirs_ref.get('date_debut'),
                    date_solde=avoirs_ref.get('date_solde'),
                    montant_debut=avoirs_ref.get('montant_debut'),
                    doc=doc)

            # --- Opérations : supprimer / reloger les lignes des comptes supprimés ---
            deleted_set = set(self._deleted_accounts)
            rehouse_set = set(self._soft_deleted_accounts)  # ops appariées → "Compte clos"
            COMPTE_CLOS = 'Compte clos'
            if deleted_set:
                rows_to_delete = []
                cursor = ws_ops.createCursor()
                cursor.gotoStartOfUsedArea(False)
                cursor.gotoEndOfUsedArea(True)
                last_row_0 = cursor.getRangeAddress().EndRow
                for row_0 in range(2, last_row_0 + 1):
                    compte = ws_ops.getCellByPosition(cr.col('OPcompte'), row_0).getString()
                    if not compte or compte.strip() not in deleted_set:
                        continue
                    name = compte.strip()
                    if name in rehouse_set:
                        # Ops appariées → reloger dans "Compte clos"
                        ref = ws_ops.getCellByPosition(cr.col('OPréf'), row_0).getString().strip()
                        cat = ws_ops.getCellByPosition(cr.col('OPcatégorie'), row_0).getString().strip()
                        if ref and ref != '-' and not cat.startswith('#'):
                            ws_ops.getCellByPosition(cr.col('OPcompte'), row_0).setString(COMPTE_CLOS)
                            continue
                    rows_to_delete.append(row_0)
                # Supprimer en ordre inverse
                for row_0 in reversed(rows_to_delete):
                    ws_ops.Rows.removeByIndex(row_0, 1)

                # Garantir la model row START_OP (row 4 = 0-indexed 3)
                # OP n'a plus qu'une seule model row depuis suppression END_OP (v3.0.0)
                cursor2 = ws_ops.createCursor()
                cursor2.gotoEndOfUsedArea(True)
                last_0 = cursor2.getRangeAddress().EndRow
                if last_0 < 3:  # rows 0-2 = en-têtes, besoin de row 3 = model row START
                    count = 3 - last_0
                    ws_ops.Rows.insertByIndex(last_0 + 1, count)
                    # Appliquer le format données (l'insertion hérite de l'en-tête)
                    from com.sun.star.table import BorderLine2
                    empty_border = BorderLine2()
                    header_0 = 2  # row 3 = en-tête colonnes (bordures hair)
                    for r0 in range(last_0 + 1, last_0 + 1 + count):
                        # B-I : copier bordures depuis en-tête, fond blanc, police 10
                        copy_row_style(ws_ops, header_0, r0, col_start=1, col_end=9)
                        for c0 in range(1, 9):
                            cell = ws_ops.getCellByPosition(c0, r0)
                            cell.CellBackColor = 0xFFFFFF
                        # Col A : fond blanc, sans bordures, police 10
                        cell_a = ws_ops.getCellByPosition(0, r0)
                        cell_a.CellBackColor = 0xFFFFFF
                        cell_a.CharHeight = 10
                        cell_a.TopBorder = empty_border
                        cell_a.BottomBorder = empty_border
                        cell_a.LeftBorder = empty_border
                        cell_a.RightBorder = empty_border
                        # Hauteur ligne données
                        ws_ops.Rows.getByIndex(r0).Height = 582  # 16.5pt

                # Créer "Compte clos" dans Avoirs si des ops ont été relogées
                if rehouse_set:
                    if self._ensure_compte_clos(ws, total_row):
                        total_row += 1
                    # Balai : supprimer les paires entièrement dans "Compte clos"
                    self._last_sweep_count = self._sweep_compte_clos_uno(ws_ops)

                had_deletions = True
                self._deleted_accounts = []
                self._soft_deleted_accounts = []
                self._deleted_ctrl_rows = []
            else:
                had_deletions = False

            # --- Recaler formule Total et named ranges ---
            # Les model rows (tête/pied) sont préservées — pas de cleanup.
            # Elles restent dans les named ranges (ancrage pour SUM, SUMIFS, etc.)
            # et sont vides donc n'affectent pas les calculs.
            if total_row and (new_accounts or had_deletions):
                # Lire START/END_AVR depuis UNO (ajustés par removeByIndex/insertByIndex)
                from inc_uno import get_table_bounds_uno
                avr_start_now, avr_end_now = get_table_bounds_uno(doc.document, 'AVR')
                avr_first = avr_start_now or self._start_avr
                last_data = avr_end_now or (total_row - 1)
                # SUM couvre les 2 model rows START_AVR..END_AVR (inclus)
                # pour éviter le collapse à SUM(L5) quand toutes les data sont supprimées
                ws.getCellByPosition(
                    cr.col('AVRmontant_solde_euro'), uno_row(total_row)
                ).setFormula('=ROUND(SUM(AVRmontant_solde_euro);2)')
                # Recaler AVR* + START/END_AVR (incluant model rows)
                avr_names = {
                    'AVRintitulé': 'A', 'AVRtype': 'B', 'AVRdomiciliation': 'C',
                    'AVRsous_type': 'D', 'AVRtitulaire': 'F', 'AVRpropriete': 'G',
                    'AVRdate_solde': 'J', 'AVRmontant_solde': 'K',
                    'AVRmontant_solde_euro': 'L',
                }
                xdoc = doc.document
                nr = xdoc.NamedRanges
                from com.sun.star.table import CellAddress
                pos = CellAddress()
                for name, cl in avr_names.items():
                    if nr.hasByName(name):
                        nr.removeByName(name)
                    nr.addNewByName(name, f'$Avoirs.${cl}${avr_first}:${cl}${last_data}', pos, 0)
                for name, val in [('START_AVR', avr_first), ('END_AVR', last_data)]:
                    if nr.hasByName(name):
                        nr.removeByName(name)
                    nr.addNewByName(name, f'$Avoirs.$A${val}', pos, 0)

                # Garde : vérifier que START/END pointent sur ✓
                for label, row in [('START_AVR', avr_first), ('END_AVR', last_data)]:
                    cell_val = ws.getCellByPosition(0, uno_row(row)).getString().strip()
                    assert cell_val == '✓', \
                        f"{label} row {row}: '{cell_val}' (attendu: '✓')"

            # Recaler PVL* + START/END_PVL
            if new_accounts or had_deletions:
                pass  # PvCol removed — using cr.col() instead
                ws_pv = doc.get_sheet(SHEET_PLUS_VALUE)
                # Lire START/END_PVL depuis les named ranges UNO (ajustés par insertByIndex)
                from inc_uno import get_named_range_pos, get_table_bounds_uno
                pvl_start_now, pvl_end_now = get_table_bounds_uno(doc.document, 'PVL')
                if pvl_start_now and pvl_end_now:
                    pvl_start = pvl_start_now
                    pvl_end = pvl_end_now
                    xdoc = doc.document
                    nr = xdoc.NamedRanges
                    from com.sun.star.table import CellAddress
                    pos = CellAddress()
                    pvl_names = {
                        'PVLcompte': 'B', 'PVLtitre': 'C',
                        'PVLdate': 'J', 'PVLmontant': 'K',
                    }
                    for name, cl in pvl_names.items():
                        if nr.hasByName(name):
                            nr.removeByName(name)
                        nr.addNewByName(name, f'$Plus_value.${cl}${pvl_start}:${cl}${pvl_end}', pos, 0)
                    for name, val in [('START_PVL', pvl_start), ('END_PVL', pvl_end)]:
                        if nr.hasByName(name):
                            nr.removeByName(name)
                        nr.addNewByName(name, f'$Plus_value.$B${val}', pos, 0)

                    # Garde : vérifier que START/END pointent sur ✓
                    for label, row in [('START_PVL', pvl_start), ('END_PVL', pvl_end)]:
                        cell_val = ws_pv.getCellByPosition(0, uno_row(row)).getString().strip()
                        assert cell_val == '✓', \
                            f"{label} row {row}: '{cell_val}' (attendu: '✓')"

            if new_ops_accounts:
                self._cleanup_model_rows_ops(ws_ops, cr=cr)

            # Patrimoine : insérer les lignes manquantes
            self._sync_patrimoine(doc)

            if owned:
                self._uno_finalize(doc)

        # Nettoyer le flag _is_new et marquer formula_j après sauvegarde réussie
        # (formula_j sert de garde pour ne pas recréer les #Solde à chaque save)
        for a in self.accounts_data:
            a.pop('_is_new', None)
            if 'formula_j' not in a:
                a['formula_j'] = True

    # --- Patrimoine : insertion automatique de lignes pour nouvelles valeurs ---

    # Mapping champ Avoirs → (nom défini AVR, header du bloc Patrimoine)
    _PATRIMOINE_BLOCKS = {
        'type':           ('AVRtype',           'par type'),
        'sous_type':      ('AVRsous_type',      'par sous-type'),
        'domiciliation':  ('AVRdomiciliation',  'par domiciliation'),
        'titulaire':      ('AVRtitulaire',      'par titulaire'),
        'propriete':      ('AVRpropriete',      'en propri'),
    }

    def _sync_patrimoine(self, doc):
        """Ajoute dans Patrimoine les lignes manquantes pour les nouvelles valeurs."""
        from inc_uno import col_of, letter_of
        ws = doc.get_sheet('Patrimoine')
        xdoc = doc.document

        # Résoudre les colonnes PAT depuis les named ranges
        cB = col_of(xdoc, 'PATlabel')
        cC = col_of(xdoc, 'PATnombre')
        cD = col_of(xdoc, 'PATvaleur')
        cE = col_of(xdoc, 'PATpoids')
        lB = letter_of(xdoc, 'PATlabel')
        lC = letter_of(xdoc, 'PATnombre')
        lD = letter_of(xdoc, 'PATvaleur')

        # Collecter toutes les valeurs actuelles des comptes
        values_by_field = {}
        for field in self._PATRIMOINE_BLOCKS:
            vals = {a.get(field, '').strip() for a in self.accounts_data}
            vals.discard('')
            values_by_field[field] = vals

        for field, (avr_name, header_prefix) in self._PATRIMOINE_BLOCKS.items():
            # Trouver le header et le TOTAL du bloc
            header_row = None
            total_row = None
            for r in range(0, 70):
                b = ws.getCellByPosition(cB, r).getString().strip()
                if header_row is None and b.lower().startswith(header_prefix.lower()):
                    header_row = r
                elif header_row is not None and b == 'TOTAL':
                    total_row = r
                    break

            if header_row is None or total_row is None:
                continue

            # Lire les valeurs existantes dans le bloc
            existing = set()
            for r in range(header_row + 1, total_row):
                val = ws.getCellByPosition(cB, r).getString().strip()
                if val:
                    existing.add(val)

            # Insérer les valeurs manquantes
            from inc_uno import copy_row_style
            for new_val in sorted(values_by_field[field] - existing):
                # Insérer avant TOTAL
                ws.Rows.insertByIndex(total_row, 1)
                # Copier le style de la ligne au-dessus
                if total_row > header_row + 1:
                    copy_row_style(ws, total_row - 1, total_row, col_start=cB, col_end=cE)
                # Écrire les formules (row = total_row + 1 en 1-indexed)
                row_1 = total_row + 1
                ws.getCellByPosition(cB, total_row).setString(new_val)
                ws.getCellByPosition(cC, total_row).setFormula(
                    f'=COUNTIF({avr_name};${lB}{row_1})')
                ws.getCellByPosition(cD, total_row).setFormula(
                    f'=SUMIF({avr_name};${lB}{row_1};AVRmontant_solde_euro)')
                # Le ratio sera corrigé en bloc après
                total_row += 1

            # Recalculer TOTAL et ratios si des lignes ont été ajoutées
            if values_by_field[field] - existing:
                first_data = header_row + 1
                last_data = total_row - 1
                total_1 = total_row + 1  # 1-indexed
                first_1 = first_data + 1

                # TOTAL SUM
                ws.getCellByPosition(cC, total_row).setFormula(
                    f'=SUM({lC}{first_1}:{lC}{total_1 - 1})')
                ws.getCellByPosition(cD, total_row).setFormula(
                    f'=ROUND(SUM({lD}{first_1}:{lD}{total_1 - 1});2)')

                # Ratios E
                for r in range(first_data, total_row):
                    ws.getCellByPosition(cE, r).setFormula(
                        f'={lD}{r + 1}/{lD}${total_1}')

    def _cleanup_patrimoine(self, keep_values=None, doc=None):
        """Supprime les lignes Patrimoine non conservées.

        Inverse de _sync_patrimoine : supprime les lignes de données dont la
        valeur B n'est pas dans keep_values[field]. Le bloc 'type' est structurel
        et n'est jamais nettoyé.

        Args:
            keep_values: dict {field: set(valeurs à conserver)}
            doc: UnoDocument ouvert (mode batch). Si None, ouvre/ferme automatiquement.
        """
        from contextlib import nullcontext
        from inc_uno import UnoDocument, col_of, letter_of

        keep_values = keep_values or {}

        owned = doc is None
        ctx = UnoDocument(self.xlsx_path) if owned else nullcontext(doc)
        with ctx as doc:
            ws = doc.get_sheet('Patrimoine')
            xdoc = doc.document

            # Résoudre les colonnes PAT depuis les named ranges
            cB = col_of(xdoc, 'PATlabel')
            cC = col_of(xdoc, 'PATnombre')
            cD = col_of(xdoc, 'PATvaleur')
            lB = letter_of(xdoc, 'PATlabel')
            lC = letter_of(xdoc, 'PATnombre')
            lD = letter_of(xdoc, 'PATvaleur')

            for field, (avr_name, header_prefix) in self._PATRIMOINE_BLOCKS.items():
                if field == 'type':
                    continue  # types structurels, jamais supprimés
                keep = keep_values.get(field, set())

                # Trouver header et TOTAL du bloc
                header_row = None
                total_row = None
                for r in range(0, 70):
                    b = ws.getCellByPosition(cB, r).getString().strip()
                    if header_row is None and b.lower().startswith(header_prefix.lower()):
                        header_row = r
                    elif header_row is not None and b == 'TOTAL':
                        total_row = r
                        break

                if header_row is None or total_row is None:
                    continue

                # Supprimer les lignes non conservées (de bas en haut)
                # Préserve : valeurs dans `keep`, ligne placeholder '-' (structurelle).
                # Supprime : lignes B vide (spacers, formules orphelines) et autres
                deleted = 0
                has_dash = False
                for r in range(total_row - 1, header_row, -1):
                    val = ws.getCellByPosition(cB, r).getString().strip()
                    if val == '-':
                        has_dash = True
                        continue
                    if not val or val not in keep:
                        ws.Rows.removeByIndex(r, 1)
                        deleted += 1
                        total_row -= 1
                # Si pas de '-' présent et bloc vidé : insérer le placeholder
                if not has_dash and total_row == header_row + 1:
                    ws.Rows.insertByIndex(total_row, 1)
                    new_dash_row = total_row
                    total_row += 1
                    ws.getCellByPosition(cB, new_dash_row).setString('-')
                    # Formules COUNTIF/SUMIF pour cohérence avec le template
                    new_dash_1 = new_dash_row + 1
                    ws.getCellByPosition(cC, new_dash_row).setFormula(
                        f'=COUNTIF({avr_name},${lB}{new_dash_1})')
                    ws.getCellByPosition(cD, new_dash_row).setFormula(
                        f'=SUMIF({avr_name},${lB}{new_dash_1},AVRmontant_solde_euro)')
                    deleted += 1  # force recalcul TOTAL ci-dessous

                if deleted:
                    # Recalculer les formules TOTAL
                    first_data = header_row + 1
                    first_1 = first_data + 1
                    total_1 = total_row + 1
                    if total_row > first_data:
                        # Bloc non vide : SUM des lignes restantes
                        ws.getCellByPosition(cC, total_row).setFormula(
                            f'=SUM({lC}{first_1}:{lC}{total_1 - 1})')
                        ws.getCellByPosition(cD, total_row).setFormula(
                            f'=ROUND(SUM({lD}{first_1}:{lD}{total_1 - 1});2)')
                    else:
                        # Bloc vide : TOTAL = 0 (éviter #REF!)
                        ws.getCellByPosition(cC, total_row).setValue(0)
                        ws.getCellByPosition(cD, total_row).setValue(0)
                    print(f"Patrimoine {field}: {deleted} lignes supprimées")

            if owned:
                doc.save()

        if owned:
            self._load_excel_data()

    def _save_config(self):
        raw = self.config_raw

        # Onglet Paramètres
        for (section, key), (vtype, var) in self.tk_vars.items():
            if section.startswith('site_'):
                continue
            if vtype == 'bool':
                val = 'true' if var.get() else 'false'
            elif vtype == 'int':
                val = str(var.get())
            else:
                val = var.get().strip()
            updated = write_config_section_key(raw, section, key, val)
            if updated is None:
                raw = _insert_key_in_section(raw, section, key, val)
            else:
                raw = updated

        # Onglet Sites — enabled list
        enabled = [s for s in self.all_sites if self.site_vars[s].get()]
        updated = write_config_section_key(raw, 'sites', 'enabled',
                                          ','.join(enabled))
        if updated is not None:
            raw = updated

        # Onglet Sites — paramètres par site (sauf clés readonly)
        readonly_keys = {'name', 'base_url', 'credential_id'}
        for (section, key), (vtype, var) in self.tk_vars.items():
            if section.startswith('site_'):
                if key in readonly_keys:
                    continue
                site = section[5:]  # strip 'site_'
                val = var.get().strip()
                if not val:
                    continue  # champ vide = pas de surcharge
                updated = write_config_section_key(raw, site, key, val)
                if updated is None:
                    # Clé absente du fichier → insérer
                    raw = _insert_key_in_section(raw, site, key, val)
                else:
                    raw = updated

        # Écrire
        with open(self.config_path, 'w', encoding='utf-8') as f:
            f.write(raw)
        self.config_raw = raw

    def _load_pipeline_json(self):
        """Charge config_pipeline.json et retourne le dict."""
        import json
        if self.pipeline_json_path.exists():
            with open(self.pipeline_json_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {'linked_operations': {}, 'solde_auto': {}}

    def _save_pipeline_config(self):
        """Sauvegarde _linked_data et _solde_auto_data dans config_pipeline.json."""
        import json
        data = {'linked_operations': {}, 'solde_auto': {}}
        if hasattr(self, '_linked_data'):
            for pattern, compte, desc in self._linked_data:
                data['linked_operations'][pattern] = {
                    'compte_cible': compte,
                    'description': desc,
                }
        if hasattr(self, '_solde_auto_data'):
            for compte, cat, devise in self._solde_auto_data:
                data['solde_auto'][compte] = {
                    'categorie_trigger': cat,
                    'devise': devise,
                }
        with open(self.pipeline_json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
            f.write('\n')

    def _save_mappings(self):
        from cpt_gui import write_mappings_json
        write_mappings_json(self.json_path, self.mappings)

