#!/usr/bin/env python3
"""tool_import.py — injection contrôlée d'un lot d'opérations MANUEL.

Passe par les BRIQUES APP réelles — `cpt_format_MANUEL.format_site` →
`ComptaExcelImport.append_to_comptes` → `generate_missing_soldes` →
`close_workbook` → `refresh_controles` — c.-à-d. exactement la séquence
d'import de `cpt_update`, MAIS sans son wrapper de collecte : pas d'archivage,
pas de session HDS, aucune écriture dans `dropbox/`. Utile pour appliquer un lot
comptable ponctuel (fusion de comptes, correction) sans polluer l'historique de
collecte, et pour valider un lot avant de l'appliquer.

Deux modes :
  - dry-run (défaut) : l'import tourne sur une COPIE temporaire du classeur →
    rapport avant/après (écarts Contrôles) → la copie est jetée. Le classeur réel
    n'est jamais ouvert en écriture.
  - --apply : backup horodaté (`comptes_backup_import_*.xlsm`, hors mécanisme HDS)
    puis import EN PLACE, conservé. Appliquer POUR DE VRAI reste ce geste ; il ne
    reconstruit pas `cpt_update` (mêmes briques, wrapper de collecte omis).

Usage :
  ./tool_import.py <dataset.xlsx> [--classeur PATH] [--apply]

Le <dataset.xlsx> est au format MANUEL (feuille « Import », colonnes
Date;Libellé;Montant;Devise;Equiv;Réf;Catégorie;Compte;Commentaire).
"""
import sys
import shutil
import argparse
import tempfile
from pathlib import Path
from datetime import datetime

import openpyxl

import inc_mode
import cpt_format_MANUEL
from inc_excel_import import ComptaExcelImport
from inc_excel_schema import Operation
from inc_logging import Logger
from inc_uno import refresh_controles

# Colonnes Contrôles lues pour le rapport (1-indexées) : A=compte (formule
# =Avoirs!A.., valeur cachée = nom), F=solde calculé, G=montant relevé, H=écart.
_CTRL_NAME, _CTRL_CALC, _CTRL_REL, _CTRL_ECART = 1, 6, 7, 8


def _dataset_accounts(dataset_path):
    """Noms de comptes distincts référencés par le dataset (colonne Compte)."""
    wb = openpyxl.load_workbook(dataset_path, read_only=True, data_only=True)
    ws = wb['Import'] if 'Import' in wb.sheetnames else wb.active
    accts = []
    for i, row in enumerate(ws.iter_rows(values_only=True, max_col=9)):
        if i == 0 and row[0] == 'Date':
            continue
        if len(row) >= 8 and row[7]:
            name = str(row[7]).strip()
            if name and name not in accts:
                accts.append(name)
    wb.close()
    return accts


def _snapshot(classeur_path, accounts):
    """Écarts Contrôles (calculé/relevé/écart) des comptes donnés + verdict A1,
    lus depuis les valeurs cachées (data_only). Le classeur doit avoir été
    recalculé (refresh_controles) pour que les valeurs soient à jour."""
    wb = openpyxl.load_workbook(classeur_path, read_only=True, data_only=True, keep_vba=True)
    ws = wb['Contrôles']
    snap = {'_A1': ws.cell(1, 1).value}
    for r in range(4, 120):
        name = ws.cell(r, _CTRL_NAME).value
        if isinstance(name, str) and name.strip() in accounts:
            snap[name.strip()] = (
                ws.cell(r, _CTRL_CALC).value,
                ws.cell(r, _CTRL_REL).value,
                ws.cell(r, _CTRL_ECART).value,
            )
    wb.close()
    return snap


def _fmt(v):
    return f"{v:>12.2f}" if isinstance(v, (int, float)) else f"{str(v):>12}"


def _report(before, after, accounts):
    print("\n" + "=" * 78)
    print("RÉCONCILIATION — écarts Contrôles avant / après import")
    print("=" * 78)
    print(f"  Verdict A1 : {before.get('_A1')!r}  →  {after.get('_A1')!r}")
    print(f"  {'Compte':32} {'calc→':>12} {'relevé→':>12} {'écart av→ap':>18}")
    print("  " + "-" * 74)
    for name in accounts:
        b = before.get(name)
        a = after.get(name)
        if a is None and b is None:
            print(f"  {name:32} {'(absent)':>12}")
            continue
        be = b[2] if b else None
        ae = a[2] if a else None
        calc = _fmt(a[0]) if a else _fmt(b[0]) if b else ' ' * 12
        rel = _fmt(a[1]) if a else _fmt(b[1]) if b else ' ' * 12
        flag = ''
        if isinstance(ae, (int, float)) and abs(ae) >= 0.015:
            flag = '  ⚠ écart'
        print(f"  {name:32} {calc} {rel} {_fmt(be)}→{_fmt(ae)}{flag}")
    print("=" * 78)


def run_import(classeur_path, dataset_path, logger):
    """Séquence d'import MANUEL sur `classeur_path` (mute le fichier)."""
    # Isole le dataset dans un dossier propre (format_site globe *.xlsx).
    with tempfile.TemporaryDirectory() as td:
        shutil.copy2(dataset_path, Path(td) / 'manuel.xlsx')
        ops, pos = cpt_format_MANUEL.format_site(td, logger=logger)
    if pos:
        logger.warning(f"{len(pos)} position(s) ignorée(s) — hors périmètre tool_import")
    operations = [Operation.from_tuple(t) for t in ops]
    logger.info(f"{len(operations)} opération(s) à injecter")

    excel = ComptaExcelImport(comptes_file=str(classeur_path), logger=logger)
    if not excel.open_workbook():
        raise RuntimeError(f"ouverture du classeur échouée : {classeur_path}")
    ok = excel.append_to_comptes(operations)
    # Séquence IDENTIQUE à cpt_update (brique inchangée) — gms compris : un lot
    # contrôlé se rend SELF-CONTAINED (fournit un #Solde pour chaque compte touché)
    # pour que gms ne se déclenche pour aucun, plutôt que de retirer gms ici.
    excel.generate_missing_soldes()
    excel.close_workbook(save=True)
    refresh_controles(str(classeur_path), logger)
    return ok


def main():
    ap = argparse.ArgumentParser(description="Injection contrôlée d'un lot MANUEL (dry-run par défaut)")
    ap.add_argument('dataset', help="Fichier .xlsx au format MANUEL (feuille Import)")
    ap.add_argument('--classeur', default=None, help="Classeur cible (défaut : celui de l'instance)")
    ap.add_argument('--apply', action='store_true', help="Applique EN PLACE (backup horodaté), sinon dry-run")
    args = ap.parse_args()

    classeur = Path(args.classeur) if args.classeur else (inc_mode.get_base_dir() / 'comptes.xlsm')
    dataset = Path(args.dataset)
    if not classeur.exists():
        sys.exit(f"Classeur introuvable : {classeur}")
    if not dataset.exists():
        sys.exit(f"Dataset introuvable : {dataset}")

    logger = Logger('tool_import', verbose=False)
    accounts = _dataset_accounts(dataset)
    print(f"Comptes touchés : {', '.join(accounts)}")

    before = _snapshot(classeur, accounts)

    if args.apply:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup = classeur.parent / f'comptes_backup_import_{ts}.xlsm'
        shutil.copy2(classeur, backup)
        print(f"Backup : {backup}")
        run_import(classeur, dataset, logger)
        after = _snapshot(classeur, accounts)
        _report(before, after, accounts)
        print(f"\n✅ APPLIQUÉ en place. Restauration éventuelle : cp '{backup}' '{classeur}'")
    else:
        with tempfile.TemporaryDirectory() as td:
            work = Path(td) / 'comptes.xlsm'
            shutil.copy2(classeur, work)
            run_import(work, dataset, logger)
            after = _snapshot(work, accounts)
        _report(before, after, accounts)
        print("\n🔍 DRY-RUN — classeur réel NON modifié. Relancer avec --apply pour appliquer.")


if __name__ == '__main__':
    main()
