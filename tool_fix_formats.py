#!/usr/bin/env python3
"""
tool_fix_formats.py — Corrige les formats de cellules dans comptes.xlsm

Usage:
  python3 tool_fix_formats.py comptes.xlsm           # dry run
  python3 tool_fix_formats.py comptes.xlsm --apply   # applique les corrections

Corrections toutes feuilles :
  - Montants : format français (virgule décimale, espace milliers)
  - Dates : DD/MM/YY
  - Avoirs : fond gris E+K pour non-EUR, format devise K, format EUR L
"""

import argparse
import shutil
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from inc_excel_schema import (
    SHEET_AVOIRS, ColResolver, uno_row,
)
# inc_uno imports done locally where needed
from inc_check_integrity import validate_structure
from inc_formats import (
    devise_format, _load_decimals, _DEFAULT_DECIMALS,
    FORMATS_DEVISE, FORMAT_EUR, FORMAT_EUR_RED, FORMAT_DATE,
    GRIS, GRIS_BLANC, GRIS_BEIGE, BLANC, BEIGE_CLAIR,
    TETE_FILL, PIED_FILL, COL_REF_FILL, DATA_FILL, JAUNE,
    HAIR_COLOR, PIED_BORDER_COLOR,
    EXC_DATA, EXC_HEAD, EXC_FOOT,
    HAIR_WIDTH_UNO, THICK_WIDTH_UNO,
)

# Cache format strings pour comparaison (LO assigne des IDs différents
# pour des formats visuellement identiques selon la locale)
_fmt_str_cache = {}


def _normalize_fmt(fmt_str):
    """Normalise un format LO pour comparaison.

    Gère deux variantes du même format visuel :
    - '\\<espace>' vs '<espace>' avant symboles devise (encodage LO)
    - US Excel (#,##0.00) vs FR UNO (#\xa0##0,00) — même rendu en locale FR
    """
    import re
    s = fmt_str.replace('\\ ', ' ')
    # US Excel → FR UNO : #,##0.00 → # ##0,00  (en protégeant les [$...])
    if '#,##0' in s:
        parts = re.split(r'(\[\$[^\]]*\])', s)
        converted = []
        for part in parts:
            if part.startswith('[$'):
                converted.append(part)
            else:
                # , → placeholder, . → , (décimal), placeholder → \xa0
                part = part.replace(',', '\x01').replace('.', ',').replace('\x01', '\xa0')
                converted.append(part)
        s = ''.join(converted)
    return s


def _fmt_eq(formats, fmt_id_a, fmt_id_b):
    """True si deux format IDs produisent le même rendu."""
    if fmt_id_a == fmt_id_b:
        return True
    for fid in (fmt_id_a, fmt_id_b):
        if fid not in _fmt_str_cache:
            _fmt_str_cache[fid] = formats.getByKey(fid).FormatString
    return _normalize_fmt(_fmt_str_cache[fmt_id_a]) == _normalize_fmt(_fmt_str_cache[fmt_id_b])


def fix_avoirs(doc, apply, cr=None):
    """Corrige les formats de la feuille Avoirs."""
    ws = doc.get_sheet(SHEET_AVOIRS)
    fixes = 0

    # Bornes via named range colonne AVRintitulé
    from inc_uno import get_col_range_bounds
    avr_bounds = get_col_range_bounds(doc.document, 'AVRintitulé')
    first_row = (avr_bounds[2] + 1) if avr_bounds else 5  # données = START+1
    end_row = avr_bounds[3] if avr_bounds else first_row + 200

    formats = doc.document.getNumberFormats()
    fmt_date = doc.register_number_format(FORMAT_DATE)
    fmt_eur = doc.register_number_format(FORMAT_EUR)
    fmt_cache = {}
    for devise, fmt_str in FORMATS_DEVISE.items():
        fmt_cache[devise] = doc.register_number_format(fmt_str)

    for r in range(first_row, end_row + 1):
        intitule = ws.getCellByPosition(cr.col('AVRintitulé'), uno_row(r)).getString().strip()
        if not intitule:
            continue

        devise = ws.getCellByPosition(cr.col('AVRdevise'), uno_row(r)).getString().strip()
        r0 = uno_row(r)
        is_non_eur = devise and devise != 'EUR'
        row_fixes = []

        # Date J
        j_cell = ws.getCellByPosition(cr.col('AVRdate_solde'), r0)
        if not _fmt_eq(formats, j_cell.NumberFormat, fmt_date):
            row_fixes.append('J:date')
            if apply:
                j_cell.NumberFormat = fmt_date

        # Montant K
        k_cell = ws.getCellByPosition(cr.col('AVRmontant_solde'), r0)
        expected_k = fmt_cache.get(devise, fmt_eur)
        if not _fmt_eq(formats, k_cell.NumberFormat, expected_k):
            row_fixes.append(f'K:fmt({devise})')
            if apply:
                k_cell.NumberFormat = expected_k

        # Equiv EUR L
        l_cell = ws.getCellByPosition(cr.col('AVRmontant_solde_euro'), r0)
        if not _fmt_eq(formats, l_cell.NumberFormat, fmt_eur):
            row_fixes.append('L:fmt(EUR)')
            if apply:
                l_cell.NumberFormat = fmt_eur

        # Col E (devise) — jamais grisée (libellé)
        # Nettoyer les anciens GRIS hérités
        e_cell = ws.getCellByPosition(cr.col('AVRdevise'), r0)
        if e_cell.CellBackColor in (GRIS, GRIS_BLANC, GRIS_BEIGE):
            row_fixes.append('E:blanc')
            if apply:
                e_cell.CellBackColor = BLANC

        # Fond GRIS_BLANC sur K (montant) pour non-EUR
        if is_non_eur:
            if k_cell.CellBackColor != GRIS_BLANC:
                row_fixes.append('K:gris')
                if apply:
                    k_cell.CellBackColor = GRIS_BLANC
        elif k_cell.CellBackColor in (GRIS, GRIS_BLANC, GRIS_BEIGE):
            row_fixes.append('K:blanc')
            if apply:
                k_cell.CellBackColor = BLANC

        if row_fixes:
            fixes += 1
            print(f"  {intitule:<35} {devise:<5} → {', '.join(row_fixes)}")

    return fixes


def fix_budget(doc, apply):
    """Corrige les formats du tableau Catégories dans Budget.

    Zones (déterminées via named range CATnom):
      - model rows {cat_start, cat_end ✓}        : montants → BLANC, pas de gris
      - data rows  cat_start+1 .. cat_end-1      : non-EUR → GRIS_BLANC ; EUR → BLANC
      - pieds      cat_end+1 .. avant Montant    : non-EUR → GRIS_BEIGE ; EUR → BEIGE_CLAIR
      - Montant Euros : format EUR uniquement, bg non touché

    Header devises = cat_start - 2 (row au-dessus des cours).
    """
    from inc_excel_schema import SHEET_BUDGET

    ws = doc.get_sheet(SHEET_BUDGET)

    cat_start, cat_end = doc.cr.rows('CATnom')
    if not cat_start or not cat_end:
        print("  (CATnom absent ou END manquant — skip)")
        return 0

    cat_col_0 = doc.cr.col('CATnom')         # 0-indexed col L
    header_row_0 = uno_row(cat_start) - 2    # 0-indexed row 29 (devise codes)
    if header_row_0 < 0:
        print("  ⚠ CATnom trop haut pour avoir un header — skip")
        return 0

    # Détection des pieds : scan col L après cat_end jusqu'à "Montant"
    pied_rows = []  # 1-indexed
    montant_euros_row = None
    for r in range(cat_end + 1, cat_end + 20):
        val = ws.getCellByPosition(cat_col_0, uno_row(r)).getString().strip()
        if val.startswith('Montant'):
            montant_euros_row = r
            break
        pied_rows.append(r)

    # Scanner les devises depuis le vrai header — restreint aux codes connus
    devises = []
    for c in range(cat_col_0 + 1, cat_col_0 + 30):
        val = ws.getCellByPosition(c, header_row_0).getString().strip()
        if val not in FORMATS_DEVISE:
            break
        devises.append((c, val))

    if not devises:
        print("  (aucune devise trouvée)")
        return 0

    # La colonne "Total € au cours du jour" (CATtotal_euro) est en EUR par définition :
    # l'ajouter comme "devise EUR" pour qu'elle reçoive le traitement EUR (BLANC / BEIGE_CLAIR)
    # et non un gris résiduel hérité d'un état antérieur.
    te_col = doc.cr.col('CATtotal_euro')
    if te_col is not None and te_col >= 0 and te_col not in {c for c, _ in devises}:
        devises.append((te_col, 'EUR'))

    formats = doc.document.getNumberFormats()
    fmt_cache = {}
    for d, fmt_str in FORMATS_DEVISE.items():
        fmt_cache[d] = doc.register_number_format(fmt_str)
    fmt_eur = doc.register_number_format(FORMAT_EUR)

    model_rows = {cat_start, cat_end}
    data_rows = list(range(cat_start + 1, cat_end))

    # Couleurs nettoyables (legacy à écraser sans toucher d'autres fonds custom)
    CLEANABLE_BG = {GRIS, GRIS_BLANC, GRIS_BEIGE, BLANC, BEIGE_CLAIR, 0xD5D5D5}

    fixes = 0
    for col_0, devise in devises:
        is_non_eur = devise != 'EUR'
        expected_fmt = fmt_cache.get(devise, fmt_eur)
        col_fixes = 0

        def set_fmt(cell, fmt):
            nonlocal col_fixes
            if not _fmt_eq(formats, cell.NumberFormat, fmt):
                col_fixes += 1
                if apply:
                    cell.NumberFormat = fmt

        def set_bg(cell, target):
            nonlocal col_fixes
            if cell.CellBackColor != target and cell.CellBackColor in CLEANABLE_BG:
                col_fixes += 1
                if apply:
                    cell.CellBackColor = target

        # 1) Model rows : pas de gris, montants → BLANC (cellules vides, pas de format)
        for r in model_rows:
            cell = ws.getCellByPosition(col_0, uno_row(r))
            set_bg(cell, BLANC)

        # 2) Ligne cours (cat_start - 1 = "année glissante" + SUMIF cours) : sur fond beige
        #    (le header devises au-dessus est BEIGE_FONCE). Non-EUR → GRIS_BEIGE pour contraste.
        cours_row = cat_start - 1
        cell_cours = ws.getCellByPosition(col_0, uno_row(cours_row))
        if is_non_eur:
            set_bg(cell_cours, GRIS_BEIGE)

        # 3) Data rows : format devise + bg
        target_data = GRIS_BLANC if is_non_eur else BLANC
        for r in data_rows:
            cell = ws.getCellByPosition(col_0, uno_row(r))
            set_fmt(cell, expected_fmt)
            set_bg(cell, target_data)

        # 4) Pied rows : format devise + bg dual tone
        target_pied = GRIS_BEIGE if is_non_eur else BEIGE_CLAIR
        for r in pied_rows:
            cell = ws.getCellByPosition(col_0, uno_row(r))
            set_fmt(cell, expected_fmt)
            set_bg(cell, target_pied)

        # 4) Montant Euros : format EUR uniquement (bg non touché)
        if montant_euros_row is not None:
            cell = ws.getCellByPosition(col_0, uno_row(montant_euros_row))
            set_fmt(cell, fmt_eur)

        if col_fixes:
            fixes += col_fixes
            print(f"  col {devise:<5} → {col_fixes} cellules")

    return fixes


def fix_plus_value(doc, apply, cr=None):
    """Corrige les formats de la feuille Plus_value.

    Pour chaque ligne avec devise :
    - Dates G et J : DD/MM/YY
    - Montants H et I : format devise
    - Solde K : format EUR (valeur en euros)
    - Devise D : fond gris si non-EUR
    """
    from inc_excel_schema import SHEET_PLUS_VALUE

    ws = doc.get_sheet(SHEET_PLUS_VALUE)

    pvl_s, pvl_e = doc.cr.rows('PVLcompte')
    if not pvl_s:
        print("  (PVLcompte absent — skip)")
        return 0

    fixes = 0

    # Marqueurs de section : label en col A (pas en col C)
    SECTION_IDS = {'portefeuilles', 'métaux', 'crypto', 'devises'}
    for r in range(1, uno_row(pvl_s) + 2):  # scanner avant START
        r0 = r - 1
        val_a = ws.getCellByPosition(cr.col('PVLsection'), r0).getString().strip()
        val_c = ws.getCellByPosition(cr.col('PVLtitre'), r0).getString().strip()
        if val_a in SECTION_IDS and val_c.startswith('Les '):
            # Déplacer label de C vers A
            print(f"  marqueur '{val_c}' : col C → col A")
            ws.getCellByPosition(cr.col('PVLsection'), r0).setString(val_c)
            ws.getCellByPosition(cr.col('PVLsection'), r0).CharWeight = 150  # BOLD
            ws.getCellByPosition(cr.col('PVLtitre'), r0).setString('')
            fixes += 1

    data_start = pvl_s + 1  # 1-indexed, skip model row START
    # end PVL peut être corrompu (row 6 au lieu de 123) → scanner jusqu'à la fin
    end_row_hint = pvl_e if pvl_e else data_start + 200  # 1-indexed
    if end_row_hint < data_start + 10:
        end_row_hint = data_start + 200
    data_end = end_row_hint + 1

    formats = doc.document.getNumberFormats()
    fmt_date = doc.register_number_format(FORMAT_DATE)
    fmt_eur = doc.register_number_format(FORMAT_EUR)
    fmt_eur_red = doc.register_number_format(FORMAT_EUR_RED)
    fmt_cache = {}
    fmt_cache_red = {'EUR': fmt_eur_red}
    for devise, fmt_str in FORMATS_DEVISE.items():
        fmt_cache[devise] = doc.register_number_format(fmt_str)
        if devise != 'EUR':
            fmt_cache_red[devise] = doc.register_number_format(f'{fmt_str};[RED]\\-{fmt_str}')

    for r in range(data_start, data_end):
        r0 = r - 1  # 0-indexed
        devise = ws.getCellByPosition(cr.col('PVLdevise'), r0).getString().strip()
        if not devise:
            continue

        ligne = ws.getCellByPosition(cr.col('PVLtitre'), r0).getString().strip()
        section = ws.getCellByPosition(cr.col('PVLsection'), r0).getString().strip()
        is_non_eur = devise != 'EUR'
        row_fixes = []

        # E (PVL) et K (Solde) : toujours en devise de la ligne (la valeur est
        # native, ex: gramme d'or pour OrPr, satoshi pour SAT, USD pour cpt USD).
        # H (Montant_init) et I (Sigma) : devise native pour portefeuilles, EUR sinon
        # (formules en equiv_euro / AVR_euro pour métaux/crypto/devises).
        is_portefeuille = section == 'portefeuilles'
        expected_devise_fmt = fmt_cache.get(devise, fmt_eur)
        expected_devise_red = fmt_cache_red.get(devise, fmt_eur_red)
        ek_label = devise
        if is_portefeuille:
            expected_hi_fmt = fmt_cache.get(devise, fmt_eur)
            hi_label = devise
        else:
            expected_hi_fmt = fmt_eur
            hi_label = 'EUR'

        # Date G (DATE_INIT)
        cell = ws.getCellByPosition(cr.col('PVLdate_init'), r0)
        if not _fmt_eq(formats, cell.NumberFormat, fmt_date):
            row_fixes.append('G:date')
            if apply:
                cell.NumberFormat = fmt_date

        # PVL E — devise native pour portefeuilles, EUR sinon (rouge négatif)
        cell = ws.getCellByPosition(cr.col('PVLpvl'), r0)
        if not _fmt_eq(formats, cell.NumberFormat, expected_devise_fmt) and \
           not _fmt_eq(formats, cell.NumberFormat, expected_devise_red):
            row_fixes.append(f'E:fmt({ek_label})')
            if apply:
                cell.NumberFormat = expected_devise_red

        # Montant H (MONTANT_INIT)
        cell = ws.getCellByPosition(cr.col('PVLmontant_init'), r0)
        if not _fmt_eq(formats, cell.NumberFormat, expected_hi_fmt):
            row_fixes.append(f'H:fmt({hi_label})')
            if apply:
                cell.NumberFormat = expected_hi_fmt

        # Sigma I
        cell = ws.getCellByPosition(cr.col('PVLsigma'), r0)
        if not _fmt_eq(formats, cell.NumberFormat, expected_hi_fmt):
            row_fixes.append(f'I:fmt({hi_label})')
            if apply:
                cell.NumberFormat = expected_hi_fmt

        # Date J (DATE_SOLDE)
        cell = ws.getCellByPosition(cr.col('PVLdate'), r0)
        if not _fmt_eq(formats, cell.NumberFormat, fmt_date):
            row_fixes.append('J:date')
            if apply:
                cell.NumberFormat = fmt_date

        # Solde K — devise native pour portefeuilles, EUR sinon (rouge négatif)
        cell = ws.getCellByPosition(cr.col('PVLmontant'), r0)
        if not _fmt_eq(formats, cell.NumberFormat, expected_devise_fmt) and \
           not _fmt_eq(formats, cell.NumberFormat, expected_devise_red):
            row_fixes.append(f'K:fmt({ek_label})')
            if apply:
                cell.NumberFormat = expected_devise_red

        # Fond blanc D-K pour lignes données (hors pieds portefeuille)
        # Étendue à C pour les lignes titres (*…*) — extension de range
        # Le gris devise se superpose au blanc → ne mettre blanc que sur les colonnes non-gris
        # Note : l'en-tête portefeuille (LIGNE='Portefeuille') n'a pas de devise
        # → skippée par le 'if not devise: continue' plus haut
        PV_PIEDS = {'Total', '#Solde Opérations', 'Retenu'}
        is_pied = ligne in PV_PIEDS
        is_titre_row = (len(ligne) >= 3
                        and ligne.startswith('*')
                        and ligne.endswith('*'))
        if not is_pied:
            WHITE_COLS = {cr.col('PVLdevise'), cr.col('PVLpvl'), cr.col('PVLpct'),
                          cr.col('PVLdate_init'), cr.col('PVLmontant_init'),
                          cr.col('PVLsigma'), cr.col('PVLdate'),
                          cr.col('PVLmontant')}
            if is_titre_row:
                WHITE_COLS.add(cr.col('PVLtitre'))
            # Colonnes montant grisées (non-EUR) — pas de blanc dessus
            # Seuls les portefeuilles non-EUR grisent (E/K/H/I en devise native) ;
            # sections métaux/crypto/devises ont tout en EUR → pas de gris.
            # NB : PVLdevise (libellé devise) n'est plus grisée — fond blanc hérité
            if is_non_eur and is_portefeuille:
                gris_set = {cr.col('PVLpvl'), cr.col('PVLmontant'),
                            cr.col('PVLmontant_init'), cr.col('PVLsigma')}
            else:
                gris_set = set()
            for c0 in sorted(WHITE_COLS - gris_set):
                cell = ws.getCellByPosition(c0, r0)
                if cell.CellBackColor != BLANC:
                    row_fixes.append(f'{chr(65+c0)}:blanc')
                    if apply:
                        cell.CellBackColor = BLANC

        # Fond gris devise sur les montants non-EUR (libellé devise non grisé)
        # Uniquement pour les blocs portefeuille — les autres sections sont en EUR
        # Data row → GRIS_BLANC (sur fond blanc forcé)
        # Pied      → GRIS_BEIGE (sur beige clair hérité du template)
        if is_non_eur and is_portefeuille:
            gris_color = GRIS_BEIGE if is_pied else GRIS_BLANC
            # Portefeuille non-EUR : E (PVL), K (SOLDE), H (MONTANT_INIT), I (SIGMA)
            GRIS_COLS = [cr.col('PVLpvl'), cr.col('PVLmontant'),
                         cr.col('PVLmontant_init'), cr.col('PVLsigma')]
            # Colonnes jamais grisées (incluant désormais PVLdevise)
            NO_GRIS_COLS = [cr.col('PVLdevise'), cr.col('PVLpct'),
                            cr.col('PVLdate_init'), cr.col('PVLdate')]

            for c0 in GRIS_COLS:
                cell = ws.getCellByPosition(c0, r0)
                if cell.CellBackColor != gris_color:
                    row_fixes.append(f'{chr(65+c0)}:gris')
                    if apply:
                        cell.CellBackColor = gris_color
            # Nettoyer les anciennes colonnes grisées (notamment D)
            # Sur data row → BLANC ; sur pied → BEIGE_CLAIR (fond hérité du template)
            for c0 in NO_GRIS_COLS:
                cell = ws.getCellByPosition(c0, r0)
                if cell.CellBackColor in (GRIS, GRIS_BLANC, GRIS_BEIGE):
                    target = BEIGE_CLAIR if is_pied else BLANC
                    row_fixes.append(f'{chr(65+c0)}:-gris')
                    if apply:
                        cell.CellBackColor = target

        if row_fixes:
            fixes += 1
            # Pour métaux/crypto/devises le nom est en col B (col C vide)
            label = ligne or ws.getCellByPosition(cr.col('PVLcompte'), r0).getString().strip()
            print(f"  {label:<20} {devise:<5} → {', '.join(row_fixes)}")

    return fixes


def fix_ctrl1(doc, apply, cr=None):
    """Corrige les formats du tableau CTRL1 dans Contrôles.

    - Col B (devise) : libellé, jamais grisé → restore BLANC
    - Cols E/F/G/H (montants) : format devise selon CTRL1devise + GRIS_BLANC pour non-EUR, BLANC pour EUR
    """
    from inc_excel_schema import SHEET_CONTROLES

    ws = doc.get_sheet(SHEET_CONTROLES)

    ctrl_s, ctrl_e = doc.cr.rows('CTRL1compte')
    if not ctrl_s:
        print("  (CTRL1compte absent — skip)")
        return 0

    fixes = 0
    MONTANT_COLS = [cr.col(n) for n in
                    ('CTRL1montant_ancrage', 'CTRL1solde_calc',
                     'CTRL1montant_releve', 'CTRL1ecart')]
    devise_col = cr.col('CTRL1devise')

    formats = doc.document.getNumberFormats()
    fmt_eur = doc.register_number_format(FORMAT_EUR)
    fmt_cache = {}
    for d, fmt_str in FORMATS_DEVISE.items():
        fmt_cache[d] = doc.register_number_format(fmt_str)

    for r in range(ctrl_s, ctrl_e + 1):
        r0 = uno_row(r)
        intitule = ws.getCellByPosition(cr.col('CTRL1compte'), r0).getString().strip()
        if not intitule or intitule in ('✓', '⚓'):
            continue

        devise = ws.getCellByPosition(devise_col, r0).getString().strip()
        is_non_eur = devise and devise != 'EUR'
        expected_fmt = fmt_cache.get(devise, fmt_eur)
        row_fixes = []

        # Col B (devise) — jamais grisée
        b_cell = ws.getCellByPosition(devise_col, r0)
        if b_cell.CellBackColor in (GRIS, GRIS_BLANC, GRIS_BEIGE):
            row_fixes.append('B:blanc')
            if apply:
                b_cell.CellBackColor = BLANC

        # Cols E/F/G/H (montants) — format devise + bg
        for c0 in MONTANT_COLS:
            cell = ws.getCellByPosition(c0, r0)
            # Format devise
            if not _fmt_eq(formats, cell.NumberFormat, expected_fmt):
                row_fixes.append(f'{chr(65+c0)}:fmt({devise or "EUR"})')
                if apply:
                    cell.NumberFormat = expected_fmt
            # Background
            target = GRIS_BLANC if is_non_eur else BLANC
            if cell.CellBackColor != target and (
                cell.CellBackColor in (GRIS, GRIS_BLANC, GRIS_BEIGE, BLANC)
            ):
                row_fixes.append(f'{chr(65+c0)}:{"gris" if is_non_eur else "blanc"}')
                if apply:
                    cell.CellBackColor = target

        if row_fixes:
            fixes += 1
            print(f"  {intitule:<30} {devise:<5} → {', '.join(row_fixes)}")

    return fixes


def fix_ctrl2(doc, apply):
    """Corrige les formats du tableau CTRL2 dans Contrôles.

    Structure CTRL2 (offsets depuis header h) :
      h+0  : header devises
      h+1  : taux
      h+2  : COMPTES (entier)
      h+3  : CATÉGORIES — montant devise, gris non-EUR
      h+4  : Date (vide — seul O général est pertinent)
      h+5  : Appariements (entier)
      h+6  : Balances (entier)
      h+7  : Virements — montant devise, gris non-EUR
      h+8  : € (equiv) — montant EUR
      h+9  : Titres — montant devise, gris non-EUR
      h+10 : € (equiv) — montant EUR
      h+11 : Changes Eq € — montant EUR
      h+12 : Total € — montant EUR
    """
    from inc_excel_schema import SHEET_CONTROLES

    ws = doc.get_sheet(SHEET_CONTROLES)

    def _read_devise(cell):
        """Lit la valeur de cellule en strippant le suffixe ' ▼' du format `@" ▼"`."""
        val = cell.getString().strip()
        if val.endswith('▼'):
            val = val[:-1].rstrip()
        return val

    # Trouver le header CTRL2 via CTRL2drill (named range qui pointe sur la
    # 1re col du header devises, ex: M62:M73 → EUR en M61).
    # v3.6 : CTRL2drill start pointe sur ⚓ (h+1), header EUR = START - 1.
    h = None
    eur_col = None
    ctrl2_drill_s, _ = doc.cr.rows('CTRL2drill')
    if ctrl2_drill_s:
        drill_col_0 = doc.cr.col('CTRL2drill')
        for offset in (1, 2):  # offset row vide intermédiaire toléré
            header_row = uno_row(ctrl2_drill_s) - offset
            if header_row < 0:
                break
            val = _read_devise(ws.getCellByPosition(drill_col_0, header_row))
            if val == 'EUR':
                h = header_row
                eur_col = drill_col_0
                break

    if h is None:
        # Fallback : scan pour "EUR" dans les colonnes K..AI
        for r in range(0, 80):
            for c in range(10, 35):
                val = _read_devise(ws.getCellByPosition(c, r))
                if val == 'EUR':
                    h = r
                    eur_col = c
                    break
            if h is not None:
                break

    if h is None:
        print("  (header CTRL2 introuvable — skip)")
        return 0

    # Scanner les devises
    devises = []
    for c in range(eur_col, eur_col + 20):
        val = _read_devise(ws.getCellByPosition(c, h))
        if not val:
            break
        devises.append((c, val))

    formats = doc.document.getNumberFormats()
    fmt_cache = {}
    for devise, fmt_str in FORMATS_DEVISE.items():
        fmt_cache[devise] = doc.register_number_format(fmt_str)
    fmt_eur = doc.register_number_format(FORMAT_EUR)
    fmt_eur_red = doc.register_number_format(FORMAT_EUR_RED)
    fmt_int = doc.register_number_format('#\xa0##0')

    # Formats rouge négatif par devise (pour Virements h+7)
    fmt_red_cache = {'EUR': fmt_eur_red}
    for devise, fmt_str in FORMATS_DEVISE.items():
        if devise == 'EUR':
            continue
        # Pattern : positif;[RED]\-négatif
        red_fmt = f'{fmt_str};[RED]\\-{fmt_str}'
        fmt_red_cache[devise] = doc.register_number_format(red_fmt)

    # Propriétés de style à propager depuis la colonne EUR
    STYLE_PROPS = ('CharHeight', 'CharWeight', 'CharColor', 'CharFontName',
                   'HoriJustify', 'VertJustify',
                   'TopBorder', 'BottomBorder', 'LeftBorder', 'RightBorder')

    # Lignes en montant devise (avec gris non-EUR)
    DEVISE_ROWS = {3, 9}        # CATÉGORIES, Titres
    DEVISE_ROWS_RED = {7}       # Virements (rouge négatif)
    # Lignes en entier
    INTEGER_ROWS = {2, 5, 6}  # COMPTES, Appariements, Balances (h+4 Dates = vide)
    # Lignes en montant EUR
    EUR_ROWS = {1, 8, 10, 11, 12}  # Taux, € equiv ×2, Changes, Total
    # Lignes pied de tableau — style seul, pas de format nombre
    # Calculer depuis end CTRL2 (2 lignes après la dernière ligne de données)
    _, ctrl2_e = doc.cr.rows('CTRL2type')
    if ctrl2_e:
        end_offset = uno_row(ctrl2_e) - h  # 0-indexed
        FOOTER_ROWS = {end_offset - 1, end_offset}
    else:
        FOOTER_ROWS = {13, 14}
    # Toutes les lignes à styler (y compris header h+0 et pied)
    ALL_DATA_ROWS = sorted({0} | DEVISE_ROWS | DEVISE_ROWS_RED | INTEGER_ROWS | EUR_ROWS | FOOTER_ROWS)

    fixes = 0
    for col_0, devise in devises:
        is_non_eur = devise != 'EUR'
        expected_fmt = fmt_cache.get(devise, fmt_eur)
        col_fixes = 0

        # 1) Propager le style de base depuis la colonne EUR
        if is_non_eur:
            for row_offset in ALL_DATA_ROWS:
                eur_cell = ws.getCellByPosition(eur_col, h + row_offset)
                cell = ws.getCellByPosition(col_0, h + row_offset)
                for prop in STYLE_PROPS:
                    eur_val = getattr(eur_cell, prop, None)
                    cur_val = getattr(cell, prop, None)
                    if eur_val is not None and str(eur_val) != str(cur_val):
                        col_fixes += 1
                        if apply:
                            setattr(cell, prop, eur_val)
                # Propager CellBackColor sauf si gris sera appliqué après
                if row_offset not in DEVISE_ROWS and row_offset not in DEVISE_ROWS_RED:
                    if eur_cell.CellBackColor != cell.CellBackColor:
                        col_fixes += 1
                        if apply:
                            cell.CellBackColor = eur_cell.CellBackColor

        # 2) Number formats spécifiques par type de ligne
        for row_offset in DEVISE_ROWS:
            cell = ws.getCellByPosition(col_0, h + row_offset)
            if not _fmt_eq(formats, cell.NumberFormat, expected_fmt):
                col_fixes += 1
                if apply:
                    cell.NumberFormat = expected_fmt
            if is_non_eur and cell.CellBackColor != GRIS_BLANC:
                col_fixes += 1
                if apply:
                    cell.CellBackColor = GRIS_BLANC

        for row_offset in DEVISE_ROWS_RED:
            cell = ws.getCellByPosition(col_0, h + row_offset)
            expected_red = fmt_red_cache.get(devise, fmt_eur_red)
            if not _fmt_eq(formats, cell.NumberFormat, expected_red):
                col_fixes += 1
                if apply:
                    cell.NumberFormat = expected_red
            if is_non_eur and cell.CellBackColor != GRIS_BLANC:
                col_fixes += 1
                if apply:
                    cell.CellBackColor = GRIS_BLANC

        for row_offset in INTEGER_ROWS:
            cell = ws.getCellByPosition(col_0, h + row_offset)
            if not _fmt_eq(formats, cell.NumberFormat, fmt_int):
                col_fixes += 1
                if apply:
                    cell.NumberFormat = fmt_int

        for row_offset in EUR_ROWS:
            cell = ws.getCellByPosition(col_0, h + row_offset)
            if not _fmt_eq(formats, cell.NumberFormat, fmt_eur):
                col_fixes += 1
                if apply:
                    cell.NumberFormat = fmt_eur

        if col_fixes:
            fixes += col_fixes
            print(f"  col {devise:<5} → {col_fixes} cellules")

    return fixes


def fix_operations(doc, apply, cr=None):
    """Corrige les formats de la feuille Opérations.

    Pour chaque ligne :
    - Montant C : format devise
    - Equiv E : format EUR
    - Fond gris C et D pour non-EUR
    """
    from inc_excel_schema import SHEET_OPERATIONS

    ws = doc.get_sheet(SHEET_OPERATIONS)

    formats = doc.document.getNumberFormats()
    fmt_cache = {}
    for devise, fmt_str in FORMATS_DEVISE.items():
        fmt_cache[devise] = doc.register_number_format(fmt_str)
    fmt_eur = doc.register_number_format(FORMAT_EUR)

    # Bornes depuis OPdevise named range (ex: $Opérations.$D$4:$D$6004)
    import re
    nr = doc.document.NamedRanges
    op_start, op_end = 4, 7000  # fallback
    if nr.hasByName('OPdevise'):
        content = nr.getByName('OPdevise').Content
        m = re.findall(r'\$(\d+)', content)
        if len(m) >= 1:
            op_start = int(m[0])
        if len(m) >= 2:
            op_end = int(m[1])

    fixes = 0
    unknown_devises = []  # liste de (row, devise) pour les codes inconnus
    for r in range(op_start, op_end + 1):
        r0 = r - 1
        devise = ws.getCellByPosition(cr.col('OPdevise'), r0).getString().strip()
        if not devise:
            continue

        # Détection devise inconnue (typo probable) — warn et skip
        if devise not in FORMATS_DEVISE:
            unknown_devises.append((r, devise))
            continue

        is_non_eur = devise != 'EUR'
        expected_fmt = fmt_cache.get(devise, fmt_eur)

        # Montant C : format devise
        c_cell = ws.getCellByPosition(cr.col('OPmontant'), r0)
        if not _fmt_eq(formats, c_cell.NumberFormat, expected_fmt):
            fixes += 1
            if apply:
                c_cell.NumberFormat = expected_fmt

        # Equiv E : format EUR
        e_cell = ws.getCellByPosition(cr.col('OPequiv_euro'), r0)
        if not _fmt_eq(formats, e_cell.NumberFormat, fmt_eur):
            fixes += 1
            if apply:
                e_cell.NumberFormat = fmt_eur

        # Col D (devise) — jamais grisée (libellé)
        # Nettoyer les anciens GRIS hérités
        d_cell = ws.getCellByPosition(cr.col('OPdevise'), r0)
        if d_cell.CellBackColor in (GRIS, GRIS_BLANC, GRIS_BEIGE):
            fixes += 1
            if apply:
                d_cell.CellBackColor = BLANC

        # Fond GRIS_BLANC sur C (montant) pour non-EUR
        if is_non_eur:
            if c_cell.CellBackColor != GRIS_BLANC:
                fixes += 1
                if apply:
                    c_cell.CellBackColor = GRIS_BLANC
        elif c_cell.CellBackColor in (GRIS, GRIS_BLANC, GRIS_BEIGE):
            fixes += 1
            if apply:
                c_cell.CellBackColor = BLANC

    if unknown_devises:
        print(f"  ⚠ {len(unknown_devises)} ligne(s) avec devise inconnue (typo ?) :")
        for r, dev in unknown_devises:
            libelle = ws.getCellByPosition(cr.col('OPlibellé'), r - 1).getString().strip()
            compte = ws.getCellByPosition(cr.col('OPcompte'), r - 1).getString().strip()
            print(f"    D{r}  '{dev}'  ({compte} — {libelle[:30]})")

    return fixes


def fix_cotations(doc, apply, cr=None):
    """Corrige les formats de la col COTcours2 (cours de l'Euro dans la devise).

    Pour chaque ligne data Cotations (entre sentinels) :
      - code = EUR : col COTcours2 fond blanc, format inchangé
      - code != EUR : col COTcours2 fond GRIS_BLANC + format `#,##0.000000000 [$CODE]`

    Le format à 9 décimales est nécessaire pour les devises où 1/cours est
    très petit (ex. XAU : 1/144 ≈ 0,007). Les autres cols Cotations sont
    laissées intactes (cours saisis user en col F).
    """
    ws = doc.get_sheet('Cotations')
    fixes = 0

    if cr is None:
        cr = doc.cr

    # Bornes data Cotations via NR COTcode (entre sentinels exclus)
    s, e = cr.rows('COTcode')
    if not s or not e:
        print("  (COTcode absent — skip)")
        return 0

    cours2_col = cr.col('COTcours2')
    code_col = cr.col('COTcode')
    if cours2_col is None or code_col is None:
        print("  (COTcours2 ou COTcode absent — skip)")
        return 0

    # Body = rows entre les sentinels (s+1 .. e-1)
    for r in range(s + 1, e):
        r0 = uno_row(r)
        code = ws.getCellByPosition(code_col, r0).getString().strip()
        if not code:
            continue
        cell = ws.getCellByPosition(cours2_col, r0)
        row_fixes = []
        if code == 'EUR':
            if cell.CellBackColor in (GRIS, GRIS_BLANC, GRIS_BEIGE):
                row_fixes.append('blanc')
                if apply:
                    cell.CellBackColor = BLANC
        else:
            # Format 9 décimales avec code devise
            fmt_str = f'#,##0.000000000\\ [${code}]'
            fmt_id = doc.register_number_format(fmt_str)
            if not _fmt_eq(doc.document.getNumberFormats(), cell.NumberFormat, fmt_id):
                row_fixes.append(f'fmt({code})')
                if apply:
                    cell.NumberFormat = fmt_id
            if cell.CellBackColor != GRIS_BLANC:
                row_fixes.append('gris')
                if apply:
                    cell.CellBackColor = GRIS_BLANC
        if row_fixes:
            fixes += 1
            print(f"  H{r} {code:<5} → {', '.join(row_fixes)}")

    return fixes


def fix_generic_sheet(doc, sheet_name, apply):
    """Corrige les formats monétaires et dates sur une feuille quelconque.

    Scanne toutes les cellules avec un format numérique et remplace
    les formats US par les équivalents français.
    """
    try:
        ws = doc.get_sheet(sheet_name)
    except Exception:
        return 0

    formats = doc.document.getNumberFormats()
    fixes = 0

    # Scanner la zone utilisée
    cursor = ws.createCursor()
    cursor.gotoStartOfUsedArea(False)
    cursor.gotoEndOfUsedArea(True)
    addr = cursor.getRangeAddress()

    for row in range(addr.StartRow, addr.EndRow + 1):
        for col in range(addr.StartColumn, addr.EndColumn + 1):
            cell = ws.getCellByPosition(col, row)
            fmt_id = cell.NumberFormat
            if fmt_id == 0:  # General
                continue

            fmt_str = formats.getByKey(fmt_id).FormatString
            new_fmt = _fix_format_string(fmt_str)

            if new_fmt and new_fmt != fmt_str:
                coord = f"{chr(65 + col) if col < 26 else chr(64 + col // 26) + chr(65 + col % 26)}{row + 1}"
                fixes += 1
                if apply:
                    cell.NumberFormat = doc.register_number_format(new_fmt)

    return fixes


def _fix_format_string(fmt_str):
    """Transforme un format US en format français.

    Retourne le nouveau format, ou None si pas de changement nécessaire.
    """
    original = fmt_str

    # Point décimal + virgule milliers → virgule décimale + espace milliers
    # Détecter si c'est un format US (point décimal)
    # Les formats US ont #,##0.00 — les formats FR ont #.##0,00 ou # ##0,00
    if '#,##0.' in fmt_str or '#,##0;' in fmt_str or fmt_str.endswith('#,##0'):
        # C'est un format US : , = milliers, . = décimal
        # Convertir en français : espace insécable = milliers, , = décimal
        # Étape 1: protéger les séquences [$...] (symboles devise)
        import re
        parts = re.split(r'(\[\$[^\]]*\])', fmt_str)
        converted = []
        for part in parts:
            if part.startswith('[$'):
                converted.append(part)
            else:
                # , → placeholder, . → , (décimal), placeholder → espace insécable
                part = part.replace(',', '\x01').replace('.', ',').replace('\x01', '\xa0')
                converted.append(part)
        fmt_str = ''.join(converted)

    # Symbole devise AVANT le montant → APRÈS
    # [$€-40C] #,##0 → # ##0 [$€-40C]
    import re
    m = re.match(r'^(\[\$[^\]]+\])\s*(.+)$', fmt_str)
    if m:
        symbol, rest = m.group(1), m.group(2)
        # Vérifier que le reste ne contient pas déjà le symbole
        if symbol not in rest:
            fmt_str = f'{rest} {symbol}'

    # Traiter les formats avec ; (positif;négatif)
    if ';' in fmt_str and not fmt_str.startswith(';'):
        parts = fmt_str.split(';')
        fixed_parts = []
        for part in parts:
            m = re.match(r'^(\\?-?)(\[\$[^\]]+\])\s*(.+)$', part.strip())
            if m:
                prefix, symbol, rest = m.group(1), m.group(2), m.group(3)
                if symbol not in rest:
                    part = f'{prefix}{rest} {symbol}'
            fixed_parts.append(part)
        fmt_str = ';'.join(fixed_parts)

    if fmt_str != original:
        return fmt_str
    return None


def check_format_coherence(doc, cr=None):
    """Vérifie la cohérence entre le format nombre et la devise indiquée.

    Détecte les cellules dont le format ne correspond pas à la devise
    de la même ligne (ex: format USD sur une ligne CHF).
    Retourne le nombre d'incohérences trouvées.
    """
    from inc_excel_schema import (
        SHEET_AVOIRS, SHEET_OPERATIONS, SHEET_PLUS_VALUE,
        ColResolver, uno_row,
    )
    if cr is None:
        cr = doc.cr
    import re

    formats = doc.document.getNumberFormats()
    issues = 0

    def _extract_devise_from_fmt(fmt_id):
        """Extrait le code devise d'un format nombre (ex: '[$USD]' → 'USD')."""
        fmt_str = formats.getByKey(fmt_id).FormatString
        # [$€-40C] → EUR, [$USD] → USD, [$CHF-40C] → CHF
        m = re.search(r'\[\$([^\]]+)\]', fmt_str)
        if m:
            code = m.group(1)
            # Retirer le suffixe locale (-40C, -409, etc.)
            code = re.sub(r'-[0-9A-Fa-f]+$', '', code)
            if code in ('€', '$'):
                return 'EUR' if code == '€' else 'USD'
            return code
        return None  # pas de symbole devise (entier, décimal pur)

    # === Avoirs : K (montant) vs E (devise) ===
    ws_av = doc.get_sheet(SHEET_AVOIRS)
    avr_s, avr_e = cr.rows('AVRintitulé')
    avr_start = (avr_s + 1) if avr_s else 5
    end_row = avr_e if avr_e else 200
    for r in range(avr_start, end_row + 1):
        r0 = uno_row(r)
        intitule = ws_av.getCellByPosition(cr.col('AVRintitulé'), r0).getString().strip()
        if not intitule:
            continue
        devise = ws_av.getCellByPosition(cr.col('AVRdevise'), r0).getString().strip()
        if not devise:
            continue
        k_fmt_devise = _extract_devise_from_fmt(
            ws_av.getCellByPosition(cr.col('AVRmontant_solde'), r0).NumberFormat)
        if k_fmt_devise and k_fmt_devise != devise:
            print(f"  ✗ Avoirs {intitule}: K formaté {k_fmt_devise}, devise={devise}")
            issues += 1

    # === Opérations : C (montant) vs D (devise) ===
    ws_op = doc.get_sheet(SHEET_OPERATIONS)
    nr = doc.document.NamedRanges
    op_start, op_end = 4, 7000
    if nr.hasByName('OPdevise'):
        content = nr.getByName('OPdevise').Content
        m = re.findall(r'\$(\d+)', content)
        if len(m) >= 1:
            op_start = int(m[0])
        if len(m) >= 2:
            op_end = int(m[1])

    op_issues = 0
    for r in range(op_start, op_end + 1):
        r0 = r - 1
        devise = ws_op.getCellByPosition(cr.col('OPdevise'), r0).getString().strip()
        if not devise:
            continue
        c_fmt_devise = _extract_devise_from_fmt(
            ws_op.getCellByPosition(cr.col('OPmontant'), r0).NumberFormat)
        if c_fmt_devise and c_fmt_devise != devise:
            op_issues += 1
    if op_issues:
        print(f"  ✗ Opérations: {op_issues} ligne(s) avec format montant ≠ devise")
        issues += op_issues

    # === PVL : H/I vs devise (section-aware) ===
    ws_pv = doc.get_sheet(SHEET_PLUS_VALUE)
    pvl_s2, pvl_e2 = cr.rows('PVLcompte')
    if pvl_s2:
        data_start = pvl_s2 + 1
        end_hint = pvl_e2 if pvl_e2 else data_start + 200
        if end_hint < data_start + 10:
            end_hint = data_start + 200

        for r in range(data_start, end_hint + 1):
            r0 = r - 1
            devise = ws_pv.getCellByPosition(cr.col('PVLdevise'), r0).getString().strip()
            if not devise:
                continue
            section = ws_pv.getCellByPosition(cr.col('PVLsection'), r0).getString().strip()
            ligne = ws_pv.getCellByPosition(cr.col('PVLtitre'), r0).getString().strip()
            # Pour les sections non-portefeuilles (métaux/crypto/devises),
            # le nom est en col B (PVLcompte), col C est vide
            label = ligne or ws_pv.getCellByPosition(cr.col('PVLcompte'), r0).getString().strip() or '?'

            # H/I : devise native pour portefeuilles, EUR pour les autres sections
            expected_hi = devise if section == 'portefeuilles' else 'EUR'
            for nr_name, col_name in (('PVLmontant_init', 'H'), ('PVLsigma', 'I')):
                fmt_devise = _extract_devise_from_fmt(
                    ws_pv.getCellByPosition(cr.col(nr_name), r0).NumberFormat)
                if fmt_devise and fmt_devise != expected_hi:
                    print(f"  ✗ PVL {label} ({devise}): {col_name} formaté {fmt_devise}, attendu {expected_hi}")
                    issues += 1

            # E (PVL) et K (SOLDE) : toujours en devise de la ligne
            for nr_name, col_name in (('PVLpvl', 'E'), ('PVLmontant', 'K')):
                fmt_devise = _extract_devise_from_fmt(
                    ws_pv.getCellByPosition(cr.col(nr_name), r0).NumberFormat)
                if fmt_devise and fmt_devise != devise:
                    print(f"  ✗ PVL {label} ({devise}): {col_name} formaté {fmt_devise}, attendu {devise}")
                    issues += 1

    return issues


# ============================================================================
# CHARTE v3.6 — fonction apply_charter
# ============================================================================
def _parse_nr_content(content):
    """Parse '$Feuille.$C$r1:$C$r2' → (sheet, col_0, row_1indexed_a, row_1indexed_b).
    Retourne None si ce n'est pas un range colonne."""
    if ':' not in content:
        return None
    left, right = content.split(':')
    left = left.lstrip('$')
    parts = left.split('.$')
    if len(parts) != 2:
        return None
    sheet_name = parts[0]
    start_ref = parts[1]
    end_ref = right.lstrip('$')
    # parse refs
    def _parse(ref):
        col = ''
        row = ''
        for ch in ref:
            if ch == '$':
                continue
            if ch.isalpha():
                col += ch
            elif ch.isdigit():
                row += ch
        if not col or not row:
            return None
        c0 = 0
        for ch in col.upper():
            c0 = c0 * 26 + (ord(ch) - ord('A') + 1)
        return (c0 - 1, int(row))
    a = _parse(start_ref)
    b = _parse(end_ref)
    if a is None or b is None or a[0] != b[0]:
        return None
    return (sheet_name, a[0], a[1], b[1])


def _collect_tables_uno(doc):
    """Regroupe les NR colonne par tableau sur (sheet, row_a, row_b).
    Retourne list[dict] : {sheet, min_col, max_col, first_row, last_row, names}.
    """
    from collections import defaultdict
    nr = doc.document.NamedRanges
    groups = defaultdict(lambda: {'cols': set(), 'names': []})
    for i in range(nr.Count):
        name_obj = nr.getByIndex(i)
        name = name_obj.Name
        parsed = _parse_nr_content(name_obj.Content)
        if parsed is None:
            continue
        sheet_name, col_0, row_a, row_b = parsed
        if row_b - row_a < 1:
            continue  # mono-cell (déjà filtré par _parse_nr_content, garde-fou)
        key = (sheet_name, row_a, row_b)
        groups[key]['cols'].add(col_0)
        groups[key]['names'].append(name)
    tables = []
    for (sheet, first_row, last_row), data in groups.items():
        cols = sorted(data['cols'])
        tables.append({
            'sheet': sheet,
            'min_col': cols[0],
            'max_col': cols[-1],
            'first_row': first_row,
            'last_row': last_row,
            'names': sorted(data['names']),
        })
    tables.sort(key=lambda t: (t['sheet'], t['first_row']))
    return tables


def _row_is_used_uno(ws, r_1, min_col, max_col):
    """True si au moins une cellule de la ligne (1-indexed) a une valeur non vide.
    Le fond seul n'est pas suffisant (certaines feuilles préformatent des milliers
    de lignes blanches dans leur NR)."""
    for c0 in range(min_col, max_col + 1):
        if ws.getCellByPosition(c0, r_1 - 1).getString():
            return True
    return False


def _row_any_content_uno(ws, r_1, min_col, max_col):
    """True si au moins une cellule a une valeur non vide."""
    for c0 in range(min_col, max_col + 1):
        if ws.getCellByPosition(c0, r_1 - 1).getString():
            return True
    return False


def _find_head_uno(ws, min_col, max_col, first_data_1):
    """Tête = ligne adjacente (first_data-1) toujours incluse + lignes au-dessus
    tant que contenu."""
    if first_data_1 <= 1:
        return None
    head_top = first_data_1 - 1
    r = head_top - 1
    while r >= 1 and _row_any_content_uno(ws, r, min_col, max_col):
        head_top = r
        r -= 1
    return head_top


def _find_foot_uno(ws, min_col, max_col, last_data_1, max_row=200):
    """Pied = ligne adjacente (last_data+1) toujours incluse + lignes en-dessous
    tant que contenu."""
    first = last_data_1 + 1
    if first > max_row:
        return None
    foot_bot = first
    r = foot_bot + 1
    while r <= max_row and _row_any_content_uno(ws, r, min_col, max_col):
        foot_bot = r
        r += 1
    return foot_bot


def _fixup_applyfill_xlsm(xlsm_path):
    """Post-traitement xl/styles.xml : ajoute applyFill='true' aux <xf> de cellXfs
    qui référencent un fillId sans l'avoir explicite.

    Corrige un effet de bord UNO → openpyxl : UNO pose les fonds via un cellXf
    dont il omet applyFill='true'. Visuellement OK (LO interprète), mais openpyxl
    lit strictement et ignore le fill → faux positifs à l'audit.
    """
    import re
    import shutil
    import zipfile

    xlsm_path = Path(xlsm_path)
    tmp = xlsm_path.with_suffix(xlsm_path.suffix + '.tmp')

    def patch_xf(match):
        tag = match.group(0)
        if 'applyFill' in tag:
            return tag
        if not re.search(r'fillId="\d+"', tag):
            return tag
        return re.sub(r'(\s*/?>)$', r' applyFill="true"\1', tag, count=1)

    def patch_cellxfs(match):
        block = match.group(0)
        return re.sub(r'<xf\b[^>]*?/?>', patch_xf, block)

    patched_any = False
    with zipfile.ZipFile(xlsm_path, 'r') as zin:
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == 'xl/styles.xml':
                    xml = data.decode('utf-8')
                    new_xml = re.sub(
                        r'<cellXfs[^>]*>.*?</cellXfs>',
                        patch_cellxfs, xml, flags=re.DOTALL)
                    if new_xml != xml:
                        patched_any = True
                        data = new_xml.encode('utf-8')
                zout.writestr(item, data)

    shutil.move(str(tmp), str(xlsm_path))
    return patched_any


def apply_charter(doc, apply, sheets=None):
    """Applique la charte graphique v3.6 à tous les tableaux du classeur.

    Palette fonds :
      - tête          → TETE_FILL (#D2C195)
      - col ref       → COL_REF_FILL (#EEEBDB)
      - data          → DATA_FILL (#FFFFFF)
      - pied          → PIED_FILL (#EEEBDB)
      - Exceptions préservées en data : jaune, gamme beige, gris devise étrangère
      - Exception universelle : jaune (jamais écrasé)

    Bordures :
      - grille hair HAIR_COLOR sur data blanche uniquement
      - BORDURE_PIED thick PIED_BORDER_COLOR sur top 1re ligne pied
      - fonds beiges : pas de grille (cosmétiquement invisible)

    Délimitation :
      - tableaux = NR colonne regroupés par (sheet, first_row, last_row)
      - tête/pied : lignes contiguës hors NR avec fond attendu (ou jaune)
      - data bornée au dernier rang effectivement utilisé.
    """
    from com.sun.star.table import BorderLine2

    def _mk(color, width):
        b = BorderLine2()
        b.Color = color
        b.LineWidth = width
        b.LineStyle = 0  # solide
        return b

    HAIR_B = _mk(HAIR_COLOR, HAIR_WIDTH_UNO)
    PIED_B = _mk(PIED_BORDER_COLOR, THICK_WIDTH_UNO)
    EMPTY_B = BorderLine2()

    def _border_eq(a, b):
        return (a.Color == b.Color and a.LineWidth == b.LineWidth
                and a.LineStyle == b.LineStyle)

    def _set_fill(cell, target):
        """Pose le fond si différent. Renvoie 1 si changement, 0 sinon."""
        if cell.CellBackColor == target:
            return 0
        if cell.CellBackColor == JAUNE:
            return 0  # intouchable
        if apply:
            cell.CellBackColor = target
        return 1

    def _set_fill_with_exc(cell, target, exceptions):
        """Pose le fond sauf si la cellule porte une exception tolérée."""
        if cell.CellBackColor in exceptions:
            return 0
        return _set_fill(cell, target)

    def _set_border(cell, side, target):
        cur = getattr(cell, side)
        if _border_eq(cur, target):
            return 0
        if apply:
            setattr(cell, side, target)
        return 1

    tables = _collect_tables_uno(doc)
    total_fixes = 0

    # Filtre feuilles (noms UI → match case-insensitive sur sheet name)
    if sheets:
        wanted = {s.lower() for s in sheets}
        # Mapping UI → noms de feuilles
        aliases = {
            'avoirs': 'avoirs', 'budget': 'budget', 'pvl': 'plus_value',
            'plus_value': 'plus_value', 'controles': 'contrôles',
            'operations': 'opérations', 'opérations': 'opérations',
            'patrimoine': 'patrimoine', 'cotations': 'cotations',
        }
        wanted_sheets = {aliases.get(w, w) for w in wanted}
    else:
        wanted_sheets = None

    for t in tables:
        if wanted_sheets is not None and t['sheet'].lower() not in wanted_sheets:
            continue
        sheet = doc.get_sheet(t['sheet'])
        if sheet is None:
            continue
        min_col = t['min_col']
        max_col = t['max_col']
        first_data_nr = t['first_row']
        last_data_nr = t['last_row']

        # Borner au dernier rang utilisé
        last_data = last_data_nr
        while last_data >= first_data_nr and not _row_is_used_uno(sheet, last_data, min_col, max_col):
            last_data -= 1
        if last_data < first_data_nr:
            last_data = first_data_nr

        head_top = _find_head_uno(sheet, min_col, max_col, first_data_nr)
        foot_bot = _find_foot_uno(sheet, min_col, max_col, last_data_nr)
        foot_first = last_data_nr + 1 if foot_bot else None

        t_fixes = 0

        def _apply_grid(cell, skip_top=False):
            """Pose hair/D2C195 sur les 4 côtés, sauf jaune (liberté user)
            et sauf top si skip_top (1re ligne pied dont top = thick)."""
            if cell.CellBackColor == JAUNE:
                return 0
            n = 0
            for side in ('TopBorder', 'BottomBorder', 'LeftBorder', 'RightBorder'):
                if skip_top and side == 'TopBorder':
                    continue
                n += _set_border(cell, side, HAIR_B)
            return n

        # TÊTE
        if head_top is not None:
            for r in range(head_top, first_data_nr):
                for c0 in range(min_col, max_col + 1):
                    cell = sheet.getCellByPosition(c0, r - 1)
                    t_fixes += _set_fill_with_exc(cell, TETE_FILL, EXC_HEAD)
                    t_fixes += _apply_grid(cell)

        # DATA
        for r in range(first_data_nr, last_data + 1):
            for c0 in range(min_col, max_col + 1):
                cell = sheet.getCellByPosition(c0, r - 1)
                is_ref = (c0 == min_col)
                target = COL_REF_FILL if is_ref else DATA_FILL
                t_fixes += _set_fill_with_exc(cell, target, EXC_DATA)
                t_fixes += _apply_grid(cell)

        # PIED
        if foot_bot is not None:
            for r in range(foot_first, foot_bot + 1):
                is_first = (r == foot_first)
                for c0 in range(min_col, max_col + 1):
                    cell = sheet.getCellByPosition(c0, r - 1)
                    t_fixes += _set_fill_with_exc(cell, PIED_FILL, EXC_FOOT)
                    t_fixes += _apply_grid(cell, skip_top=is_first)
                    if is_first:
                        t_fixes += _set_border(cell, 'TopBorder', PIED_B)

        if t_fixes:
            col_lbl = f"col{min_col+1}→col{max_col+1}"
            names = ', '.join(t['names'][:3])
            if len(t['names']) > 3:
                names += f" +{len(t['names'])-3}"
            print(f"  {t['sheet']} [{col_lbl}] {names} : {t_fixes} fix(es)")
        total_fixes += t_fixes

    return total_fixes


def _read_alarm_sqrefs(xlsm_path):
    """Pré-pass openpyxl : retourne {sheet_name: set((col_letter, row))} pour
    les cellules ciblées par une CF alarme.

    Alarme = dxf dont le fill (fgColor ou bgColor) est FFC7CE (rouge clair) ou
    FFEB9C (jaune-orange). Identifié par scan du styles.xml puis croisement
    avec les `dxfId` des règles CF de chaque feuille.
    """
    import zipfile
    import re as _re
    from openpyxl import load_workbook
    ALARM_FILLS = {'FFC7CE', 'FFEB9C'}

    with zipfile.ZipFile(xlsm_path) as z:
        styles = z.read('xl/styles.xml').decode('utf-8')

    m = _re.search(r'<dxfs[^>]*>(.*?)</dxfs>', styles, _re.DOTALL)
    dxfs_xml = m.group(1) if m else ''
    dxf_blocks = _re.findall(r'<dxf>(.*?)</dxf>', dxfs_xml, _re.DOTALL)
    alarm_dxf_set = set()
    for i, blk in enumerate(dxf_blocks):
        fills = _re.findall(r'(?:fg|bg)Color[^/]*rgb="(?:FF)?([0-9A-Fa-f]{6})"', blk)
        if any(f.upper() in ALARM_FILLS for f in fills):
            alarm_dxf_set.add(i)
    if not alarm_dxf_set:
        return {}

    def _col_to_idx(letter):
        n = 0
        for c in letter:
            n = n * 26 + (ord(c) - 64)
        return n

    def _idx_to_col(n):
        cl = ''
        while n > 0:
            n, r = divmod(n - 1, 26)
            cl = chr(65 + r) + cl
        return cl

    def _parse_sqref(s):
        cells = []
        for part in str(s).split():
            mm = _re.match(r'^([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?$', part)
            if not mm:
                continue
            c1 = _col_to_idx(mm.group(1))
            r1 = int(mm.group(2))
            c2 = _col_to_idx(mm.group(3) or mm.group(1))
            r2 = int(mm.group(4) or mm.group(2))
            for c in range(c1, c2 + 1):
                cl = _idx_to_col(c)
                for r in range(r1, r2 + 1):
                    cells.append((cl, r))
        return cells

    wb = load_workbook(xlsm_path, keep_vba=True)
    out = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        s = set()
        for cfr in ws.conditional_formatting:
            for rule in ws.conditional_formatting[cfr]:
                if rule.dxfId in alarm_dxf_set:
                    s.update(_parse_sqref(cfr.sqref))
        if s:
            out[sheet_name] = s
    return out


def apply_alarm_bold(doc, alarm_sqrefs, apply, sheets=None):
    """Pose CharWeight=BOLD (150) sur les cells de contrôle non-vides hors col drill.

    Cell de contrôle = ciblée par une CF alarme (cf. `_read_alarm_sqrefs`).
    Drill cols exclues = CATmontant (Budget) et CTRL2drill (Contrôles), lus
    via NRs (layout-agnostic).

    Skip si :
      - cell vide (pas de valeur ni formule) — sqref élargi
      - cell dans col drill (gérée par dxf devise séparément)
    """
    BOLD = 150
    import re as _re

    nr = doc.document.NamedRanges
    drill_cols = {}  # {sheet_name: col_0}
    for sheet, nm in [('Budget', 'CATmontant'), ('Contrôles', 'CTRL2drill')]:
        if nr.hasByName(nm):
            content = nr.getByName(nm).Content
            mm = _re.match(r"\$([^.]+)\.\$([A-Z]+)\$", content)
            if mm:
                col_letter = mm.group(2)
                col_0 = 0
                for c in col_letter:
                    col_0 = col_0 * 26 + (ord(c) - 64)
                drill_cols[sheet] = col_0 - 1

    if sheets:
        wanted = {s.lower() for s in sheets}
        aliases = {
            'avoirs': 'avoirs', 'budget': 'budget', 'pvl': 'plus_value',
            'plus_value': 'plus_value', 'controles': 'contrôles',
            'operations': 'opérations', 'opérations': 'opérations',
            'patrimoine': 'patrimoine', 'cotations': 'cotations',
        }
        wanted_sheets = {aliases.get(w, w) for w in wanted}
    else:
        wanted_sheets = None

    def _col_to_idx(letter):
        n = 0
        for c in letter:
            n = n * 26 + (ord(c) - 64)
        return n - 1

    def _idx_to_col(col_0):
        col_1 = col_0 + 1
        letters = ''
        while col_1:
            col_1, rem = divmod(col_1 - 1, 26)
            letters = chr(65 + rem) + letters
        return letters

    # Cells de gras-alarme additionnelles non couvertes par les CF du classeur
    # (cf. inventaire témoin) — ciblées via NR pour rester layout-agnostic.
    additional = {}
    cr_doc = doc.cr if hasattr(doc, 'cr') else None

    def _add(sheet_name, col_0, row_1):
        additional.setdefault(sheet_name, set()).add((_idx_to_col(col_0), row_1))

    # POSTES ligne écart (Budget) : col POSTESmontant, row pos_end + 4
    pos_s, pos_e = doc.cr.rows('POSTESnom') if cr_doc else (None, None)
    if pos_s and pos_e:
        col0 = doc.cr.col('POSTESmontant')
        if col0 is not None and col0 >= 0:
            _add('Budget', col0, pos_e + 4)

    # CAT ligne écart Total euro (Budget) : col CATtotal_euro, row cat_end + 3
    # (pied CAT : +1 Total, +2 Somme opérations, +3 Écart, +4 Total hors C&V)
    cat_s, cat_e = doc.cr.rows('CATnom') if cr_doc else (None, None)
    if cat_s and cat_e:
        col0 = doc.cr.col('CATtotal_euro')
        if col0 is not None and col0 >= 0:
            _add('Budget', col0, cat_e + 3)

    # Fusionner avec alarm_sqrefs avant la boucle principale
    merged_sqrefs = {s: set(cells) for s, cells in alarm_sqrefs.items()}
    for s, extra in additional.items():
        merged_sqrefs.setdefault(s, set()).update(extra)

    total = 0
    for sheet_name, cells in merged_sqrefs.items():
        if wanted_sheets and sheet_name.lower() not in wanted_sheets:
            continue
        try:
            ws = doc.get_sheet(sheet_name)
        except (ValueError, KeyError):
            continue
        drill_col = drill_cols.get(sheet_name)
        n_changed = 0
        for col_letter, r in sorted(cells, key=lambda x: (x[0], x[1])):
            col_0 = _col_to_idx(col_letter)
            if drill_col is not None and col_0 == drill_col:
                continue  # drill col exclue
            cell = ws.getCellByPosition(col_0, r - 1)
            # Skip cellules vides (sqref élargi sans contenu)
            if not cell.getString() and not cell.getFormula().startswith('='):
                continue
            if cell.CharWeight != BOLD:
                if apply:
                    cell.CharWeight = BOLD
                n_changed += 1
        if n_changed:
            print(f"  {sheet_name} : {n_changed} cell(s) contrôle bold")
        total += n_changed
    return total


def fix_formats(xlsm_path, apply=False, sheets=None, charter=False):
    """Scanne et corrige les formats des feuilles.

    Args:
        sheets: liste de noms de feuilles à traiter (None = toutes).
                Noms courts acceptés : avoirs, budget, pvl, ctrl2, operations, controles, patrimoine, cotations.
        charter: si True, applique en plus la charte graphique v3.6 (palette fonds + grille hair + BORDURE_PIED).
    """
    from inc_uno import UnoDocument

    xlsm_path = Path(xlsm_path).resolve()
    if not xlsm_path.exists():
        print(f"❌ Fichier introuvable : {xlsm_path}")
        return False

    if apply:
        bak = xlsm_path.with_suffix('.xlsm.bak')
        shutil.copy2(xlsm_path, bak)
        print(f"Backup : {bak.name}")

    total_fixes = 0

    with UnoDocument(xlsm_path) as doc:
        cr = doc.cr
        doc.cr = cr

        # === Validation structurelle ===
        ok, errors, warnings = validate_structure(doc.document, cr=cr)
        if warnings:
            for w in warnings:
                print(f"  ⚠ {w}")
        if not ok:
            for e in errors:
                print(f"  ❌ {e}")
            print("\nAbandon : structure invalide")
            return False
        print("✓ Structure validée")

        # === Cohérence format ↔ devise ===
        # Audit informatif (ne modifie pas), comptabilisé séparément du total des
        # corrections appliquées par les routines fix_*.
        print("\n🔍 Cohérence format/devise...", flush=True)
        n_coherence = check_format_coherence(doc, cr=cr)
        if not n_coherence:
            print("  ✓ OK")

        # Mapping noms courts → liste de (label, fonction)
        # Une feuille peut avoir plusieurs passes (ex: controles = CTRL2 + generic)
        SHEET_FIXES = {
            'avoirs':     [('📋 Avoirs',              lambda: fix_avoirs(doc, apply, cr))],
            'budget':     [('📋 Budget (catégories)',  lambda: fix_budget(doc, apply)),
                           ('📋 Budget (global)',      lambda: fix_generic_sheet(doc, 'Budget', apply))],
            'pvl':        [('📋 Plus_value',           lambda: fix_plus_value(doc, apply, cr))],
            'controles':  [('📋 Contrôles (CTRL1)',    lambda: fix_ctrl1(doc, apply, cr)),
                           ('📋 Contrôles (CTRL2)',    lambda: fix_ctrl2(doc, apply)),
                           ('📋 Contrôles (global)',   lambda: fix_generic_sheet(doc, 'Contrôles', apply))],
            'operations': [('📋 Opérations',           lambda: fix_operations(doc, apply, cr))],
            'patrimoine': [('📋 Patrimoine',           lambda: fix_generic_sheet(doc, 'Patrimoine', apply))],
            'cotations':  [('📋 Cotations (COTcours2)', lambda: fix_cotations(doc, apply, cr)),
                           ('📋 Cotations (global)',    lambda: fix_generic_sheet(doc, 'Cotations', apply))],
        }

        # Normaliser les noms de feuilles demandés
        if sheets:
            selected = {s.lower().replace('plus_value', 'pvl').replace('opérations', 'operations')
                        for s in sheets}
        else:
            selected = None  # toutes

        for key, passes in SHEET_FIXES.items():
            if selected and key not in selected:
                continue
            for label, fix_fn in passes:
                print(f"\n{label}...", flush=True)
                n = fix_fn()
                total_fixes += n
                if not n:
                    print("  ✓ OK")
                elif n and 'global' in label.lower():
                    print(f"  {n} cellule(s)")

        # === Charte graphique v3.6 ===
        if charter:
            print("\n🎨 Charte graphique v3.6...", flush=True)
            n_charter = apply_charter(doc, apply, sheets=sheets)
            total_fixes += n_charter
            if not n_charter:
                print("  ✓ OK")

            # === Bold direct sur cells contrôle (CF alarme) hors drill ===
            print("\n🔔 Bold cells contrôle (CF alarme)...", flush=True)
            alarm_sqrefs = _read_alarm_sqrefs(xlsm_path)
            n_bold = apply_alarm_bold(doc, alarm_sqrefs, apply, sheets=sheets)
            total_fixes += n_bold
            if not n_bold:
                print("  ✓ OK")

        if apply and total_fixes:
            doc.calculate_all()
            doc.save()

    # Post-traitement xlsm : fix applyFill manquant (UNO → openpyxl)
    if charter and apply:
        if _fixup_applyfill_xlsm(xlsm_path):
            print("✓ styles.xml patché : applyFill='true' ajouté sur xf concernés")

    if total_fixes == 0 and not n_coherence:
        print("\nAucune anomalie détectée")
    elif total_fixes == 0:
        print(f"\n{n_coherence} anomalie(s) signalée(s) — non auto-corrigeable(s)")
    else:
        action = 'corrigée(s)' if apply else 'à corriger'
        msg = f"Total : {total_fixes} anomalie(s) {action}"
        if n_coherence:
            msg += f" + {n_coherence} signalée(s) (non auto-corrigeable(s))"
        print(f"\n{msg}")
    return True


def main():
    parser = argparse.ArgumentParser(
        description='Corrige les formats de cellules dans comptes.xlsm')
    parser.add_argument('xlsm', help='Fichier xlsm à corriger')
    parser.add_argument('--apply', action='store_true',
                        help='Appliquer les corrections (défaut: dry run)')
    parser.add_argument('--sheet', '-s', nargs='+', metavar='FEUILLE',
                        help='Feuille(s) à traiter (avoirs, budget, pvl, controles, operations, '
                             'patrimoine, cotations). Défaut: toutes.')
    parser.add_argument('--charter', action='store_true',
                        help='Applique en plus la charte graphique v3.6 '
                             '(palette fonds de zone + grille hair + BORDURE_PIED)')
    args = parser.parse_args()

    fix_formats(args.xlsm, apply=args.apply, sheets=args.sheet, charter=args.charter)


if __name__ == '__main__':
    main()
