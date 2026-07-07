#!/usr/bin/env python3-uno
"""Migration one-shot : ancrage PVL au PREMIER #Solde (MIN), sans re-ancrage.

Corrige `tool_migrate_pvl_ancrage.py` : la formule d'ancrage (date_init /
date_anter) utilisait `MAXIFS(...;OPcatégorie;Solde;OPequiv_euro;"<>")` pour
re-ancrer sur le *dernier* #Solde VALORISÉ — mais un #Solde ne porte jamais
d'`OPequiv_euro` → condition **morte** → l'ancrage tombait en epoch (dégradé).

Le re-ancrage n'est PAS une logique de la formule PVL : c'est une conséquence
de la **purge** (qui, en condensant l'historique, pose un #Solde valorisé
devenant le premier présent → le MIN l'ancre, H récupère sa valorisation).
→ On remet l'ancrage sur MIN(#Solde), comme la section *portefeuilles*.

Réécrit (chirurgical : SEULE la cellule d'ancrage ; H/I/montant_anter la
référencent et se recalculent) :
- Plus_value (métaux/crypto/devises) : `PVLdate_init` → MINIFS sans equiv
- Avoirs (comptes à devise)          : `AVRdate_anter` → MINIFS sans equiv

Détection par la FORMULE (présence de `equiv` + `MAX...`) → ne touche que les
lignes posées par l'ancienne migration ; les valorisations manuelles (autre
formule / valeur statique) sont laissées intactes.

Usage:
    python3 tool_migrate_pvl_min_ancrage.py ~/Compta/comptes.xlsm [--dry-run]
"""
import argparse
import sys
from pathlib import Path

from inc_uno import UnoDocument, check_lock_file
from inc_excel_schema import SHEET_PLUS_VALUE, SHEET_AVOIRS, ColResolver, uno_row


def _is_old_anchor(formula):
    """Vrai si la formule est l'ancien ancrage MAX + condition equiv (morte)."""
    if not formula:
        return False
    f = str(formula).lower()
    return 'equiv' in f and ('maxifs' in f or 'max.si' in f)


def _inspect(xlsm_path):
    """Sonde openpyxl SANS LibreOffice : cellules d'ancrage encore en MAX+equiv
    (Plus_value.date_init + Avoirs.date_anter). Rend list ou None (introuvable).

    Scanne UNIQUEMENT la colonne d'ancrage, bornée par son named range (délimité
    par les sentinelles ⚓) via ColResolver — pas la grille déclarée (1048576 ×
    16384) qu'openpyxl suivrait aveuglément."""
    import openpyxl
    from inc_excel_schema import ColResolver
    p = Path(xlsm_path).expanduser().resolve()
    if not p.exists():
        return None
    wb = openpyxl.load_workbook(p, data_only=False)
    cr = ColResolver.from_openpyxl(wb)
    hits = []
    for sh, name in (('Plus_value', 'PVLdate_init'), ('Avoirs', 'AVRdate_anter')):
        if sh not in wb.sheetnames:
            continue
        s, e = cr.rows(name)          # bornes du named range (sentinelles)
        if not s:
            continue
        col = cr.col(name)            # 1-indexed
        ws = wb[sh]
        for r in range(s, (e or s) + 1):
            cell = ws.cell(r, col)
            if _is_old_anchor(cell.value):
                hits.append((sh, cell.coordinate))
    wb.close()
    return hits


def _pvl_rows(ws, cr):
    rows = []
    for r in range(1, 400):
        f = ws.getCellByPosition(cr.col('PVLdate_init'), uno_row(r)).getFormula()
        if _is_old_anchor(f):
            rows.append(r)
    return rows


def _avr_rows(ws, cr):
    rows = []
    s, e = cr.rows('AVRintitulé')
    if not s:
        return rows
    for r in range(s, (e or s + 60) + 1):
        f = ws.getCellByPosition(cr.col('AVRdate_anter'), uno_row(r)).getFormula()
        if _is_old_anchor(f):
            rows.append(r)
    return rows


def _snapshot(ws, cr, rows, cols):
    snap = {}
    for r in rows:
        r0 = uno_row(r)
        snap[r] = {c: ws.getCellByPosition(cr.col(c), r0).getValue() for c in cols}
    return snap


def _deltas(before, after, cols, label):
    lines = []
    for r in sorted(before):
        for c in cols:
            b, a = before[r][c], after[r][c]
            if abs(b - a) > 1e-6:
                lines.append(f"    {label} L{r} {c}: {b:.2f} → {a:.2f} (Δ {a-b:+.2f})")
    return lines


PVL_COLS = ('PVLpvl', 'PVLdate_init', 'PVLmontant_init', 'PVLsigma', 'PVLmontant')
AVR_COLS = ('AVRdate_anter', 'AVRmontant_anter')


def migrate(xlsm_path, dry_run=False):
    p = Path(xlsm_path).expanduser().resolve()
    if not p.exists():
        print(f"❌ Fichier introuvable : {p}")
        return 1
    if check_lock_file(p):
        print(f"❌ Fichier verrouillé (LibreOffice ouvert) : {p}")
        return 1

    with UnoDocument(str(p)) as doc:
        xdoc = doc.document
        cr = ColResolver.from_uno(xdoc)
        ws_pv = doc.get_sheet(SHEET_PLUS_VALUE)
        ws_av = doc.get_sheet(SHEET_AVOIRS)

        pv = _pvl_rows(ws_pv, cr)
        av = _avr_rows(ws_av, cr)

        print(f"Plus_value : {len(pv)} ligne(s) d'ancrage MAX+equiv (date_init)")
        for r in pv:
            nom = ws_pv.getCellByPosition(cr.col('PVLcompte'), uno_row(r)).getString()
            print(f"  L{r}: {nom}")
        print(f"Avoirs : {len(av)} ligne(s) d'ancrage MAX+equiv (date_anter)")
        for r in av:
            nom = ws_av.getCellByPosition(cr.col('AVRintitulé'), uno_row(r)).getString()
            print(f"  L{r}: {nom}")

        if not pv and not av:
            print("\nAucune ligne à migrer (déjà en MIN ?).")
            return 0

        # Snapshot AVANT
        doc.calculate_all()
        pv_before = _snapshot(ws_pv, cr, pv, PVL_COLS)
        av_before = _snapshot(ws_av, cr, av, AVR_COLS)

        # Réécriture chirurgicale : seule la cellule d'ancrage (MIN, sans equiv)
        cB = cr.letter('PVLcompte'); cD = cr.letter('PVLdevise')
        for r in pv:
            ws_pv.getCellByPosition(cr.col('PVLdate_init'), uno_row(r)).setFormula(
                f'=MINIFS(OPdate;OPcompte;{cB}{r};OPdevise;{cD}{r};OPcatégorie;Solde)')
        aA = cr.letter('AVRintitulé'); aE = cr.letter('AVRdevise')
        for r in av:
            ws_av.getCellByPosition(cr.col('AVRdate_anter'), uno_row(r)).setFormula(
                f'=MINIFS(OPdate;OPcompte;${aA}{r};OPdevise;${aE}{r};OPcatégorie;Solde)')

        # Snapshot APRÈS
        doc.calculate_all()
        pv_after = _snapshot(ws_pv, cr, pv, PVL_COLS)
        av_after = _snapshot(ws_av, cr, av, AVR_COLS)

        deltas = (_deltas(pv_before, pv_after, PVL_COLS, 'PV')
                  + _deltas(av_before, av_after, AVR_COLS, 'AVR'))
        if deltas:
            print("\n⚠️  Écarts (ancrage MAX-dégradé → MIN) :")
            for line in deltas:
                print(line)
        else:
            print("\n✓ Aucune valeur ne change (ancrage n'alimentait rien).")

        if dry_run:
            print("\n[dry-run] pas de sauvegarde")
            return 0

        doc.save()
        print(f"\n✓ Sauvé : {p}")
    return 0


def main():
    ap = argparse.ArgumentParser(description=__doc__.split('\n')[0])
    ap.add_argument('xlsm', help='Chemin du classeur comptes.xlsm')
    ap.add_argument('--dry-run', action='store_true',
                    help="Sonde openpyxl (SANS LibreOffice) : rc 3 = changerait, "
                         "0 = rien à faire")
    args = ap.parse_args()
    if args.dry_run:
        # Sonde SANS LibreOffice (convention upgrade : rc 3 = travail, 0 = rien).
        hits = _inspect(args.xlsm)
        if hits is None:
            print(f"❌ Fichier introuvable : {args.xlsm}")
            return 1
        if hits:
            print(f"{len(hits)} cellule(s) d'ancrage MAX+equiv à corriger "
                  "(Plus_value.date_init / Avoirs.date_anter).")
            return 3
        print("✓ Rien à faire (ancrage déjà en MIN).")
        return 0
    return migrate(args.xlsm, dry_run=False)


if __name__ == '__main__':
    sys.exit(main())
