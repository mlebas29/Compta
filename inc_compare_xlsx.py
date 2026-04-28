"""
inc_compare_xlsx.py — Comparaison de fichiers Excel (Opérations, Plus_value, Avoirs)

Fonctions partagées : compare_xlsx(), compare_xlsx_with_prev(), SHEETS_CONFIG.
Utilisé par tool_compare_xlsx (CLI), cpt_update, tnr_lib.
"""

import argparse
import re
import sys
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook
from inc_excel_schema import (
    SHEET_OPERATIONS, SHEET_PLUS_VALUE, SHEET_AVOIRS,
    SHEET_BUDGET, SHEET_COTATIONS, SHEET_CONTROLES,
    ColResolver,
)


def _build_sheets_config(cr):
    """Construit la config de comparaison avec les colonnes résolues dynamiquement."""
    return {
    SHEET_OPERATIONS: {
        'skip_rows': 2,
        'ignore_cols': {cr.col('OPréf'), cr.col('OPcatégorie'), cr.col('OPcommentaire')},
        'max_cols': cr.col('OPcommentaire'),  # A-I
        'exclude_re': r'#Solde|#Info|#Balance',
    },
    SHEET_PLUS_VALUE: {
        'skip_rows': 4,
        'ignore_cols': {cr.col('PVLpct'), cr.col('PVLdate'), cr.col('PVLmontant')},
        'max_cols': cr.col('PVLmontant'),  # A-K
        'brutal_ignore_cols': {cr.col('PVLdate')},
        'brutal_tolerance': 0.01,
        'warn_only': True,
        'warn_threshold': 0.10,
        'value_col': cr.col('PVLmontant'),
    },
    SHEET_AVOIRS: {
        'skip_rows': 2,
        'max_cols': 12,              # A-L
        'brutal_ignore_cols': set(),
        'brutal_tolerance': 0.01,
        'warn_only': True,
    },
    SHEET_BUDGET: {
        'skip_rows': 0,
        'max_cols': 30,
        'brutal_ignore_cols': set(),
        'brutal_ignore_cells': {(1, 3), (2, 3)},  # TODAY() et TODAY()-365
        'brutal_tolerance': 0.01,
        'brutal_only': True,
    },
    SHEET_COTATIONS: {
        'skip_rows': 1,
        'max_cols': 8,
        'brutal_ignore_cols': {cr.col('COTdate')},
        'brutal_tolerance': 0.01,
        'brutal_only': True,
    },
    SHEET_CONTROLES: {
        'skip_rows': 1,
        'max_cols': 20,
        'brutal_ignore_cols': set(),
        'brutal_tolerance': 0.01,
        'brutal_only': True,
    },
    'Patrimoine': {
        'skip_rows': 0,
        'max_cols': 8,
        'brutal_ignore_cols': set(),
        'brutal_tolerance': 0.01,
        'brutal_only': True,
    },
    }




def normalize_value(val):
    """Normalise une valeur pour comparaison."""
    # None et chaîne vide sont équivalents
    if val is None or val == '':
        return None

    # Dates : convertir en string ISO
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')

    # Floats : arrondir à 2 décimales
    if isinstance(val, float):
        return round(val, 2)

    # Strings : strip espaces
    if isinstance(val, str):
        return val.strip()

    return val


def format_value(val):
    """Formate une valeur pour affichage."""
    if val is None or val == '':
        return '(vide)'

    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')

    if isinstance(val, float):
        return f'{val:.2f}'

    # Tronquer les chaînes longues
    s = str(val)
    if len(s) > 40:
        return s[:37] + '...'
    return s


def extract_row(ws, row_idx, max_col, ignore_cols):
    """Extrait une ligne comme tuple de valeurs normalisées."""
    values = []
    for col in range(1, max_col + 1):
        if col in ignore_cols:
            continue
        values.append(normalize_value(ws.cell(row=row_idx, column=col).value))
    return tuple(values)


def format_row(ws, row_idx, max_col, ignore_cols):
    """Formate une ligne pour affichage."""
    values = []
    for col in range(1, max_col + 1):
        if col in ignore_cols:
            continue
        values.append(format_value(ws.cell(row=row_idx, column=col).value))
    return ' | '.join(values)


def row_matches_regex(ws, row_idx, max_col, regex_pattern, invert=False):
    """Vérifie si une ligne matche la regex."""
    if regex_pattern is None:
        return True
    # Construire une chaîne avec toutes les valeurs de la ligne
    row_text = ' '.join(str(ws.cell(row=row_idx, column=c).value or '') for c in range(1, max_col + 1))
    match = regex_pattern.search(row_text) is not None
    return not match if invert else match


def compare_values_with_threshold(ws_result, ws_expected, skip_rows, value_col, threshold, max_display=5):
    """Compare les valeurs numériques d'une colonne avec un seuil de variation.

    Retourne liste de warnings pour variations > threshold (ex: 0.10 = 10%)
    """
    warnings = []

    # Construire un index par clé (colonnes A-C = compte, type, titre)
    def get_key(ws, row):
        return tuple(str(ws.cell(row, c).value or '') for c in range(1, 4))

    def get_value(ws, row):
        val = ws.cell(row, value_col).value
        if val is None:
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    # Indexer expected
    expected_values = {}
    for row in range(skip_rows + 1, ws_expected.max_row + 1):
        key = get_key(ws_expected, row)
        val = get_value(ws_expected, row)
        if key and val is not None:
            expected_values[key] = (row, val)

    # Comparer avec result
    variations = []
    for row in range(skip_rows + 1, ws_result.max_row + 1):
        key = get_key(ws_result, row)
        result_val = get_value(ws_result, row)

        if key in expected_values and result_val is not None:
            exp_row, exp_val = expected_values[key]
            if exp_val != 0:
                variation = abs(result_val - exp_val) / abs(exp_val)
                if variation > threshold:
                    variations.append((key, exp_val, result_val, variation))

    if variations:
        variations.sort(key=lambda x: -x[3])  # Trier par variation décroissante
        warnings.append(f"  Variations > {threshold*100:.0f}% ({len(variations)} lignes):")
        for key, exp_val, res_val, var in variations[:max_display]:
            label = ' | '.join(k for k in key if k)[:50]
            warnings.append(f"    {label}: {exp_val:.2f} → {res_val:.2f} ({var*100:+.1f}%)")
        if len(variations) > max_display:
            warnings.append(f"    ... et {len(variations) - max_display} autres")

    return warnings


def compare_sheet_brutal(ws_result, ws_expected, config):
    """Compare deux feuilles cellule par cellule (formules et données).

    Détecte: formule écrasée par constante, valeur différente, formule modifiée.
    Les workbooks doivent être ouverts SANS data_only pour lire les formules.
    """
    skip_rows = config.get('skip_rows', 0)
    max_cols = config.get('max_cols', 11)
    ignore_cols = config.get('brutal_ignore_cols', set())
    ignore_cells = config.get('brutal_ignore_cells', set())
    tolerance = config.get('brutal_tolerance', 0.01)

    diffs = []
    cell_count = 0

    max_rows = max(ws_result.max_row, ws_expected.max_row)

    for row in range(skip_rows + 1, max_rows + 1):
        for col in range(1, max_cols + 1):
            if col in ignore_cols:
                continue
            if (row, col) in ignore_cells:
                continue

            val_r = ws_result.cell(row=row, column=col).value
            val_e = ws_expected.cell(row=row, column=col).value

            cell_count += 1

            # Both None/empty → OK
            if (val_r is None or val_r == '') and (val_e is None or val_e == ''):
                continue

            is_formula_r = isinstance(val_r, str) and val_r.startswith('=')
            is_formula_e = isinstance(val_e, str) and val_e.startswith('=')

            col_letter = chr(64 + col)

            # Formula vs data = erreur (formule écrasée par une constante)
            if is_formula_e and not is_formula_r:
                diffs.append(f'    L{row} col {col_letter}: formule "{val_e}" remplacée par valeur {format_value(val_r)}')
                continue
            if is_formula_r and not is_formula_e:
                diffs.append(f'    L{row} col {col_letter}: valeur {format_value(val_e)} remplacée par formule "{val_r}"')
                continue

            # Two formulas → compare strings
            if is_formula_r and is_formula_e:
                if val_r != val_e:
                    diffs.append(f'    L{row} col {col_letter}: formule "{val_e}" ≠ "{val_r}"')
                continue

            # Two data → compare values
            nr = normalize_value(val_r)
            ne = normalize_value(val_e)

            if nr == ne:
                continue

            # Float tolerance
            if isinstance(nr, (int, float)) and isinstance(ne, (int, float)):
                if abs(nr - ne) <= tolerance:
                    continue

            diffs.append(f'    L{row} col {col_letter}: {format_value(val_e)} ≠ {format_value(val_r)}')

    return diffs, cell_count


def compare_sheets(ws_result, ws_expected, sheet_name, skip_rows=0, ignore_cols=None, max_display=10,
                   regex_pattern=None, invert=False, max_cols=None, exclude_re=None, since_date=None, cr=None,
                   approx_tolerance=None, labels=None):
    """Compare deux feuilles par contenu de ligne."""
    if labels is None:
        labels = ('RESULT', 'EXPECTED')
    if ignore_cols is None:
        ignore_cols = set()

    diffs = []

    def row_date_ok(ws, row):
        """Vérifie si la date de la ligne est >= since_date."""
        if since_date is None:
            return True
        cell_date = ws.cell(row, cr.col('OPdate') if cr else 1).value
        if cell_date is None:
            return True  # Pas de date = inclure
        if isinstance(cell_date, datetime):
            return cell_date >= since_date
        if isinstance(cell_date, date):
            return datetime.combine(cell_date, datetime.min.time()) >= since_date
        return True  # Type inconnu = inclure

    # Dimensions
    result_rows = ws_result.max_row
    expected_rows = ws_expected.max_row
    # Limiter les colonnes si max_cols est défini
    detected_max_col = max(ws_result.max_column, ws_expected.max_column)
    max_col = min(detected_max_col, max_cols) if max_cols else detected_max_col

    # Extraire les lignes comme ensembles (avec comptage pour doublons)
    result_lines = {}
    for row in range(skip_rows + 1, result_rows + 1):
        if not row_date_ok(ws_result, row):
            continue
        line = extract_row(ws_result, row, max_col, ignore_cols)
        if line not in result_lines:
            result_lines[line] = []
        result_lines[line].append(row)

    expected_lines = {}
    for row in range(skip_rows + 1, expected_rows + 1):
        if not row_date_ok(ws_expected, row):
            continue
        line = extract_row(ws_expected, row, max_col, ignore_cols)
        if line not in expected_lines:
            expected_lines[line] = []
        expected_lines[line].append(row)

    # Comparer les ensembles
    only_in_result = []
    only_in_expected = []

    for line, rows in result_lines.items():
        expected_count = len(expected_lines.get(line, []))
        result_count = len(rows)
        if result_count > expected_count:
            for row in rows[expected_count:]:
                only_in_result.append((row, line))

    for line, rows in expected_lines.items():
        result_count = len(result_lines.get(line, []))
        expected_count = len(rows)
        if expected_count > result_count:
            for row in rows[result_count:]:
                only_in_expected.append((row, line))

    # Trier par numéro de ligne
    only_in_result.sort(key=lambda x: x[0])
    only_in_expected.sort(key=lambda x: x[0])

    # Matching approché sur Equiv (col E) si demandé
    approx_count = 0
    if approx_tolerance is not None:
        equiv_col = 5  # E=Equiv
        if equiv_col not in ignore_cols:
            # Position de la col Equiv dans le tuple (colonnes avant E non ignorées)
            equiv_pos = sum(1 for c in range(1, equiv_col) if c not in ignore_cols)

            matched_r = set()
            matched_e = set()
            for ri, (r_row, r_line) in enumerate(only_in_result):
                for ei, (e_row, e_line) in enumerate(only_in_expected):
                    if ei in matched_e:
                        continue
                    # Tout identique sauf Equiv ?
                    r_key = r_line[:equiv_pos] + r_line[equiv_pos+1:]
                    e_key = e_line[:equiv_pos] + e_line[equiv_pos+1:]
                    if r_key != e_key:
                        continue
                    # Comparer Equiv avec tolérance
                    r_val = r_line[equiv_pos]
                    e_val = e_line[equiv_pos]
                    if r_val is not None and e_val is not None:
                        try:
                            r_f = float(r_val) if not isinstance(r_val, (int, float)) else r_val
                            e_f = float(e_val) if not isinstance(e_val, (int, float)) else e_val
                            if e_f != 0 and abs(r_f - e_f) / abs(e_f) <= approx_tolerance:
                                matched_r.add(ri)
                                matched_e.add(ei)
                                break
                        except (ValueError, TypeError):
                            pass

            approx_count = len(matched_r)
            only_in_result = [x for i, x in enumerate(only_in_result) if i not in matched_r]
            only_in_expected = [x for i, x in enumerate(only_in_expected) if i not in matched_e]

    # Construire le rapport
    if ignore_cols:
        col_names = {6: 'Réf', 7: 'Catégorie', 10: 'Commentaire'}
        col_labels = [col_names.get(c, chr(64+c)) for c in sorted(ignore_cols)]
        ignored_str = f" (sans {', '.join(col_labels)})"
    else:
        ignored_str = ""

    # Filtrer par regex si fournie
    if regex_pattern is not None:
        only_in_result = [(row, line) for row, line in only_in_result
                          if row_matches_regex(ws_result, row, max_col, regex_pattern, invert)]
        only_in_expected = [(row, line) for row, line in only_in_expected
                            if row_matches_regex(ws_expected, row, max_col, regex_pattern, invert)]

    # Exclure les méta-opérations (config exclude_re)
    if exclude_re is not None:
        exclude_pattern = re.compile(exclude_re)
        only_in_result = [(row, line) for row, line in only_in_result
                          if not row_matches_regex(ws_result, row, max_col, exclude_pattern, False)]
        only_in_expected = [(row, line) for row, line in only_in_expected
                            if not row_matches_regex(ws_expected, row, max_col, exclude_pattern, False)]

    # max_display=0 signifie pas de limite
    display_limit = max_display if max_display > 0 else len(only_in_result) + len(only_in_expected)

    if only_in_result:
        diffs.append(f"  Uniquement dans {labels[0]} ({len(only_in_result)} lignes){ignored_str}:")
        for row, line in only_in_result[:display_limit]:
            formatted = format_row(ws_result, row, max_col, ignore_cols)
            diffs.append(f"    L{row}: {formatted}")
        if max_display > 0 and len(only_in_result) > max_display:
            diffs.append(f"    ... et {len(only_in_result) - max_display} autres")

    if only_in_expected:
        diffs.append(f"  Uniquement dans {labels[1]} ({len(only_in_expected)} lignes){ignored_str}:")
        for row, line in only_in_expected[:display_limit]:
            formatted = format_row(ws_expected, row, max_col, ignore_cols)
            diffs.append(f"    L{row}: {formatted}")
        if max_display > 0 and len(only_in_expected) > max_display:
            diffs.append(f"    ... et {len(only_in_expected) - max_display} autres")

    return diffs, approx_count


def compare_tuples(ws_result, ws_expected, skip_rows, max_cols, ignore_cols, max_display=10,
                   approx_tolerance=None, labels=None, cr=None):
    """Compare les groupes d'appariement (opérations partageant la même Ref).

    Vérifie que les mêmes opérations sont regroupées ensemble,
    indépendamment des numéros de Ref (qui varient entre runs).
    """
    if labels is None:
        labels = ('RESULT', 'EXPECTED')
    REF_COL = cr.col('OPréf') if cr else 6
    EQUIV_COL = cr.col('OPequiv_euro') if cr else 5
    tuple_ignore = ignore_cols | {REF_COL}
    # Position de Equiv dans le tuple (colonnes avant E non ignorées)
    equiv_pos = sum(1 for c in range(1, EQUIV_COL) if c not in tuple_ignore)

    def extract_groups(ws):
        """Extrait les groupes ref → [lignes normalisées]."""
        groups = {}  # ref → [(row_idx, row_tuple), ...]
        for row in range(skip_rows + 1, ws.max_row + 1):
            ref = ws.cell(row, REF_COL).value
            if not ref or ref == '-' or ref == 'Réf.':
                continue
            row_tuple = extract_row(ws, row, max_cols, tuple_ignore)
            if ref not in groups:
                groups[ref] = []
            groups[ref].append((row, row_tuple))
        return groups

    def group_signature(rows_tuples):
        """Signature d'un groupe = frozenset trié des lignes (indépendant de l'ordre)."""
        return frozenset(rt for _, rt in rows_tuples)

    def row_without_equiv(row_tuple):
        """Retourne le tuple sans la colonne Equiv."""
        return row_tuple[:equiv_pos] + row_tuple[equiv_pos + 1:]

    def groups_match_approx(r_rows, e_rows, tolerance):
        """Vérifie si deux groupes matchent à l'Equiv près."""
        if len(r_rows) != len(e_rows):
            return False
        # Tenter un matching 1:1 par clé sans Equiv
        r_by_key = {}
        for _, rt in r_rows:
            key = row_without_equiv(rt)
            if key not in r_by_key:
                r_by_key[key] = []
            r_by_key[key].append(rt)
        e_by_key = {}
        for _, rt in e_rows:
            key = row_without_equiv(rt)
            if key not in e_by_key:
                e_by_key[key] = []
            e_by_key[key].append(rt)
        if r_by_key.keys() != e_by_key.keys():
            return False
        for key in r_by_key:
            if len(r_by_key[key]) != len(e_by_key[key]):
                return False
            # Vérifier Equiv avec tolérance pour chaque paire
            for r_rt, e_rt in zip(sorted(r_by_key[key]), sorted(e_by_key[key])):
                r_val = r_rt[equiv_pos]
                e_val = e_rt[equiv_pos]
                if r_val == e_val:
                    continue
                if r_val is None or e_val is None:
                    return False
                try:
                    r_f = float(r_val) if not isinstance(r_val, (int, float)) else r_val
                    e_f = float(e_val) if not isinstance(e_val, (int, float)) else e_val
                    if e_f == 0 or abs(r_f - e_f) / abs(e_f) > tolerance:
                        return False
                except (ValueError, TypeError):
                    return False
        return True

    def format_group(ws, rows_tuples):
        """Formate un groupe pour affichage."""
        lines = []
        for row_idx, _ in sorted(rows_tuples):
            lines.append(f"      L{row_idx}: {format_row(ws, row_idx, max_cols, tuple_ignore)}")
        return lines

    groups_r = extract_groups(ws_result)
    groups_e = extract_groups(ws_expected)

    # Construire les multisets de signatures
    sigs_r = {}  # signature → [(ref, rows_tuples), ...]
    for ref, rows_tuples in groups_r.items():
        if len(rows_tuples) < 2:
            continue
        sig = group_signature(rows_tuples)
        if sig not in sigs_r:
            sigs_r[sig] = []
        sigs_r[sig].append((ref, rows_tuples))

    sigs_e = {}
    for ref, rows_tuples in groups_e.items():
        if len(rows_tuples) < 2:
            continue
        sig = group_signature(rows_tuples)
        if sig not in sigs_e:
            sigs_e[sig] = []
        sigs_e[sig].append((ref, rows_tuples))

    # Comparer : retirer les signatures communes (en respectant les multiplicités)
    only_r = []  # groupes uniquement dans RESULT
    only_e = []  # groupes uniquement dans EXPECTED
    all_sigs = set(sigs_r.keys()) | set(sigs_e.keys())
    matched_count = 0

    for sig in all_sigs:
        r_groups = sigs_r.get(sig, [])
        e_groups = sigs_e.get(sig, [])
        matched_count += min(len(r_groups), len(e_groups))
        if len(r_groups) > len(e_groups):
            for ref, rows in r_groups[len(e_groups):]:
                only_r.append((ref, rows))
        elif len(e_groups) > len(r_groups):
            for ref, rows in e_groups[len(r_groups):]:
                only_e.append((ref, rows))

    # Matching approché sur Equiv si demandé
    approx_count = 0
    if approx_tolerance is not None and EQUIV_COL not in tuple_ignore and only_r and only_e:
        matched_ri = set()
        matched_ei = set()
        for ri, (r_ref, r_rows) in enumerate(only_r):
            for ei, (e_ref, e_rows) in enumerate(only_e):
                if ei in matched_ei:
                    continue
                if groups_match_approx(r_rows, e_rows, approx_tolerance):
                    matched_ri.add(ri)
                    matched_ei.add(ei)
                    approx_count += 1
                    break
        only_r = [x for i, x in enumerate(only_r) if i not in matched_ri]
        only_e = [x for i, x in enumerate(only_e) if i not in matched_ei]

    diffs = []
    display_limit = max_display if max_display > 0 else len(only_r) + len(only_e)

    if only_r:
        diffs.append(f"  Tuples uniquement dans {labels[0]} ({len(only_r)} groupes):")
        for ref, rows_tuples in only_r[:display_limit]:
            diffs.append(f"    [{ref}] ({len(rows_tuples)} lignes):")
            diffs.extend(format_group(ws_result, rows_tuples))
        if max_display > 0 and len(only_r) > max_display:
            diffs.append(f"    ... et {len(only_r) - max_display} autres")

    if only_e:
        diffs.append(f"  Tuples uniquement dans {labels[1]} ({len(only_e)} groupes):")
        for ref, rows_tuples in only_e[:display_limit]:
            diffs.append(f"    [{ref}] ({len(rows_tuples)} lignes):")
            diffs.extend(format_group(ws_expected, rows_tuples))
        if max_display > 0 and len(only_e) > max_display:
            diffs.append(f"    ... et {len(only_e) - max_display} autres")

    return diffs, matched_count, approx_count


def compare_xlsx(result_path, expected_path, max_display=10, sheet_filter=None,
                 regex_pattern=None, invert=False, override_ignore_cols=None, since_date=None,
                 approx_tolerance=None, compare_tuples_flag=False, brutal=False,
                 prev_mode=False, warn_threshold_override=None, labels=None,
                 skip_sheets=None):
    """Compare deux fichiers Excel.

    prev_mode=True : mode PROD (informatif, pas d'échec).
    warn_threshold_override : surcharge le seuil warn_threshold pour Plus_value.
    labels : tuple (label_result, label_expected) pour l'affichage contextuel.
             Défaut : ('ACTUEL', 'PRÉCÉDENT') si prev_mode, sinon noms de fichiers.
    """
    if labels is None:
        if prev_mode:
            labels = ('ACTUEL', 'PRÉCÉDENT')
        else:
            labels = (Path(result_path).name, Path(expected_path).name)
    pad = max(len(labels[0]), len(labels[1]))
    print(f"🔍 Comparaison{' (vs archive)' if prev_mode else ''}:")
    print(f"   {labels[0]:{pad}}: {result_path}")
    print(f"   {labels[1]:{pad}}: {expected_path}")
    if sheet_filter:
        print(f"   FEUILLE:  {sheet_filter}")
    if regex_pattern:
        print(f"   FILTRE:   /{regex_pattern.pattern}/ {'(inversé)' if invert else ''}")
    if since_date:
        print(f"   DEPUIS:   {since_date.strftime('%Y-%m-%d')}")
    print()

    # Vérifier existence
    if not Path(result_path).exists():
        print(f"❌ Fichier introuvable: {result_path}")
        return False

    if not Path(expected_path).exists():
        print(f"❌ Fichier introuvable: {expected_path}")
        return False

    # Charger les workbooks
    wb_result = load_workbook(result_path, data_only=True)
    wb_expected = load_workbook(expected_path, data_only=True)

    # Workbooks sans data_only pour comparaison brutale (formules lues comme strings)
    wb_result_formulas = None
    wb_expected_formulas = None
    if brutal:
        wb_result_formulas = load_workbook(result_path, data_only=False)
        wb_expected_formulas = load_workbook(expected_path, data_only=False)

    all_ok = True
    sheets_processed = 0

    cr = ColResolver.from_openpyxl(wb_result)
    sheets_config = _build_sheets_config(cr)

    for sheet_name, config in sheets_config.items():
        # Filtrer par feuille si demandé
        if sheet_filter and sheet_name != sheet_filter:
            continue
        if skip_sheets and sheet_name in skip_sheets:
            print(f"⏭  {sheet_name} (ignoré)")
            continue

        if sheet_name not in wb_result.sheetnames:
            print(f"❌ Feuille '{sheet_name}' absente du fichier result")
            all_ok = False
            continue

        if sheet_name not in wb_expected.sheetnames:
            print(f"❌ Feuille '{sheet_name}' absente du fichier expected")
            all_ok = False
            continue

        # Feuilles brutal_only : ignorées si --brutal n'est pas activé
        if config.get('brutal_only') and not brutal:
            continue

        sheets_processed += 1

        # Comparaison brutale cellule par cellule (mode TNR)
        if brutal and 'brutal_ignore_cols' in config:
            brutal_diffs, brutal_cells = compare_sheet_brutal(
                wb_result_formulas[sheet_name],
                wb_expected_formulas[sheet_name],
                config
            )
            if brutal_diffs:
                print(f"❌ {sheet_name} (brutal): {len(brutal_diffs)} différence(s)")
                for d in brutal_diffs:
                    print(d)
                all_ok = False
            else:
                print(f"✓ {sheet_name} (brutal): {brutal_cells} cellules, 0 différence")

            # brutal_only : pas de comparaison supplémentaire
            if config.get('brutal_only'):
                continue

        # Mode warn_only : avertissement sans échec (feuilles dynamiques)
        if config.get('warn_only'):
            threshold = warn_threshold_override if warn_threshold_override is not None else config.get('warn_threshold', 0.10)
            value_col = config.get('value_col', 8)
            warnings = compare_values_with_threshold(
                wb_result[sheet_name],
                wb_expected[sheet_name],
                skip_rows=config.get('skip_rows', 0),
                value_col=value_col,
                threshold=threshold,
                max_display=max_display
            )
            if warnings:
                print(f"⚠️  {sheet_name} (variations):")
                for w in warnings:
                    print(w)
            elif not prev_mode:
                print(f"✓ {sheet_name}: variations < {threshold*100:.0f}%")
            continue  # Pas d'échec pour warn_only

        # Colonnes ignorées : -x remplace les défauts, sinon config
        if override_ignore_cols is not None:
            ignore_cols = override_ignore_cols
        else:
            ignore_cols = config.get('ignore_cols', set()).copy()

        diffs = compare_sheets(
            wb_result[sheet_name],
            wb_expected[sheet_name],
            sheet_name,
            skip_rows=config.get('skip_rows', 0),
            ignore_cols=ignore_cols,
            max_display=max_display,
            regex_pattern=regex_pattern,
            invert=invert,
            max_cols=config.get('max_cols'),
            exclude_re=config.get('exclude_re'),
            since_date=since_date,
            approx_tolerance=approx_tolerance,
            labels=labels,
            cr=cr
        )

        diffs, approx_count = diffs

        if prev_mode:
            # Mode PROD : informatif, pas d'échec
            if diffs:
                for diff in diffs:
                    if labels[1] in diff and 'lignes)' in diff:
                        print(f"⚠️  {sheet_name}: {diff.strip()}")
                    elif labels[0] in diff and 'lignes)' in diff:
                        print(f"ℹ️  {sheet_name}: {diff.strip()}")
                    else:
                        print(diff)
            else:
                print(f"✓ {sheet_name}: identique")
        elif diffs:
            print(f"❌ {sheet_name}:")
            for diff in diffs:
                print(diff)
            if approx_count > 0:
                print(f"  ≈ {approx_count} ligne(s) identiques à l'Equiv près (tolérance {approx_tolerance*100:.0f}%)")
            all_ok = False
        elif approx_count > 0:
            print(f"✓ {sheet_name}: identique (≈ {approx_count} ligne(s) à l'Equiv près, tolérance {approx_tolerance*100:.0f}%)")
        else:
            print(f"✓ {sheet_name}: identique")

        # Comparaison des tuples d'appariement si demandé (pas en prev_mode)
        if compare_tuples_flag and not prev_mode and sheet_name == SHEET_OPERATIONS:
            tuple_diffs, tuple_matched, tuple_approx = compare_tuples(
                wb_result[sheet_name], wb_expected[sheet_name],
                skip_rows=config.get('skip_rows', 0),
                max_cols=config.get('max_cols', 10),
                ignore_cols=ignore_cols,
                max_display=max_display,
                approx_tolerance=approx_tolerance,
                labels=labels,
                cr=cr
            )
            approx_msg = f", ≈ {tuple_approx} à l'Equiv près" if tuple_approx > 0 else ""
            if tuple_diffs:
                print(f"❌ Appariements ({tuple_matched} identiques{approx_msg}):")
                for diff in tuple_diffs:
                    print(diff)
                all_ok = False
            else:
                print(f"✓ Appariements: {tuple_matched} tuples identiques{approx_msg}")

    if sheet_filter and sheets_processed == 0:
        print(f"⚠ Feuille '{sheet_filter}' introuvable dans la config.")
        print(f"  Feuilles disponibles : {', '.join(sheets_config.keys())}")
        all_ok = False

    wb_result.close()
    wb_expected.close()
    if wb_result_formulas:
        wb_result_formulas.close()
    if wb_expected_formulas:
        wb_expected_formulas.close()

    return all_ok


def find_prev_archive(archives_dir, prev=1):
    """Trouve le Nème backup le plus récent de comptes.xlsm dans archives/.

    prev=1 : le plus récent (= backup juste avant le dernier run).
    Retourne le Path ou None si pas assez d'archives.
    """
    archives_dir = Path(archives_dir)
    backups = sorted(
        list(archives_dir.glob('comptes_HDS_*.xlsx')) + list(archives_dir.glob('comptes_HDS_*.xlsm')),
        reverse=True)
    if prev <= 0 or prev > len(backups):
        return None
    return backups[prev - 1]


def compare_xlsx_with_prev(comptes_file, archives_dir, prev=1, warn_threshold=None):
    """Compare comptes.xlsm avec son archive précédente (mode PROD).

    Purement informatif — retour sans impact sur le pipeline.
    warn_threshold : surcharge le seuil Plus_value (None = défaut config.ini ou 10%).
    """
    archive = find_prev_archive(archives_dir, prev)
    if archive is None:
        print(f"\nℹ️  Pas d'archive précédente (prev={prev}), comparaison ignorée")
        return

    # Lire le seuil depuis config.ini si pas de surcharge
    threshold = warn_threshold
    if threshold is None:
        threshold = _read_config_threshold()

    print()
    compare_xlsx(
        str(comptes_file), str(archive),
        prev_mode=True,
        warn_threshold_override=threshold,
        labels=('ACTUEL', 'PRÉCÉDENT')
    )


def _read_config_threshold():
    """Lit le seuil warn_threshold depuis config.ini [comparison]."""
    import configparser
    config_path = Path(__file__).parent / 'config.ini'
    if not config_path.exists():
        return None
    config = configparser.ConfigParser()
    config.read(config_path)
    try:
        pct = config.getfloat('comparison', 'warn_threshold')
        return pct / 100.0  # config en %, interne en ratio
    except (configparser.NoSectionError, configparser.NoOptionError, ValueError):
        return None

